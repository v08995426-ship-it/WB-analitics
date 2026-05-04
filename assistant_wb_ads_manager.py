
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import io
import json
import math
import os
import re
import time
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import boto3
import numpy as np
import pandas as pd
import requests
from botocore.client import Config as BotoConfig
from botocore.exceptions import ClientError
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

STORE_NAME = "TOPFACE"
TARGET_SUBJECTS = {"кисти косметические", "блески", "помады", "косметические карандаши"}
GROWTH_SUBJECTS = {"блески", "помады", "косметические карандаши"}
CATEGORY_DRR_LIMITS = {
    "кисти косметические": 0.14,
    "косметические карандаши": 0.14,
    "помады": 0.16,
    "блески": 0.16,
}
SUBJECT_DISPLAY_NAMES = {
    "кисти косметические": "Кисти косметические",
    "косметические карандаши": "Косметические карандаши",
    "помады": "Помады",
    "блески": "Блески",
}

SUBJECT_FIXED_BUYOUT_RATES = {
    "кисти косметические": 0.85,
    "косметические карандаши": 0.95,
    "помады": 0.93,
    "блески": 0.90,
}
FUNNEL_SALES_CANDIDATES = [
    "ordersSumRub",
    "ordersSum",
    "ordersSumRur",
    "ordersAmountRub",
    "ordersAmount",
    "sumOrdersRub",
    "sumOrders",
    "sum_orders_rub",
    "sum_orders",
    "salesRub",
    "salesSumRub",
    "salesSum",
    "salesAmount",
    "revenueRub",
    "revenue",
    "Продажи, ₽",
    "Продажи",
    "Сумма продаж, ₽",
    "Сумма продаж",
    "Сумма заказов, ₽",
    "Сумма заказов",
    "Заказано на сумму, ₽",
    "Заказано на сумму",
]

FUNNEL_BUYOUT_CANDIDATES = [
    "buyoutPercent",
    "buyoutRate",
    "buyout_rate",
    "Процент выкупа",
    "% выкупа",
    "Выкуп, %",
]
ECON_BUYOUT_CANDIDATES = [
    "Процент выкупа",
    "% выкупа",
    "buyoutPercent",
    "buyout_rate",
]
ECON_GP_UNIT_CANDIDATES = [
    "Валовая прибыль, руб/ед",
    "Валовая прибыль на 1 товар, ₽",
    "Валовая прибыль на 1 товар",
    "ВП на 1 товар, ₽",
    "ВП на 1 товар",
    "gp_unit",
]
ECON_NP_UNIT_CANDIDATES = [
    "Чистая прибыль, руб/ед",
    "Чистая прибыль на 1 товар, ₽",
    "Чистая прибыль на 1 товар",
    "ЧП на 1 товар, ₽",
    "ЧП на 1 товар",
    "np_unit",
]
ECON_WEEK_CANDIDATES = [
    "Неделя",
    "week",
    "Дата",
    "date",
    "Период",
]

def get_category_drr_limit(subject_norm: Any) -> float:
    return CATEGORY_DRR_LIMITS.get(canonical_subject(subject_norm), 0.15)

def get_subject_display_name(subject_norm: Any) -> str:
    norm = canonical_subject(subject_norm)
    return SUBJECT_DISPLAY_NAMES.get(norm, str(subject_norm or "").strip() or norm)


def get_subject_buyout_rate(subject_norm: Any, default: float = 0.90) -> float:
    return SUBJECT_FIXED_BUYOUT_RATES.get(canonical_subject(subject_norm), default)

def get_bid_step_rub(payment_type: Any) -> float:
    return 1.0 if canonical_payment_type(payment_type) == 'cpc' else 6.0

def apply_bid_step(current_bid: float, payment_type: Any, direction: str, floor_bid: float, max_bid: float = 0.0) -> float:
    current_bid = safe_float(current_bid)
    step = get_bid_step_rub(payment_type)
    if direction == 'up':
        target = current_bid + step
        if max_bid > 0:
            target = min(target, max_bid)
        return round(max(target, floor_bid), 2)
    return round(max(current_bid - step, floor_bid), 2)

def is_drop_explained_by_demand(order_growth_pct: float, demand_growth_pct: float, tolerance_pp: float = 5.0) -> bool:
    order_growth_pct = safe_float(order_growth_pct)
    demand_growth_pct = safe_float(demand_growth_pct)
    if order_growth_pct >= 0 or demand_growth_pct >= 0:
        return False
    return demand_growth_pct <= order_growth_pct + tolerance_pp


def build_category_window_diagnostics(
    orders: pd.DataFrame,
    ads_daily: pd.DataFrame,
    funnel: pd.DataFrame,
    keywords: pd.DataFrame,
    master: pd.DataFrame,
    econ_latest: pd.DataFrame,
    window: Dict[str, date],
) -> pd.DataFrame:
    subjects = pd.DataFrame({'subject_norm': sorted(TARGET_SUBJECTS)})
    key_map = master[['nmId', 'subject_norm']].drop_duplicates() if not master.empty else pd.DataFrame(columns=['nmId', 'subject_norm'])
    key_map = key_map.copy()
    if 'nmId' in key_map.columns:
        key_map['nmId'] = pd.to_numeric(key_map['nmId'], errors='coerce')
        key_map = key_map.dropna(subset=['nmId']).copy()
    nm_to_subject: Dict[Any, Any] = {}
    if not key_map.empty:
        km = key_map.dropna(subset=['subject_norm']).copy()
        km['subject_norm'] = km['subject_norm'].map(canonical_subject)
        nm_to_subject = dict(zip(km['nmId'].tolist(), km['subject_norm'].tolist()))

    def _orders_slice(start: date, end: date, suffix: str) -> pd.DataFrame:
        if orders.empty:
            return pd.DataFrame(columns=['subject_norm', f'category_orders_{suffix}', f'category_gp_before_ads_{suffix}'])
        ords = orders[(orders['date'] >= start) & (orders['date'] <= end) & (~orders['isCancel'])].copy()
        if ords.empty:
            return pd.DataFrame(columns=['subject_norm', f'category_orders_{suffix}', f'category_gp_before_ads_{suffix}'])
        gp_map = econ_latest[['nmId', 'gp_realized']].drop_duplicates() if not econ_latest.empty else pd.DataFrame(columns=['nmId', 'gp_realized'])
        if not key_map.empty:
            ords = ords.merge(key_map, on='nmId', how='left', suffixes=('', '_m'))
        ords = with_resolved_subject_norm(ords, nm_to_subject)
        ords = ords.merge(gp_map, on='nmId', how='left')
        ords['gp_realized'] = pd.to_numeric(ords.get('gp_realized'), errors='coerce').fillna(0.0)
        ords = ords[ords['subject_norm'].isin(TARGET_SUBJECTS)].copy()
        if ords.empty:
            return pd.DataFrame(columns=['subject_norm', f'category_orders_{suffix}', f'category_gp_before_ads_{suffix}'])
        return ords.groupby('subject_norm', as_index=False).agg(
            **{
                f'category_orders_{suffix}': ('nmId', 'count'),
                f'category_gp_before_ads_{suffix}': ('gp_realized', 'sum'),
            }
        )

    def _spend_slice(start: date, end: date, suffix: str) -> pd.DataFrame:
        if ads_daily.empty:
            return pd.DataFrame(columns=['subject_norm', f'category_spend_{suffix}'])
        ad = ads_daily[(ads_daily['date'] >= start) & (ads_daily['date'] <= end)].copy()
        if ad.empty:
            return pd.DataFrame(columns=['subject_norm', f'category_spend_{suffix}'])
        if not key_map.empty:
            ad = ad.merge(key_map, on='nmId', how='left', suffixes=('', '_m'))
        ad = with_resolved_subject_norm(ad, nm_to_subject)
        ad['Расход'] = pd.to_numeric(ad.get('Расход'), errors='coerce').fillna(0.0)
        ad = ad[ad['subject_norm'].isin(TARGET_SUBJECTS)].copy()
        if ad.empty:
            return pd.DataFrame(columns=['subject_norm', f'category_spend_{suffix}'])
        return ad.groupby('subject_norm', as_index=False).agg(**{f'category_spend_{suffix}': ('Расход', 'sum')})

    def _demand_slice(start: date, end: date, suffix: str) -> pd.DataFrame:
        if keywords.empty:
            return pd.DataFrame(columns=['subject_norm', f'category_demand_{suffix}'])
        kw = keywords[(keywords['date'] >= start) & (keywords['date'] <= end)].copy()
        if kw.empty:
            return pd.DataFrame(columns=['subject_norm', f'category_demand_{suffix}'])
        kw = with_resolved_subject_norm(kw, nm_to_subject)
        kw['demand_week'] = pd.to_numeric(kw.get('demand_week'), errors='coerce').fillna(0.0)
        kw = kw[kw['subject_norm'].isin(TARGET_SUBJECTS)].copy()
        if kw.empty:
            return pd.DataFrame(columns=['subject_norm', f'category_demand_{suffix}'])
        return kw.groupby('subject_norm', as_index=False).agg(**{f'category_demand_{suffix}': ('demand_week', 'sum')})

    def _realized_sales_slice(start: date, end: date) -> pd.DataFrame:
        if funnel.empty:
            return pd.DataFrame(columns=['subject_norm', 'category_funnel_sales_cur', 'category_realized_sales_cur'])
        fw = funnel[(funnel['date'] >= start) & (funnel['date'] <= end)].copy()
        if fw.empty:
            return pd.DataFrame(columns=['subject_norm', 'category_funnel_sales_cur', 'category_realized_sales_cur'])
        sales_col = find_matching_column(fw, FUNNEL_SALES_CANDIDATES)
        if not key_map.empty:
            fw = fw.merge(key_map, on='nmId', how='left', suffixes=('', '_m'))
        fw = with_resolved_subject_norm(fw, nm_to_subject)
        fw['funnel_sales'] = pd.to_numeric(fw.get(sales_col if sales_col else 'ordersSumRub', 0), errors='coerce').fillna(0.0)
        fw = fw[fw['subject_norm'].isin(TARGET_SUBJECTS)].copy()
        if fw.empty:
            return pd.DataFrame(columns=['subject_norm', 'category_funnel_sales_cur', 'category_realized_sales_cur'])
        fw['buyout_rate'] = fw['subject_norm'].map(get_subject_buyout_rate)
        fw['realized_sales'] = fw['funnel_sales'] * fw['buyout_rate']
        return fw.groupby('subject_norm', as_index=False).agg(
            category_funnel_sales_cur=('funnel_sales', 'sum'),
            category_realized_sales_cur=('realized_sales', 'sum'),
        )

    out = subjects
    for df in [
        _orders_slice(window['cur_start'], window['cur_end'], 'cur'),
        _orders_slice(window['base_start'], window['base_end'], 'base'),
        _spend_slice(window['cur_start'], window['cur_end'], 'cur'),
        _spend_slice(window['base_start'], window['base_end'], 'base'),
        _demand_slice(window['cur_start'], window['cur_end'], 'cur'),
        _demand_slice(window['base_start'], window['base_end'], 'base'),
        _realized_sales_slice(window['cur_start'], window['cur_end']),
    ]:
        out = out.merge(df, on='subject_norm', how='left')
    for col in [
        'category_orders_cur', 'category_gp_before_ads_cur', 'category_orders_base', 'category_gp_before_ads_base',
        'category_spend_cur', 'category_spend_base', 'category_demand_cur', 'category_demand_base',
        'category_funnel_sales_cur', 'category_realized_sales_cur'
    ]:
        if col not in out.columns:
            out[col] = 0.0
        out[col] = pd.to_numeric(out[col], errors='coerce').fillna(0.0)
    out['category_gp_cur'] = out['category_gp_before_ads_cur'] - out['category_spend_cur']
    out['category_gp_base'] = out['category_gp_before_ads_base'] - out['category_spend_base']
    out['category_drr_cur'] = np.where(out['category_realized_sales_cur'] > 0, out['category_spend_cur'] / out['category_realized_sales_cur'], 0.0)
    out['category_orders_growth_pct'] = np.where(out['category_orders_base'] > 0, (out['category_orders_cur'] / out['category_orders_base'] - 1.0) * 100.0, np.where(out['category_orders_cur'] > 0, 100.0, 0.0))
    out['category_gp_growth_pct'] = np.where(out['category_gp_base'] != 0, (out['category_gp_cur'] - out['category_gp_base']) / np.abs(out['category_gp_base']) * 100.0, np.where(out['category_gp_cur'] > 0, 100.0, 0.0))
    out['category_demand_growth_pct'] = np.where(out['category_demand_base'] > 0, (out['category_demand_cur'] / out['category_demand_base'] - 1.0) * 100.0, np.where(out['category_demand_cur'] > 0, 100.0, 0.0))
    out['category_limit_drr'] = out['subject_norm'].map(get_category_drr_limit)
    out['Категория'] = out['subject_norm'].map(get_subject_display_name)
    out['Фиксированный % выкупа'] = out['subject_norm'].map(lambda x: round(get_subject_buyout_rate(x) * 100.0, 1))
    return out

def _normalize_name_for_match(value: Any) -> str:
    return re.sub(r"[^0-9a-zа-я]+", "", str(value or "").strip().lower())

def find_matching_column(df: pd.DataFrame, candidates: List[str]) -> str:
    if df is None or df.empty:
        return ""
    cols = list(df.columns)
    exact = {str(c): str(c) for c in cols}
    for cand in candidates:
        if cand in exact:
            return exact[cand]
    normalized = {_normalize_name_for_match(c): str(c) for c in cols}
    for cand in candidates:
        key = _normalize_name_for_match(cand)
        if key in normalized:
            return normalized[key]
    for cand in candidates:
        key = _normalize_name_for_match(cand)
        if not key:
            continue
        for norm_col, original_col in normalized.items():
            if key in norm_col or norm_col in key:
                return original_col
    return ""


def to_buyout_rate(series: pd.Series, default: float = 0.0) -> pd.Series:
    s = pd.to_numeric(series, errors="coerce")
    s = np.where(pd.Series(s).fillna(0) > 1, pd.Series(s).fillna(0) / 100.0, pd.Series(s).fillna(0))
    return pd.Series(s).fillna(default).clip(lower=0.0, upper=1.0)

def resolve_buyout_rate_from_funnel(df: pd.DataFrame, default: float = 0.85) -> pd.Series:
    if df is None or df.empty:
        return pd.Series(dtype=float)
    rate_col = find_matching_column(df, FUNNEL_BUYOUT_CANDIDATES)
    if rate_col:
        rate = to_buyout_rate(df[rate_col], default=np.nan)
    else:
        rate = pd.Series(np.nan, index=df.index, dtype=float)
    sales_col = find_matching_column(df, FUNNEL_SALES_CANDIDATES)
    if sales_col and "buyoutsSumRub" in df.columns:
        sales = pd.to_numeric(df[sales_col], errors="coerce").fillna(0.0)
        buyout_sum = pd.to_numeric(df["buyoutsSumRub"], errors="coerce").fillna(0.0)
        ratio_sum = np.where(sales > 0, buyout_sum / sales, np.nan)
        rate = rate.where(rate.fillna(0) > 0, pd.Series(ratio_sum, index=df.index))
    if "ordersCount" in df.columns and "buyoutsCount" in df.columns:
        orders_cnt = pd.to_numeric(df["ordersCount"], errors="coerce").fillna(0.0)
        buyouts_cnt = pd.to_numeric(df["buyoutsCount"], errors="coerce").fillna(0.0)
        ratio_cnt = np.where(orders_cnt > 0, buyouts_cnt / orders_cnt, np.nan)
        rate = rate.where(rate.fillna(0) > 0, pd.Series(ratio_cnt, index=df.index))
    return pd.to_numeric(rate, errors="coerce").fillna(default).clip(lower=0.0, upper=1.0)

def latest_econ_rows(econ: pd.DataFrame, columns: List[str]) -> pd.DataFrame:
    if econ is None or econ.empty:
        return pd.DataFrame(columns=columns)
    cols = [c for c in columns if c in econ.columns]
    if "nmId" not in cols:
        cols = ["nmId"] + cols
    work = econ.copy()
    week_col = find_matching_column(work, ECON_WEEK_CANDIDATES)
    if week_col:
        raw = work[week_col].fillna("").astype(str).str.strip()
        parsed = pd.Series(pd.NaT, index=work.index, dtype="datetime64[ns]")

        mask_iso_week = raw.str.fullmatch(r"\d{4}-W\d{2}")
        if mask_iso_week.any():
            parsed.loc[mask_iso_week] = pd.to_datetime(
                raw.loc[mask_iso_week] + "-1",
                format="%G-W%V-%u",
                errors="coerce",
            )

        mask_iso_date = raw.str.fullmatch(r"\d{4}-\d{2}-\d{2}")
        if mask_iso_date.any():
            parsed.loc[mask_iso_date] = pd.to_datetime(
                raw.loc[mask_iso_date],
                format="%Y-%m-%d",
                errors="coerce",
            )

        mask_dot_date = raw.str.fullmatch(r"\d{2}\.\d{2}\.\d{4}")
        if mask_dot_date.any():
            parsed.loc[mask_dot_date] = pd.to_datetime(
                raw.loc[mask_dot_date],
                format="%d.%m.%Y",
                errors="coerce",
            )

        mask_slash_date = raw.str.fullmatch(r"\d{2}/\d{2}/\d{4}")
        if mask_slash_date.any():
            parsed.loc[mask_slash_date] = pd.to_datetime(
                raw.loc[mask_slash_date],
                format="%d/%m/%Y",
                errors="coerce",
            )

        other_mask = parsed.isna() & raw.ne("")
        if other_mask.any():
            parsed.loc[other_mask] = pd.to_datetime(raw.loc[other_mask], errors="coerce")

        work["_econ_order"] = parsed.fillna(pd.Timestamp("1900-01-01"))
    else:
        work["_econ_order"] = np.arange(len(work))
    cols = [c for c in cols if c in work.columns]
    cols.append("_econ_order")
    return work[cols].sort_values("_econ_order").drop_duplicates("nmId", keep="last").drop(columns=["_econ_order"], errors="ignore")

def round_output_value(value: Any) -> Any:
    if isinstance(value, (pd.Timestamp, datetime, date)):
        return value
    if isinstance(value, (bool, np.bool_)):
        return bool(value)
    if isinstance(value, (np.integer, int)):
        return int(value)
    if isinstance(value, (np.floating, float)):
        if pd.isna(value):
            return np.nan
        abs_v = abs(float(value))
        if abs_v == 0:
            return 0
        if abs_v < 1:
            return round(float(value), 2)
        return int(round(float(value)))
    return value

def normalize_output_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return df
    out = df.copy()
    for col in out.columns:
        if pd.api.types.is_numeric_dtype(out[col]):
            out[col] = out[col].map(round_output_value)
    return out

def trim_to_columns(df: pd.DataFrame, columns: List[str]) -> pd.DataFrame:
    if df is None or not isinstance(df, pd.DataFrame):
        return pd.DataFrame(columns=columns)
    out = df.copy()
    for c in columns:
        if c not in out.columns:
            out[c] = "" if "Дата" in c or "Предмет" in c or "Артикул" in c or "Тип" in c or "Плейсмент" in c or "Комментарий" in c or "Причина" in c else 0
    return out[columns]

def is_active_campaign_status(value: Any) -> bool:
    text = str(value or "").strip().lower()
    return text in {"активна", "active", "running", "started", "в работе"}


def explain_limit_reason(row: pd.Series) -> str:
    gp_realized = safe_float(row.get("gp_realized"))
    local_clicks = safe_float(row.get("Клики"))
    local_orders = safe_float(row.get("Заказы"))
    inherited_clicks = safe_float(row.get("item_clicks_cur", row.get("control_ad_clicks", 0.0)))
    inherited_orders = safe_float(row.get("total_orders"))
    parts: List[str] = []
    if gp_realized <= 0:
        parts.append("нет положительной unit-экономики после выкупа")
    if local_clicks >= 50 and local_orders >= 3:
        parts.append(f"лимит рассчитан по факту: клики={local_clicks:.0f}, заказы РК={local_orders:.0f}")
    elif inherited_clicks >= 50 and inherited_orders >= 5:
        parts.append(f"лимит рассчитан по товару: клики={inherited_clicks:.0f}, все заказы={inherited_orders:.0f}")
    else:
        parts.append(
            "недостаточно данных для CPO: "
            f"по кампании клики={local_clicks:.0f} (<50) или заказы РК={local_orders:.0f} (<3); "
            f"по товару клики={inherited_clicks:.0f} (<50) или все заказы={inherited_orders:.0f} (<5)"
        )
    return "; ".join(parts)


def with_resolved_subject_norm(df: pd.DataFrame, nm_to_subject: Dict[Any, Any]) -> pd.DataFrame:
    out = df.copy()
    if out.empty:
        if "subject_norm" not in out.columns:
            out["subject_norm"] = pd.Series(dtype="object")
        return out

    resolved = pd.Series([""] * len(out), index=out.index, dtype="object")

    for col in ["subject_norm", "subject_norm_x", "subject_norm_y", "subject", "Предмет", "Название предмета"]:
        if col in out.columns:
            normalized = out[col].fillna("").astype(str).map(canonical_subject)
            resolved = resolved.where(resolved.astype(str).str.strip() != "", normalized)

    if "nmId" in out.columns and nm_to_subject:
        nm_series = pd.to_numeric(out["nmId"], errors="coerce")
        mapped = nm_series.map(nm_to_subject).fillna("").astype(str).map(canonical_subject)
        resolved = resolved.where(resolved.astype(str).str.strip() != "", mapped)

    out["subject_norm"] = resolved.fillna("").astype(str).map(canonical_subject)
    return out


def build_nm_to_subject_map(master: pd.DataFrame) -> Dict[Any, Any]:
    if master is None or master.empty:
        return {}
    work = master.copy()
    if 'nmId' not in work.columns:
        return {}
    work = with_resolved_subject_norm(work, {})
    nm_series = pd.to_numeric(work['nmId'], errors='coerce')
    subj_series = work.get('subject_norm', pd.Series('', index=work.index)).fillna('').astype(str).map(canonical_subject)
    mask = nm_series.notna() & subj_series.ne('')
    if not mask.any():
        return {}
    dedup = pd.DataFrame({'nmId': nm_series[mask].astype('int64'), 'subject_norm': subj_series[mask]})
    dedup = dedup.drop_duplicates(subset=['nmId'], keep='first')
    return dict(zip(dedup['nmId'], dedup['subject_norm']))
WB_BIDS_URL = "https://advert-api.wildberries.ru/api/advert/v1/bids"
WB_BIDS_MIN_URL = "https://advert-api.wildberries.ru/api/advert/v1/bids/min"
WB_NMS_URL = "https://advert-api.wildberries.ru/adv/v0/auction/nms"

ADS_ANALYSIS_KEY = f"Отчёты/Реклама/{STORE_NAME}/Анализ рекламы.xlsx"
ECONOMICS_KEY = f"Отчёты/Финансовые показатели/{STORE_NAME}/Экономика.xlsx"
FUNNEL_KEY = f"Отчёты/Воронка продаж/{STORE_NAME}/Воронка продаж.xlsx"
ORDERS_WEEKLY_PREFIX = f"Отчёты/Заказы/{STORE_NAME}/Недельные/"
KEYWORDS_WEEKLY_PREFIX = f"Отчёты/Поисковые запросы/{STORE_NAME}/Недельные/"
ABC_PREFIX = "Отчёты/ABC/"
DYNAMICS_PREFIX = "Отчёты/ABC/"

SERVICE_ROOT = f"Служебные файлы/Ассистент WB/{STORE_NAME}/"
OUT_PREVIEW = SERVICE_ROOT + "Предпросмотр_последнего_запуска.xlsx"
OUT_SUMMARY = SERVICE_ROOT + "Сводка_последнего_запуска.json"
OUT_ARCHIVE = SERVICE_ROOT + "Архив_решений.xlsx"
OUT_BID_HISTORY = SERVICE_ROOT + "История_ставок.xlsx"
OUT_LIMITS = SERVICE_ROOT + "Лимиты_ставок_ежедневно.xlsx"
OUT_PRODUCT = SERVICE_ROOT + "Метрики_по_товарам.xlsx"
OUT_EFF = SERVICE_ROOT + "Эффективность_ставки_ежедневно.xlsx"
OUT_WEAK = SERVICE_ROOT + "Слабые_позиции_приоритет.xlsx"
OUT_EFFECTS = SERVICE_ROOT + "Эффект_изменений.xlsx"
# Блок работы с оттенками отключён: каждый оттенок должен вестись отдельной рекламной кампанией.
OUT_SHADE_ACTIONS = SERVICE_ROOT + "Рекомендации_по_оттенкам.xlsx"
OUT_SHADE_PORTFOLIO = SERVICE_ROOT + "Состав_кампаний_по_оттенкам.xlsx"
OUT_SHADE_TESTS = SERVICE_ROOT + "Тесты_оттенков.xlsx"
OUT_BENCHMARK = SERVICE_ROOT + "Сравнение_с_сильными_РК.xlsx"

# Единый итоговый файл. Все отчёты пишем только сюда.
OUT_SINGLE_REPORT = SERVICE_ROOT + "Итог_последнего_запуска.xlsx"

MIN_RATING_SHADE = 4.6
MATURE_START_OFFSET = 7
MATURE_END_OFFSET = 3
WINDOW_LEN = 5

API_CALL_LOGS: List[Dict[str, Any]] = []
MIN_BID_ROWS: List[Dict[str, Any]] = []
_LAST_API_CALL_AT: Dict[str, float] = {}
_API_MIN_INTERVAL_SEC = {
    WB_BIDS_MIN_URL: 3.1,   # 20 req/min, interval 3 sec
    WB_NMS_URL: 1.05,       # 1 req/sec
    WB_BIDS_URL: 0.25,      # 5 req/sec
}

def now_ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def json_dumps_safe(value: Any) -> str:
    try:
        return json.dumps(value, ensure_ascii=False, default=str)
    except Exception:
        return str(value)

def truncate_text(value: Any, limit: int = 4000) -> str:
    text_value = value if isinstance(value, str) else json_dumps_safe(value)
    return text_value[:limit]

def canonical_payment_type(value: Any) -> str:
    v = str(value or "").strip().lower()
    return "cpc" if v == "cpc" else "cpm"

def normalize_internal_placement(value: Any) -> str:
    v = str(value or "").strip().lower()
    mapping = {
        "combined": "combined",
        "search": "search",
        "recommendation": "recommendation",
        "recommendations": "recommendation",
    }
    return mapping.get(v, "search")

def placement_for_min_endpoint(value: Any) -> str:
    v = normalize_internal_placement(value)
    return "recommendation" if v == "recommendation" else v

def placement_for_bids_endpoint(value: Any) -> str:
    v = normalize_internal_placement(value)
    return "recommendations" if v == "recommendation" else v

def wait_for_rate_limit(url: str) -> None:
    delay = _API_MIN_INTERVAL_SEC.get(url, 0.0)
    if delay <= 0:
        return
    last = _LAST_API_CALL_AT.get(url, 0.0)
    now = time.time()
    sleep_for = delay - (now - last)
    if sleep_for > 0:
        time.sleep(sleep_for)

def extract_request_id(response_text: str) -> str:
    if not response_text:
        return ""
    try:
        data = json.loads(response_text)
        return str(data.get("requestId") or data.get("request_id") or "")
    except Exception:
        return ""

def append_api_log(
    *,
    method_name: str,
    http_method: str,
    url: str,
    request_body: Any,
    response_status: Any = "",
    response_text: Any = "",
    status: str = "",
    context: Optional[Dict[str, Any]] = None,
) -> None:
    row: Dict[str, Any] = {
        "timestamp": now_ts(),
        "Метод": method_name,
        "HTTP метод": http_method.upper(),
        "URL": url,
        "status": status,
        "http_status": response_status,
        "request_id": extract_request_id(str(response_text)),
        "request_body": truncate_text(request_body, 8000),
        "response": truncate_text(response_text, 8000),
    }
    if context:
        for k, v in context.items():
            row[k] = v
    API_CALL_LOGS.append(row)

def wb_api_request(
    http_method: str,
    url: str,
    api_key: str,
    body: Any,
    *,
    method_name: str,
    timeout: int = 120,
    dry_run: bool = False,
    context: Optional[Dict[str, Any]] = None,
) -> Optional[requests.Response]:
    if not api_key:
        append_api_log(
            method_name=method_name,
            http_method=http_method,
            url=url,
            request_body=body,
            response_status="",
            response_text="Нет WB_PROMO_KEY_TOPFACE, вызов не выполнен",
            status="skipped",
            context=context,
        )
        return None
    if dry_run:
        append_api_log(
            method_name=method_name,
            http_method=http_method,
            url=url,
            request_body=body,
            response_status="",
            response_text="dry-run",
            status="dry-run",
            context=context,
        )
        return None

    wait_for_rate_limit(url)
    headers = {"Authorization": api_key.strip(), "Content-Type": "application/json"}
    try:
        resp = requests.request(http_method.upper(), url, headers=headers, json=body, timeout=timeout)
        _LAST_API_CALL_AT[url] = time.time()
        append_api_log(
            method_name=method_name,
            http_method=http_method,
            url=url,
            request_body=body,
            response_status=resp.status_code,
            response_text=resp.text,
            status="ok" if resp.status_code == 200 else "failed",
            context=context,
        )
        return resp
    except Exception as e:
        _LAST_API_CALL_AT[url] = time.time()
        append_api_log(
            method_name=method_name,
            http_method=http_method,
            url=url,
            request_body=body,
            response_status="",
            response_text=str(e),
            status="failed",
            context=context,
        )
        return None


def log(msg: str) -> None:
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}", flush=True)

def safe_float(v: Any, default: float = 0.0) -> float:
    try:
        if pd.isna(v):
            return default
        if isinstance(v, str):
            v = v.replace("\xa0", " ").replace("%", "").replace(",", ".").strip()
            if not v:
                return default
        return float(v)
    except Exception:
        return default



def ensure_business_keys(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or not isinstance(df, pd.DataFrame):
        return df
    out = df.copy()
    if 'nmId' not in out.columns:
        for c in ['Артикул WB', 'nm_id', 'nmID']:
            if c in out.columns:
                out['nmId'] = out[c]
                break
    if 'supplier_article' not in out.columns:
        for c in ['Артикул продавца', 'supplier_article_x', 'supplier_article_y', 'supplierArticle', 'supplierArticle_x', 'supplierArticle_y', 'control_key']:
            if c in out.columns:
                out['supplier_article'] = out[c]
                break
    if 'subject' not in out.columns:
        for c in ['Предмет', 'subject_norm']:
            if c in out.columns:
                out['subject'] = out[c]
                break
    if 'Артикул WB' not in out.columns and 'nmId' in out.columns:
        out['Артикул WB'] = out['nmId']
    if 'Артикул продавца' not in out.columns and 'supplier_article' in out.columns:
        out['Артикул продавца'] = out['supplier_article']
    if 'Предмет' not in out.columns and 'subject' in out.columns:
        out['Предмет'] = out['subject']
    return out



def normalize_core_columns(df: pd.DataFrame) -> pd.DataFrame:
    return ensure_business_keys(df)

def safe_int(v: Any, default: int = 0) -> int:
    try:
        if pd.isna(v):
            return default
        if isinstance(v, str):
            v = v.replace("\xa0", " ").replace(",", ".").strip()
        return int(float(v))
    except Exception:
        return default


def series_or_default(df: pd.DataFrame, column: str, default: Any = 0.0) -> pd.Series:
    if isinstance(df, pd.DataFrame) and column in df.columns:
        return df[column]
    if isinstance(default, pd.Series):
        try:
            return default.reindex(df.index)
        except Exception:
            return default
    if isinstance(df, pd.DataFrame):
        return pd.Series(default, index=df.index)
    return pd.Series([default])


def numeric_series(df: pd.DataFrame, column: str, default: float = 0.0) -> pd.Series:
    return pd.to_numeric(series_or_default(df, column, default), errors="coerce").fillna(default)

def canonical_subject(v: Any) -> str:
    return str(v or "").strip().lower()

def product_root_from_supplier_article(v: Any) -> str:
    s = str(v or "").strip()
    if not s or s.lower() in {"nan", "none"}:
        return ""
    root = s.split("/")[0].strip()
    root = re.sub(r"[^0-9A-Za-zА-Яа-я_-]+", "", root)
    root = re.sub(r"[_-]+$", "", root)
    return root.upper()

def pct(a: float, b: float) -> float:
    return (safe_float(a) / safe_float(b) * 100.0) if safe_float(b) else 0.0

def growth_pct(cur: float, base: float) -> float:
    cur = safe_float(cur)
    base = safe_float(base)
    if base <= 0:
        return 100.0 if cur > 0 else 0.0
    return (cur / base - 1.0) * 100.0

def clamp(x: float, low: float, high: float) -> float:
    return max(low, min(high, x))

def daterange(start: date, end: date) -> Iterable[date]:
    cur = start
    while cur <= end:
        yield cur
        cur += timedelta(days=1)

def sanitize_sheet_name(name: str, used: Optional[set] = None) -> str:
    name = re.sub(r'[:\\/?*\[\]]', '_', str(name))
    name = re.sub(r'\s+', '_', name)
    name = re.sub(r'_+', '_', name).strip('_')
    name = name[:31] if len(name) > 31 else name
    if used is None:
        return name or "Лист"
    base = name or "Лист"
    candidate = base
    i = 2
    while candidate in used:
        suffix = f"_{i}"
        candidate = (base[:31-len(suffix)] + suffix) if len(base)+len(suffix) > 31 else base + suffix
        i += 1
    used.add(candidate)
    return candidate


def style_workbook(path: Path) -> None:
    try:
        wb = load_workbook(path)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            max_widths: Dict[int, int] = {}
            for row_idx, row in enumerate(ws.iter_rows(), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    if row_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                    else:
                        if isinstance(cell.value, (datetime, date)):
                            cell.number_format = "yyyy-mm-dd"
                        elif isinstance(cell.value, (int, float, np.integer, np.floating)) and not isinstance(cell.value, bool):
                            value = float(cell.value)
                            if value == 0:
                                cell.number_format = "#,##0"
                            elif abs(value) < 1:
                                cell.number_format = "0.00"
                            else:
                                cell.number_format = "#,##0"
                    val = "" if cell.value is None else str(cell.value)
                    width = min(max(len(val) + 2, 10), 42)
                    max_widths[col_idx] = max(max_widths.get(col_idx, 0), width)
            for col_idx, width in max_widths.items():
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            ws.row_dimensions[1].height = 34
        wb.save(path)
    except Exception:
        pass

class BaseProvider:
    def read_excel(self, key: str, sheet_name: Any = 0) -> pd.DataFrame:
        raise NotImplementedError
    def read_excel_all_sheets(self, key: str) -> Dict[str, pd.DataFrame]:
        raise NotImplementedError
    def write_excel(self, key: str, sheets: Dict[str, pd.DataFrame]) -> None:
        raise NotImplementedError
    def read_text(self, key: str) -> str:
        raise NotImplementedError
    def write_text(self, key: str, text: str) -> None:
        raise NotImplementedError
    def file_exists(self, key: str) -> bool:
        raise NotImplementedError
    def list_keys(self, prefix: str) -> List[str]:
        raise NotImplementedError

class S3Provider(BaseProvider):
    def __init__(self, access_key: str, secret_key: str, bucket_name: str):
        self.bucket = bucket_name
        self.s3 = boto3.client(
            "s3",
            endpoint_url="https://storage.yandexcloud.net",
            aws_access_key_id=access_key,
            aws_secret_access_key=secret_key,
            region_name="ru-central1",
            config=BotoConfig(signature_version="s3v4", read_timeout=300, connect_timeout=60, retries={"max_attempts": 5}),
        )
    def read_bytes(self, key: str) -> bytes:
        obj = self.s3.get_object(Bucket=self.bucket, Key=key)
        return obj["Body"].read()
    def read_excel(self, key: str, sheet_name: Any = 0) -> pd.DataFrame:
        return pd.read_excel(io.BytesIO(self.read_bytes(key)), sheet_name=sheet_name)
    def read_excel_all_sheets(self, key: str) -> Dict[str, pd.DataFrame]:
        data = self.read_bytes(key)
        xls = pd.ExcelFile(io.BytesIO(data))
        return {sh: pd.read_excel(io.BytesIO(data), sheet_name=sh) for sh in xls.sheet_names}
    def write_excel(self, key: str, sheets: Dict[str, pd.DataFrame]) -> None:
        tmp = Path("/tmp") / f"{int(time.time()*1000)}.xlsx"
        with pd.ExcelWriter(tmp, engine="openpyxl") as writer:
            for sh, df in sheets.items():
                (df if df is not None else pd.DataFrame()).to_excel(writer, sheet_name=sanitize_sheet_name(sh), index=False)
        style_workbook(tmp)
        self.s3.put_object(Bucket=self.bucket, Key=key, Body=tmp.read_bytes())
        tmp.unlink(missing_ok=True)
    def read_text(self, key: str) -> str:
        return self.read_bytes(key).decode("utf-8")
    def write_text(self, key: str, text: str) -> None:
        self.s3.put_object(Bucket=self.bucket, Key=key, Body=text.encode("utf-8"))
    def file_exists(self, key: str) -> bool:
        try:
            self.s3.head_object(Bucket=self.bucket, Key=key)
            return True
        except ClientError:
            return False
    def list_keys(self, prefix: str) -> List[str]:
        out: List[str] = []
        token = None
        while True:
            kwargs = {"Bucket": self.bucket, "Prefix": prefix, "MaxKeys": 1000}
            if token:
                kwargs["ContinuationToken"] = token
            resp = self.s3.list_objects_v2(**kwargs)
            out.extend([x["Key"] for x in resp.get("Contents", [])])
            if not resp.get("IsTruncated"):
                break
            token = resp.get("NextContinuationToken")
        return out

class LocalProvider(BaseProvider):
    def __init__(self, base_dir: str):
        self.base_dir = Path(base_dir)
    def _search(self, patterns: List[str]) -> List[Path]:
        out = []
        for child in self.base_dir.iterdir():
            if child.is_file():
                for p in patterns:
                    if re.search(p, child.name, flags=re.I):
                        out.append(child)
                        break
        return sorted(out)
    def _resolve(self, key: str) -> Path:
        p = Path(key)
        if p.exists():
            return p
        mappings = [
            (ADS_ANALYSIS_KEY, [r"^Анализ рекламы.*\.xlsx$"]),
            (ECONOMICS_KEY, [r"^Экономика.*\.xlsx$"]),
            (FUNNEL_KEY, [r"^Воронка продаж.*\.xlsx$"]),
            (OUT_BID_HISTORY, [r"^История_ставок.*\.xlsx$", r"^bid_history.*\.xlsx$"]),
            (OUT_PREVIEW, [r"^Предпросмотр_последнего_запуска.*\.xlsx$", r"^preview_last_run.*\.xlsx$"]),
            (OUT_SINGLE_REPORT, [r"^Итог_последнего_запуска.*\.xlsx$", r"^Предпросмотр_последнего_запуска.*\.xlsx$", r"^preview_last_run.*\.xlsx$"]),
            (OUT_SUMMARY, [r"^Сводка_последнего_запуска.*\.json$", r"^last_run_summary.*\.json$"]),
            (OUT_ARCHIVE, [r"^Архив_решений.*\.xlsx$", r"^decision_archive.*\.xlsx$"]),
        ]
        for logical, pats in mappings:
            if key == logical:
                found = self._search(pats)
                if found:
                    return found[0]
        return self.base_dir / Path(key).name
    def read_excel(self, key: str, sheet_name: Any = 0) -> pd.DataFrame:
        return pd.read_excel(self._resolve(key), sheet_name=sheet_name)
    def read_excel_all_sheets(self, key: str) -> Dict[str, pd.DataFrame]:
        path = self._resolve(key)
        xls = pd.ExcelFile(path)
        return {sh: pd.read_excel(path, sheet_name=sh) for sh in xls.sheet_names}
    def write_excel(self, key: str, sheets: Dict[str, pd.DataFrame]) -> None:
        path = self._resolve(key)
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for sh, df in sheets.items():
                (df if df is not None else pd.DataFrame()).to_excel(writer, sheet_name=sanitize_sheet_name(sh), index=False)
        style_workbook(path)
    def read_text(self, key: str) -> str:
        return self._resolve(key).read_text(encoding="utf-8")
    def write_text(self, key: str, text: str) -> None:
        self._resolve(key).write_text(text, encoding="utf-8")
    def file_exists(self, key: str) -> bool:
        return self._resolve(key).exists()
    def list_keys(self, prefix: str) -> List[str]:
        if prefix == ORDERS_WEEKLY_PREFIX:
            return [str(p) for p in self._search([r"^Заказы_\d{4}-W\d{2}.*\.xlsx$"])]
        if prefix == KEYWORDS_WEEKLY_PREFIX:
            return [str(p) for p in self._search([r"^Неделя .*\.xlsx$", r"^W\d+.*\.xlsx$"])]
        if prefix == ABC_PREFIX:
            return [str(p) for p in self._search([r"^wb_abc_report_goods__.*\.xlsx$", r"^wb_dynamics__.*\.xlsx$"])]
        if prefix == DYNAMICS_PREFIX:
            return [str(p) for p in self._search([r"^wb_dynamics__.*\.xlsx$"])]
        return []

@dataclass
class Config:
    comfort_drr_min: float = 0.10
    comfort_drr_max: float = 0.12
    max_drr: float = 0.15
    max_up_step: float = 0.08
    test_up_step: float = 0.05
    down_step: float = 0.08

def compute_analysis_window(as_of_date: date) -> Dict[str, date]:
    cur_end = as_of_date - timedelta(days=MATURE_END_OFFSET)
    cur_start = cur_end - timedelta(days=WINDOW_LEN-1)
    base_end = cur_start - timedelta(days=1)
    base_start = base_end - timedelta(days=WINDOW_LEN-1)
    return {"cur_start": cur_start, "cur_end": cur_end, "base_start": base_start, "base_end": base_end}

def parse_date_col(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce").dt.date

def choose_provider(local_data_dir: str = "") -> BaseProvider:
    if local_data_dir:
        return LocalProvider(local_data_dir)
    access = os.getenv("YC_ACCESS_KEY_ID", "")
    secret = os.getenv("YC_SECRET_ACCESS_KEY", "")
    bucket = os.getenv("YC_BUCKET_NAME", "")
    if not (access and secret and bucket):
        raise RuntimeError("Не заданы YC_ACCESS_KEY_ID / YC_SECRET_ACCESS_KEY / YC_BUCKET_NAME")
    return S3Provider(access, secret, bucket)


def load_ads(provider: BaseProvider) -> Tuple[pd.DataFrame, pd.DataFrame]:
    sheets = provider.read_excel_all_sheets(ADS_ANALYSIS_KEY)
    daily = sheets.get("Статистика_Ежедневно", pd.DataFrame()).copy()
    campaigns = sheets.get("Список_кампаний", pd.DataFrame()).copy()
    if daily.empty:
        return daily, campaigns
    daily = daily.rename(columns={
        "ID кампании": "id_campaign",
        "Артикул WB": "nmId",
        "Название предмета": "subject",
        "Дата": "date",
    })
    daily["date"] = parse_date_col(daily["date"])
    for c in ["Показы", "Клики", "Заказы", "Расход", "Сумма заказов", "CTR", "CR", "ДРР"]:
        if c not in daily.columns:
            daily[c] = 0
    daily["Показы"] = daily["Показы"].map(safe_float)
    daily["Клики"] = daily["Клики"].map(safe_float)
    daily["Заказы"] = daily["Заказы"].map(safe_float)
    daily["Расход"] = daily["Расход"].map(safe_float)
    daily["Сумма заказов"] = daily["Сумма заказов"].map(safe_float)
    daily["subject_norm"] = daily["subject"].map(canonical_subject)
    daily = daily[daily["subject_norm"].isin(TARGET_SUBJECTS)].copy()

    if not campaigns.empty:
        campaigns = campaigns.rename(columns={"ID кампании": "id_campaign", "Артикул WB": "nmId", "Название предмета": "subject"})
        campaigns["subject_norm"] = campaigns["subject"].map(canonical_subject)
        campaigns = campaigns[campaigns["subject_norm"].isin(TARGET_SUBJECTS)].copy()
        campaigns["payment_type"] = campaigns["Тип оплаты"].astype(str).str.lower().str.strip()
        campaigns["bid_search_rub"] = campaigns.get("Ставка в поиске (руб)", 0).map(safe_float)
        campaigns["bid_reco_rub"] = campaigns.get("Ставка в рекомендациях (руб)", 0).map(safe_float)

        def _placement(r):
            s = safe_float(r["bid_search_rub"])
            rr = safe_float(r["bid_reco_rub"])
            if s > 0 and rr > 0:
                return "combined"
            if s > 0:
                return "search"
            if rr > 0:
                return "recommendation"
            return "search"

        campaigns["placement"] = campaigns.apply(_placement, axis=1)
        campaigns["current_bid_rub"] = campaigns.apply(lambda r: r["bid_search_rub"] if r["placement"] in {"search", "combined"} else r["bid_reco_rub"], axis=1)
        campaigns["campaign_status"] = campaigns.get("Статус", "").astype(str)
        campaigns["campaign_is_active"] = campaigns["campaign_status"].map(is_active_campaign_status)
        campaigns = campaigns[campaigns["campaign_is_active"]].copy()

        if not daily.empty:
            active_pairs = campaigns[["id_campaign", "nmId"]].drop_duplicates()
            daily = daily.merge(active_pairs, on=["id_campaign", "nmId"], how="inner")
    return daily, campaigns
    daily = daily.rename(columns={
        "ID кампании": "id_campaign",
        "Артикул WB": "nmId",
        "Название предмета": "subject",
        "Дата": "date",
    })
    daily["date"] = parse_date_col(daily["date"])
    for c in ["Показы","Клики","Заказы","Расход","Сумма заказов","CTR","CR","ДРР"]:
        if c not in daily.columns:
            daily[c] = 0
    daily["Показы"] = daily["Показы"].map(safe_float)
    daily["Клики"] = daily["Клики"].map(safe_float)
    daily["Заказы"] = daily["Заказы"].map(safe_float)
    daily["Расход"] = daily["Расход"].map(safe_float)
    daily["Сумма заказов"] = daily["Сумма заказов"].map(safe_float)
    daily["subject_norm"] = daily["subject"].map(canonical_subject)
    daily = daily[daily["subject_norm"].isin(TARGET_SUBJECTS)].copy()

    if not campaigns.empty:
        campaigns = campaigns.rename(columns={"ID кампании":"id_campaign","Артикул WB":"nmId","Название предмета":"subject"})
        campaigns["subject_norm"] = campaigns["subject"].map(canonical_subject)
        campaigns = campaigns[campaigns["subject_norm"].isin(TARGET_SUBJECTS)].copy()
        campaigns["payment_type"] = campaigns["Тип оплаты"].astype(str).str.lower().str.strip()
        campaigns["bid_search_rub"] = campaigns.get("Ставка в поиске (руб)", 0).map(safe_float)
        campaigns["bid_reco_rub"] = campaigns.get("Ставка в рекомендациях (руб)", 0).map(safe_float)
        def _placement(r):
            s = safe_float(r["bid_search_rub"])
            rr = safe_float(r["bid_reco_rub"])
            if s > 0 and rr > 0:
                return "combined"
            if s > 0:
                return "search"
            if rr > 0:
                return "recommendation"
            return "search"
        campaigns["placement"] = campaigns.apply(_placement, axis=1)
        campaigns["current_bid_rub"] = campaigns.apply(lambda r: r["bid_search_rub"] if r["placement"] in {"search","combined"} else r["bid_reco_rub"], axis=1)
        campaigns["campaign_status"] = campaigns.get("Статус", "").astype(str)
    return daily, campaigns


def load_economics(provider: BaseProvider) -> pd.DataFrame:
    df = provider.read_excel(ECONOMICS_KEY, sheet_name="Юнит экономика").copy()
    df = df.rename(columns={"Артикул WB": "nmId", "Артикул продавца": "supplier_article", "Предмет": "subject"})
    df["subject_norm"] = df.get("subject", "").map(canonical_subject)
    df = df[df["subject_norm"].isin(TARGET_SUBJECTS)].copy()
    df["product_root"] = df["supplier_article"].map(product_root_from_supplier_article)

    buyout_col = find_matching_column(df, ECON_BUYOUT_CANDIDATES)
    gp_col = find_matching_column(df, ECON_GP_UNIT_CANDIDATES)
    np_col = find_matching_column(df, ECON_NP_UNIT_CANDIDATES)

    df["buyout_rate"] = df["subject_norm"].map(get_subject_buyout_rate)
    df["gp_unit"] = pd.to_numeric(df[gp_col], errors="coerce").fillna(0.0) if gp_col else pd.Series(0.0, index=df.index)
    df["np_unit"] = pd.to_numeric(df[np_col], errors="coerce").fillna(0.0) if np_col else pd.Series(0.0, index=df.index)
    df["gp_realized"] = df["gp_unit"] * df["buyout_rate"].clip(lower=0, upper=1)

    week_col = find_matching_column(df, ECON_WEEK_CANDIDATES)
    if week_col:
        df["Неделя"] = df[week_col]
    elif "Неделя" not in df.columns:
        df["Неделя"] = np.arange(len(df))
    return df

def load_orders(provider: BaseProvider) -> pd.DataFrame:
    keys = provider.list_keys(ORDERS_WEEKLY_PREFIX)
    frames = []
    for key in keys:
        try:
            df = provider.read_excel(key).copy()
            if df.empty:
                continue
            df = df.rename(columns={"nmID":"nmId"})
            df["date"] = parse_date_col(df["date"])
            df["supplier_article"] = df.get("supplierArticle", "")
            df["subject"] = df.get("subject", "")
            df["subject_norm"] = df["subject"].map(canonical_subject)
            df = df[df["subject_norm"].isin(TARGET_SUBJECTS)].copy()
            df["product_root"] = df["supplier_article"].map(product_root_from_supplier_article)
            df["finishedPrice"] = df.get("finishedPrice", 0).map(safe_float)
            df["isCancel"] = df.get("isCancel", False).fillna(False).astype(bool)
            frames.append(df)
        except Exception:
            continue
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def load_funnel(provider: BaseProvider) -> pd.DataFrame:
    try:
        df = provider.read_excel(FUNNEL_KEY).copy()
    except Exception:
        return pd.DataFrame()
    df = df.rename(columns={"nmID": "nmId", "dt": "date"})
    df["date"] = parse_date_col(df["date"])

    numeric_cols = [
        "openCardCount", "addToCartCount", "ordersCount", "ordersSumRub",
        "buyoutsCount", "buyoutsSumRub", "cancelCount", "cancelSumRub",
        "addToCartConversion", "cartToOrderConversion", "buyoutPercent",
    ]
    for c in numeric_cols:
        if c in df.columns:
            df[c] = df[c].map(safe_float)

    sales_col = find_matching_column(df, FUNNEL_SALES_CANDIDATES)
    if sales_col and sales_col != "ordersSumRub":
        df["ordersSumRub"] = pd.to_numeric(df[sales_col], errors="coerce").fillna(0.0)
    elif "ordersSumRub" not in df.columns:
        df["ordersSumRub"] = 0.0

    df["buyoutPercent"] = resolve_buyout_rate_from_funnel(df, default=0.85)
    return df

def load_keywords(provider: BaseProvider) -> pd.DataFrame:
    keys = provider.list_keys(KEYWORDS_WEEKLY_PREFIX)
    frames = []
    for key in keys:
        try:
            xls = provider.read_excel_all_sheets(key)
            sheet = xls.get("Позиции по Ключам", next(iter(xls.values())))
            df = sheet.copy()
            if df.empty:
                continue
            df = df.rename(columns={
                "Дата":"date",
                "Артикул WB":"nmId",
                "Артикул продавца":"supplier_article",
                "Предмет":"subject",
                "Рейтинг отзывов":"rating_reviews",
                "Рейтинг карточки":"rating_card",
                "Частота запросов":"query_freq",
                "Частота за неделю":"demand_week",
                "Медианная позиция":"median_position",
                "Переходы в карточку":"clicks_to_card",
                "Заказы":"keyword_orders",
                "Конверсия в заказ %":"keyword_conversion",
                "Видимость %":"visibility_pct",
            })
            df["date"] = parse_date_col(df["date"])
            df["subject_norm"] = df["subject"].map(canonical_subject)
            df = df[df["subject_norm"].isin(TARGET_SUBJECTS)].copy()
            df["product_root"] = df["supplier_article"].map(product_root_from_supplier_article)
            for c in ["query_freq","demand_week","median_position","clicks_to_card","keyword_orders","keyword_conversion","visibility_pct","rating_reviews","rating_card"]:
                if c in df.columns:
                    df[c] = df[c].map(safe_float)
            frames.append(df)
        except Exception:
            continue
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

def load_bid_history(provider: BaseProvider) -> pd.DataFrame:
    if not provider.file_exists(OUT_BID_HISTORY):
        return pd.DataFrame()
    try:
        df = provider.read_excel(OUT_BID_HISTORY).copy()
    except Exception:
        return pd.DataFrame()
    df = df.rename(columns={"Дата запуска":"run_ts","ID кампании":"id_campaign","Артикул WB":"nmId","Тип кампании":"campaign_type"})
    if df.empty:
        return df

    df["run_ts"] = pd.to_datetime(df["run_ts"], errors="coerce")
    df["date"] = df["run_ts"].dt.normalize().astype("datetime64[ns]")

    search_col = pd.to_numeric(df.get("Ставка поиск, коп", 0), errors="coerce") if "Ставка поиск, коп" in df.columns else pd.Series(0, index=df.index, dtype=float)
    reco_col = pd.to_numeric(df.get("Ставка рекомендации, коп", 0), errors="coerce") if "Ставка рекомендации, коп" in df.columns else pd.Series(0, index=df.index, dtype=float)
    bid_kop = search_col.where(search_col.fillna(0) > 0, reco_col)
    df["bid_rub"] = (bid_kop.fillna(0) / 100.0).astype(float)

    df["id_campaign"] = pd.to_numeric(df.get("id_campaign"), errors="coerce")
    df["nmId"] = pd.to_numeric(df.get("nmId"), errors="coerce")
    df = df.dropna(subset=["run_ts", "date", "id_campaign", "nmId"]).copy()
    df["id_campaign"] = df["id_campaign"].astype("int64")
    df["nmId"] = df["nmId"].astype("int64")
    return df

def build_master(econ: pd.DataFrame, orders: pd.DataFrame, keywords: pd.DataFrame, campaigns: pd.DataFrame) -> pd.DataFrame:
    frames = []
    if not econ.empty:
        frames.append(econ[["nmId","supplier_article","product_root","subject","subject_norm","buyout_rate","gp_realized"]].copy())
    if not orders.empty:
        t = orders[["nmId","supplier_article","product_root","subject","subject_norm"]].copy()
        frames.append(t)
    if not keywords.empty:
        t = keywords[["nmId","supplier_article","product_root","subject","subject_norm","rating_reviews","rating_card"]].copy()
        frames.append(t)
    if not campaigns.empty:
        nm_map = campaigns[["id_campaign","nmId","subject","subject_norm"]].copy()
        frames.append(nm_map.rename(columns={"id_campaign":"_drop"}).drop(columns=["_drop"]))
    if not frames:
        return pd.DataFrame(columns=["nmId","supplier_article","product_root","subject","subject_norm","buyout_rate","gp_realized","rating_reviews","rating_card"])
    master = pd.concat(frames, ignore_index=True, sort=False)
    def first_non_empty(s):
        for v in s:
            if pd.notna(v) and str(v) != "":
                return v
        return None
    agg = master.groupby("nmId", as_index=False).agg({
        "supplier_article": first_non_empty,
        "product_root": first_non_empty,
        "subject": first_non_empty,
        "subject_norm": first_non_empty,
        "buyout_rate": "max",
        "gp_realized": "max",
        "rating_reviews": "max",
        "rating_card": "max",
    })
    agg["product_root"] = agg["product_root"].fillna(agg["supplier_article"].map(product_root_from_supplier_article))
    return agg

def aggregate_orders(orders: pd.DataFrame, start: date, end: date, control_field: str) -> pd.DataFrame:
    if orders.empty:
        return pd.DataFrame(columns=[control_field, "total_orders", "total_revenue", "total_orders_raw"])
    df = orders[(orders["date"] >= start) & (orders["date"] <= end) & (~orders["isCancel"])].copy()
    if df.empty:
        return pd.DataFrame(columns=[control_field, "total_orders", "total_revenue", "total_orders_raw"])
    out = df.groupby(control_field, as_index=False).agg(
        total_orders=("nmId", "count"),
        total_revenue=("finishedPrice", "sum"),
    )
    return out

def aggregate_ads_control(ads_daily: pd.DataFrame, start: date, end: date, mapping: pd.DataFrame, control_field: str) -> pd.DataFrame:
    if ads_daily.empty:
        return pd.DataFrame(columns=[control_field, "ad_spend", "ad_clicks", "ad_orders", "ad_impressions", "ad_revenue"])
    df = ads_daily[(ads_daily["date"] >= start) & (ads_daily["date"] <= end)].copy()
    if df.empty:
        return pd.DataFrame(columns=[control_field, "ad_spend", "ad_clicks", "ad_orders", "ad_impressions", "ad_revenue"])
    df = df.merge(mapping[["nmId", control_field]].drop_duplicates(), on="nmId", how="left")
    df = df[df[control_field].notna()].copy()
    if df.empty:
        return pd.DataFrame(columns=[control_field, "ad_spend", "ad_clicks", "ad_orders", "ad_impressions", "ad_revenue"])
    return df.groupby(control_field, as_index=False).agg(
        ad_spend=("Расход", "sum"),
        ad_clicks=("Клики", "sum"),
        ad_orders=("Заказы", "sum"),
        ad_impressions=("Показы", "sum"),
        ad_revenue=("Сумма заказов", "sum"),
    )

def aggregate_keyword_item(keywords: pd.DataFrame, start: date, end: date) -> pd.DataFrame:
    if keywords.empty:
        return pd.DataFrame(columns=["nmId","supplier_article","demand_week","median_position","visibility_pct","rating_reviews","rating_card"])
    df = keywords[(keywords["date"] >= start) & (keywords["date"] <= end)].copy()
    if df.empty:
        return pd.DataFrame(columns=["nmId","supplier_article","demand_week","median_position","visibility_pct","rating_reviews","rating_card"])
    return df.groupby(["nmId","supplier_article"], as_index=False).agg(
        demand_week=("demand_week", "sum"),
        median_position=("median_position", "median"),
        visibility_pct=("visibility_pct", "mean"),
        rating_reviews=("rating_reviews", "max"),
        rating_card=("rating_card", "max"),
        keyword_orders=("keyword_orders", "sum"),
        keyword_clicks=("clicks_to_card", "sum"),
    )

def aggregate_keyword_daily(keywords: pd.DataFrame) -> pd.DataFrame:
    if keywords.empty:
        return pd.DataFrame(columns=["date","nmId","supplier_article","demand","median_position","visibility_pct"])
    return keywords.groupby(["date","nmId","supplier_article"], as_index=False).agg(
        demand=("demand_week", "sum"),
        median_position=("median_position", "median"),
        visibility_pct=("visibility_pct", "mean"),
    )

def build_funnel_item(funnel: pd.DataFrame, master: pd.DataFrame, start: date, end: date) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if funnel.empty:
        cols1 = ["nmId","addToCartConversion","cartToOrderConversion","buyoutPercent"]
        cols2 = ["subject_norm","subj_addToCart","subj_cartToOrder"]
        return pd.DataFrame(columns=cols1), pd.DataFrame(columns=cols2)
    df = funnel[(funnel["date"] >= start) & (funnel["date"] <= end)].copy()
    if df.empty:
        cols1 = ["nmId","addToCartConversion","cartToOrderConversion","buyoutPercent"]
        cols2 = ["subject_norm","subj_addToCart","subj_cartToOrder"]
        return pd.DataFrame(columns=cols1), pd.DataFrame(columns=cols2)
    item = df.groupby("nmId", as_index=False).agg(
        addToCartConversion=("addToCartConversion", "mean"),
        cartToOrderConversion=("cartToOrderConversion", "mean"),
        buyoutPercent=("buyoutPercent", "mean"),
    )
    subj = item.merge(master[["nmId","subject_norm"]].drop_duplicates(), on="nmId", how="left")
    subj = subj.groupby("subject_norm", as_index=False).agg(
        subj_addToCart=("addToCartConversion", "median"),
        subj_cartToOrder=("cartToOrderConversion", "median"),
    )
    return item, subj

def compute_required_growth(blended_drr: float, spend_growth: float, subject_norm: str) -> float:
    sg = max(0.0, safe_float(spend_growth))
    if subject_norm in GROWTH_SUBJECTS:
        if blended_drr <= 0.12:
            return min(max(3.0, sg * 0.40), 15.0)
        if blended_drr <= 0.15:
            return min(max(6.0, sg * 0.60), 20.0)
        return min(max(10.0, sg * 0.80), 25.0)
    else:
        if blended_drr <= 0.12:
            return min(max(3.0, sg * 0.50), 12.0)
        if blended_drr <= 0.15:
            return min(max(6.0, sg * 0.75), 18.0)
        return min(max(10.0, sg * 1.00), 25.0)

def choose_control_key(subject_norm: str, supplier_article: str, product_root: str) -> str:
    supplier_article = str(supplier_article or '').strip()
    if supplier_article:
        return supplier_article
    return str(product_root or '').strip()

def build_subject_benchmarks(rows: pd.DataFrame) -> pd.DataFrame:
    if rows.empty:
        return pd.DataFrame(columns=["subject_norm","placement","bench_ctr","bench_capture_imp","bench_capture_click"])
    df = rows.copy()
    if "subject_norm" not in df.columns:
        df["subject_norm"] = df.get("subject", "")
    if "placement" not in df.columns:
        df["placement"] = ""
    if "ctr_pct" not in df.columns:
        if "Показы" in df.columns and "Клики" in df.columns:
            df["ctr_pct"] = np.where(pd.to_numeric(df["Показы"], errors="coerce").fillna(0) > 0,
                                     pd.to_numeric(df["Клики"], errors="coerce").fillna(0) / pd.to_numeric(df["Показы"], errors="coerce").fillna(0) * 100.0,
                                     0.0)
        else:
            df["ctr_pct"] = 0.0
    if "capture_imp" not in df.columns:
        if "Показы" in df.columns and "demand_week" in df.columns:
            df["capture_imp"] = np.where(pd.to_numeric(df["demand_week"], errors="coerce").fillna(0) > 0,
                                         pd.to_numeric(df["Показы"], errors="coerce").fillna(0) / pd.to_numeric(df["demand_week"], errors="coerce").fillna(0),
                                         0.0)
        else:
            df["capture_imp"] = 0.0
    if "capture_click" not in df.columns:
        base = "keyword_clicks" if "keyword_clicks" in df.columns else ("demand_week" if "demand_week" in df.columns else None)
        if "Клики" in df.columns and base:
            df["capture_click"] = np.where(pd.to_numeric(df[base], errors="coerce").fillna(0) > 0,
                                           pd.to_numeric(df["Клики"], errors="coerce").fillna(0) / pd.to_numeric(df[base], errors="coerce").fillna(0),
                                           0.0)
        else:
            df["capture_click"] = 0.0
    if "total_orders" not in df.columns:
        df["total_orders"] = pd.to_numeric(df.get("Заказы", 0), errors="coerce").fillna(0.0)
    df["capture_imp"] = pd.to_numeric(df["capture_imp"], errors="coerce").fillna(0.0)
    df["capture_click"] = pd.to_numeric(df["capture_click"], errors="coerce").fillna(0.0)
    df["ctr_pct"] = pd.to_numeric(df["ctr_pct"], errors="coerce").fillna(0.0)
    df["total_orders"] = pd.to_numeric(df["total_orders"], errors="coerce").fillna(0.0)
    eligible = df[df["total_orders"] > 0].copy()
    if eligible.empty:
        eligible = df.copy()
    out = eligible.groupby(["subject_norm","placement"], as_index=False).agg(
        bench_ctr=("ctr_pct","median"),
        bench_capture_imp=("capture_imp","median"),
        bench_capture_click=("capture_click","median"),
    )
    return out

def compute_bid_limits(row: pd.Series, subject_benchmarks: pd.DataFrame) -> Tuple[Optional[float], Optional[float], Optional[float], str]:
    subject_norm = row["subject_norm"]
    gp_realized = safe_float(row.get("gp_realized"))
    payment_type = str(row.get("payment_type","cpm"))
    placement = str(row.get("placement","search"))

    local_clicks = safe_float(row.get("Клики"))
    local_orders = safe_float(row.get("Заказы"))
    inherited_clicks = safe_float(row.get("item_clicks_cur", row.get("control_ad_clicks", 0.0)))
    inherited_orders = safe_float(row.get("total_orders"))

    limit_type = "Нет данных"
    cpo = None
    if local_clicks >= 50 and local_orders >= 3:
        cpo = local_clicks / max(local_orders, 1.0)
        limit_type = "Фактический"
    elif inherited_clicks >= 50 and inherited_orders >= 5:
        cpo = inherited_clicks / max(inherited_orders, 1.0)
        limit_type = "Наследуемый"

    if gp_realized <= 0 or cpo is None or cpo <= 0:
        return None, None, None, limit_type

    comfort_share, max_share = (0.50, 0.80) if subject_norm in GROWTH_SUBJECTS else (0.40, 0.65)
    comfort_cpo = gp_realized * comfort_share
    max_cpo = gp_realized * max_share
    comfort_cpc = comfort_cpo / cpo
    max_cpc = max_cpo / cpo

    if payment_type == "cpc":
        comfort_bid = round(comfort_cpc, 2)
        max_bid = round(max_cpc, 2)
    else:
        ctr = safe_float(row.get("ctr_pct")) / 100.0
        if ctr <= 0:
            bench = subject_benchmarks[
                (subject_benchmarks["subject_norm"] == subject_norm) &
                (subject_benchmarks["placement"] == placement)
            ]
            ctr = safe_float(bench["bench_ctr"].iloc[0]) / 100.0 if not bench.empty else 0.02
        ctr = max(ctr, 0.005)
        comfort_bid = round(comfort_cpc * 1000 * ctr, 2)
        max_bid = round(max_cpc * 1000 * ctr, 2)

    if payment_type == "cpc":
        comfort_bid = clamp(comfort_bid, 4.0, 150.0)
        max_bid = clamp(max_bid, 4.0, 300.0)
    else:
        low = 80.0
        high = 700.0 if placement in {"search","combined"} else 1200.0
        comfort_bid = clamp(comfort_bid, low, high)
        max_bid = clamp(max_bid, low, high)

    experiment_bid = round(max_bid, 2)
    return round(comfort_bid, 2), round(max_bid, 2), round(experiment_bid, 2), limit_type

def determine_action(row: pd.Series, cfg: Config) -> Tuple[str, float, str, bool]:
    subject_norm = row["subject_norm"]
    current_bid = safe_float(row["current_bid_rub"])
    comfort_bid = row.get("comfort_bid_rub")
    max_bid = row.get("max_bid_rub")
    total_orders = safe_float(row.get("total_orders"))
    ad_orders = safe_float(row.get("Заказы"))
    blended_drr = safe_float(row.get("blended_drr"))
    order_growth = safe_float(row.get("order_growth_pct"))
    required_growth = safe_float(row.get("required_growth_pct"))
    position = safe_float(row.get("median_position"))
    demand = safe_float(row.get("demand_week"))
    rating = safe_float(row.get("rating_reviews"))
    buyout = safe_float(row.get("buyout_rate"))
    gp_realized = safe_float(row.get("gp_realized"))
    weak_card = bool(row.get("card_issue"))
    weak_eff = safe_float(row.get("eff_index_click")) < 0.7 if pd.notna(row.get("eff_index_click")) else False
    growth = subject_norm in GROWTH_SUBJECTS
    rate_limit = False

    if pd.notna(max_bid) and safe_float(max_bid) > 0:
        rate_limit = current_bid >= safe_float(max_bid) * 0.95

    # If no reliable limits and no sales, collect data only
    if (pd.isna(max_bid) or safe_float(max_bid) <= 0) and total_orders <= 0 and ad_orders <= 0:
        return "Без изменений", current_bid, "Недостаточно данных для расчёта лимитов, собираем статистику", rate_limit

    # Final hard filter by blended DRR > 15%
    if blended_drr > cfg.max_drr:
        if rate_limit or weak_eff:
            return "Предел эффективности ставки", current_bid, f"Общий ДРР {blended_drr*100:.1f}% выше 15%: дальше ставкой расти нецелесообразно", True
        if current_bid > 0 and order_growth < required_growth:
            new_bid = round(current_bid * (1 - cfg.down_step), 2)
            return "Снизить", max(new_bid, 4.0 if str(row.get("payment_type")) == "cpc" else 80.0), f"Общий ДРР {blended_drr*100:.1f}% выше 15% и рост заказов слабый", rate_limit
        return "Без изменений", current_bid, f"Общий ДРР {blended_drr*100:.1f}% выше 15%: рост запрещён финальным фильтром", rate_limit

    if gp_realized <= 0 or rating and rating < 4.5 or buyout and buyout < 0.70:
        if growth:
            return "Без изменений", current_bid, "Локальная экономика слабая: для growth-товара не режем автоматически, наблюдаем", rate_limit
        if current_bid > 0:
            new_bid = round(current_bid * (1 - cfg.down_step), 2)
            return "Снизить", max(new_bid, 4.0 if str(row.get("payment_type")) == "cpc" else 80.0), "Негативная экономика / рейтинг / выкуп", rate_limit
        return "Без изменений", current_bid, "Негативная экономика / рейтинг / выкуп", rate_limit

    weak_position = position <= 0 or position > 15
    demand_high = demand >= 3000
    can_raise = pd.notna(max_bid) and safe_float(max_bid) > current_bid + 0.01

    # Strong sign that ставка уже не помогает
    if weak_eff and rate_limit and weak_position:
        return "Предел эффективности ставки", current_bid, "Ставка близка к максимуму, а трафик/позиция не улучшаются", True

    if growth:
        # default to HOLD for growth categories
        if weak_position and demand_high and can_raise and not weak_card:
            step = cfg.test_up_step if blended_drr >= cfg.comfort_drr_max else cfg.max_up_step
            proposed = round(current_bid * (1 + step), 2)
            new_bid = min(round(safe_float(max_bid), 2), proposed)
            if blended_drr <= cfg.comfort_drr_max:
                return "Повысить", new_bid, "Есть запас по max-ставке и потенциал роста позиции", rate_limit
            return "Тест роста", new_bid, "Запускаем осторожный тест роста в зоне 12–15%", rate_limit
        if weak_card and order_growth < required_growth:
            return "Предел эффективности ставки", current_bid, "Проблема в карточке / воронке: ставкой дальше не лечится", True
        if current_bid > safe_float(max_bid) > 0 and order_growth < required_growth:
            return "Без изменений", current_bid, "Ставка выше расчётного max, но товар ростовый: не режем автоматически", rate_limit
        return "Без изменений", current_bid, "Growth-товар: удерживаем ставку, пока нет сильного сигнала на снижение", rate_limit

    # Brushes and others
    severe = 0
    severe += 1 if weak_card else 0
    severe += 1 if weak_eff else 0
    severe += 1 if order_growth < required_growth else 0
    severe += 1 if weak_position and demand_high else 0

    if weak_position and demand_high and can_raise and order_growth >= 0:
        proposed = round(current_bid * (1 + cfg.max_up_step), 2)
        return "Повысить", min(round(safe_float(max_bid), 2), proposed), "Слабая позиция: подтягиваем ставку к комфортной", rate_limit
    if severe >= 3 and current_bid > 0:
        new_bid = round(current_bid * (1 - cfg.down_step), 2)
        return "Снизить", max(new_bid, 4.0 if str(row.get("payment_type")) == "cpc" else 80.0), "Проблема в карточке / воронке или рост заказов слабый", rate_limit
    return "Без изменений", current_bid, "Без изменений", rate_limit

def build_shade_portfolio(campaigns: pd.DataFrame, master: pd.DataFrame, orders_60: pd.DataFrame) -> pd.DataFrame:
    if campaigns.empty:
        return pd.DataFrame()
    df = campaigns[campaigns["subject_norm"].isin(GROWTH_SUBJECTS)].copy()
    if df.empty:
        return pd.DataFrame()
    df = df.merge(master[["nmId","supplier_article","product_root","rating_reviews"]].drop_duplicates(), on="nmId", how="left")
    ord_map = orders_60.groupby("supplier_article", as_index=False).agg(total_orders_60=("nmId", "count"))
    df = df.merge(ord_map, on="supplier_article", how="left")
    df["total_orders_60"] = df["total_orders_60"].fillna(0)
    core_rows = []
    for advert_id, g in df.groupby("id_campaign"):
        g = g.sort_values(["total_orders_60","rating_reviews"], ascending=[False, False]).copy()
        core_article = g["supplier_article"].iloc[0] if not g.empty else ""
        g["роль"] = g["supplier_article"].eq(core_article).map({True:"CORE", False:"WORKING"})
        core_rows.append(g)
    return pd.concat(core_rows, ignore_index=True) if core_rows else pd.DataFrame()


def build_shade_actions(campaigns: pd.DataFrame, portfolio: pd.DataFrame, master: pd.DataFrame, orders_60: pd.DataFrame, product_metrics: pd.DataFrame, api_key: str = "") -> Tuple[pd.DataFrame, pd.DataFrame]:
    if campaigns.empty or portfolio.empty:
        return pd.DataFrame([{"Комментарий":"Нет подходящих кампаний для анализа оттенков"}]), pd.DataFrame([{"Комментарий":"Нет активных тестов оттенков"}])

    actions: List[Dict[str, Any]] = []

    order_stats = pd.DataFrame()
    if not orders_60.empty:
        order_stats = orders_60.groupby("supplier_article", as_index=False).agg(
            total_orders_60=("nmId", "count"),
            revenue_60=("finishedPrice", "sum"),
        )

    universe = master[["supplier_article", "product_root", "nmId", "rating_reviews", "subject"]].dropna(subset=["supplier_article", "nmId"]).drop_duplicates().copy()
    if not order_stats.empty:
        universe = universe.merge(order_stats, on="supplier_article", how="left")
    universe["total_orders_60"] = pd.to_numeric(universe.get("total_orders_60"), errors="coerce").fillna(0)
    universe["revenue_60"] = pd.to_numeric(universe.get("revenue_60"), errors="coerce").fillna(0)
    universe["rating_reviews"] = pd.to_numeric(universe.get("rating_reviews"), errors="coerce").fillna(0)

    control_drr = product_metrics[["control_key", "blended_drr", "subject_norm"]].drop_duplicates().copy()
    control_drr["blended_drr"] = pd.to_numeric(control_drr.get("blended_drr"), errors="coerce").fillna(0)

    for advert_id, g in portfolio.groupby("id_campaign"):
        current = g.iloc[0]
        product_root = current["product_root"]
        control = control_drr[control_drr["control_key"] == product_root]
        blended = safe_float(control["blended_drr"].iloc[0]) if not control.empty else 0.0

        if blended > 0.15:
            actions.append({
                "ID кампании": safe_int(advert_id),
                "Товар": product_root,
                "Предмет": current.get("subject", ""),
                "Текущий CORE": current.get("supplier_article", ""),
                "Новый оттенок": "",
                "Артикул WB": "",
                "Действие": "Нет действий",
                "Минимальная ставка WB, ₽": None,
                "Причина": "Общий ДРР товара выше 15%, новые оттенки не добавляем",
                "Действие API": "",
                "Статус применения": "не требуется",
                "Тип кампании": f'{current.get("payment_type", "cpm")}_{current.get("placement", "combined")}',
            })
            continue

        used_articles = set(g["supplier_article"].dropna().astype(str))
        candidates = universe[(universe["product_root"] == product_root) & (~universe["supplier_article"].astype(str).isin(used_articles))].copy()
        candidates = candidates[candidates["rating_reviews"] >= MIN_RATING_SHADE].copy()

        if candidates.empty:
            actions.append({
                "ID кампании": safe_int(advert_id),
                "Товар": product_root,
                "Предмет": current.get("subject", ""),
                "Текущий CORE": current.get("supplier_article", ""),
                "Новый оттенок": "",
                "Артикул WB": "",
                "Действие": "Нет действий",
                "Минимальная ставка WB, ₽": None,
                "Причина": "Нет подходящих оттенков с рейтингом >= 4.6",
                "Действие API": "",
                "Статус применения": "не требуется",
                "Тип кампании": f'{current.get("payment_type", "cpm")}_{current.get("placement", "combined")}',
            })
            continue

        candidates = candidates.sort_values(["total_orders_60", "revenue_60", "rating_reviews"], ascending=[False, False, False])
        best = candidates.iloc[0]
        actions.append({
            "ID кампании": safe_int(advert_id),
            "Товар": product_root,
            "Предмет": current.get("subject", ""),
            "Текущий CORE": current.get("supplier_article", ""),
            "Новый оттенок": best["supplier_article"],
            "Артикул WB": safe_int(best["nmId"]),
            "Действие": "Добавить тестовый оттенок",
            "Минимальная ставка WB, ₽": None,
            "Причина": "Расширяем охват товара новым оттенком, старт с минимальной ставкой WB",
            "Действие API": "add",
            "Статус применения": "готово к применению",
            "Тип кампании": f'{current.get("payment_type", "cpm")}_{current.get("placement", "combined")}',
            "Заказы оттенка за 60 дней": round(safe_float(best.get("total_orders_60")), 2),
            "Выручка оттенка за 60 дней, ₽": round(safe_float(best.get("revenue_60")), 2),
            "Рейтинг оттенка": round(safe_float(best.get("rating_reviews")), 2),
        })

    actions_df = pd.DataFrame(actions)
    if actions_df.empty:
        actions_df = pd.DataFrame([{"Комментарий":"Нет действий по оттенкам"}])
    return actions_df, pd.DataFrame([{"Комментарий":"История тестов оттенков начнёт копиться после первого успешного добавления"}])


def apply_shade_actions(actions_df: pd.DataFrame, api_key: str, dry_run: bool) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if actions_df.empty or "Действие API" not in actions_df.columns:
        empty_log = pd.DataFrame([{"Комментарий":"Нет действий по оттенкам"}])
        empty_tests = pd.DataFrame([{"Комментарий":"Нет активных тестов оттенков"}])
        return empty_log, actions_df.copy(), empty_tests

    work = actions_df.copy()
    add_rows = work[work["Действие API"] == "add"].copy()
    add_rows["ID кампании"] = pd.to_numeric(add_rows.get("ID кампании"), errors="coerce")
    add_rows["Артикул WB"] = pd.to_numeric(add_rows.get("Артикул WB"), errors="coerce")
    add_rows = add_rows.dropna(subset=["ID кампании", "Артикул WB"]).copy()

    if add_rows.empty:
        empty_log = pd.DataFrame([{"Комментарий":"Нет валидных оттенков для добавления"}])
        empty_tests = pd.DataFrame([{"Комментарий":"Нет активных тестов оттенков"}])
        return empty_log, work, empty_tests

    logs: List[Dict[str, Any]] = []
    tests_rows: List[Dict[str, Any]] = []

    for advert_id, g in add_rows.groupby("ID кампании"):
        nm_ids = sorted({safe_int(x) for x in g["Артикул WB"].tolist() if safe_int(x) > 0})
        if not nm_ids:
            continue

        payload = {
            "nms": [
                {
                    "advert_id": safe_int(advert_id),
                    "nms": {"add": nm_ids, "delete": []},
                }
            ]
        }
        context = {
            "advert_id": safe_int(advert_id),
            "nm_ids": ",".join(map(str, nm_ids)),
        }

        resp = wb_api_request(
            "PATCH",
            WB_NMS_URL,
            api_key,
            payload,
            method_name="Изменение оттенков",
            timeout=120,
            dry_run=dry_run,
            context=context,
        )

        if dry_run or not api_key:
            logs.append({
                "timestamp": now_ts(),
                "advert_id": safe_int(advert_id),
                "status": "dry-run" if api_key else "skipped",
                "http_status": "",
                "nm_count": len(nm_ids),
                "request_body": json_dumps_safe(payload),
                "response": "dry-run" if api_key else "Нет WB_PROMO_KEY_TOPFACE",
            })
            for idx in g.index:
                work.at[idx, "Статус применения"] = "dry-run" if api_key else "пропущено: нет ключа"
            continue

        ok = bool(resp is not None and resp.status_code == 200)
        response_text = resp.text if resp is not None else ""
        logs.append({
            "timestamp": now_ts(),
            "advert_id": safe_int(advert_id),
            "status": "ok" if ok else "failed",
            "http_status": resp.status_code if resp is not None else "",
            "nm_count": len(nm_ids),
            "request_body": json_dumps_safe(payload),
            "response": truncate_text(response_text, 4000),
        })

        added_set: set[int] = set()
        if ok:
            try:
                data = resp.json()
                for row in data.get("nms", []) or []:
                    if safe_int(row.get("advert_id")) == safe_int(advert_id):
                        added_set = {safe_int(x) for x in ((row.get("nms") or {}).get("added") or [])}
                        break
            except Exception:
                added_set = set()

        for idx in g.index:
            nm_id = safe_int(work.at[idx, "Артикул WB"])
            if ok and (not added_set or nm_id in added_set):
                work.at[idx, "Статус применения"] = "успешно"
                tests_rows.append({
                    "Дата запуска": now_ts(),
                    "ID кампании": safe_int(advert_id),
                    "Артикул WB": nm_id,
                    "Новый оттенок": work.at[idx, "Новый оттенок"],
                    "Минимальная ставка WB, ₽": work.at[idx, "Минимальная ставка WB, ₽"],
                    "Статус": "добавлен",
                })
            else:
                work.at[idx, "Статус применения"] = "ошибка"

    log_df = pd.DataFrame(logs) if logs else pd.DataFrame([{"Комментарий":"Нет оттенков для применения"}])
    tests_df = pd.DataFrame(tests_rows) if tests_rows else pd.DataFrame([{"Комментарий":"Нет успешных добавлений оттенков в этом запуске"}])
    return log_df, work, tests_df


def fetch_wb_min_bids(api_key: str, advert_id: int, nm_ids: List[int], payment_type: str, placement_types: List[str]) -> Dict[int, Dict[str, float]]:
    if not nm_ids:
        return {}
    placement_types = [placement_for_min_endpoint(x) for x in placement_types if str(x).strip()]
    placement_types = list(dict.fromkeys(placement_types))
    body = {
        "advert_id": safe_int(advert_id),
        "nm_ids": [safe_int(x) for x in nm_ids[:100] if safe_int(x) > 0],
        "payment_type": canonical_payment_type(payment_type),
        "placement_types": placement_types or ["combined"],
    }
    resp = wb_api_request(
        "POST",
        WB_BIDS_MIN_URL,
        api_key,
        body,
        method_name="Минимальные ставки",
        timeout=60,
        dry_run=False,
        context={
            "advert_id": safe_int(advert_id),
            "payment_type": canonical_payment_type(payment_type),
            "placement_types": ",".join(body["placement_types"]),
            "nm_count": len(body["nm_ids"]),
        },
    )
    if resp is None or resp.status_code != 200:
        return {}

    out: Dict[int, Dict[str, float]] = {}
    try:
        data = resp.json()
    except Exception:
        return {}

    for item in data.get("bids", []) or []:
        nm_id = safe_int(item.get("nm_id"))
        if nm_id <= 0:
            continue
        by_type: Dict[str, float] = {}
        for bid in item.get("bids", []) or []:
            ptype = placement_for_min_endpoint(bid.get("type"))
            val = safe_float(bid.get("value"))
            if val > 0:
                by_type[ptype] = round(val / 100.0, 2)
                MIN_BID_ROWS.append({
                    "ID кампании": safe_int(advert_id),
                    "Артикул WB": nm_id,
                    "Тип оплаты": canonical_payment_type(payment_type),
                    "Плейсмент": ptype,
                    "Минимальная ставка WB, ₽": round(val / 100.0, 2),
                })
        if by_type:
            out[nm_id] = by_type
    return out

def enrich_with_min_bids(results: Dict[str, Any], api_key: str) -> Dict[str, Any]:
    decisions = results.get("decisions", pd.DataFrame()).copy()
    shade_actions = results.get("shade_actions", pd.DataFrame()).copy()
    MIN_BID_ROWS.clear()

    requests_rows: List[Dict[str, Any]] = []

    if not decisions.empty:
        d = decisions.copy()
        if "Активна для API" in d.columns:
            d = d[d["Активна для API"].fillna(False).astype(bool)].copy()
        elif "Статус кампании" in d.columns:
            d = d[d["Статус кампании"].map(is_active_campaign_status)].copy()
        d["Тип оплаты"] = d["Тип кампании"].map(lambda x: "cpc" if "cpc" in str(x).lower() else "cpm")
        d["Плейсмент API min"] = d["Плейсмент"].map(placement_for_min_endpoint)
        for _, r in d.iterrows():
            advert_id = safe_int(r.get("ID кампании"))
            nm_id = safe_int(r.get("Артикул WB"))
            if advert_id > 0 and nm_id > 0:
                requests_rows.append({
                    "source": "решения",
                    "advert_id": advert_id,
                    "nm_id": nm_id,
                    "payment_type": canonical_payment_type(r.get("Тип оплаты")),
                    "placement_type": placement_for_min_endpoint(r.get("Плейсмент")),
                })

    if not shade_actions.empty and "Артикул WB" in shade_actions.columns:
        s = shade_actions.copy()
        s["Тип оплаты"] = s["Тип кампании"].map(lambda x: "cpc" if "cpc" in str(x).lower() else "cpm")
        s["Плейсмент API min"] = s["Тип кампании"].map(
            lambda x: "combined" if "combined" in str(x).lower() else ("search" if "search" in str(x).lower() else "recommendation")
        )
        action_series = s["Действие API"].astype(str) if "Действие API" in s.columns else pd.Series("", index=s.index)
        for _, r in s[action_series.eq("add")].iterrows():
            advert_id = safe_int(r.get("ID кампании"))
            nm_id = safe_int(r.get("Артикул WB"))
            if advert_id > 0 and nm_id > 0:
                requests_rows.append({
                    "source": "оттенки",
                    "advert_id": advert_id,
                    "nm_id": nm_id,
                    "payment_type": canonical_payment_type(r.get("Тип оплаты")),
                    "placement_type": placement_for_min_endpoint(r.get("Плейсмент API min")),
                })

    if not api_key or not requests_rows:
        results["decisions"] = decisions
        if not shade_actions.empty and "Статус применения" in shade_actions.columns:
            action_series = shade_actions["Действие API"].astype(str) if "Действие API" in shade_actions.columns else pd.Series("", index=shade_actions.index)
            mask = action_series.eq("add")
            shade_actions.loc[mask & shade_actions["Статус применения"].astype(str).isin(["ожидает", ""]), "Статус применения"] = "готово к применению"
        results["shade_actions"] = shade_actions
        results["min_bids_df"] = pd.DataFrame(MIN_BID_ROWS)
        return results

    req_df = pd.DataFrame(requests_rows).drop_duplicates()
    for (advert_id, payment_type), grp in req_df.groupby(["advert_id", "payment_type"]):
        nm_ids = sorted({safe_int(x) for x in grp["nm_id"].tolist() if safe_int(x) > 0})
        placement_types = sorted({placement_for_min_endpoint(x) for x in grp["placement_type"].tolist() if str(x).strip()})
        for i in range(0, len(nm_ids), 100):
            fetch_wb_min_bids(api_key, safe_int(advert_id), nm_ids[i:i+100], payment_type, placement_types)

    min_df = pd.DataFrame(MIN_BID_ROWS).drop_duplicates() if MIN_BID_ROWS else pd.DataFrame(columns=["ID кампании", "Артикул WB", "Тип оплаты", "Плейсмент", "Минимальная ставка WB, ₽"])
    if not min_df.empty:
        min_df["ID кампании"] = pd.to_numeric(min_df["ID кампании"], errors="coerce")
        min_df["Артикул WB"] = pd.to_numeric(min_df["Артикул WB"], errors="coerce")
        lookup = {
            (safe_int(r["ID кампании"]), safe_int(r["Артикул WB"]), canonical_payment_type(r["Тип оплаты"]), placement_for_min_endpoint(r["Плейсмент"])): safe_float(r["Минимальная ставка WB, ₽"])
            for _, r in min_df.iterrows()
        }

        if not decisions.empty:
            decisions["Тип оплаты"] = decisions["Тип кампании"].map(lambda x: "cpc" if "cpc" in str(x).lower() else "cpm")
            decisions["Плейсмент API min"] = decisions["Плейсмент"].map(placement_for_min_endpoint)
            decisions["Минимальная ставка WB, ₽"] = decisions.apply(
                lambda r: lookup.get((safe_int(r["ID кампании"]), safe_int(r["Артикул WB"]), canonical_payment_type(r["Тип оплаты"]), placement_for_min_endpoint(r["Плейсмент"]))),
                axis=1,
            )
            for idx, row in decisions.iterrows():
                min_bid = safe_float(row.get("Минимальная ставка WB, ₽"), default=-1)
                new_bid = safe_float(row.get("Новая ставка, ₽"))
                if min_bid > 0 and new_bid > 0 and new_bid < min_bid:
                    decisions.at[idx, "Новая ставка, ₽"] = round(min_bid, 2)
                    reason = str(decisions.at[idx, "Причина"])
                    suffix = f" | Подняли до минимума WB {min_bid:.2f} ₽"
                    if suffix not in reason:
                        decisions.at[idx, "Причина"] = reason + suffix

        if not shade_actions.empty and "Артикул WB" in shade_actions.columns:
            shade_actions["Тип оплаты"] = shade_actions["Тип кампании"].map(lambda x: "cpc" if "cpc" in str(x).lower() else "cpm")
            shade_actions["Плейсмент API min"] = shade_actions["Тип кампании"].map(
                lambda x: "combined" if "combined" in str(x).lower() else ("search" if "search" in str(x).lower() else "recommendation")
            )
            shade_actions["Минимальная ставка WB, ₽"] = shade_actions.apply(
                lambda r: lookup.get((safe_int(r["ID кампании"]), safe_int(r["Артикул WB"]), canonical_payment_type(r["Тип оплаты"]), placement_for_min_endpoint(r["Плейсмент API min"]))),
                axis=1,
            )
    if not shade_actions.empty and "Статус применения" in shade_actions.columns:
        action_series = shade_actions["Действие API"].astype(str) if "Действие API" in shade_actions.columns else pd.Series("", index=shade_actions.index)
        mask = action_series.eq("add")
        shade_actions.loc[mask & shade_actions["Статус применения"].astype(str).isin(["ожидает", "", "готово к применению"]), "Статус применения"] = "готово к применению"

    results["decisions"] = decisions
    results["shade_actions"] = shade_actions
    results["min_bids_df"] = min_df
    return results

def build_efficiency_history(ads_daily: pd.DataFrame, campaigns: pd.DataFrame, keywords_daily: pd.DataFrame, master: pd.DataFrame, bid_history: pd.DataFrame, as_of_date: date) -> Dict[str, pd.DataFrame]:
    if ads_daily.empty:
        return {"Нет данных": pd.DataFrame([{"Комментарий":"Нет рекламной дневной статистики"}])}
    hist = ads_daily.merge(campaigns[["id_campaign","nmId","placement","payment_type","current_bid_rub"]].drop_duplicates(), on=["id_campaign","nmId"], how="left")
    hist = hist.merge(master[["nmId","supplier_article","subject"]].drop_duplicates(), on="nmId", how="left")
    hist = hist.merge(keywords_daily, on=["date","nmId","supplier_article"], how="left")
    hist["demand"] = hist.get("demand", 0).map(safe_float)
    hist["current_bid_rub"] = hist["current_bid_rub"].map(safe_float)
    hist["id_campaign"] = pd.to_numeric(hist.get("id_campaign"), errors="coerce")
    hist["nmId"] = pd.to_numeric(hist.get("nmId"), errors="coerce")
    hist["date"] = pd.to_datetime(hist["date"], errors="coerce").dt.normalize().astype("datetime64[ns]")
    hist = hist.dropna(subset=["date", "id_campaign", "nmId"]).copy()
    hist["id_campaign"] = hist["id_campaign"].astype("int64")
    hist["nmId"] = hist["nmId"].astype("int64")

    # bid history merge_asof: only datetime64 is valid here
    if not bid_history.empty:
        events = bid_history[["id_campaign","nmId","date","bid_rub"]].copy()
        events["id_campaign"] = pd.to_numeric(events.get("id_campaign"), errors="coerce")
        events["nmId"] = pd.to_numeric(events.get("nmId"), errors="coerce")
        events["date"] = pd.to_datetime(events["date"], errors="coerce").dt.normalize().astype("datetime64[ns]")
        events["bid_rub"] = pd.to_numeric(events.get("bid_rub"), errors="coerce")
        events = events.dropna(subset=["id_campaign", "nmId", "date", "bid_rub"]).copy()
        if not events.empty:
            events["id_campaign"] = events["id_campaign"].astype("int64")
            events["nmId"] = events["nmId"].astype("int64")
        out_parts = []
        for (cid, nm), g in hist.groupby(["id_campaign","nmId"], dropna=False):
            gg = g.sort_values("date").copy()
            ev = events[(events["id_campaign"] == cid) & (events["nmId"] == nm)].copy() if not events.empty else pd.DataFrame()
            if not ev.empty:
                gg = pd.merge_asof(
                    gg.sort_values("date"),
                    ev[["date","bid_rub"]].sort_values("date"),
                    on="date",
                    direction="backward",
                    allow_exact_matches=True,
                )
                gg["bid_rub"] = gg["bid_rub"].fillna(gg["current_bid_rub"])
            else:
                gg["bid_rub"] = gg["current_bid_rub"]
            out_parts.append(gg)
        hist = pd.concat(out_parts, ignore_index=True) if out_parts else hist.assign(bid_rub=hist["current_bid_rub"])
    else:
        hist["bid_rub"] = hist["current_bid_rub"]
    hist["ctr_pct"] = hist["CTR"].map(safe_float)
    hist["capture_imp"] = hist.apply(lambda r: safe_float(r["Показы"]) / safe_float(r["demand"]) if safe_float(r["demand"]) else math.nan, axis=1)
    hist["capture_click"] = hist.apply(lambda r: safe_float(r["Клики"]) / safe_float(r["demand"]) if safe_float(r["demand"]) else math.nan, axis=1)
    hist["eff_imp"] = hist.apply(lambda r: (safe_float(r["Показы"]) / safe_float(r["demand"]) / safe_float(r["bid_rub"])) if safe_float(r["demand"]) and safe_float(r["bid_rub"]) else math.nan, axis=1)
    hist["eff_click"] = hist.apply(lambda r: (safe_float(r["Клики"]) / safe_float(r["demand"]) / safe_float(r["bid_rub"])) if safe_float(r["demand"]) and safe_float(r["bid_rub"]) else math.nan, axis=1)
    hist["Тип кампании"] = hist["payment_type"].astype(str) + "_" + hist["placement"].astype(str)
    hist = hist.sort_values(["supplier_article","date","id_campaign"])

    # conclusions
    out_sheets: Dict[str, pd.DataFrame] = {}
    used_names = set()
    for article, g in hist.groupby("supplier_article"):
        if not str(article):
            continue
        g = g.copy().sort_values(["date","id_campaign"])
        conclusions = []
        prev_eff = {}
        for _, r in g.iterrows():
            key = (r["id_campaign"], r["Тип кампании"])
            cur = safe_float(r["eff_click"], math.nan)
            if math.isnan(cur):
                conclusions.append("Нет спроса или данных")
                continue
            prior = prev_eff.get(key, [])
            prior_valid = [x for x in prior if not math.isnan(x)]
            if len(prior_valid) >= 3:
                base = float(pd.Series(prior_valid[-7:]).median())
                if base > 0:
                    ratio = cur / base
                    if ratio >= 1.10:
                        conclusions.append("За ту же ставку начали получать больше кликов")
                    elif ratio <= 0.90:
                        conclusions.append("Эффективность ставки снижается")
                    else:
                        conclusions.append("Без существенных изменений")
                else:
                    conclusions.append("Недостаточно истории")
            else:
                conclusions.append("Недостаточно истории")
            prev_eff.setdefault(key, []).append(cur)
        sheet = pd.DataFrame({
            "Дата": g["date"],
            "ID кампании": g["id_campaign"],
            "Тип кампании": g["Тип кампании"],
            "Плейсмент": g["placement"],
            "Ставка, ₽": g["bid_rub"].round(2),
            "Показы": g["Показы"].round(0),
            "Клики": g["Клики"].round(0),
            "CTR, %": g["ctr_pct"].round(2),
            "Спрос": g["demand"].round(0),
            "Доля показов": (g["capture_imp"] * 100).round(4),
            "Доля кликов": (g["capture_click"] * 100).round(4),
            "Эффективность ставки по показам": g["eff_imp"].round(6),
            "Эффективность ставки по кликам": g["eff_click"].round(6),
            "Вывод": conclusions,
        })
        out_sheets[sanitize_sheet_name(str(article), used_names)] = sheet
    if not out_sheets:
        out_sheets = {"Нет данных": pd.DataFrame([{"Комментарий":"Нет истории эффективности ставки"}])}
    return out_sheets

def prepare_metrics(provider: BaseProvider, cfg: Config, as_of_date: date) -> Dict[str, Any]:
    window = compute_analysis_window(as_of_date)
    log(f"📅 Анализируем зрелое окно {window['cur_start']} .. {window['cur_end']}; база сравнения {window['base_start']} .. {window['base_end']}")
    ads_daily, campaigns = load_ads(provider)
    econ = load_economics(provider)
    orders = load_orders(provider)
    funnel = load_funnel(provider)
    keywords = load_keywords(provider)
    bid_history = load_bid_history(provider)
    log(f"📣 Реклама: {len(ads_daily):,} строк; кампании: {campaigns['id_campaign'].nunique() if not campaigns.empty else 0}; placement-строк: {len(campaigns):,}")
    log(f"💰 Экономика: {len(econ):,} SKU; Заказы: {len(orders):,} строк; Воронка: {len(funnel):,}; Keywords: {len(keywords):,}")

    master = build_master(econ, orders, keywords, campaigns)
    keywords_current = aggregate_keyword_item(keywords, window["cur_start"], window["cur_end"])
    keywords_daily = aggregate_keyword_daily(keywords)
    funnel_item, funnel_subject = build_funnel_item(funnel, master, window["cur_start"], window["cur_end"])

    econ_latest = econ.sort_values("Неделя").drop_duplicates("nmId", keep="last")[["nmId","supplier_article","product_root","subject","subject_norm","buyout_rate","gp_realized"]]
    campaign_base = campaigns.merge(master[["nmId","supplier_article","product_root","subject","subject_norm","rating_reviews","rating_card"]].drop_duplicates(), on="nmId", how="left")
    campaign_base = campaign_base.merge(econ_latest[["nmId","buyout_rate","gp_realized"]], on="nmId", how="left")
    if campaign_base.empty:
        raise RuntimeError("Нет кампаний целевых предметов в файле рекламы")

    campaign_cur = ads_daily[(ads_daily["date"] >= window["cur_start"]) & (ads_daily["date"] <= window["cur_end"])].groupby(["id_campaign","nmId"], as_index=False).agg(
        Показы=("Показы","sum"), Клики=("Клики","sum"), Заказы=("Заказы","sum"), Расход=("Расход","sum"), Сумма_заказов=("Сумма заказов","sum")
    )
    campaign_base_stats = ads_daily[(ads_daily["date"] >= window["base_start"]) & (ads_daily["date"] <= window["base_end"])].groupby(["id_campaign","nmId"], as_index=False).agg(
        base_Показы=("Показы","sum"), base_Клики=("Клики","sum"), base_Заказы=("Заказы","sum"), base_Расход=("Расход","sum"), base_Сумма_заказов=("Сумма заказов","sum")
    )
    rows = campaign_base.merge(campaign_cur, on=["id_campaign","nmId"], how="left").merge(campaign_base_stats, on=["id_campaign","nmId"], how="left").fillna(0)

    # robustly restore key descriptive columns after merges
    # restore subject and subject_norm after merges
    if "subject" not in rows.columns:
        subject_cols = [c for c in ["subject_x", "subject_y"] if c in rows.columns]
        if subject_cols:
            rows["subject"] = rows[subject_cols[0]]
            for c in subject_cols[1:]:
                rows["subject"] = rows["subject"].where(rows["subject"].astype(str).str.strip() != "", rows[c])
        else:
            rows["subject"] = ""
    else:
        rows["subject"] = rows["subject"].fillna("")

    if "subject_norm" not in rows.columns:
        subject_candidates = [c for c in ["subject_norm_x", "subject_norm_y"] if c in rows.columns]
        if subject_candidates:
            rows["subject_norm"] = rows[subject_candidates[0]]
            for c in subject_candidates[1:]:
                rows["subject_norm"] = rows["subject_norm"].where(rows["subject_norm"].astype(str).str.strip() != "", rows[c])
        else:
            rows["subject_norm"] = rows["subject"].map(canonical_subject)
    else:
        rows["subject_norm"] = rows["subject_norm"].fillna("")
        mask_empty = rows["subject_norm"].astype(str).str.strip() == ""
        rows.loc[mask_empty, "subject_norm"] = rows.loc[mask_empty, "subject"].map(canonical_subject)

    if "supplier_article" not in rows.columns:
        for c in ["supplier_article_x", "supplier_article_y", "supplierArticle", "supplierArticle_x", "supplierArticle_y"]:
            if c in rows.columns:
                rows["supplier_article"] = rows[c]
                break
        else:
            rows["supplier_article"] = ""
    rows["supplier_article"] = rows["supplier_article"].fillna("").astype(str)

    if "product_root" not in rows.columns:
        for c in ["product_root_x", "product_root_y"]:
            if c in rows.columns:
                rows["product_root"] = rows[c]
                break
        else:
            rows["product_root"] = rows["supplier_article"].map(product_root_from_supplier_article)
    missing_root = rows["product_root"].isna() | (rows["product_root"].astype(str).str.strip() == "")
    rows.loc[missing_root, "product_root"] = rows.loc[missing_root, "supplier_article"].map(product_root_from_supplier_article)

    # control metrics
    rows["control_key"] = rows.apply(lambda r: choose_control_key(r.get("subject_norm", ""), r.get("supplier_article", ""), r.get("product_root", "")), axis=1)
    orders_cur_root = aggregate_orders(orders, window["cur_start"], window["cur_end"], "product_root")
    orders_base_root = aggregate_orders(orders, window["base_start"], window["base_end"], "product_root").rename(columns={"total_orders":"base_total_orders","total_revenue":"base_total_revenue"})
    orders_cur_article = aggregate_orders(orders, window["cur_start"], window["cur_end"], "supplier_article")
    orders_base_article = aggregate_orders(orders, window["base_start"], window["base_end"], "supplier_article").rename(columns={"total_orders":"base_total_orders","total_revenue":"base_total_revenue"})

    ads_cur_root = aggregate_ads_control(ads_daily, window["cur_start"], window["cur_end"], master, "product_root")
    ads_base_root = aggregate_ads_control(ads_daily, window["base_start"], window["base_end"], master, "product_root").rename(columns={
        "ad_spend":"base_ad_spend","ad_clicks":"base_ad_clicks","ad_orders":"base_ad_orders","ad_impressions":"base_ad_impressions","ad_revenue":"base_ad_revenue"})
    ads_cur_article = aggregate_ads_control(ads_daily, window["cur_start"], window["cur_end"], master, "supplier_article")
    ads_base_article = aggregate_ads_control(ads_daily, window["base_start"], window["base_end"], master, "supplier_article").rename(columns={
        "ad_spend":"base_ad_spend","ad_clicks":"base_ad_clicks","ad_orders":"base_ad_orders","ad_impressions":"base_ad_impressions","base_ad_revenue":"ad_revenue","ad_revenue":"base_ad_revenue"})

    # attach based on control type (safe merges without duplicate key columns)
    root_rows = rows["subject_norm"].isin(GROWTH_SUBJECTS)

    growth_part = rows[root_rows].copy()
    growth_part = growth_part.merge(orders_cur_root.rename(columns={"product_root": "control_key"}), on="control_key", how="left")
    growth_part = growth_part.merge(orders_base_root.rename(columns={"product_root": "control_key"}), on="control_key", how="left")
    growth_part = growth_part.merge(ads_cur_root.rename(columns={"product_root": "control_key"}), on="control_key", how="left")
    growth_part = growth_part.merge(ads_base_root.rename(columns={"product_root": "control_key"}), on="control_key", how="left")

    brush_part = rows[~root_rows].copy()
    brush_part = brush_part.merge(orders_cur_article.rename(columns={"supplier_article": "control_key"}), on="control_key", how="left")
    brush_part = brush_part.merge(orders_base_article.rename(columns={"supplier_article": "control_key"}), on="control_key", how="left")
    brush_part = brush_part.merge(ads_cur_article.rename(columns={"supplier_article": "control_key"}), on="control_key", how="left")
    brush_part = brush_part.merge(ads_base_article.rename(columns={"supplier_article": "control_key"}), on="control_key", how="left")

    rows = pd.concat([growth_part, brush_part], ignore_index=True, sort=False).fillna(0)

    rows = rows.merge(keywords_current, on=["nmId","supplier_article"], how="left")
    rows = rows.merge(funnel_item, on="nmId", how="left").merge(funnel_subject, on="subject_norm", how="left")
    rows["ctr_pct"] = rows.apply(lambda r: pct(r["Клики"], r["Показы"]), axis=1)
    rows["capture_imp"] = rows.apply(lambda r: safe_float(r["Показы"]) / safe_float(r["demand_week"]) if safe_float(r["demand_week"]) else 0.0, axis=1)
    rows["capture_click"] = rows.apply(lambda r: safe_float(r["Клики"]) / safe_float(r["demand_week"]) if safe_float(r["demand_week"]) else 0.0, axis=1)
    rows["blended_drr"] = rows.apply(lambda r: safe_float(r["ad_spend"]) / safe_float(r["total_revenue"]) if safe_float(r["total_revenue"]) else 0.0, axis=1)
    rows["ad_drr"] = rows.apply(lambda r: safe_float(r["Расход"]) / safe_float(r["Сумма_заказов"]) if safe_float(r["Сумма_заказов"]) else 0.0, axis=1)
    rows["order_growth_pct"] = rows.apply(lambda r: growth_pct(r["total_orders"], r["base_total_orders"]), axis=1)
    rows["spend_growth_pct"] = rows.apply(lambda r: growth_pct(r["ad_spend"], r["base_ad_spend"]), axis=1)
    rows["drr_growth_pp"] = rows.apply(lambda r: (safe_float(r["blended_drr"]) - (safe_float(r["base_ad_spend"]) / safe_float(r["base_total_revenue"]) if safe_float(r["base_total_revenue"]) else 0.0))*100.0, axis=1)
    rows["required_growth_pct"] = rows.apply(lambda r: compute_required_growth(safe_float(r["blended_drr"]), safe_float(r["spend_growth_pct"]), r["subject_norm"]), axis=1)
    rows["card_issue"] = rows.apply(lambda r: (safe_float(r.get("addToCartConversion")) > 0 and safe_float(r.get("subj_addToCart")) > 0 and safe_float(r["addToCartConversion"]) < safe_float(r["subj_addToCart"]) * 0.7) or (safe_float(r.get("cartToOrderConversion")) > 0 and safe_float(r.get("subj_cartToOrder")) > 0 and safe_float(r["cartToOrderConversion"]) < safe_float(r["subj_cartToOrder"]) * 0.7), axis=1)

    # preliminary rows for benchmarks
    rows["bid_eff_imp"] = rows.apply(lambda r: (safe_float(r["capture_imp"]) / safe_float(r["current_bid_rub"])) if safe_float(r["current_bid_rub"]) else 0.0, axis=1)
    rows["bid_eff_click"] = rows.apply(lambda r: (safe_float(r["capture_click"]) / safe_float(r["current_bid_rub"])) if safe_float(r["current_bid_rub"]) else 0.0, axis=1)
    subject_benchmarks = build_subject_benchmarks(rows)
    rows = rows.merge(subject_benchmarks, on=["subject_norm","placement"], how="left")
    rows["eff_index_imp"] = rows.apply(lambda r: safe_float(r["capture_imp"]) / safe_float(r["bench_capture_imp"]) if safe_float(r["bench_capture_imp"]) else 1.0, axis=1)
    rows["eff_index_click"] = rows.apply(lambda r: safe_float(r["capture_click"]) / safe_float(r["bench_capture_click"]) if safe_float(r["bench_capture_click"]) else 1.0, axis=1)

    # limits and decisions
    limits = rows.apply(lambda r: pd.Series(compute_bid_limits(r, subject_benchmarks), index=["comfort_bid_rub","max_bid_rub","experiment_bid_rub","limit_type"]), axis=1)
    rows = pd.concat([rows, limits], axis=1)
    decisions = []
    for _, r in rows.iterrows():
        action, new_bid, reason, rate_limit = determine_action(r, cfg)
        decisions.append({
            "Дата запуска": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "ID кампании": safe_int(r["id_campaign"]),
            "Артикул WB": safe_int(r["nmId"]),
            "Артикул продавца": r["supplier_article"],
            "Товар": r["control_key"],
            "Предмет": r.get("subject", ""),
            "Плейсмент": r["placement"],
            "Тип кампании": f'{r["payment_type"]}_{r["placement"]}',
            "Текущая ставка, ₽": round(safe_float(r["current_bid_rub"]), 2),
            "Комфортная ставка, ₽": round(safe_float(r["comfort_bid_rub"]), 2) if pd.notna(r["comfort_bid_rub"]) else None,
            "Максимальная ставка, ₽": round(safe_float(r["max_bid_rub"]), 2) if pd.notna(r["max_bid_rub"]) else None,
            "Экспериментальная ставка, ₽": round(safe_float(r["experiment_bid_rub"]), 2) if pd.notna(r["experiment_bid_rub"]) else None,
            "Тип лимита": r["limit_type"],
            "Действие": action,
            "Новая ставка, ₽": round(safe_float(new_bid), 2),
            "Причина": reason,
            "Показы": round(safe_float(r["Показы"]), 0),
            "Клики": round(safe_float(r["Клики"]), 0),
            "CTR, %": round(safe_float(r["ctr_pct"]), 2),
            "Заказы РК": round(safe_float(r["Заказы"]), 2),
            "Все заказы товара": round(safe_float(r["total_orders"]), 2),
            "Расход РК, ₽": round(safe_float(r["Расход"]), 2),
            "Выручка РК, ₽": round(safe_float(r["Сумма_заказов"]), 2),
            "Выручка товара, ₽": round(safe_float(r["total_revenue"]), 2),
            "Общий ДРР товара, %": round(safe_float(r["blended_drr"]) * 100, 2),
            "Рекламный ДРР, %": round(safe_float(r["ad_drr"]) * 100, 2),
            "Рост заказов, %": round(safe_float(r["order_growth_pct"]), 2),
            "Рост расходов, %": round(safe_float(r["spend_growth_pct"]), 2),
            "Требуемый рост заказов, %": round(safe_float(r["required_growth_pct"]), 2),
            "Спрос за окно": round(safe_float(r["demand_week"]), 0),
            "Медианная позиция": round(safe_float(r["median_position"]), 2),
            "Видимость, %": round(safe_float(r["visibility_pct"]), 2),
            "Индекс эффективности ставки по показам": round(safe_float(r["eff_index_imp"]), 4),
            "Индекс эффективности ставки по кликам": round(safe_float(r["eff_index_click"]), 4),
            "Предел эффективности": "Да" if rate_limit or action == "Предел эффективности ставки" else "Нет",
            "Проблема карточки": "Да" if bool(r["card_issue"]) else "Нет",
        })
    decisions_df = pd.DataFrame(decisions)
    # weak positions simple
    weak = decisions_df[(decisions_df["Действие"].isin(["Снизить","Предел эффективности ставки"])) | (decisions_df["Медианная позиция"] > 20)].copy()
    weak["Комментарий"] = weak["Причина"]
    weak = weak[["Артикул продавца","Артикул WB","ID кампании","Тип кампании","Плейсмент","Действие","Комментарий"]].drop_duplicates()

    # product metrics
    product_metrics = rows.groupby(["control_key","subject_norm"], as_index=False).agg(
        total_orders=("total_orders","max"),
        total_revenue=("total_revenue","max"),
        ad_spend=("ad_spend","max"),
        ad_orders=("ad_orders","max"),
        ad_clicks=("ad_clicks","max"),
        blended_drr=("blended_drr","max"),
        order_growth_pct=("order_growth_pct","max"),
        spend_growth_pct=("spend_growth_pct","max"),
        required_growth_pct=("required_growth_pct","max"),
    ).rename(columns={"control_key":"Товар","subject_norm":"Предмет код"})
    product_metrics["Общий ДРР товара, %"] = (product_metrics["blended_drr"]*100).round(2)

    # benchmark comparison clean
    bench_cmp = decisions_df.merge(subject_benchmarks, left_on=["Предмет","Плейсмент"], right_on=["subject_norm","placement"], how="left")
    bench_cmp = bench_cmp[["Артикул продавца","ID кампании","Тип кампании","Плейсмент","CTR, %","Индекс эффективности ставки по показам","Индекс эффективности ставки по кликам","Причина","bench_ctr","bench_capture_imp","bench_capture_click"]].copy()
    bench_cmp = bench_cmp.rename(columns={"bench_ctr":"Эталон CTR, %","bench_capture_imp":"Эталон доля показов","bench_capture_click":"Эталон доля кликов"})

    # effects: simple from changed decisions
    changed = decisions_df[decisions_df["Действие"].isin(["Повысить","Снизить","Тест роста"]) & (decisions_df["Текущая ставка, ₽"] != decisions_df["Новая ставка, ₽"])].copy()
    if changed.empty:
        effects = pd.DataFrame([{"Комментарий":"В этом запуске не было изменений ставок"}])
    else:
        effects = changed[["Дата запуска","Артикул продавца","ID кампании","Тип кампании","Текущая ставка, ₽","Новая ставка, ₽","Действие","Причина"]].copy()
        effects["Комментарий"] = "Ожидаем накопление зрелых данных после изменения"

    orders_60 = orders[(orders["date"] >= as_of_date - timedelta(days=60)) & (orders["date"] <= as_of_date) & (~orders["isCancel"])].copy() if not orders.empty else pd.DataFrame()
    shade_portfolio = build_shade_portfolio(campaigns, master, orders_60)
    shade_actions, shade_tests = build_shade_actions(campaigns, shade_portfolio, master, orders_60, product_metrics.rename(columns={"Товар":"control_key","Предмет код":"subject_norm","Общий ДРР товара, %":"blended_drr"}), api_key=os.getenv("WB_PROMO_KEY_TOPFACE",""))
    if shade_actions.empty:
        shade_actions = pd.DataFrame([{"Комментарий":"Нет действий по оттенкам"}])

    return {
        "rows": rows,
        "decisions": decisions_df,
        "weak": weak,
        "product_metrics": product_metrics,
        "bench_cmp": bench_cmp,
        "effects": effects,
        "shade_portfolio": shade_portfolio if not shade_portfolio.empty else pd.DataFrame([{"Комментарий":"Нет кампаний по оттенкам"}]),
        "shade_actions": shade_actions,
        "shade_tests": shade_tests,
        "eff_history_sheets": build_efficiency_history(ads_daily, campaigns, keywords_daily, master, bid_history, as_of_date),
        "window": window,
    }

def normalize_bid_for_wb(value_rub: float, payment_type: str, placement: str) -> int:
    value_rub = safe_float(value_rub)
    if payment_type == "cpc":
        return int(round(value_rub * 100))
    # cpm in WB examples also in kopecks
    return int(round(value_rub * 100))

def decisions_to_payload(decisions_df: pd.DataFrame) -> Dict[str, Any]:
    changed = decisions_df[decisions_df["Действие"].isin(["Повысить","Снизить","Тест роста"]) & (decisions_df["Новая ставка, ₽"] != decisions_df["Текущая ставка, ₽"])].copy()
    grouped = {}
    for _, r in changed.iterrows():
        advert = safe_int(r["ID кампании"])
        nm_id = safe_int(r["Артикул WB"])
        payment_type = "cpc" if "cpc" in str(r["Тип кампании"]).lower() else "cpm"
        placement = str(r["Плейсмент"])
        grouped.setdefault((advert, payment_type), []).append({
            "nm_id": nm_id,
            "placement": placement_for_bids_endpoint(placement),
            "bid_kopecks": normalize_bid_for_wb(r["Новая ставка, ₽"], payment_type, placement),
        })
    out = []
    for (advert, payment_type), items in grouped.items():
        out.append({"advert_id": advert, "payment_type": payment_type, "nm_bids": items})
    return {"bids": out}


def send_payload(payload: Dict[str, Any], api_key: str, dry_run: bool) -> pd.DataFrame:
    logs: List[Dict[str, Any]] = []
    for block in payload.get("bids", []):
        advert_id = safe_int(block["advert_id"])
        nm_bids = []
        for item in block.get("nm_bids", []):
            nm_bids.append({
                "nm_id": safe_int(item.get("nm_id")),
                "bid_kopecks": safe_int(item.get("bid_kopecks")),
                "placement": placement_for_bids_endpoint(item.get("placement")),
            })
        body = {"bids": [{"advert_id": advert_id, "nm_bids": nm_bids}]}
        resp = wb_api_request(
            "PATCH",
            WB_BIDS_URL,
            api_key,
            body,
            method_name="Изменение ставок",
            timeout=120,
            dry_run=dry_run,
            context={"advert_id": advert_id, "nm_count": len(nm_bids)},
        )
        logs.append({
            "timestamp": now_ts(),
            "advert_id": advert_id,
            "status": "dry-run" if dry_run and api_key else ("skipped" if not api_key else ("ok" if resp is not None and resp.status_code == 200 else "failed")),
            "http_status": resp.status_code if resp is not None else "",
            "request_body": json_dumps_safe(body),
            "response": truncate_text(resp.text if resp is not None else ("dry-run" if api_key else "Нет WB_PROMO_KEY_TOPFACE"), 4000),
        })
    return pd.DataFrame(logs)

def save_outputs(provider: BaseProvider, results: Dict[str, Any], run_mode: str, bid_send_log: Optional[pd.DataFrame], history_append: pd.DataFrame) -> None:
    decisions = results["decisions"].copy()
    limits_df = decisions[["Артикул продавца","ID кампании","Тип кампании","Текущая ставка, ₽","Комфортная ставка, ₽","Максимальная ставка, ₽","Экспериментальная ставка, ₽","Тип лимита"]].copy()

    min_bids_df = results.get("min_bids_df", pd.DataFrame()).copy()
    if not min_bids_df.empty:
        sort_cols = [c for c in ["ID кампании", "Артикул WB", "Плейсмент"] if c in min_bids_df.columns]
        min_bids_df = min_bids_df.sort_values(sort_cols).drop_duplicates()

    summary = {
        "Режим": run_mode,
        "Дата формирования": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Всего рекомендаций": int(len(decisions)),
        "Изменённых ставок": int(len(decisions[(decisions["Действие"].isin(["Повысить","Снизить","Тест роста"])) & (decisions["Новая ставка, ₽"] != decisions["Текущая ставка, ₽"])])),
        "Достигнут предел эффективности": int((decisions["Действие"] == "Предел эффективности ставки").sum()) if "Действие" in decisions.columns else 0,
        "Слабых позиций": int(len(results["weak"])),
        "Рекомендаций по оттенкам": 0 if results["shade_actions"].empty else int(len(results["shade_actions"])),
        "Блоков отправки ставок": 0 if bid_send_log is None or bid_send_log.empty else int(len(bid_send_log)),
        "Блоков применения оттенков": 0 if shade_apply_log is None or shade_apply_log.empty else int(len(shade_apply_log)),
        "Текущее окно с": results["window"]["cur_start"],
        "Текущее окно по": results["window"]["cur_end"],
        "База с": results["window"]["base_start"],
        "База по": results["window"]["base_end"],
    }
    summary_df = pd.DataFrame([summary])

    # История решений и ставок храним внутри того же единого файла.
    old_sheets = {}
    try:
        if provider.file_exists(OUT_SINGLE_REPORT):
            old_sheets = provider.read_excel_all_sheets(OUT_SINGLE_REPORT)
    except Exception:
        old_sheets = {}

    old_archive = old_sheets.get("Архив решений", old_sheets.get("Архив_решений", pd.DataFrame()))
    new_archive = pd.concat([old_archive, decisions], ignore_index=True) if not old_archive.empty else decisions.copy()

    old_bid_hist = old_sheets.get("История_ставок", pd.DataFrame())
    if history_append is not None and not history_append.empty:
        history_append = history_append.copy()
        new_bid_hist = pd.concat([old_bid_hist, history_append], ignore_index=True) if not old_bid_hist.empty else history_append
    else:
        new_bid_hist = old_bid_hist

    api_log_df = pd.DataFrame(API_CALL_LOGS).copy() if API_CALL_LOGS else pd.DataFrame()

    single_report_sheets = {
        "Решения": decisions,
        "Сводка": summary_df,
        "Минимальные ставки WB": min_bids_df if not min_bids_df.empty else pd.DataFrame([{"Комментарий": "Минимальные ставки не получены"}]),
        "Лимиты ставок": limits_df if not limits_df.empty else pd.DataFrame([{"Комментарий": "Нет данных"}]),
        "Расчёт логики": results["rows"],
        "Метрики по товарам": results["product_metrics"],
        "Слабые позиции": results["weak"] if not results["weak"].empty else pd.DataFrame([{"Комментарий":"Нет слабых позиций"}]),
        "Рекомендации по оттенкам": results["shade_actions"] if not results["shade_actions"].empty else pd.DataFrame([{"Комментарий":"Нет рекомендаций"}]),
        "Состав кампаний по оттенкам": results["shade_portfolio"] if not results["shade_portfolio"].empty else pd.DataFrame([{"Комментарий":"Нет данных"}]),
        "Тесты оттенков": results["shade_tests"] if not results["shade_tests"].empty else pd.DataFrame([{"Комментарий":"Нет данных"}]),
        "Сравнение с сильными РК": results["bench_cmp"] if not results["bench_cmp"].empty else pd.DataFrame([{"Комментарий":"Нет данных"}]),
        "Эффект изменений": results["effects"] if not results["effects"].empty else pd.DataFrame([{"Комментарий":"Нет данных"}]),
        "Эффективность ставки": pd.DataFrame([{"Комментарий":"См. листы ниже по истории эффективности"}]),
        "Лог API": api_log_df if not api_log_df.empty else pd.DataFrame([{"Комментарий":"API-вызовы в этом запуске не выполнялись"}]),
        "Архив решений": new_archive,
        "История ставок": new_bid_hist if new_bid_hist is not None and not new_bid_hist.empty else pd.DataFrame([{"Комментарий":"История ставок пока пуста"}]),
        "Окно анализа": pd.DataFrame([{
            "Текущее окно с": results["window"]["cur_start"],
            "Текущее окно по": results["window"]["cur_end"],
            "База с": results["window"]["base_start"],
            "База по": results["window"]["base_end"],
            "Режим": run_mode,
        }]),
    }

    # Добавляем листы ежедневной эффективности в конец единого файла.
    eff_sheets = results.get("eff_history_sheets", {}) or {}
    for sh_name, sh_df in eff_sheets.items():
        single_report_sheets[f"Эффективность {sh_name}"] = sh_df

    provider.write_excel(OUT_SINGLE_REPORT, single_report_sheets)



def _parse_abc_period_from_key(key: str) -> Tuple[Optional[date], Optional[date]]:
    name = Path(str(key)).name
    m = re.search(r'__(\d{2}\.\d{2}\.\d{4})-(\d{2}\.\d{2}\.\d{4})__', name)
    if not m:
        return None, None
    try:
        s = datetime.strptime(m.group(1), "%d.%m.%Y").date()
        e = datetime.strptime(m.group(2), "%d.%m.%Y").date()
        return s, e
    except Exception:
        return None, None

def _month_bounds(any_day: date) -> Tuple[date, date]:
    start = any_day.replace(day=1)
    if start.month == 12:
        next_month = date(start.year + 1, 1, 1)
    else:
        next_month = date(start.year, start.month + 1, 1)
    end = next_month - timedelta(days=1)
    return start, end

def _previous_month_bounds(as_of_date: date) -> Tuple[date, date]:
    first_cur = as_of_date.replace(day=1)
    prev_end = first_cur - timedelta(days=1)
    return _month_bounds(prev_end)

def _days_in_month(any_day: date) -> int:
    s, e = _month_bounds(any_day)
    return (e - s).days + 1

def _normalize_abc_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["nmId","supplier_article","subject","subject_norm","product_root","gross_profit","promotion","orders","sales_count","gross_revenue","buyout_pct","open_card","add_to_cart","cart_conv_pct","order_conv_pct"])
    out = df.copy()
    out = out.rename(columns={
        "Артикул WB":"nmId",
        "Артикул продавца":"supplier_article",
        "Предмет":"subject",
        "Валовая прибыль":"gross_profit",
        "Продвижение":"promotion",
        "Заказы":"orders",
        "Кол-во продаж":"sales_count",
        "Валовая выручка":"gross_revenue",
        "Процент выкупов, %":"buyout_pct",
        "Открытие карточки":"open_card",
        "Добавлени в корзину":"add_to_cart",
        "Конверсия в корзину, %":"cart_conv_pct",
        "Конверсия в заказ (из корзины), %":"order_conv_pct",
    })
    out["nmId"] = pd.to_numeric(out.get("nmId"), errors="coerce")
    out["supplier_article"] = out.get("supplier_article", "").fillna("").astype(str)
    out["subject"] = out.get("subject", "").fillna("").astype(str)
    out["subject_norm"] = out["subject"].map(canonical_subject)
    out = out[out["subject_norm"].isin(TARGET_SUBJECTS)].copy()
    out["product_root"] = out["supplier_article"].map(product_root_from_supplier_article)
    for c in ["gross_profit","promotion","orders","sales_count","gross_revenue","buyout_pct","open_card","add_to_cart","cart_conv_pct","order_conv_pct"]:
        out[c] = pd.to_numeric(out.get(c), errors="coerce").fillna(0.0)
    out["buyout_rate"] = np.where(out["buyout_pct"] > 1, out["buyout_pct"] / 100.0, out["buyout_pct"])
    out["gp_per_buyout"] = np.where(out["sales_count"] > 0, out["gross_profit"] / out["sales_count"], 0.0)
    out["gp_after_ads_est"] = out["orders"] * out["buyout_rate"] * out["gp_per_buyout"] + out["promotion"]
    return out

def load_abc_month_plan(provider: BaseProvider, as_of_date: date) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    prev_start, prev_end = _previous_month_bounds(as_of_date)
    keys = provider.list_keys(ABC_PREFIX)
    abc_keys = [k for k in keys if "wb_abc_report_goods__" in Path(str(k)).name]
    chosen_key = None
    weekly_keys: List[str] = []
    for key in abc_keys:
        s, e = _parse_abc_period_from_key(key)
        if s == prev_start and e == prev_end:
            chosen_key = key
            break
    plan_source = "monthly_file"
    frames: List[pd.DataFrame] = []
    if chosen_key:
        try:
            frames.append(_normalize_abc_df(provider.read_excel(chosen_key).copy()))
        except Exception:
            frames = []
    else:
        plan_source = "weekly_sum"
        for key in abc_keys:
            s, e = _parse_abc_period_from_key(key)
            if s is None or e is None:
                continue
            if s >= prev_start and e <= prev_end:
                weekly_keys.append(key)
        for key in sorted(weekly_keys):
            try:
                frames.append(_normalize_abc_df(provider.read_excel(key).copy()))
            except Exception:
                continue
    if not frames:
        return pd.DataFrame(), {"source": "none", "period_start": prev_start, "period_end": prev_end, "file": ""}
    abc = pd.concat(frames, ignore_index=True)
    group_cols = ["nmId","supplier_article","subject","subject_norm","product_root"]
    plan = abc.groupby(group_cols, as_index=False).agg(
        gross_profit=("gross_profit","sum"),
        promotion=("promotion","sum"),
        orders=("orders","sum"),
        sales_count=("sales_count","sum"),
        gross_revenue=("gross_revenue","sum"),
        buyout_num=("buyout_rate","sum"),
        row_count=("nmId","count"),
        open_card=("open_card","sum"),
        add_to_cart=("add_to_cart","sum"),
        cart_conv_pct_num=("cart_conv_pct","sum"),
        order_conv_pct_num=("order_conv_pct","sum"),
    )
    plan["buyout_rate"] = np.where(plan["row_count"] > 0, plan["buyout_num"] / plan["row_count"], 0.0)
    plan["gp_per_buyout"] = np.where(plan["sales_count"] > 0, plan["gross_profit"] / plan["sales_count"], 0.0)
    plan["plan_gp_after_ads_month"] = plan["orders"] * plan["buyout_rate"] * plan["gp_per_buyout"] + plan["promotion"]
    plan["plan_cpo_month"] = np.where(plan["orders"] > 0, np.abs(plan["promotion"]) / plan["orders"], 0.0)
    plan["plan_drr_month"] = np.where(plan["gross_revenue"] > 0, np.abs(plan["promotion"]) / plan["gross_revenue"], 0.0)
    plan["plan_cart_conv_pct"] = np.where(plan["open_card"] > 0, plan["add_to_cart"] / plan["open_card"] * 100.0, 0.0)
    plan["plan_order_conv_pct"] = np.where(plan["add_to_cart"] > 0, plan["orders"] / plan["add_to_cart"] * 100.0, 0.0)
    meta = {"source": plan_source, "period_start": prev_start, "period_end": prev_end, "file": Path(chosen_key).name if chosen_key else ",".join(Path(k).name for k in weekly_keys)}
    return plan, meta

def build_demand_ratio_table(keywords: pd.DataFrame, as_of_date: date, cur_end: date) -> pd.DataFrame:
    if keywords.empty:
        return pd.DataFrame(columns=["control_key","demand_ratio"])
    prev_start, prev_end = _previous_month_bounds(as_of_date)
    cur_start = as_of_date.replace(day=1)
    same_len = max(1, (cur_end - cur_start).days + 1)
    prev_same_end = min(prev_end, prev_start + timedelta(days=same_len - 1))
    kw = keywords.copy()
    kw["control_key"] = kw.apply(lambda r: choose_control_key(r.get("subject_norm",""), r.get("supplier_article",""), r.get("product_root","")), axis=1)
    kw["query_freq"] = pd.to_numeric(kw.get("query_freq"), errors="coerce").fillna(0.0)
    cur = kw[(kw["date"] >= cur_start) & (kw["date"] <= cur_end)].groupby("control_key", as_index=False).agg(cur_freq=("query_freq","sum"), cur_days=("date","nunique"))
    prev = kw[(kw["date"] >= prev_start) & (kw["date"] <= prev_same_end)].groupby("control_key", as_index=False).agg(prev_freq=("query_freq","sum"), prev_days=("date","nunique"))
    out = cur.merge(prev, on="control_key", how="outer").fillna(0)
    out["cur_avg_freq"] = np.where(out["cur_days"] > 0, out["cur_freq"] / out["cur_days"], 0.0)
    out["prev_avg_freq"] = np.where(out["prev_days"] > 0, out["prev_freq"] / out["prev_days"], 0.0)
    out["demand_ratio"] = np.where(out["prev_avg_freq"] > 0, out["cur_avg_freq"] / out["prev_avg_freq"], 1.0)
    out["demand_ratio"] = out["demand_ratio"].replace([np.inf, -np.inf], np.nan).fillna(1.0).clip(lower=0.5, upper=1.5)
    return out[["control_key","demand_ratio","cur_avg_freq","prev_avg_freq"]]

def build_channel_balance(ads_daily: pd.DataFrame, campaigns: pd.DataFrame, master: pd.DataFrame, econ_latest: pd.DataFrame, window: Dict[str, date]) -> pd.DataFrame:
    if ads_daily.empty or campaigns.empty:
        return pd.DataFrame(columns=["control_key","cpo_cpc","cpo_cpm","drr_cpc","drr_cpm","gp_after_ads_cpc","gp_after_ads_cpm","orders_cpc","orders_cpm","better_channel","worse_channel"])
    meta = campaigns[["id_campaign","nmId","payment_type"]].drop_duplicates()
    m = master[["nmId","supplier_article","product_root","subject_norm"]].drop_duplicates()
    gp = econ_latest[["nmId","gp_realized"]].drop_duplicates()
    df = ads_daily[(ads_daily["date"] >= window["cur_start"]) & (ads_daily["date"] <= window["cur_end"])].merge(meta, on=["id_campaign","nmId"], how="left").merge(m, on="nmId", how="left").merge(gp, on="nmId", how="left")
    if df.empty:
        return pd.DataFrame(columns=["control_key","cpo_cpc","cpo_cpm","drr_cpc","drr_cpm","gp_after_ads_cpc","gp_after_ads_cpm","orders_cpc","orders_cpm","better_channel","worse_channel"])
    df["control_key"] = df.apply(lambda r: choose_control_key(r.get("subject_norm",""), r.get("supplier_article",""), r.get("product_root","")), axis=1)
    df["channel"] = np.where(df["payment_type"].astype(str).str.lower().eq("cpc"), "CPC", "CPM")
    grp = df.groupby(["control_key","channel"], as_index=False).agg(
        spend=("Расход","sum"),
        orders=("Заказы","sum"),
        revenue=("Сумма заказов","sum"),
        gp_realized=("gp_realized","median"),
        clicks=("Клики","sum"),
        impressions=("Показы","sum"),
    )
    grp["cpo"] = np.where(grp["orders"] > 0, grp["spend"] / grp["orders"], 0.0)
    grp["drr"] = np.where(grp["revenue"] > 0, grp["spend"] / grp["revenue"], 0.0)
    grp["gp_after_ads"] = grp["orders"] * grp["gp_realized"] - grp["spend"]
    wide = grp.pivot_table(index="control_key", columns="channel", values=["cpo","drr","gp_after_ads","orders","spend","clicks","impressions"], aggfunc="first")
    wide.columns = [f"{a}_{b.lower()}" for a,b in wide.columns]
    wide = wide.reset_index()
    for c in ["cpo_cpc","cpo_cpm","drr_cpc","drr_cpm","gp_after_ads_cpc","gp_after_ads_cpm","orders_cpc","orders_cpm","spend_cpc","spend_cpm","clicks_cpc","clicks_cpm","impressions_cpc","impressions_cpm"]:
        if c not in wide.columns:
            wide[c] = 0.0
    def _better(r):
        cpo_cpc, cpo_cpm = safe_float(r["cpo_cpc"]), safe_float(r["cpo_cpm"])
        gp_cpc, gp_cpm = safe_float(r["gp_after_ads_cpc"]), safe_float(r["gp_after_ads_cpm"])
        orders_cpc, orders_cpm = safe_float(r["orders_cpc"]), safe_float(r["orders_cpm"])
        score_cpc = (1 if gp_cpc > gp_cpm else 0) + (1 if (cpo_cpc > 0 and (cpo_cpm <= 0 or cpo_cpc < cpo_cpm)) else 0) + (1 if orders_cpc >= orders_cpm else 0)
        score_cpm = (1 if gp_cpm > gp_cpc else 0) + (1 if (cpo_cpm > 0 and (cpo_cpc <= 0 or cpo_cpm < cpo_cpc)) else 0) + (1 if orders_cpm > orders_cpc else 0)
        if score_cpc > score_cpm:
            return "CPC"
        if score_cpm > score_cpc:
            return "CPM"
        return "BALANCED"
    wide["better_channel"] = wide.apply(_better, axis=1)
    wide["worse_channel"] = wide["better_channel"].map({"CPC":"CPM","CPM":"CPC","BALANCED":"BALANCED"})
    return wide

def build_daily_metrics_history(orders: pd.DataFrame, ads_daily: pd.DataFrame, campaigns: pd.DataFrame, master: pd.DataFrame, econ_latest: pd.DataFrame, funnel: pd.DataFrame, keywords: pd.DataFrame, as_of_date: date, abc_plan: pd.DataFrame) -> pd.DataFrame:
    meta = campaigns[["id_campaign","nmId","placement","payment_type","current_bid_rub"]].drop_duplicates() if not campaigns.empty else pd.DataFrame(columns=["id_campaign","nmId","placement","payment_type","current_bid_rub"])
    m = master[["nmId","supplier_article","product_root","subject_norm","subject"]].drop_duplicates()
    gp = econ_latest[["nmId","buyout_rate","gp_realized"]].drop_duplicates()
    abc_small = pd.DataFrame()
    if not abc_plan.empty:
        abc_small = abc_plan[["nmId","gp_per_buyout","buyout_rate"]].drop_duplicates()
    ad = ads_daily.merge(meta, on=["id_campaign","nmId"], how="left").merge(m, on="nmId", how="left").merge(gp, on="nmId", how="left")
    if not abc_small.empty:
        ad = ad.merge(abc_small.rename(columns={"buyout_rate":"abc_buyout_rate"}), on="nmId", how="left")
    ords = orders[(~orders["isCancel"])].groupby(["date","nmId"], as_index=False).agg(
        total_orders=("nmId","count"),
        revenue_total=("finishedPrice","sum"),
    ) if not orders.empty else pd.DataFrame(columns=["date","nmId","total_orders","revenue_total"])
    hist = ad.merge(ords, on=["date","nmId"], how="left")
    hist["total_orders"] = pd.to_numeric(hist.get("total_orders"), errors="coerce").fillna(0.0)
    hist["revenue_total"] = pd.to_numeric(hist.get("revenue_total"), errors="coerce").fillna(0.0)
    if not funnel.empty:
        f = funnel.groupby(["date","nmId"], as_index=False).agg(
            openCardCount=("openCardCount","sum"),
            addToCartCount=("addToCartCount","sum"),
            ordersCount=("ordersCount","sum"),
            addToCartConversion=("addToCartConversion","mean"),
            cartToOrderConversion=("cartToOrderConversion","mean"),
        )
        hist = hist.merge(f, on=["date","nmId"], how="left")
    if not keywords.empty:
        kw = keywords.groupby(["date","nmId"], as_index=False).agg(
            query_freq=("query_freq","sum"),
            demand_week=("demand_week","sum"),
            keyword_orders=("keyword_orders","sum"),
            median_position=("median_position","median"),
            visibility_pct=("visibility_pct","mean"),
            keyword_clicks=("clicks_to_card","sum"),
        )
        hist = hist.merge(kw, on=["date","nmId"], how="left")
    hist["control_key"] = hist.apply(lambda r: choose_control_key(r.get("subject_norm",""), r.get("supplier_article",""), r.get("product_root","")), axis=1)
    abc_buyout_rate = numeric_series(hist, "abc_buyout_rate", 0.0)
    hist_buyout_rate = numeric_series(hist, "buyout_rate", 0.0)
    hist["expected_buyout_orders"] = hist["total_orders"] * np.where(abc_buyout_rate > 0, abc_buyout_rate, hist_buyout_rate)
    gp_unit_for_day = np.where(hist.get("gp_per_buyout", pd.Series(0, index=hist.index)).fillna(0) > 0, hist.get("gp_per_buyout", pd.Series(0, index=hist.index)).fillna(0), np.where(hist.get("buyout_rate", pd.Series(0, index=hist.index)).fillna(0) > 0, hist.get("gp_realized", pd.Series(0, index=hist.index)).fillna(0) / hist.get("buyout_rate", pd.Series(0, index=hist.index)).replace(0, np.nan), 0))
    gp_unit_for_day = pd.to_numeric(pd.Series(gp_unit_for_day, index=hist.index), errors="coerce").fillna(0.0)
    hist["gross_profit_before_ads"] = hist["expected_buyout_orders"] * gp_unit_for_day
    hist["gp_after_ads"] = hist["gross_profit_before_ads"] - hist["Расход"].fillna(0.0)
    hist["DRR, %"] = np.where(hist["revenue_total"] > 0, (hist["Расход"] / hist["revenue_total"]) * 100.0, 0.0)
    hist["CPO, ₽"] = np.where(hist["Заказы"] > 0, hist["Расход"] / hist["Заказы"], 0.0)
    hist["CTR, %"] = np.where(hist["Показы"] > 0, hist["Клики"] / hist["Показы"] * 100.0, 0.0)
    hist["День зрелый"] = hist["date"] <= (as_of_date - timedelta(days=MATURE_END_OFFSET))
    hist["Дата"] = pd.to_datetime(hist["date"]).dt.strftime("%Y-%m-%d")
    hist["campaign_gross_profit_before_ads"] = hist["Заказы"].fillna(0.0) * gp_unit_for_day * np.where(abc_buyout_rate > 0, abc_buyout_rate, hist_buyout_rate)
    hist["campaign_gp_after_ads"] = hist["campaign_gross_profit_before_ads"] - hist["Расход"].fillna(0.0)
    out = hist[["Дата","date","id_campaign","nmId","supplier_article","control_key","subject","placement","payment_type","current_bid_rub","Показы","Клики","CTR, %","Заказы","Расход","Сумма заказов","campaign_gross_profit_before_ads","campaign_gp_after_ads","total_orders","revenue_total","gross_profit_before_ads","gp_after_ads","DRR, %","CPO, ₽","openCardCount","addToCartCount","ordersCount","addToCartConversion","cartToOrderConversion","query_freq","demand_week","keyword_orders","median_position","visibility_pct","День зрелый"]].copy()
    out = out.rename(columns={
        "id_campaign":"ID кампании",
        "nmId":"Артикул WB",
        "supplier_article":"Артикул продавца",
        "control_key":"Товар",
        "subject":"Предмет",
        "placement":"Плейсмент",
        "payment_type":"Тип оплаты",
        "current_bid_rub":"Ставка, ₽",
        "Расход":"Расходы РК, ₽",
        "Сумма заказов":"Доход РК, ₽",
        "campaign_gross_profit_before_ads":"Валовая прибыль кампании до рекламы, ₽",
        "campaign_gp_after_ads":"Валовая прибыль кампании после рекламы, ₽",
        "total_orders":"Все заказы товара",
        "revenue_total":"Выручка товара, ₽",
        "gross_profit_before_ads":"Валовая прибыль до рекламы, ₽",
        "gp_after_ads":"Валовая прибыль после рекламы, ₽",
        "addToCartConversion":"Конверсия в корзину, %",
        "cartToOrderConversion":"Конверсия в заказ, %",
        "query_freq":"Частотность ключей",
        "demand_week":"Спрос по ключам",
        "keyword_orders":"Заказы по ключам",
        "median_position":"Медианная позиция",
        "visibility_pct":"Видимость, %",
        "День зрелый":"Данные зрелые",
    })
    return out.sort_values(["date","Артикул продавца","ID кампании"]).drop(columns=["date"])

def build_plan_vs_fact(abc_plan: pd.DataFrame, keywords: pd.DataFrame, daily_history: pd.DataFrame, as_of_date: date, cur_end: date) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if abc_plan.empty:
        return pd.DataFrame([{"Комментарий":"Нет ABC-отчётов для расчёта плана"}]), pd.DataFrame([{"Комментарий":"Нет ABC-отчётов для расчёта плана по категории"}])
    demand_ratio = build_demand_ratio_table(keywords, as_of_date, cur_end)
    prev_start, prev_end = _previous_month_bounds(as_of_date)
    cur_start = as_of_date.replace(day=1)
    elapsed_days = max(1, (cur_end - cur_start).days + 1)
    month_days = _days_in_month(prev_start)
    plan = abc_plan.copy()
    plan["control_key"] = plan.apply(lambda r: choose_control_key(r.get("subject_norm",""), r.get("supplier_article",""), r.get("product_root","")), axis=1)
    plan = plan.merge(demand_ratio, on="control_key", how="left")
    plan["demand_ratio"] = pd.to_numeric(plan.get("demand_ratio"), errors="coerce").fillna(1.0)
    plan["plan_gp_after_ads_mtd"] = plan["plan_gp_after_ads_month"] * elapsed_days / month_days * plan["demand_ratio"]
    plan["plan_orders_mtd"] = plan["orders"] * elapsed_days / month_days * plan["demand_ratio"]
    plan["plan_revenue_mtd"] = plan["gross_revenue"] * elapsed_days / month_days * plan["demand_ratio"]
    plan["plan_promo_mtd"] = np.abs(plan["promotion"]) * elapsed_days / month_days * plan["demand_ratio"]
    if daily_history.empty:
        fact = pd.DataFrame(columns=["control_key","fact_gp_after_ads_mtd","fact_orders_mtd","fact_revenue_mtd","fact_spend_mtd","fact_add_to_cart_conv","fact_cart_to_order_conv"])
    else:
        dh = daily_history.copy()
        dh["Дата"] = pd.to_datetime(dh["Дата"], errors="coerce").dt.date
        dh = dh[(dh["Дата"] >= cur_start) & (dh["Дата"] <= cur_end)].copy()
        fact = dh.groupby("Товар", as_index=False).agg(
            fact_gp_after_ads_mtd=("Валовая прибыль после рекламы, ₽","sum"),
            fact_orders_mtd=("Все заказы товара","sum"),
            fact_revenue_mtd=("Выручка товара, ₽","sum"),
            fact_spend_mtd=("Расходы РК, ₽","sum"),
            fact_atc_conv=("Конверсия в корзину, %","mean"),
            fact_ord_conv=("Конверсия в заказ, %","mean"),
        ).rename(columns={"Товар":"control_key"})
    out = plan.merge(fact, on="control_key", how="left").fillna(0)
    out["Темп плана ВП, %"] = np.where(out["plan_gp_after_ads_mtd"] != 0, out["fact_gp_after_ads_mtd"] / out["plan_gp_after_ads_mtd"] * 100.0, 0.0)
    out["Отклонение ВП, ₽"] = out["fact_gp_after_ads_mtd"] - out["plan_gp_after_ads_mtd"]
    out["Темп заказов к плану, %"] = np.where(out["plan_orders_mtd"] != 0, out["fact_orders_mtd"] / out["plan_orders_mtd"] * 100.0, 0.0)
    out["fact_cpo_mtd"] = np.where(out["fact_orders_mtd"] > 0, out["fact_spend_mtd"] / out["fact_orders_mtd"], 0.0)
    out["Проблема плана"] = np.select(
        [
            (out["Темп плана ВП, %"] < 90) & (out["fact_orders_mtd"] < out["plan_orders_mtd"]),
            (out["Темп плана ВП, %"] < 90) & (out["fact_cpo_mtd"] > out["plan_cpo_month"] * 1.1),
            (out["Темп плана ВП, %"] < 90) & (out["fact_atc_conv"] < out["plan_cart_conv_pct"] * 0.9),
            (out["Темп плана ВП, %"] < 90) & (out["fact_ord_conv"] < out["plan_order_conv_pct"] * 0.9 if "plan_order_conv_pct" in out.columns else False),
        ],
        ["Недобор заказов","Слишком дорогой трафик","Просадка в корзину","Просадка в заказ"],
        default=np.where(out["Темп плана ВП, %"] >= 100, "План выполняется", "Требует анализа"),
    )
    out["Категория укрупнённо"] = np.where(out["supplier_article"].astype(str).str.upper().str.startswith("901"), "901", out["subject"])
    prod_cols = ["nmId","supplier_article","control_key","subject","Категория укрупнённо","plan_gp_after_ads_month","plan_gp_after_ads_mtd","fact_gp_after_ads_mtd","Отклонение ВП, ₽","Темп плана ВП, %","orders","plan_orders_mtd","fact_orders_mtd","Темп заказов к плану, %","plan_cpo_month","fact_cpo_mtd","plan_drr_month","plan_cart_conv_pct","plan_order_conv_pct","demand_ratio","Проблема плана"]
    prod = out[prod_cols].copy()
    prod = prod.rename(columns={
        "nmId":"Артикул WB",
        "supplier_article":"Артикул продавца",
        "control_key":"Товар",
        "subject":"Предмет",
        "plan_gp_after_ads_month":"План ВП после рекламы, ₽",
        "plan_gp_after_ads_mtd":"План ВП MTD, ₽",
        "fact_gp_after_ads_mtd":"Факт ВП MTD, ₽",
        "orders":"План заказов месяца",
        "plan_orders_mtd":"План заказов MTD",
        "fact_orders_mtd":"Факт заказов MTD",
        "plan_cpo_month":"План CPO, ₽",
        "fact_cpo_mtd":"Факт CPO MTD, ₽",
        "plan_drr_month":"План ДРР, доля",
        "plan_cart_conv_pct":"План конверсии в корзину, %",
        "plan_order_conv_pct":"План конверсии в заказ, %",
        "demand_ratio":"Коррекция плана по частотности",
    })
    cat = out.groupby("Категория укрупнённо", as_index=False).agg(
        plan_gp_after_ads_month=("plan_gp_after_ads_month","sum"),
        plan_gp_after_ads_mtd=("plan_gp_after_ads_mtd","sum"),
        fact_gp_after_ads_mtd=("fact_gp_after_ads_mtd","sum"),
        plan_orders_mtd=("plan_orders_mtd","sum"),
        fact_orders_mtd=("fact_orders_mtd","sum"),
        fact_spend_mtd=("fact_spend_mtd","sum"),
    )
    cat["Темп плана ВП, %"] = np.where(cat["plan_gp_after_ads_mtd"] != 0, cat["fact_gp_after_ads_mtd"] / cat["plan_gp_after_ads_mtd"] * 100.0, 0.0)
    cat["Факт CPO MTD, ₽"] = np.where(cat["fact_orders_mtd"] > 0, cat["fact_spend_mtd"] / cat["fact_orders_mtd"], 0.0)
    cat = cat.rename(columns={"Категория укрупнённо":"Категория","plan_gp_after_ads_month":"План ВП после рекламы, ₽","plan_gp_after_ads_mtd":"План ВП MTD, ₽","fact_gp_after_ads_mtd":"Факт ВП MTD, ₽","plan_orders_mtd":"План заказов MTD","fact_orders_mtd":"Факт заказов MTD","fact_spend_mtd":"Факт расходы РК MTD, ₽"})
    return prod.sort_values(["Темп плана ВП, %","Факт ВП MTD, ₽"], ascending=[True, False]), cat.sort_values(["Категория"])

def determine_action(row: pd.Series, cfg: Config) -> Tuple[str, float, str, bool]:
    current_bid = safe_float(row.get("current_bid_rub"))
    max_bid = safe_float(row.get("max_bid_rub"))
    payment_type = canonical_payment_type(row.get("payment_type"))
    floor_bid = 4.0 if payment_type == "cpc" else 80.0
    blended_drr = safe_float(row.get("blended_drr"))
    gp_growth_pct = safe_float(row.get("gp_growth_pct"))
    order_growth = safe_float(row.get("order_growth_pct"))
    required_growth = safe_float(row.get("required_growth_pct"))
    campaign_gp_cur = safe_float(row.get("campaign_gp_after_ads_cur"))
    campaign_gp_base = safe_float(row.get("campaign_gp_after_ads_base"))
    campaign_gp_growth_pct = safe_float(row.get("campaign_gp_growth_pct"))
    campaign_cpo = safe_float(row.get("campaign_cpo"))
    better_channel = str(row.get("better_channel") or "BALANCED")
    channel = "CPC" if payment_type == "cpc" else "CPM"
    better_gp = safe_float(row.get("gp_after_ads_cpc" if channel == "CPC" else "gp_after_ads_cpm"))
    alt_gp = safe_float(row.get("gp_after_ads_cpm" if channel == "CPC" else "gp_after_ads_cpc"))
    on_plan_pct = safe_float(row.get("plan_attainment_pct"))
    issue = str(row.get("plan_issue") or "").strip()
    rate_limit = max_bid > 0 and current_bid >= max_bid * 0.95

    def cap_raise(proposed: float) -> float:
        # за один шаг не поднимаем выше max_bid и выше разрешённого шага роста
        proposed = min(proposed, current_bid * (1.0 + cfg.max_up_step))
        if max_bid > 0:
            proposed = min(proposed, max_bid)
        return round(max(proposed, floor_bid), 2)

    def cap_down(proposed: float) -> float:
        return round(max(floor_bid, proposed), 2)

    worse_than_alt = better_channel not in {"BALANCED", channel}
    campaign_profit_negative = campaign_gp_cur <= 0
    campaign_profit_weak = campaign_gp_cur > 0 and campaign_gp_growth_pct < -10
    can_raise_more = max_bid > current_bid + 0.01

    # 1) если план выполняется, держим позицию и очень осторожно тестируем рост только на прибыльном канале
    if on_plan_pct >= 95:
        if better_channel == channel and campaign_gp_cur > 0 and campaign_gp_growth_pct >= -5 and order_growth >= max(0.0, required_growth * 0.5) and can_raise_more:
            new_bid = cap_raise(current_bid * (1.0 + cfg.test_up_step))
            return "Тест роста", new_bid, f"План ВП выполняется ({on_plan_pct:.0f}%), канал прибыльный; тестируем следующий шаг", rate_limit
        return "Без изменений", round(current_bid, 2), f"План ВП выполняется ({on_plan_pct:.0f}%), удерживаем ставку", rate_limit

    # 2) если по кампании прибыль отрицательная и канал слабее альтернативного — режем именно его
    if campaign_profit_negative and worse_than_alt and current_bid > floor_bid:
        new_bid = cap_down(current_bid * (1.0 - cfg.down_step))
        return "Снизить", new_bid, f"Кампания убыточна ({campaign_gp_cur:.0f} ₽), канал хуже альтернативного по CPO/ВП; причина: {issue}", rate_limit

    # 3) если план не выполняется, но канал лучший и прибыль кампании положительная — пытаемся добрать объём им
    if on_plan_pct < 95 and better_channel == channel and campaign_gp_cur > 0 and campaign_gp_growth_pct >= -5 and can_raise_more:
        new_bid = cap_raise(current_bid * (1.0 + cfg.test_up_step))
        return "Повысить", new_bid, f"Ниже плана ВП ({on_plan_pct:.0f}%), канал прибыльный; пробуем добрать объём", rate_limit

    # 4) если общий ДРР перегрет и канал хуже по весам — снижаем
    if blended_drr > cfg.max_drr and worse_than_alt and current_bid > floor_bid:
        new_bid = cap_down(current_bid * (1.0 - cfg.down_step))
        return "Снизить", new_bid, f"Общий ДРР {blended_drr*100:.1f}% выше порога, а канал слабее альтернативного", rate_limit

    # 5) если по кампании прибыль падает и канал хуже альтернативного — снижаем даже при ещё положительной ВП
    if campaign_profit_weak and worse_than_alt and current_bid > floor_bid:
        new_bid = cap_down(current_bid * (1.0 - cfg.down_step))
        return "Снизить", new_bid, f"Прибыль кампании ухудшается ({campaign_gp_growth_pct:.1f}%), а лучший канал сейчас другой", rate_limit

    # 6) если кампания прибыльная, но канал не лучший — просто тормозим рост, не тащим автоматически к базе
    if better_channel != "BALANCED" and worse_than_alt and campaign_gp_cur > 0:
        return "Без изменений", round(current_bid, 2), f"Канал не лучший по весам, но кампания прибыльная ({campaign_gp_cur:.0f} ₽): рост тормозим, ставку не режем автоматически", rate_limit

    if rate_limit:
        return "Предел эффективности ставки", round(current_bid, 2), "Ставка близка к расчётному максимуму", True

    return "Без изменений", round(current_bid, 2), f"Недостаточно сигнала для изменения; план {on_plan_pct:.0f}%, прибыль кампании {campaign_gp_cur:.0f} ₽ ({issue})", rate_limit

def prepare_metrics(provider: BaseProvider, cfg: Config, as_of_date: date) -> Dict[str, Any]:
    window = compute_analysis_window(as_of_date)
    log(f"📅 Анализируем зрелое окно {window['cur_start']} .. {window['cur_end']}; база сравнения {window['base_start']} .. {window['base_end']}")
    ads_daily, campaigns = load_ads(provider)
    econ = load_economics(provider)
    orders = load_orders(provider)
    funnel = load_funnel(provider)
    keywords = load_keywords(provider)
    bid_history = load_bid_history(provider)
    abc_plan, abc_meta = load_abc_month_plan(provider, as_of_date)
    log(f"📣 Реклама: {len(ads_daily):,} строк; кампании: {campaigns['id_campaign'].nunique() if not campaigns.empty else 0}; placement-строк: {len(campaigns):,}")
    log(f"💰 Экономика: {len(econ):,} SKU; Заказы: {len(orders):,} строк; Воронка: {len(funnel):,}; Keywords: {len(keywords):,}; ABC plan rows: {len(abc_plan):,}")

    master = build_master(econ, orders, keywords, campaigns)
    keywords_current = aggregate_keyword_item(keywords, window["cur_start"], window["cur_end"])
    keywords_daily = aggregate_keyword_daily(keywords)
    funnel_item, funnel_subject = build_funnel_item(funnel, master, window["cur_start"], window["cur_end"])

    econ_latest = econ.sort_values("Неделя").drop_duplicates("nmId", keep="last")[["nmId","supplier_article","product_root","subject","subject_norm","buyout_rate","gp_realized"]]
    campaign_base = campaigns.merge(master[["nmId","supplier_article","product_root","subject","subject_norm","rating_reviews","rating_card"]].drop_duplicates(), on="nmId", how="left")
    campaign_base = campaign_base.merge(econ_latest[["nmId","buyout_rate","gp_realized"]], on="nmId", how="left")
    if campaign_base.empty:
        raise RuntimeError("Нет кампаний целевых предметов в файле рекламы")

    campaign_cur = ads_daily[(ads_daily["date"] >= window["cur_start"]) & (ads_daily["date"] <= window["cur_end"])].groupby(["id_campaign","nmId"], as_index=False).agg(
        Показы=("Показы","sum"), Клики=("Клики","sum"), Заказы=("Заказы","sum"), Расход=("Расход","sum"), Сумма_заказов=("Сумма заказов","sum")
    )
    campaign_base_stats = ads_daily[(ads_daily["date"] >= window["base_start"]) & (ads_daily["date"] <= window["base_end"])].groupby(["id_campaign","nmId"], as_index=False).agg(
        base_Показы=("Показы","sum"), base_Клики=("Клики","sum"), base_Заказы=("Заказы","sum"), base_Расход=("Расход","sum"), base_Сумма_заказов=("Сумма заказов","sum")
    )
    rows = campaign_base.merge(campaign_cur, on=["id_campaign","nmId"], how="left").merge(campaign_base_stats, on=["id_campaign","nmId"], how="left").fillna(0)

    if "subject" not in rows.columns:
        subject_cols = [c for c in ["subject_x", "subject_y"] if c in rows.columns]
        rows["subject"] = rows[subject_cols[0]] if subject_cols else ""
    rows["subject"] = rows["subject"].fillna("")
    if "subject_norm" not in rows.columns:
        subject_candidates = [c for c in ["subject_norm_x", "subject_norm_y"] if c in rows.columns]
        rows["subject_norm"] = rows[subject_candidates[0]] if subject_candidates else rows["subject"].map(canonical_subject)
    rows["subject_norm"] = rows["subject_norm"].fillna("").astype(str)
    if "supplier_article" not in rows.columns:
        for c in ["supplier_article_x", "supplier_article_y", "supplierArticle", "supplierArticle_x", "supplierArticle_y"]:
            if c in rows.columns:
                rows["supplier_article"] = rows[c]
                break
        else:
            rows["supplier_article"] = ""
    rows["supplier_article"] = rows["supplier_article"].fillna("").astype(str)
    if "product_root" not in rows.columns:
        for c in ["product_root_x", "product_root_y"]:
            if c in rows.columns:
                rows["product_root"] = rows[c]
                break
        else:
            rows["product_root"] = rows["supplier_article"].map(product_root_from_supplier_article)
    missing_root = rows["product_root"].isna() | (rows["product_root"].astype(str).str.strip() == "")
    rows.loc[missing_root, "product_root"] = rows.loc[missing_root, "supplier_article"].map(product_root_from_supplier_article)

    rows["control_key"] = rows.apply(lambda r: choose_control_key(r.get("subject_norm", ""), r.get("supplier_article", ""), r.get("product_root", "")), axis=1)
    orders_cur_root = aggregate_orders(orders, window["cur_start"], window["cur_end"], "product_root")
    orders_base_root = aggregate_orders(orders, window["base_start"], window["base_end"], "product_root").rename(columns={"total_orders":"base_total_orders","total_revenue":"base_total_revenue"})
    orders_cur_article = aggregate_orders(orders, window["cur_start"], window["cur_end"], "supplier_article")
    orders_base_article = aggregate_orders(orders, window["base_start"], window["base_end"], "supplier_article").rename(columns={"total_orders":"base_total_orders","total_revenue":"base_total_revenue"})

    ads_cur_root = aggregate_ads_control(ads_daily, window["cur_start"], window["cur_end"], master, "product_root")
    ads_base_root = aggregate_ads_control(ads_daily, window["base_start"], window["base_end"], master, "product_root").rename(columns={"ad_spend":"base_ad_spend","ad_clicks":"base_ad_clicks","ad_orders":"base_ad_orders","ad_impressions":"base_ad_impressions","ad_revenue":"base_ad_revenue"})
    ads_cur_article = aggregate_ads_control(ads_daily, window["cur_start"], window["cur_end"], master, "supplier_article")
    ads_base_article = aggregate_ads_control(ads_daily, window["base_start"], window["base_end"], master, "supplier_article").rename(columns={"ad_spend":"base_ad_spend","ad_clicks":"base_ad_clicks","ad_orders":"base_ad_orders","ad_impressions":"base_ad_impressions","ad_revenue":"base_ad_revenue"})

    root_rows = rows["subject_norm"].isin(GROWTH_SUBJECTS)
    growth_part = rows[root_rows].copy()
    growth_part = growth_part.merge(orders_cur_root.rename(columns={"product_root":"control_key"}), on="control_key", how="left")
    growth_part = growth_part.merge(orders_base_root.rename(columns={"product_root":"control_key"}), on="control_key", how="left")
    growth_part = growth_part.merge(ads_cur_root.rename(columns={"product_root":"control_key"}), on="control_key", how="left")
    growth_part = growth_part.merge(ads_base_root.rename(columns={"product_root":"control_key"}), on="control_key", how="left")
    brush_part = rows[~root_rows].copy()
    brush_part = brush_part.merge(orders_cur_article.rename(columns={"supplier_article":"control_key"}), on="control_key", how="left")
    brush_part = brush_part.merge(orders_base_article.rename(columns={"supplier_article":"control_key"}), on="control_key", how="left")
    brush_part = brush_part.merge(ads_cur_article.rename(columns={"supplier_article":"control_key"}), on="control_key", how="left")
    brush_part = brush_part.merge(ads_base_article.rename(columns={"supplier_article":"control_key"}), on="control_key", how="left")
    rows = pd.concat([growth_part, brush_part], ignore_index=True, sort=False).fillna(0)

    rows = rows.merge(keywords_current, on=["nmId","supplier_article"], how="left")
    rows = rows.merge(funnel_item, on="nmId", how="left").merge(funnel_subject, on="subject_norm", how="left")
    rows["ctr_pct"] = rows.apply(lambda r: pct(r["Клики"], r["Показы"]), axis=1)
    rows["capture_imp"] = rows.apply(lambda r: safe_float(r["Показы"]) / safe_float(r["demand_week"]) if safe_float(r["demand_week"]) else 0.0, axis=1)
    rows["capture_click"] = rows.apply(lambda r: safe_float(r["Клики"]) / safe_float(r["demand_week"]) if safe_float(r["demand_week"]) else 0.0, axis=1)
    rows["blended_drr"] = rows.apply(lambda r: safe_float(r["ad_spend"]) / safe_float(r["total_revenue"]) if safe_float(r["total_revenue"]) else 0.0, axis=1)
    rows["ad_drr"] = rows.apply(lambda r: safe_float(r["Расход"]) / safe_float(r["Сумма_заказов"]) if safe_float(r["Сумма_заказов"]) else 0.0, axis=1)
    rows["order_growth_pct"] = rows.apply(lambda r: growth_pct(r["total_orders"], r["base_total_orders"]), axis=1)
    rows["spend_growth_pct"] = rows.apply(lambda r: growth_pct(r["ad_spend"], r["base_ad_spend"]), axis=1)
    rows["required_growth_pct"] = rows.apply(lambda r: compute_required_growth(safe_float(r["blended_drr"]), safe_float(r["spend_growth_pct"]), r["subject_norm"]), axis=1)
    rows["card_issue"] = rows.apply(lambda r: (safe_float(r.get("addToCartConversion")) > 0 and safe_float(r.get("subj_addToCart")) > 0 and safe_float(r["addToCartConversion"]) < safe_float(r["subj_addToCart"]) * 0.7) or (safe_float(r.get("cartToOrderConversion")) > 0 and safe_float(r.get("subj_cartToOrder")) > 0 and safe_float(r["cartToOrderConversion"]) < safe_float(r["subj_cartToOrder"]) * 0.7), axis=1)
    rows["bid_eff_imp"] = rows.apply(lambda r: (safe_float(r["capture_imp"]) / safe_float(r["current_bid_rub"])) if safe_float(r["current_bid_rub"]) else 0.0, axis=1)
    rows["bid_eff_click"] = rows.apply(lambda r: (safe_float(r["capture_click"]) / safe_float(r["current_bid_rub"])) if safe_float(r["current_bid_rub"]) else 0.0, axis=1)
    subject_benchmarks = build_subject_benchmarks(rows)
    rows = rows.merge(subject_benchmarks, on=["subject_norm","placement"], how="left")
    rows["eff_index_imp"] = rows.apply(lambda r: safe_float(r["capture_imp"]) / safe_float(r["bench_capture_imp"]) if safe_float(r["bench_capture_imp"]) else 1.0, axis=1)
    rows["eff_index_click"] = rows.apply(lambda r: safe_float(r["capture_click"]) / safe_float(r["bench_capture_click"]) if safe_float(r["bench_capture_click"]) else 1.0, axis=1)

    daily_history = build_daily_metrics_history(orders, ads_daily, campaigns, master, econ_latest, funnel, keywords, as_of_date, abc_plan)
    plan_vs_fact, category_plan = build_plan_vs_fact(abc_plan, keywords, daily_history, as_of_date, window["cur_end"])
    channel_balance = build_channel_balance(ads_daily, campaigns, master, econ_latest, window)
    rows = rows.merge(channel_balance, on="control_key", how="left")
    if not plan_vs_fact.empty and "Товар" in plan_vs_fact.columns:
        rows = rows.merge(plan_vs_fact[["Товар","План ВП MTD, ₽","Факт ВП MTD, ₽","Темп плана ВП, %","Проблема плана"]].rename(columns={"Товар":"control_key","Темп плана ВП, %":"plan_attainment_pct","Проблема плана":"plan_issue","План ВП MTD, ₽":"plan_gp_mtd","Факт ВП MTD, ₽":"fact_gp_mtd"}), on="control_key", how="left")
    rows["plan_attainment_pct"] = numeric_series(rows, "plan_attainment_pct", 100.0)
    rows["gp_after_ads_cur"] = rows["total_orders"] * rows["gp_realized"] - rows["ad_spend"]
    rows["gp_after_ads_base"] = rows["base_total_orders"] * rows["gp_realized"] - rows["base_ad_spend"]
    rows["gp_growth_pct"] = rows.apply(lambda r: growth_pct(r["gp_after_ads_cur"], r["gp_after_ads_base"]), axis=1)
    # кампанийная прибыль: считаем по прямым заказам кампании, без multi-touch атрибуции
    rows["campaign_gp_after_ads_cur"] = rows["Заказы"] * rows["gp_realized"] - rows["Расход"]
    rows["campaign_gp_after_ads_base"] = rows["base_Заказы"] * rows["gp_realized"] - rows["base_Расход"]
    rows["campaign_gp_growth_pct"] = rows.apply(lambda r: growth_pct(r["campaign_gp_after_ads_cur"], r["campaign_gp_after_ads_base"]), axis=1)
    rows["campaign_cpo"] = np.where(rows["Заказы"] > 0, rows["Расход"] / rows["Заказы"], 0.0)

    limits = rows.apply(lambda r: pd.Series(compute_bid_limits(r, subject_benchmarks), index=["comfort_bid_rub","max_bid_rub","experiment_bid_rub","limit_type"]), axis=1)
    rows = pd.concat([rows, limits], axis=1)

    decisions = []
    for _, r in rows.iterrows():
        action, new_bid, reason, rate_limit = determine_action(r, cfg)
        decisions.append({
            "Дата запуска": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "ID кампании": safe_int(r["id_campaign"]),
            "Артикул WB": safe_int(r["nmId"]),
            "Артикул продавца": r["supplier_article"],
            "Товар": r["control_key"],
            "Предмет": r.get("subject", ""),
            "Плейсмент": r["placement"],
            "Тип кампании": f'{r["payment_type"]}_{r["placement"]}',
            "Текущая ставка, ₽": round(safe_float(r["current_bid_rub"]), 2),
            "Комфортная ставка, ₽": round(safe_float(r["comfort_bid_rub"]), 2) if pd.notna(r["comfort_bid_rub"]) else None,
            "Максимальная ставка, ₽": round(safe_float(r["max_bid_rub"]), 2) if pd.notna(r["max_bid_rub"]) else None,
            "Экспериментальная ставка, ₽": round(safe_float(r["experiment_bid_rub"]), 2) if pd.notna(r["experiment_bid_rub"]) else None,
            "Тип лимита": r["limit_type"],
            "Действие": action,
            "Новая ставка, ₽": round(min(safe_float(new_bid), safe_float(r["max_bid_rub"])) if safe_float(r["max_bid_rub"]) > 0 and safe_float(new_bid) > safe_float(r["max_bid_rub"]) else safe_float(new_bid), 2),
            "Причина": reason,
            "Показы": round(safe_float(r["Показы"]), 0),
            "Клики": round(safe_float(r["Клики"]), 0),
            "CTR, %": round(safe_float(r["ctr_pct"]), 2),
            "Заказы РК": round(safe_float(r["Заказы"]), 2),
            "Все заказы товара": round(safe_float(r["total_orders"]), 2),
            "Расход РК, ₽": round(safe_float(r["Расход"]), 2),
            "Выручка РК, ₽": round(safe_float(r["Сумма_заказов"]), 2),
            "Выручка товара, ₽": round(safe_float(r["total_revenue"]), 2),
            "ВП кампании текущее окно после рекламы, ₽": round(safe_float(r.get("campaign_gp_after_ads_cur")), 2),
            "ВП кампании базовое окно после рекламы, ₽": round(safe_float(r.get("campaign_gp_after_ads_base")), 2),
            "Рост ВП кампании, %": round(safe_float(r.get("campaign_gp_growth_pct")), 2),
            "CPO кампании, ₽": round(safe_float(r.get("campaign_cpo")), 2),
            "ВП текущее окно после рекламы, ₽": round(safe_float(r["gp_after_ads_cur"]), 2),
            "ВП базовое окно после рекламы, ₽": round(safe_float(r["gp_after_ads_base"]), 2),
            "Рост ВП, %": round(safe_float(r["gp_growth_pct"]), 2),
            "План ВП MTD, ₽": round(safe_float(r.get("plan_gp_mtd")), 2),
            "Факт ВП MTD, ₽": round(safe_float(r.get("fact_gp_mtd")), 2),
            "Темп плана ВП, %": round(safe_float(r.get("plan_attainment_pct")), 2),
            "Проблема плана": r.get("plan_issue",""),
            "Общий ДРР товара, %": round(safe_float(r["blended_drr"]) * 100, 2),
            "Рекламный ДРР, %": round(safe_float(r["ad_drr"]) * 100, 2),
            "Рост заказов, %": round(safe_float(r["order_growth_pct"]), 2),
            "Рост расходов, %": round(safe_float(r["spend_growth_pct"]), 2),
            "Требуемый рост заказов, %": round(safe_float(r["required_growth_pct"]), 2),
            "Конверсия в корзину, %": round(safe_float(r.get("addToCartConversion")), 2),
            "Конверсия в заказ, %": round(safe_float(r.get("cartToOrderConversion")), 2),
            "Спрос за окно": round(safe_float(r["demand_week"]), 0),
            "Медианная позиция": round(safe_float(r["median_position"]), 2),
            "Видимость, %": round(safe_float(r["visibility_pct"]), 2),
            "CPO CPC, ₽": round(safe_float(r.get("cpo_cpc")), 2),
            "CPO Полок, ₽": round(safe_float(r.get("cpo_cpm")), 2),
            "ДРР CPC, %": round(safe_float(r.get("drr_cpc")) * 100, 2),
            "ДРР Полок, %": round(safe_float(r.get("drr_cpm")) * 100, 2),
            "ВП после рекламы CPC, ₽": round(safe_float(r.get("gp_after_ads_cpc")), 2),
            "ВП после рекламы Полок, ₽": round(safe_float(r.get("gp_after_ads_cpm")), 2),
            "Лучший канал": r.get("better_channel",""),
            "Индекс эффективности ставки по показам": round(safe_float(r["eff_index_imp"]), 4),
            "Индекс эффективности ставки по кликам": round(safe_float(r["eff_index_click"]), 4),
            "Предел эффективности": "Да" if rate_limit or action == "Предел эффективности ставки" else "Нет",
            "Проблема карточки": "Да" if bool(r["card_issue"]) else "Нет",
        })
    decisions_df = pd.DataFrame(decisions)

    weak = decisions_df[(decisions_df["Действие"].isin(["Снизить","Предел эффективности ставки"])) | (decisions_df["Медианная позиция"] > 20)].copy()
    weak["Комментарий"] = weak["Причина"]
    weak = weak[["Артикул продавца","Артикул WB","ID кампании","Тип кампании","Плейсмент","Действие","Комментарий"]].drop_duplicates()

    product_metrics = rows.groupby(["control_key","subject_norm"], as_index=False).agg(
        total_orders=("total_orders","max"),
        total_revenue=("total_revenue","max"),
        ad_spend=("ad_spend","max"),
        ad_orders=("ad_orders","max"),
        ad_clicks=("ad_clicks","max"),
        blended_drr=("blended_drr","max"),
        order_growth_pct=("order_growth_pct","max"),
        spend_growth_pct=("spend_growth_pct","max"),
        required_growth_pct=("required_growth_pct","max"),
        gp_after_ads_cur=("gp_after_ads_cur","max"),
        gp_after_ads_base=("gp_after_ads_base","max"),
    ).rename(columns={"control_key":"Товар","subject_norm":"Предмет код"})
    product_metrics["Общий ДРР товара, %"] = (product_metrics["blended_drr"]*100).round(2)
    product_metrics["Рост ВП, %"] = product_metrics.apply(lambda r: growth_pct(r["gp_after_ads_cur"], r["gp_after_ads_base"]), axis=1).round(2)

    bench_cmp = decisions_df.merge(subject_benchmarks, left_on=["Предмет","Плейсмент"], right_on=["subject_norm","placement"], how="left")
    bench_cmp = bench_cmp[["Артикул продавца","ID кампании","Тип кампании","Плейсмент","CTR, %","Индекс эффективности ставки по показам","Индекс эффективности ставки по кликам","Причина","bench_ctr","bench_capture_imp","bench_capture_click"]].copy()
    bench_cmp = bench_cmp.rename(columns={"bench_ctr":"Эталон CTR, %","bench_capture_imp":"Эталон доля показов","bench_capture_click":"Эталон доля кликов"})

    changed = decisions_df[decisions_df["Действие"].isin(["Повысить","Снизить","Тест роста"]) & (decisions_df["Текущая ставка, ₽"] != decisions_df["Новая ставка, ₽"])].copy()
    if changed.empty:
        effects = pd.DataFrame([{"Комментарий":"В этом запуске не было изменений ставок"}])
    else:
        effects = changed[["Дата запуска","Артикул продавца","ID кампании","Тип кампании","Текущая ставка, ₽","Новая ставка, ₽","Действие","Причина","План ВП MTD, ₽","Факт ВП MTD, ₽","Темп плана ВП, %","Рост ВП, %"]].copy()
        effects["Комментарий"] = "Ожидаем накопление зрелых данных после изменения"

    orders_60 = orders[(orders["date"] >= as_of_date - timedelta(days=60)) & (orders["date"] <= as_of_date) & (~orders["isCancel"])].copy() if not orders.empty else pd.DataFrame()
    shade_portfolio = build_shade_portfolio(campaigns, master, orders_60)
    product_metrics_for_shades = product_metrics[["Товар","Предмет код","blended_drr"]].copy().rename(columns={"Товар":"control_key","Предмет код":"subject_norm"})
    product_metrics_for_shades = product_metrics_for_shades.drop_duplicates(["control_key","subject_norm"])
    shade_actions, shade_tests = build_shade_actions(campaigns, shade_portfolio, master, orders_60, product_metrics_for_shades, api_key=os.getenv("WB_PROMO_KEY_TOPFACE",""))

    return {
        "rows": rows,
        "decisions": decisions_df,
        "weak": weak,
        "product_metrics": product_metrics,
        "bench_cmp": bench_cmp,
        "effects": effects,
        "shade_portfolio": shade_portfolio if not shade_portfolio.empty else pd.DataFrame([{"Комментарий":"Нет кампаний по оттенкам"}]),
        "shade_actions": shade_actions if not shade_actions.empty else pd.DataFrame([{"Комментарий":"Нет действий по оттенкам"}]),
        "shade_tests": shade_tests,
        "eff_history_sheets": build_efficiency_history(ads_daily, campaigns, keywords_daily, master, bid_history, as_of_date),
        "window": window,
        "daily_history": daily_history,
        "abc_plan_month": abc_plan,
        "plan_vs_fact": plan_vs_fact,
        "category_plan": category_plan,
        "abc_meta": pd.DataFrame([abc_meta]),
    }

def build_history_append(changed: pd.DataFrame, as_of_date: date) -> pd.DataFrame:
    if changed.empty:
        return pd.DataFrame()
    rows = []
    week = f"{as_of_date.isocalendar().year}-W{as_of_date.isocalendar().week:02d}"
    for _, r in changed.iterrows():
        placement = normalize_internal_placement(r.get("Плейсмент"))
        bid_kop = normalize_bid_for_wb(r.get("Новая ставка, ₽"), "cpc" if "cpc" in str(r.get("Тип кампании", "")).lower() else "cpm", placement)
        rows.append({
            "Дата запуска": now_ts(),
            "Неделя": week,
            "ID кампании": safe_int(r.get("ID кампании")),
            "Артикул WB": safe_int(r.get("Артикул WB")),
            "Артикул продавца": r.get("Артикул продавца"),
            "Тип кампании": r.get("Тип кампании"),
            "Плейсмент": r.get("Плейсмент"),
            "Старая ставка, ₽": safe_float(r.get("Текущая ставка, ₽")),
            "Новая ставка, ₽": safe_float(r.get("Новая ставка, ₽")),
            "Действие": r.get("Действие"),
            "Причина": r.get("Причина"),
            "Расход РК, ₽": safe_float(r.get("Расход РК, ₽")),
            "Выручка товара, ₽": safe_float(r.get("Выручка товара, ₽")),
            "ВП текущее окно после рекламы, ₽": safe_float(r.get("ВП текущее окно после рекламы, ₽")),
            "Рост ВП, %": safe_float(r.get("Рост ВП, %")),
            "Ставка поиск, коп": bid_kop if placement in {"search", "combined"} else 0,
            "Ставка рекомендации, коп": bid_kop if placement in {"recommendation", "combined"} else 0,
            "Стратегия": "RUN_ONLY_VP_PLAN",
        })
    return pd.DataFrame(rows)

def save_outputs(provider: BaseProvider, results: Dict[str, Any], run_mode: str, bid_send_log: Optional[pd.DataFrame], shade_apply_log: Optional[pd.DataFrame], history_append: pd.DataFrame) -> None:
    decisions = results["decisions"].copy()
    limits_df = decisions[["Артикул продавца","ID кампании","Тип кампании","Текущая ставка, ₽","Комфортная ставка, ₽","Максимальная ставка, ₽","Экспериментальная ставка, ₽","Тип лимита"]].copy() if not decisions.empty else pd.DataFrame()
    min_bids_df = results.get("min_bids_df", pd.DataFrame()).copy()
    summary = {
        "Режим": "run",
        "Дата формирования": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Всего рекомендаций": int(len(decisions)),
        "Изменённых ставок": int(len(decisions[(decisions["Действие"].isin(["Повысить","Снизить","Тест роста"])) & (decisions["Новая ставка, ₽"] != decisions["Текущая ставка, ₽"])])) if not decisions.empty else 0,
        "Блоков отправки ставок": 0 if bid_send_log is None or bid_send_log.empty else int(len(bid_send_log)),
        "Блоков применения оттенков": 0 if shade_apply_log is None or shade_apply_log.empty else int(len(shade_apply_log)),
        "Текущее окно с": results["window"]["cur_start"],
        "Текущее окно по": results["window"]["cur_end"],
        "База с": results["window"]["base_start"],
        "База по": results["window"]["base_end"],
    }
    summary_df = pd.DataFrame([summary])

    old_sheets = {}
    for candidate in [OUT_SINGLE_REPORT, OUT_PREVIEW]:
        try:
            if provider.file_exists(candidate):
                old_sheets = provider.read_excel_all_sheets(candidate)
                if old_sheets:
                    break
        except Exception:
            pass

    old_archive = old_sheets.get("Архив решений", old_sheets.get("Архив_решений", pd.DataFrame()))
    new_archive = pd.concat([old_archive, decisions], ignore_index=True) if not old_archive.empty else decisions.copy()

    old_bid_hist = old_sheets.get("История ставок", old_sheets.get("История_ставок", pd.DataFrame()))
    if history_append is not None and not history_append.empty:
        new_bid_hist = pd.concat([old_bid_hist, history_append], ignore_index=True) if not old_bid_hist.empty else history_append.copy()
    else:
        new_bid_hist = old_bid_hist.copy() if not old_bid_hist.empty else pd.DataFrame()

    old_daily = old_sheets.get("История день", old_sheets.get("История_день", pd.DataFrame()))
    daily_history = results.get("daily_history", pd.DataFrame()).copy()
    if not daily_history.empty:
        if not old_daily.empty:
            daily_history = pd.concat([old_daily, daily_history], ignore_index=True)
        dedup_cols = [c for c in ["Дата","ID кампании","Артикул WB","Плейсмент"] if c in daily_history.columns]
        if dedup_cols:
            daily_history = daily_history.drop_duplicates(subset=dedup_cols, keep="last")
        daily_history = daily_history.sort_values([c for c in ["Дата","Артикул продавца","ID кампании"] if c in daily_history.columns])

    api_log = pd.DataFrame(API_CALL_LOGS) if API_CALL_LOGS else pd.DataFrame([{"Комментарий":"Нет вызовов API"}])

    sheets = {
        "Сводка": summary_df,
        "Решения": decisions,
        "История день": daily_history if not daily_history.empty else pd.DataFrame([{"Комментарий":"История будет копиться после первого запуска"}]),
        "План ВП месяца": results.get("abc_plan_month", pd.DataFrame()),
        "План vs Факт MTD": results.get("plan_vs_fact", pd.DataFrame()),
        "План категории": results.get("category_plan", pd.DataFrame()),
        "Метаданные ABC": results.get("abc_meta", pd.DataFrame()),
        "Лимиты ставок": limits_df,
        "Метрики по товарам": results["product_metrics"],
        "Минимальные ставки WB": min_bids_df if not min_bids_df.empty else pd.DataFrame([{"Комментарий":"Нет данных WB min bids"}]),
        "Слабые позиции": results["weak"],
        "Эффект изменений": results["effects"],
        "Сравнение с сильными": results["bench_cmp"],
        "Лог API": api_log,
        "Архив решений": new_archive,
        "История ставок": new_bid_hist,
        "Состав оттенков": results["shade_portfolio"],
        "Рекомендации оттенков": results["shade_actions"],
        "Тесты оттенков": results["shade_tests"],
    }
    provider.write_excel(OUT_SINGLE_REPORT, sheets)
    # legacy compatibility
    try:
        provider.write_excel(OUT_PREVIEW, sheets)
    except Exception:
        pass
    provider.write_text(OUT_SUMMARY, json.dumps(summary, ensure_ascii=False, indent=2, default=str))
    eff_sheets = results.get("eff_history_sheets", {})
    if eff_sheets:
        provider.write_excel(OUT_EFF, eff_sheets)

def run_manager(args: argparse.Namespace) -> None:
    API_CALL_LOGS.clear()
    MIN_BID_ROWS.clear()
    provider = choose_provider(args.local_data_dir)
    as_of_date = datetime.strptime(args.as_of_date, "%Y-%m-%d").date() if args.as_of_date else datetime.now().date()
    cfg = Config()
    results = prepare_metrics(provider, cfg, as_of_date)
    api_key = os.getenv("WB_PROMO_KEY_TOPFACE","").strip()
    results = enrich_with_min_bids(results, api_key)
    decisions = results["decisions"].copy()
    log(f"✅ Всего строк решений: {len(decisions)}")
    changed = decisions[(decisions["Действие"].isin(["Повысить","Снизить","Тест роста"])) & (decisions["Текущая ставка, ₽"] != decisions["Новая ставка, ₽"])].copy()
    log(f"🔁 Изменённых ставок: {len(changed)}")
    if not changed.empty:
        print(changed[["Товар","Артикул продавца","Предмет","ID кампании","Плейсмент","Текущая ставка, ₽","Новая ставка, ₽","Действие","Причина"]].head(30).to_string(index=False), flush=True)
    bid_send_log = pd.DataFrame()
    shade_apply_log = pd.DataFrame()
    history_append = pd.DataFrame()
    payload = decisions_to_payload(decisions)
    bid_send_log = send_payload(payload, api_key, dry_run=not bool(api_key))
    log(f"📤 Отправлено блоков в WB: {len(payload.get('bids', []))}")
    history_append = build_history_append(changed, as_of_date)
    if args.apply_shades:
        shade_apply_log, updated_shade_actions, tests_df = apply_shade_actions(results["shade_actions"], api_key, dry_run=not bool(api_key))
        results["shade_actions"] = updated_shade_actions
        results["shade_tests"] = tests_df
        log(f"🎨 Блоков оттенков к применению: {0 if shade_apply_log.empty else len(shade_apply_log)}")
    else:
        log("🎨 Применение оттенков отключено")
    save_outputs(provider, results, "run", bid_send_log, shade_apply_log, history_append)

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Боевой менеджер ставок WB для TOPFACE")
    p.add_argument("mode", nargs="?", default="run", choices=["run"], help="Всегда боевой запуск")
    p.add_argument("--apply-shades", dest="apply_shades", action="store_true", default=True, help="Применять рекомендации по оттенкам через API")
    p.add_argument("--skip-shades", dest="apply_shades", action="store_false", help="Не применять рекомендации по оттенкам")
    p.add_argument("--local-data-dir", default="", help="Локальная папка с файлами")
    p.add_argument("--as-of-date", default="", help="Дата расчёта YYYY-MM-DD")
    return p



# ===================== OVERRIDES: VP-PLAN / HARD-RUN / 3-LEVEL HIERARCHY =====================

class Config:
    comfort_drr_min: float = 0.10
    comfort_drr_max: float = 0.12
    campaign_target_drr: float = 0.12
    campaign_hard_drr: float = 0.16
    category_target_drr: float = 0.10
    category_soft_drr: float = 0.14
    max_up_step: float = 0.08
    test_up_step: float = 0.05
    down_step: float = 0.08
    hard_down_step: float = 0.15
    settle_days: int = 2
    eval_days: int = 5
    orders_growth_per_1pp_over_12: float = 4.0
    roi_9016_target: float = 0.10


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description='TOPFACE WB Ads Manager')
    p.add_argument('mode', nargs='?', default='run', choices=['run'])
    p.add_argument('--local-data-dir', default=None)
    p.add_argument('--as-of-date', default=None)
    return p


def load_keywords(provider: BaseProvider) -> pd.DataFrame:
    keys = provider.list_keys(KEYWORDS_WEEKLY_PREFIX)
    frames = []
    for key in keys:
        try:
            xls = provider.read_excel_all_sheets(key)
            sheet = xls.get('Позиции по Ключам', next(iter(xls.values())))
            df = sheet.copy()
            if df.empty:
                continue
            df = df.rename(columns={
                'Дата':'date',
                'Артикул WB':'nmId',
                'Артикул продавца':'supplier_article',
                'Предмет':'subject',
                'Рейтинг отзывов':'rating_reviews',
                'Рейтинг карточки':'rating_card',
                'Частота запросов':'query_freq',
                'Частота за неделю':'demand_week',
                'Медианная позиция':'median_position',
                'Переходы в карточку':'clicks_to_card',
                'Добавления в корзину':'keyword_add_to_cart',
                'Заказы':'keyword_orders',
                'Конверсия в заказ %':'keyword_conversion',
                'Конверсия в корзину %':'keyword_cart_conversion',
                'Видимость %':'visibility_pct',
                'Поисковый запрос':'query_text',
                'Фильтр':'query_filter',
            })
            df['date'] = parse_date_col(df['date'])
            df['subject_norm'] = df['subject'].map(canonical_subject)
            df = df[df['subject_norm'].isin(TARGET_SUBJECTS)].copy()
            df['product_root'] = df['supplier_article'].map(product_root_from_supplier_article)
            for c in ['query_freq','demand_week','median_position','clicks_to_card','keyword_add_to_cart','keyword_orders','keyword_conversion','keyword_cart_conversion','visibility_pct','rating_reviews','rating_card']:
                if c not in df.columns:
                    df[c] = 0
                df[c] = df[c].map(safe_float)
            df['query_text'] = df.get('query_text', '').fillna('').astype(str)
            df['query_filter'] = df.get('query_filter', '').fillna('').astype(str)
            frames.append(df)
        except Exception:
            continue
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def _filter_priority(v: str) -> int:
    s = str(v).strip().lower()
    if s == 'opencard':
        return 0
    if s == 'addtocart':
        return 1
    if s == 'orders':
        return 2
    return 9


def dedupe_keyword_rows(keywords: pd.DataFrame) -> pd.DataFrame:
    if keywords.empty:
        return pd.DataFrame()
    df = keywords.copy()
    df['query_filter_priority'] = df['query_filter'].map(_filter_priority)
    sort_cols = ['date','supplier_article','query_text','query_filter_priority']
    if 'nmId' in df.columns:
        sort_cols.insert(2, 'nmId')
    df = df.sort_values(sort_cols)
    grp = ['date','supplier_article','query_text']
    if 'nmId' in df.columns:
        grp = ['date','nmId','supplier_article','query_text']
    df = df.drop_duplicates(grp, keep='first')
    return df.drop(columns=['query_filter_priority'], errors='ignore')


def build_query_traffic_coefficients(keywords: pd.DataFrame, as_of_date: date, cur_end: date) -> pd.DataFrame:
    if keywords.empty:
        return pd.DataFrame(columns=['supplier_article','traffic_coeff','query_freq_cur','query_freq_prev_scaled'])
    df = dedupe_keyword_rows(keywords)
    if df.empty:
        return pd.DataFrame(columns=['supplier_article','traffic_coeff','query_freq_cur','query_freq_prev_scaled'])

    prev_start, prev_end = _previous_month_bounds(as_of_date)
    cur_start = as_of_date.replace(day=1)
    elapsed_days = max(1, (cur_end - cur_start).days + 1)
    prev_days = _days_in_month(prev_start)

    prev_df = df[(df['date'] >= prev_start) & (df['date'] <= prev_end)].copy()
    cur_df = df[(df['date'] >= cur_start) & (df['date'] <= cur_end)].copy()

    prev = prev_df.groupby('supplier_article', as_index=False).agg(query_freq_prev=('query_freq','sum'), prev_days_seen=('date','nunique'))
    cur = cur_df.groupby('supplier_article', as_index=False).agg(query_freq_cur=('query_freq','sum'), cur_days_seen=('date','nunique'))

    out = prev.merge(cur, on='supplier_article', how='outer').fillna(0)
    out['prev_days_seen'] = out['prev_days_seen'].replace(0, prev_days)
    out['cur_days_seen'] = out['cur_days_seen'].replace(0, elapsed_days)
    out['prev_daily_avg'] = out['query_freq_prev'] / out['prev_days_seen']
    out['cur_scaled'] = (out['query_freq_cur'] / out['cur_days_seen']) * elapsed_days
    out['query_freq_prev_scaled'] = out['prev_daily_avg'] * elapsed_days
    out['traffic_coeff'] = np.where(out['query_freq_prev_scaled'] > 0, out['cur_scaled'] / out['query_freq_prev_scaled'], 1.0)
    out['traffic_coeff'] = out['traffic_coeff'].replace([np.inf, -np.inf], np.nan).fillna(1.0).clip(lower=0.5, upper=1.5)
    return out[['supplier_article','traffic_coeff','cur_scaled','query_freq_prev_scaled']].rename(columns={'cur_scaled':'query_freq_cur'})



def build_previous_month_plan(orders: pd.DataFrame, funnel: pd.DataFrame, ads_daily: pd.DataFrame, keywords: pd.DataFrame, econ: pd.DataFrame, as_of_date: date, master: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    prev_start, prev_end = _previous_month_bounds(as_of_date)
    cur_start = as_of_date.replace(day=1)
    cur_end = as_of_date - timedelta(days=MATURE_END_OFFSET)
    elapsed_days = max(1, (cur_end - cur_start).days + 1)
    prev_days = _days_in_month(prev_start)
    window = compute_analysis_window(as_of_date)

    ord_prev = orders[(orders['date'] >= prev_start) & (orders['date'] <= prev_end) & (~orders['isCancel'])].copy() if not orders.empty else pd.DataFrame()
    ord_cur = orders[(orders['date'] >= cur_start) & (orders['date'] <= cur_end) & (~orders['isCancel'])].copy() if not orders.empty else pd.DataFrame()

    funnel_prev = funnel[(funnel['date'] >= prev_start) & (funnel['date'] <= prev_end)].copy() if not funnel.empty else pd.DataFrame()
    funnel_cur = funnel[(funnel['date'] >= cur_start) & (funnel['date'] <= cur_end)].copy() if not funnel.empty else pd.DataFrame()
    funnel_window = funnel[(funnel['date'] >= window['cur_start']) & (funnel['date'] <= window['cur_end'])].copy() if not funnel.empty else pd.DataFrame()

    ads_prev = ads_daily[(ads_daily['date'] >= prev_start) & (ads_daily['date'] <= prev_end)].copy() if not ads_daily.empty else pd.DataFrame()
    ads_cur = ads_daily[(ads_daily['date'] >= cur_start) & (ads_daily['date'] <= cur_end)].copy() if not ads_daily.empty else pd.DataFrame()
    ads_window = ads_daily[(ads_daily['date'] >= window['cur_start']) & (ads_daily['date'] <= window['cur_end'])].copy() if not ads_daily.empty else pd.DataFrame()

    econ_latest = latest_econ_rows(econ, ['nmId', 'supplier_article', 'np_unit']).copy() if not econ.empty else pd.DataFrame(columns=['nmId', 'supplier_article', 'np_unit'])
    key_map = master[['nmId', 'supplier_article', 'product_root', 'subject', 'subject_norm']].drop_duplicates().copy() if not master.empty else pd.DataFrame(columns=['nmId', 'supplier_article', 'product_root', 'subject', 'subject_norm'])

    if ord_prev.empty:
        empty = pd.DataFrame([{'Комментарий': 'Нет данных для плана прошлого месяца'}])
        empty_cat = pd.DataFrame([{'Комментарий': 'Нет данных для расчёта категорий'}])
        return empty, empty_cat

    prev_item = ord_prev.groupby(['nmId', 'supplier_article'], as_index=False).agg(
        orders_prev=('nmId', 'count'),
        revenue_prev=('finishedPrice', 'sum'),
    )
    cur_item = ord_cur.groupby(['nmId', 'supplier_article'], as_index=False).agg(
        orders_cur=('nmId', 'count'),
        revenue_cur=('finishedPrice', 'sum'),
    ) if not ord_cur.empty else pd.DataFrame(columns=['nmId', 'supplier_article', 'orders_cur', 'revenue_cur'])

    prev_buyout = funnel_prev.groupby('nmId', as_index=False).agg(buyout_prev=('buyoutPercent', 'mean')) if not funnel_prev.empty else pd.DataFrame(columns=['nmId', 'buyout_prev'])
    cur_buyout = funnel_cur.groupby('nmId', as_index=False).agg(
        buyout_cur=('buyoutPercent', 'mean'),
        atc_cur=('addToCartConversion', 'mean'),
        ord_conv_cur=('cartToOrderConversion', 'mean'),
    ) if not funnel_cur.empty else pd.DataFrame(columns=['nmId', 'buyout_cur', 'atc_cur', 'ord_conv_cur'])

    prev_spend = ads_prev.groupby('nmId', as_index=False).agg(ad_spend_prev=('Расход', 'sum')) if not ads_prev.empty else pd.DataFrame(columns=['nmId', 'ad_spend_prev'])
    cur_spend = ads_cur.groupby('nmId', as_index=False).agg(ad_spend_cur=('Расход', 'sum')) if not ads_cur.empty else pd.DataFrame(columns=['nmId', 'ad_spend_cur'])

    traffic = build_query_traffic_coefficients(keywords, as_of_date, cur_end)

    plan = (
        prev_item
        .merge(prev_buyout, on='nmId', how='left')
        .merge(prev_spend, on='nmId', how='left')
        .merge(cur_item, on=['nmId', 'supplier_article'], how='left')
        .merge(cur_buyout, on='nmId', how='left')
        .merge(cur_spend, on='nmId', how='left')
        .merge(key_map, on=['nmId', 'supplier_article'], how='left')
        .merge(econ_latest, on='nmId', how='left', suffixes=('', '_econ'))
        .merge(traffic, on='supplier_article', how='left')
    )
    subject_buyout_default = plan.get('subject_norm', pd.Series(index=plan.index, dtype=object)).map(get_subject_buyout_rate).fillna(0.90)
    for c in ['buyout_prev', 'buyout_cur']:
        plan[c] = to_buyout_rate(plan.get(c, pd.Series(index=plan.index, dtype=float)), default=np.nan)
        plan[c] = plan[c].where(plan[c].notna(), subject_buyout_default)
    for c in ['ad_spend_prev', 'ad_spend_cur', 'orders_cur', 'revenue_cur', 'np_unit', 'traffic_coeff', 'query_freq_cur', 'query_freq_prev_scaled', 'atc_cur', 'ord_conv_cur']:
        if c not in plan.columns:
            plan[c] = 0.0
        plan[c] = pd.to_numeric(plan[c], errors='coerce').fillna(0.0)

    plan['traffic_coeff'] = plan['traffic_coeff'].replace(0, np.nan).fillna(1.0).clip(0.5, 1.5)
    plan['vp_after_ads_prev_month'] = plan['revenue_prev'] * plan['buyout_prev'] - plan['ad_spend_prev']
    plan['daily_plan_vp_after_ads'] = plan['vp_after_ads_prev_month'] / prev_days
    plan['plan_vp_after_ads_mtd'] = plan['daily_plan_vp_after_ads'] * elapsed_days * plan['traffic_coeff']
    plan['fact_vp_after_ads_mtd'] = plan['revenue_cur'] * plan['buyout_cur'] - plan['ad_spend_cur']
    plan['plan_orders_mtd'] = plan['orders_prev'] / prev_days * elapsed_days * plan['traffic_coeff']
    plan['buyout_revenue_mtd'] = plan['revenue_cur'] * plan['buyout_cur']
    plan['fact_drr_mtd'] = np.where(plan['buyout_revenue_mtd'] > 0, plan['ad_spend_cur'] / plan['buyout_revenue_mtd'], 0.0)
    plan['forecast_net_profit_mtd'] = plan['orders_cur'] * plan['buyout_cur'] * plan['np_unit']
    plan['plan_attainment_pct'] = np.where(plan['plan_vp_after_ads_mtd'] != 0, plan['fact_vp_after_ads_mtd'] / plan['plan_vp_after_ads_mtd'] * 100.0, 0.0)
    plan['gp_gap_rub'] = plan['fact_vp_after_ads_mtd'] - plan['plan_vp_after_ads_mtd']
    plan['problem_reason'] = np.where(plan['gp_gap_rub'] < 0, 'Ниже плана', 'План выполняется')

    out = plan[[
        'nmId', 'supplier_article', 'subject', 'subject_norm', 'orders_prev', 'revenue_prev', 'buyout_prev', 'ad_spend_prev',
        'vp_after_ads_prev_month', 'daily_plan_vp_after_ads', 'traffic_coeff', 'query_freq_cur', 'query_freq_prev_scaled',
        'plan_vp_after_ads_mtd', 'fact_vp_after_ads_mtd', 'plan_attainment_pct', 'gp_gap_rub', 'orders_cur', 'revenue_cur',
        'buyout_cur', 'ad_spend_cur', 'fact_drr_mtd', 'forecast_net_profit_mtd', 'np_unit', 'atc_cur', 'ord_conv_cur', 'problem_reason'
    ]].copy()
    out = out.rename(columns={
        'nmId': 'Артикул WB',
        'supplier_article': 'Артикул продавца',
        'subject': 'Предмет',
        'orders_prev': 'Заказы прошлого месяца, шт',
        'revenue_prev': 'Сумма заказов прошлого месяца, ₽',
        'buyout_prev': '% выкупа прошлого месяца',
        'ad_spend_prev': 'Расходы РК прошлого месяца, ₽',
        'vp_after_ads_prev_month': 'ВП после рекламы прошлого месяца, ₽',
        'daily_plan_vp_after_ads': 'Дневной план ВП после рекламы, ₽',
        'traffic_coeff': 'Коэффициент трафика',
        'query_freq_cur': 'Частота запросов текущий период',
        'query_freq_prev_scaled': 'Ожидаемая частота по прошлому периоду',
        'plan_vp_after_ads_mtd': 'План ВП MTD, ₽',
        'fact_vp_after_ads_mtd': 'Факт ВП MTD, ₽',
        'plan_attainment_pct': 'Темп плана ВП, %',
        'gp_gap_rub': 'Отклонение ВП от плана, ₽',
        'orders_cur': 'Заказы MTD, шт',
        'revenue_cur': 'Сумма заказов MTD, ₽',
        'buyout_cur': '% выкупа MTD',
        'ad_spend_cur': 'Расходы РК MTD, ₽',
        'fact_drr_mtd': 'ДРР MTD, доля',
        'forecast_net_profit_mtd': 'Прогнозная чистая прибыль MTD, ₽',
        'np_unit': 'Чистая прибыль на 1 товар, ₽',
        'atc_cur': 'Конверсия в корзину, %',
        'ord_conv_cur': 'Конверсия в заказ, %',
        'problem_reason': 'Причина плана',
    })

    plan_subject = out.groupby('subject_norm', as_index=False).agg(
        **{
            'План ВП MTD, ₽': ('План ВП MTD, ₽', 'sum'),
            'Факт ВП MTD, ₽': ('Факт ВП MTD, ₽', 'sum'),
            'Отклонение ВП от плана, ₽': ('Отклонение ВП от плана, ₽', 'sum'),
            'Расходы РК MTD, ₽': ('Расходы РК MTD, ₽', 'sum'),
            'Сумма заказов MTD, ₽': ('Сумма заказов MTD, ₽', 'sum'),
        }
    ) if not out.empty else pd.DataFrame(columns=['subject_norm', 'План ВП MTD, ₽', 'Факт ВП MTD, ₽', 'Отклонение ВП от плана, ₽', 'Расходы РК MTD, ₽', 'Сумма заказов MTD, ₽'])
    if not plan_subject.empty:
        plan_subject['Темп плана ВП, %'] = np.where(plan_subject['План ВП MTD, ₽'] != 0, plan_subject['Факт ВП MTD, ₽'] / plan_subject['План ВП MTD, ₽'] * 100.0, 0.0)

    ads_window_subject = pd.DataFrame(columns=['subject_norm', 'Расходы рекламы окно, ₽'])
    if not ads_window.empty:
        ads_window_subject = (
            ads_window.merge(key_map[['nmId', 'subject_norm']].drop_duplicates(), on='nmId', how='left', suffixes=('', '_m'))
        )
        ads_window_subject['subject_norm'] = ads_window_subject.get('subject_norm', '').where(
            ads_window_subject.get('subject_norm', '').astype(str).str.len() > 0,
            ads_window_subject.get('subject_norm_m', '')
        )
        ads_window_subject['subject_norm'] = ads_window_subject['subject_norm'].map(canonical_subject)
        ads_window_subject['Расход'] = pd.to_numeric(ads_window_subject.get('Расход'), errors='coerce').fillna(0.0)
        ads_window_subject = (
            ads_window_subject[ads_window_subject['subject_norm'].isin(TARGET_SUBJECTS)]
            .groupby('subject_norm', as_index=False)
            .agg(**{'Расходы рекламы окно, ₽': ('Расход', 'sum')})
        )

    funnel_sales_col = find_matching_column(funnel_window, FUNNEL_SALES_CANDIDATES)
    funnel_window_subject = pd.DataFrame(columns=['subject_norm', 'Продажи воронка окно, ₽', 'Выкупленная выручка окно, ₽'])
    if not funnel_window.empty:
        fw = funnel_window.merge(key_map[['nmId', 'subject_norm']].drop_duplicates(), on='nmId', how='left')
        fw['subject_norm'] = fw['subject_norm'].map(canonical_subject)
        fw['Продажи воронка строка, ₽'] = pd.to_numeric(fw.get(funnel_sales_col if funnel_sales_col else 'ordersSumRub', 0), errors='coerce').fillna(0.0)
        fw['buyout_rate_row'] = resolve_buyout_rate_from_funnel(fw, default=np.nan)
        fw['buyout_rate_row'] = fw['buyout_rate_row'].where(fw['buyout_rate_row'].notna(), fw['subject_norm'].map(get_subject_buyout_rate).fillna(0.90))
        fw['Выкупленная выручка строка, ₽'] = fw['Продажи воронка строка, ₽'] * fw['buyout_rate_row']
        fw = fw[fw['subject_norm'].isin(TARGET_SUBJECTS)].copy()
        if not fw.empty:
            funnel_window_subject = fw.groupby('subject_norm', as_index=False).agg(
                **{
                    'Продажи воронка окно, ₽': ('Продажи воронка строка, ₽', 'sum'),
                    'Выкупленная выручка окно, ₽': ('Выкупленная выручка строка, ₽', 'sum'),
                }
            )

    category_out = pd.DataFrame({'subject_norm': sorted(TARGET_SUBJECTS)})
    if not plan_subject.empty:
        category_out = category_out.merge(plan_subject, on='subject_norm', how='left')
    if not ads_window_subject.empty:
        category_out = category_out.merge(ads_window_subject, on='subject_norm', how='left')
    if not funnel_window_subject.empty:
        category_out = category_out.merge(funnel_window_subject, on='subject_norm', how='left')

    for col in ['План ВП MTD, ₽', 'Факт ВП MTD, ₽', 'Отклонение ВП от плана, ₽', 'Расходы РК MTD, ₽', 'Сумма заказов MTD, ₽', 'Темп плана ВП, %', 'Расходы рекламы окно, ₽', 'Продажи воронка окно, ₽', 'Выкупленная выручка окно, ₽']:
        if col not in category_out.columns:
            category_out[col] = 0.0
        category_out[col] = pd.to_numeric(category_out[col], errors='coerce').fillna(0.0)

    category_out['Категория'] = category_out['subject_norm'].map(get_subject_display_name)
    category_out['Лимит ДРР категории, доля'] = category_out['subject_norm'].map(get_category_drr_limit)
    category_out['Лимит ДРР категории, %'] = category_out['Лимит ДРР категории, доля'] * 100.0
    category_out['Рабочее окно с'] = window['cur_start']
    category_out['Рабочее окно по'] = window['cur_end']
    category_out['Источник продаж для ДРР'] = funnel_sales_col if funnel_sales_col else 'ordersSumRub'
    category_out['ДРР категории, доля'] = np.where(category_out['Выкупленная выручка окно, ₽'] > 0, category_out['Расходы рекламы окно, ₽'] / category_out['Выкупленная выручка окно, ₽'], 0.0)
    category_out['Комментарий'] = np.where(category_out['Выкупленная выручка окно, ₽'] > 0, '', 'Проверь привязку nmId -> предмет в воронке')
    category_out = category_out[[
        'Категория', 'subject_norm', 'Лимит ДРР категории, доля', 'Лимит ДРР категории, %', 'Рабочее окно с', 'Рабочее окно по',
        'Расходы рекламы окно, ₽', 'Продажи воронка окно, ₽', 'Выкупленная выручка окно, ₽', 'ДРР категории, доля',
        'План ВП MTD, ₽', 'Факт ВП MTD, ₽', 'Отклонение ВП от плана, ₽', 'Темп плана ВП, %', 'Расходы РК MTD, ₽',
        'Сумма заказов MTD, ₽', 'Источник продаж для ДРР', 'Комментарий'
    ]].sort_values('Категория')

    return out.sort_values(['Предмет', 'Артикул продавца']), category_out



def build_daily_item_history(orders: pd.DataFrame, ads_daily: pd.DataFrame, funnel: pd.DataFrame, econ: pd.DataFrame, master: pd.DataFrame) -> pd.DataFrame:
    if orders.empty and funnel.empty and ads_daily.empty:
        return pd.DataFrame([{'Комментарий': 'Нет данных для истории товара'}])

    key_map = master[['nmId', 'supplier_article', 'product_root', 'subject', 'subject_norm']].drop_duplicates() if not master.empty else pd.DataFrame(columns=['nmId', 'supplier_article', 'product_root', 'subject', 'subject_norm'])
    econ_latest = latest_econ_rows(econ, ['nmId', 'gp_realized', 'np_unit']).copy() if not econ.empty else pd.DataFrame(columns=['nmId', 'gp_realized', 'np_unit'])

    ords = orders[~orders['isCancel']].copy() if not orders.empty else pd.DataFrame(columns=['date', 'nmId', 'supplier_article', 'finishedPrice'])
    item_orders = ords.groupby(['date', 'nmId', 'supplier_article'], as_index=False).agg(Заказы=('nmId', 'count'), Сумма_заказов=('finishedPrice', 'sum')) if not ords.empty else pd.DataFrame(columns=['date', 'nmId', 'supplier_article', 'Заказы', 'Сумма_заказов'])

    funnel_sales_col = find_matching_column(funnel, FUNNEL_SALES_CANDIDATES)
    if not funnel.empty:
        f = funnel.merge(key_map[['nmId', 'supplier_article', 'subject', 'subject_norm']].drop_duplicates(), on='nmId', how='left')
        f['Продажи_воронки'] = pd.to_numeric(f.get(funnel_sales_col if funnel_sales_col else 'ordersSumRub', 0), errors='coerce').fillna(0.0)
        f['buyout_rate'] = f['subject_norm'].map(get_subject_buyout_rate)
        funnel_item = f.groupby(['date', 'nmId', 'supplier_article'], as_index=False).agg(
            Продажи_воронки=('Продажи_воронки', 'sum'),
            buyout_rate=('buyout_rate', 'mean'),
            addToCartConversion=('addToCartConversion', 'mean'),
            cartToOrderConversion=('cartToOrderConversion', 'mean'),
        )
    else:
        funnel_item = pd.DataFrame(columns=['date', 'nmId', 'supplier_article', 'Продажи_воронки', 'buyout_rate', 'addToCartConversion', 'cartToOrderConversion'])

    base = item_orders.merge(funnel_item, on=['date', 'nmId', 'supplier_article'], how='outer')
    if base.empty:
        return pd.DataFrame([{'Комментарий': 'Нет данных для истории товара'}])

    base = base.merge(key_map, on=['nmId', 'supplier_article'], how='left')
    base['Товар'] = base['supplier_article']

    ad = ads_daily.groupby(['date', 'nmId'], as_index=False).agg(
        Расходы_РК=('Расход', 'sum'), Клики=('Клики', 'sum'), Показы=('Показы', 'sum'), Заказы_РК=('Заказы', 'sum'), Выручка_РК=('Сумма заказов', 'sum')
    ) if not ads_daily.empty else pd.DataFrame(columns=['date', 'nmId', 'Расходы_РК', 'Клики', 'Показы', 'Заказы_РК', 'Выручка_РК'])

    kw = dedupe_keyword_rows(keywords_global_for_history) if 'keywords_global_for_history' in globals() and isinstance(globals().get('keywords_global_for_history'), pd.DataFrame) else pd.DataFrame()
    if not kw.empty:
        kw = kw.groupby(['date', 'nmId'], as_index=False).agg(
            query_freq=('query_freq', 'sum'), keyword_orders=('keyword_orders', 'sum'), median_position=('median_position', 'median'), visibility_pct=('visibility_pct', 'mean'), demand_week=('demand_week', 'sum')
        )
    else:
        kw = pd.DataFrame(columns=['date', 'nmId', 'query_freq', 'keyword_orders', 'median_position', 'visibility_pct', 'demand_week'])

    out = base.merge(ad, on=['date', 'nmId'], how='left').merge(econ_latest, on='nmId', how='left').merge(kw, on=['date', 'nmId'], how='left')
    for c in ['Заказы', 'Сумма_заказов', 'Продажи_воронки', 'buyout_rate', 'addToCartConversion', 'cartToOrderConversion', 'Расходы_РК', 'Клики', 'Показы', 'Заказы_РК', 'Выручка_РК', 'gp_realized', 'np_unit', 'query_freq', 'median_position', 'visibility_pct', 'demand_week']:
        if c not in out.columns:
            out[c] = 0.0
        out[c] = pd.to_numeric(out[c], errors='coerce').fillna(0.0)

    out['buyout_rate'] = np.where(out['buyout_rate'] > 0, out['buyout_rate'], out['subject_norm'].map(get_subject_buyout_rate)).clip(0.0, 1.0)
    out['Выручка_для_расчёта'] = np.where(out['Продажи_воронки'] > 0, out['Продажи_воронки'], out['Сумма_заказов'])
    out['Валовая прибыль после рекламы, ₽'] = out['Заказы'] * out['gp_realized'] - out['Расходы_РК']
    out['ДРР, доля'] = np.where(out['Выручка_для_расчёта'] * out['buyout_rate'] > 0, out['Расходы_РК'] / (out['Выручка_для_расчёта'] * out['buyout_rate']), 0.0)
    out['CPO, ₽'] = np.where(out['Заказы_РК'] > 0, out['Расходы_РК'] / out['Заказы_РК'], 0.0)
    out['Прогнозная чистая прибыль, ₽'] = out['Заказы'] * out['buyout_rate'] * out['np_unit'] - out['Расходы_РК']
    out['День зрелый'] = pd.to_datetime(out['date']).dt.date <= (datetime.now().date() - timedelta(days=MATURE_END_OFFSET))

    if 'subject' not in out.columns:
        out['subject'] = out.get('subject_norm', '')

    out = out.rename(columns={
        'date': 'Дата', 'nmId': 'Артикул WB', 'supplier_article': 'Артикул продавца', 'subject': 'Предмет', 'query_freq': 'Частота запросов',
        'median_position': 'Медианная позиция', 'visibility_pct': 'Видимость, %', 'demand_week': 'Спрос по ключам',
        'addToCartConversion': 'Конверсия в корзину, %', 'cartToOrderConversion': 'Конверсия в заказ, %', 'buyout_rate': '% выкупа',
    })
    wanted = ['Дата', 'Товар', 'Артикул WB', 'Артикул продавца', 'Предмет', 'Заказы', 'Сумма_заказов', '% выкупа', 'Расходы_РК', 'Валовая прибыль после рекламы, ₽', 'Клики', 'CPO, ₽', 'Прогнозная чистая прибыль, ₽', 'Показы', 'Заказы_РК', 'Выручка_РК', 'ДРР, доля', 'Конверсия в корзину, %', 'Конверсия в заказ, %', 'Частота запросов', 'Спрос по ключам', 'Медианная позиция', 'Видимость, %', 'День зрелый']
    return trim_to_columns(out, wanted)


def build_daily_campaign_history(ads_daily: pd.DataFrame, campaigns: pd.DataFrame, funnel: pd.DataFrame, econ: pd.DataFrame, master: pd.DataFrame) -> pd.DataFrame:
    if ads_daily.empty:
        return pd.DataFrame([{'Комментарий': 'Нет данных рекламы'}])

    meta = campaigns[['id_campaign', 'nmId', 'placement', 'payment_type', 'current_bid_rub', 'campaign_status']].drop_duplicates()
    key_map = master[['nmId', 'supplier_article', 'subject', 'subject_norm']].drop_duplicates()
    econ_latest = latest_econ_rows(econ, ['nmId', 'gp_realized', 'np_unit']).copy() if not econ.empty else pd.DataFrame(columns=['nmId', 'gp_realized', 'np_unit'])
    df = ads_daily.groupby(['date', 'id_campaign', 'nmId'], as_index=False).agg(
        Показы=('Показы', 'sum'), Клики=('Клики', 'sum'), Заказы_РК=('Заказы', 'sum'), Расходы_РК=('Расход', 'sum'), Выручка_РК=('Сумма заказов', 'sum')
    )

    kw = dedupe_keyword_rows(keywords_global_for_history) if 'keywords_global_for_history' in globals() and isinstance(globals().get('keywords_global_for_history'), pd.DataFrame) else pd.DataFrame()
    if not kw.empty:
        kw = kw.groupby(['date', 'nmId'], as_index=False).agg(demand_week=('demand_week', 'sum'), query_freq=('query_freq', 'sum'))
    else:
        kw = pd.DataFrame(columns=['date', 'nmId', 'demand_week', 'query_freq'])

    df = df.merge(meta, on=['id_campaign', 'nmId'], how='left').merge(key_map, on='nmId', how='left').merge(econ_latest, on='nmId', how='left').merge(kw, on=['date', 'nmId'], how='left')
    df['buyout_rate'] = df['subject_norm'].map(get_subject_buyout_rate)
    df['gp_realized'] = pd.to_numeric(df.get('gp_realized'), errors='coerce').fillna(0.0)
    df['np_unit'] = pd.to_numeric(df.get('np_unit'), errors='coerce').fillna(0.0)
    df['demand_week'] = pd.to_numeric(df.get('demand_week'), errors='coerce').fillna(0.0)
    df['query_freq'] = pd.to_numeric(df.get('query_freq'), errors='coerce').fillna(0.0)
    df['Выручка_РК'] = pd.to_numeric(df.get('Выручка_РК'), errors='coerce').fillna(0.0)

    df['ДРР кампании, доля'] = np.where(df['Выручка_РК'] * df['buyout_rate'] > 0, df['Расходы_РК'] / (df['Выручка_РК'] * df['buyout_rate']), 0.0)
    df['ВП кампании после рекламы, ₽'] = df['Заказы_РК'] * df['gp_realized'] - df['Расходы_РК']
    df['ROI кампании'] = np.where(df['Расходы_РК'] > 0, ((df['Заказы_РК'] * df['buyout_rate'] * df['np_unit']) - df['Расходы_РК']) / df['Расходы_РК'], 0.0)
    df['CPO кампании, ₽'] = np.where(df['Заказы_РК'] > 0, df['Расходы_РК'] / df['Заказы_РК'], 0.0)
    df['Тип'] = np.where(df['payment_type'].astype(str).str.lower().eq('cpc'), 'Поиск', 'Полки')
    df['Тип кампании'] = np.where(df['payment_type'].astype(str).str.lower().eq('cpc'), 'CPC', 'CPM')
    df['День зрелый'] = pd.to_datetime(df['date']).dt.date <= (datetime.now().date() - timedelta(days=MATURE_END_OFFSET))

    if 'subject' not in df.columns:
        df['subject'] = df.get('subject_norm', '')

    df = df.rename(columns={
        'date': 'Дата', 'nmId': 'Артикул WB', 'supplier_article': 'Артикул продавца', 'subject': 'Предмет', 'id_campaign': 'ID кампании',
        'current_bid_rub': 'Ставка, ₽', 'placement': 'Плейсмент', 'demand_week': 'Спрос по ключам', 'query_freq': 'Частота запросов',
    })
    wanted = ['Дата', 'ID кампании', 'Артикул WB', 'Артикул продавца', 'Предмет', 'Тип кампании', 'Тип', 'Плейсмент', 'Ставка, ₽', 'Показы', 'Клики', 'Заказы_РК', 'Выручка_РК', 'Расходы_РК', 'ДРР кампании, доля', 'ВП кампании после рекламы, ₽', 'ROI кампании', 'CPO кампании, ₽', 'Спрос по ключам', 'Частота запросов', 'День зрелый']
    return trim_to_columns(df, wanted)


def build_channel_balance(ads_daily: pd.DataFrame, campaigns: pd.DataFrame, master: pd.DataFrame, econ_latest: pd.DataFrame, window: Dict[str, date], funnel: Optional[pd.DataFrame]=None) -> pd.DataFrame:
    if ads_daily.empty or campaigns.empty:
        return pd.DataFrame(columns=['supplier_article'])
    meta_cols = [c for c in ['id_campaign', 'nmId', 'payment_type'] if c in campaigns.columns]
    meta = campaigns[meta_cols].drop_duplicates().copy()
    key_cols = [c for c in ['nmId', 'supplier_article', 'subject_norm', 'subject', 'Предмет', 'Название предмета'] if c in master.columns]
    keys = master[key_cols].drop_duplicates().copy() if key_cols else pd.DataFrame(columns=['nmId'])
    npu_cols = [c for c in ['nmId', 'gp_realized', 'np_unit'] if c in econ_latest.columns]
    npu = econ_latest[npu_cols].drop_duplicates().copy() if npu_cols else pd.DataFrame(columns=['nmId'])

    df = ads_daily[(ads_daily['date'] >= window['cur_start']) & (ads_daily['date'] <= window['cur_end'])].copy()
    if not meta.empty:
        merge_on = [c for c in ['id_campaign', 'nmId'] if c in df.columns and c in meta.columns]
        if merge_on:
            df = df.merge(meta, on=merge_on, how='left')
    if not keys.empty and 'nmId' in df.columns and 'nmId' in keys.columns:
        df = df.merge(keys, on='nmId', how='left')
    if not npu.empty and 'nmId' in df.columns and 'nmId' in npu.columns:
        df = df.merge(npu, on='nmId', how='left')

    nm_to_subject = build_nm_to_subject_map(master)
    df = with_resolved_subject_norm(df, nm_to_subject)

    if 'supplier_article' not in df.columns:
        df['supplier_article'] = ''
    df['supplier_article'] = df['supplier_article'].fillna('').astype(str)
    if 'subject' not in df.columns:
        df['subject'] = df.get('subject_norm', '')

    df['buyout_rate'] = pd.to_numeric(df['subject_norm'].map(get_subject_buyout_rate), errors='coerce').fillna(0.0)
    df['payment_type'] = df.get('payment_type', '').astype(str)
    df['channel'] = np.where(df['payment_type'].str.lower().eq('cpc'), 'CPC', 'CPM')

    for col in ['Расход', 'Заказы', 'Сумма заказов', 'Клики', 'Показы', 'np_unit']:
        if col not in df.columns:
            df[col] = 0.0
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)

    grp = df.groupby(['supplier_article', 'channel'], as_index=False).agg(
        spend=('Расход', 'sum'),
        orders=('Заказы', 'sum'),
        revenue=('Сумма заказов', 'sum'),
        clicks=('Клики', 'sum'),
        shows=('Показы', 'sum'),
        buyout_rate=('buyout_rate', 'mean'),
        np_unit=('np_unit', 'mean')
    )
    grp['cpo'] = np.where(grp['orders'] > 0, grp['spend'] / grp['orders'], 0.0)
    grp['drr'] = np.where(grp['revenue'] * grp['buyout_rate'] > 0, grp['spend'] / (grp['revenue'] * grp['buyout_rate']), 0.0)
    grp['gp_after_ads'] = grp['revenue'] * grp['buyout_rate'] - grp['spend']

    if grp.empty:
        return pd.DataFrame(columns=['supplier_article'])

    wide = grp.pivot(index='supplier_article', columns='channel', values=['cpo', 'drr', 'gp_after_ads', 'orders', 'spend']).copy()
    wide.columns = [f'{a.lower()}_{b.lower()}' for a, b in wide.columns]
    wide = wide.reset_index()
    for c in ['cpo_cpc', 'cpo_cpm', 'drr_cpc', 'drr_cpm', 'gp_after_ads_cpc', 'gp_after_ads_cpm', 'orders_cpc', 'orders_cpm', 'spend_cpc', 'spend_cpm']:
        if c not in wide.columns:
            wide[c] = 0.0
        wide[c] = pd.to_numeric(wide[c], errors='coerce').fillna(0.0)

    def _better(r):
        cpc_score = (1 if (r['gp_after_ads_cpc'] > 0) else 0) + (1 if (r['cpo_cpc'] > 0 and (r['cpo_cpm'] == 0 or r['cpo_cpc'] <= r['cpo_cpm'])) else 0) + (1 if r['gp_after_ads_cpc'] >= r['gp_after_ads_cpm'] else 0)
        cpm_score = (1 if (r['gp_after_ads_cpm'] > 0) else 0) + (1 if (r['cpo_cpm'] > 0 and (r['cpo_cpc'] == 0 or r['cpo_cpm'] <= r['cpo_cpc'])) else 0) + (1 if r['gp_after_ads_cpm'] > r['gp_after_ads_cpc'] else 0)
        return 'CPC' if cpc_score >= cpm_score else 'CPM'

    wide['better_channel'] = wide.apply(_better, axis=1)
    wide['worse_channel'] = np.where(wide['better_channel'].eq('CPC'), 'CPM', 'CPC')
    return wide


def build_item_current_metrics(item_history: pd.DataFrame, cur_start: date, cur_end: date, base_start: date, base_end: date) -> pd.DataFrame:
    if item_history.empty or 'Комментарий' in item_history.columns:
        return pd.DataFrame(columns=['Артикул продавца'])
    ih = item_history.copy()
    ih['Дата'] = pd.to_datetime(ih['Дата']).dt.date
    cur = ih[(ih['Дата'] >= cur_start) & (ih['Дата'] <= cur_end)].groupby('Артикул продавца', as_index=False).agg(
        item_orders_cur=('Заказы','sum'), item_revenue_cur=('Сумма_заказов','sum'), item_spend_cur=('Расходы_РК','sum'), item_gp_cur=('Валовая прибыль после рекламы, ₽','sum'), item_clicks_cur=('Клики','sum'), item_cpo_cur=('CPO, ₽','mean'), item_drr_cur=('ДРР, доля','mean'), item_net_profit_cur=('Прогнозная чистая прибыль, ₽','sum'), item_atc_cur=('Конверсия в корзину, %','mean'), item_ord_conv_cur=('Конверсия в заказ, %','mean')
    )
    base = ih[(ih['Дата'] >= base_start) & (ih['Дата'] <= base_end)].groupby('Артикул продавца', as_index=False).agg(
        item_orders_base=('Заказы','sum'), item_gp_base=('Валовая прибыль после рекламы, ₽','sum')
    )
    out = cur.merge(base, on='Артикул продавца', how='left').fillna(0)
    out['item_order_growth_pct'] = np.where(out['item_orders_base']>0, (out['item_orders_cur']/out['item_orders_base']-1)*100.0, 0.0)
    out['item_gp_growth_pct'] = np.where(out['item_gp_base']!=0, (out['item_gp_cur']-out['item_gp_base'])/abs(out['item_gp_base'])*100.0, 0.0)
    return out


def determine_action(row: pd.Series, cfg: Config) -> Tuple[str, float, str, bool]:
    current_bid = safe_float(row.get('current_bid_rub'))
    max_bid = safe_float(row.get('max_bid_rub'))
    payment_type = canonical_payment_type(row.get('payment_type'))
    min_bid = safe_float(row.get('Минимальная ставка WB, ₽')) if pd.notna(row.get('Минимальная ставка WB, ₽')) else (4.0 if payment_type == 'cpc' else 80.0)
    floor_bid = max(min_bid, 4.0 if payment_type == 'cpc' else 80.0)
    step = get_bid_step_rub(payment_type)
    subject_norm = canonical_subject(row.get('subject_norm', row.get('subject', '')))
    subject_name = get_subject_display_name(subject_norm)

    category_drr = safe_float(row.get('category_drr_cur'))
    category_limit = safe_float(row.get('category_limit_drr')) or get_category_drr_limit(subject_norm)
    category_orders_growth = safe_float(row.get('category_orders_growth_pct'))
    category_gp_growth = safe_float(row.get('category_gp_growth_pct'))
    category_demand_growth = safe_float(row.get('category_demand_growth_pct'))
    category_plan_att = safe_float(row.get('category_plan_attainment_pct', row.get('plan_attainment_pct', 100.0)))

    campaign_drr = safe_float(row.get('campaign_drr_cur'))
    campaign_gp = safe_float(row.get('campaign_gp_cur'))
    campaign_gp_growth = safe_float(row.get('campaign_gp_growth_pct'))
    campaign_orders_growth = safe_float(row.get('campaign_order_growth_pct'))
    campaign_click_growth = safe_float(row.get('campaign_click_growth_pct'))
    campaign_impression_growth = safe_float(row.get('campaign_impression_growth_pct'))
    campaign_roi = safe_float(row.get('campaign_roi_cur'))

    item_gp_growth = safe_float(row.get('item_gp_growth_pct'))
    item_orders_growth = safe_float(row.get('item_order_growth_pct'))
    better_channel = str(row.get('better_channel', '') or '').strip().upper()
    row_channel = 'CPC' if payment_type == 'cpc' else 'CPM'
    campaign_is_active = bool(row.get('campaign_is_active', False))
    supplier_article = str(row.get('supplier_article', '') or '').strip()

    can_raise_more = (max_bid <= 0) or (current_bid + step <= max_bid + 1e-9)
    rate_limit = max_bid > 0 and current_bid >= max_bid - 1e-9
    positive_traffic = campaign_impression_growth > 0 or campaign_click_growth > 0
    positive_orders = campaign_orders_growth > 0
    positive_gp = campaign_gp_growth > 0
    trend_confirmed = positive_traffic and positive_orders and positive_gp
    doubtful_trend = (positive_traffic and not positive_orders) or (positive_orders and not positive_gp)
    demand_explains_drop = is_drop_explained_by_demand(category_orders_growth, category_demand_growth)

    def raise_bid() -> float:
        return apply_bid_step(current_bid, payment_type, 'up', floor_bid, max_bid)

    def lower_bid() -> float:
        return apply_bid_step(current_bid, payment_type, 'down', floor_bid, max_bid)

    level1 = f"Уровень 1 Категория: {subject_name}; план {category_plan_att:.0f}%, ДРР {category_drr*100:.1f}% при лимите {category_limit*100:.1f}%, заказы {category_orders_growth:.1f}%, ВП {category_gp_growth:.1f}%, спрос {category_demand_growth:.1f}%"
    level2 = f"Уровень 2 Кампания: ДРР {campaign_drr*100:.1f}%, ВП {campaign_gp:.0f} ₽, рост показов {campaign_impression_growth:.1f}%, кликов {campaign_click_growth:.1f}%, заказов {campaign_orders_growth:.1f}%, ВП {campaign_gp_growth:.1f}%"
    level3 = f"Уровень 3 Товар: рост заказов {item_orders_growth:.1f}%, рост чистой/валовой прибыли {item_gp_growth:.1f}%"

    if not campaign_is_active:
        return 'Без изменений', round(current_bid, 2), f"{level1}; {level2}; {level3}; кампания не активна — ставку не меняем", False

    if supplier_article in {'901_/6', '901/6'} and campaign_roi < cfg.roi_9016_target and current_bid > floor_bid:
        return 'Снизить', lower_bid(), f"{level1}; {level2}; {level3}; ROI 901/6 ниже целевого {cfg.roi_9016_target*100:.0f}%", True

    if campaign_gp <= 0 and current_bid > floor_bid:
        return 'Снизить', lower_bid(), f"{level1}; {level2}; {level3}; кампания убыточна — снижаем на 1 шаг ({step:.0f} ₽)", True

    if category_drr > category_limit:
        if campaign_drr > 0.10 and current_bid > floor_bid:
            return 'Снизить', lower_bid(), f"{level1}; {level2}; {level3}; категория выше лимита, эта кампания тоже >10% ДРР — снижаем на 1 шаг ({step:.0f} ₽)", True
        return 'Без изменений', round(current_bid, 2), f"{level1}; {level2}; {level3}; категория выше лимита, но у кампании ДРР <= 10% — не режем автоматически", False

    if category_orders_growth < 0 and not demand_explains_drop and category_drr <= category_limit:
        if campaign_drr <= 0.10 and campaign_gp > 0 and can_raise_more:
            return 'Повысить', raise_bid(), f"{level1}; {level2}; {level3}; заказы категории падают быстрее спроса — усиливаем эффективную кампанию <=10% ДРР на 1 шаг ({step:.0f} ₽)", False

    if category_orders_growth > 0 and category_gp_growth < 0:
        if campaign_drr > 0.10 and current_bid > floor_bid:
            return 'Снизить', lower_bid(), f"{level1}; {level2}; {level3}; заказы категории растут, а ВП падает — режем кампанию >10% ДРР на 1 шаг ({step:.0f} ₽)", True

    if campaign_drr <= category_limit and campaign_gp > 0:
        if doubtful_trend:
            return 'Без изменений', round(current_bid, 2), f"{level1}; {level2}; {level3}; трафик меняется, но заказ/ВП не подтверждены — ждём 3 дня", False
        if trend_confirmed and can_raise_more:
            return 'Повысить', raise_bid(), f"{level1}; {level2}; {level3}; кампания прибыльная и растёт — повышаем на 1 шаг ({step:.0f} ₽)", False
        if category_plan_att < 95 and item_gp_growth <= 0 and can_raise_more:
            return 'Повысить', raise_bid(), f"{level1}; {level2}; {level3}; отстаём от плана, ДРР в норме, прибыль товара не растёт — повышаем на 1 шаг ({step:.0f} ₽)", False
        if better_channel and better_channel != row_channel:
            return 'Без изменений', round(current_bid, 2), f"{level1}; {level2}; {level3}; альтернативный канал сильнее, но эта кампания прибыльная — не режем автоматически", False
        return 'Без изменений', round(current_bid, 2), f"{level1}; {level2}; {level3}; недостаточно сигнала для изменения", rate_limit

    if campaign_drr > category_limit:
        if (campaign_gp_growth <= 0 or campaign_orders_growth <= 0) and current_bid > floor_bid:
            return 'Снизить', lower_bid(), f"{level1}; {level2}; {level3}; ДРР кампании выше верхнего значения и рост не подтверждён — снижаем на 1 шаг ({step:.0f} ₽)", True
        return 'Без изменений', round(current_bid, 2), f"{level1}; {level2}; {level3}; ДРР кампании выше верхнего значения, но кампания пока прибыльна — ждём подтверждение", False

    if rate_limit:
        return 'Предел эффективности ставки', round(current_bid, 2), f"{level1}; {level2}; {level3}; ставка упёрлась в расчётный максимум", True

    return 'Без изменений', round(current_bid, 2), f"{level1}; {level2}; {level3}; без изменений", False


def prepare_metrics(provider: BaseProvider, cfg: Config, as_of_date: date) -> Dict[str, Any]:
    window = compute_analysis_window(as_of_date)
    log(f"📅 Анализируем зрелое окно {window['cur_start']} .. {window['cur_end']}; база сравнения {window['base_start']} .. {window['base_end']}")
    ads_daily, campaigns = load_ads(provider)
    econ = load_economics(provider)
    orders = load_orders(provider)
    funnel = load_funnel(provider)
    keywords = load_keywords(provider)
    globals()['keywords_global_for_history'] = keywords.copy()
    bid_history = load_bid_history(provider)
    master = build_master(econ, orders, keywords, campaigns)
    log(f"📣 Реклама: {len(ads_daily):,} строк; кампании: {campaigns['id_campaign'].nunique() if not campaigns.empty else 0}; placement-строк: {len(campaigns):,}")
    plan_df, category_plan = build_previous_month_plan(orders, funnel, ads_daily, keywords, econ, as_of_date, master)
    log(f"💰 Экономика: {len(econ):,} SKU; Заказы: {len(orders):,} строк; Воронка: {len(funnel):,}; Keywords: {len(keywords):,}; План строк: {0 if 'Комментарий' in plan_df.columns else len(plan_df):,}")

    daily_item = build_daily_item_history(orders, ads_daily, funnel, econ, master)
    daily_campaign = build_daily_campaign_history(ads_daily, campaigns, funnel, econ, master)

    econ_latest = latest_econ_rows(econ, ['nmId','supplier_article','product_root','subject','subject_norm','buyout_rate','gp_realized','np_unit']).copy() if not econ.empty else pd.DataFrame(columns=['nmId','supplier_article','product_root','subject','subject_norm','buyout_rate','gp_realized','np_unit'])
    category_diag = build_category_window_diagnostics(orders, ads_daily, funnel, keywords, master, econ_latest, window)
    keywords_current = aggregate_keyword_item(keywords, window['cur_start'], window['cur_end'])
    funnel_item, funnel_subject = build_funnel_item(funnel, master, window['cur_start'], window['cur_end'])

    campaigns_base = campaigns[['id_campaign','nmId','placement','payment_type','current_bid_rub','campaign_status']].drop_duplicates().merge(master[['nmId','supplier_article','product_root','subject','subject_norm','rating_reviews','rating_card']].drop_duplicates(), on='nmId', how='left').merge(econ_latest[['nmId','buyout_rate','gp_realized','np_unit']], on='nmId', how='left').drop_duplicates(['id_campaign','nmId','placement','payment_type'])
    campaigns_base['buyout_rate'] = campaigns_base['subject_norm'].map(get_subject_buyout_rate)

    cur = ads_daily[(ads_daily['date'] >= window['cur_start']) & (ads_daily['date'] <= window['cur_end'])].groupby(['id_campaign','nmId'], as_index=False).agg(
        Показы=('Показы','sum'), Клики=('Клики','sum'), Заказы=('Заказы','sum'), Расход=('Расход','sum'), Сумма_заказов=('Сумма заказов','sum')
    ) if not ads_daily.empty else pd.DataFrame(columns=['id_campaign','nmId','Показы','Клики','Заказы','Расход','Сумма_заказов'])
    base = ads_daily[(ads_daily['date'] >= window['base_start']) & (ads_daily['date'] <= window['base_end'])].groupby(['id_campaign','nmId'], as_index=False).agg(
        base_Показы=('Показы','sum'), base_Клики=('Клики','sum'), base_Заказы=('Заказы','sum'), base_Расход=('Расход','sum'), base_Сумма_заказов=('Сумма заказов','sum')
    ) if not ads_daily.empty else pd.DataFrame(columns=['id_campaign','nmId','base_Показы','base_Клики','base_Заказы','base_Расход','base_Сумма_заказов'])

    rows = campaigns_base.merge(cur, on=['id_campaign','nmId'], how='left').merge(base, on=['id_campaign','nmId'], how='left').fillna(0)
    rows = ensure_business_keys(rows)
    rows['supplier_article'] = series_or_default(rows, 'supplier_article', '').fillna('').astype(str)
    rows['subject'] = series_or_default(rows, 'subject', series_or_default(rows, 'subject_norm', '')).fillna('').astype(str)
    rows['subject_norm'] = series_or_default(rows, 'subject_norm', rows['subject'].map(canonical_subject)).fillna('').astype(str).map(canonical_subject)
    rows['campaign_status'] = series_or_default(rows, 'campaign_status', '').fillna('').astype(str)
    rows['campaign_is_active'] = rows['campaign_status'].map(is_active_campaign_status)

    rows = rows.merge(keywords_current[['nmId','supplier_article','demand_week','median_position','visibility_pct','keyword_orders','keyword_clicks']].drop_duplicates(), on=['nmId','supplier_article'], how='left')
    rows = rows.merge(funnel_item[['nmId','addToCartConversion','cartToOrderConversion','buyoutPercent']], on='nmId', how='left')

    if not daily_item.empty and 'Комментарий' not in daily_item.columns:
        daily_item = ensure_business_keys(daily_item)
        item_metrics = build_item_current_metrics(daily_item, window['cur_start'], window['cur_end'], window['base_start'], window['base_end'])
        item_metrics = ensure_business_keys(item_metrics)
        if 'Артикул продавца' in rows.columns and 'Артикул продавца' in item_metrics.columns:
            rows = rows.merge(item_metrics, on='Артикул продавца', how='left')
        elif 'supplier_article' in rows.columns and 'supplier_article' in item_metrics.columns:
            rows = rows.merge(item_metrics, on='supplier_article', how='left')

    if not plan_df.empty and 'Комментарий' not in plan_df.columns:
        plan_df = ensure_business_keys(plan_df)
        plan_cols = [c for c in ['Артикул WB','Артикул продавца','План ВП MTD, ₽','Факт ВП MTD, ₽','Темп плана ВП, %','Причина плана'] if c in plan_df.columns]
        rows = rows.merge(plan_df[plan_cols].drop_duplicates(), on=[c for c in ['Артикул WB','Артикул продавца'] if c in plan_cols], how='left')
    else:
        rows['План ВП MTD, ₽'] = 0.0
        rows['Факт ВП MTD, ₽'] = 0.0
        rows['Темп плана ВП, %'] = 0.0
        rows['Причина плана'] = ''

    rows['ctr_pct'] = np.where(rows['Показы'] > 0, rows['Клики'] / rows['Показы'] * 100.0, 0.0)
    rows['campaign_drr_cur'] = np.where(rows['Сумма_заказов'] * rows['buyout_rate'] > 0, rows['Расход'] / (rows['Сумма_заказов'] * rows['buyout_rate']), 0.0)
    rows['campaign_drr_base'] = np.where(rows['base_Сумма_заказов'] * rows['buyout_rate'] > 0, rows['base_Расход'] / (rows['base_Сумма_заказов'] * rows['buyout_rate']), 0.0)
    rows['campaign_gp_cur'] = rows['Заказы'] * rows['gp_realized'] - rows['Расход']
    rows['campaign_gp_base'] = rows['base_Заказы'] * rows['gp_realized'] - rows['base_Расход']
    rows['campaign_gp_growth_pct'] = np.where(rows['campaign_gp_base'] != 0, (rows['campaign_gp_cur'] - rows['campaign_gp_base']) / np.abs(rows['campaign_gp_base']) * 100.0, np.where(rows['campaign_gp_cur'] > 0, 100.0, 0.0))
    rows['campaign_order_growth_pct'] = np.where(rows['base_Заказы'] > 0, (rows['Заказы'] / rows['base_Заказы'] - 1) * 100.0, np.where(rows['Заказы'] > 0, 100.0, 0.0))
    rows['campaign_click_growth_pct'] = np.where(rows['base_Клики'] > 0, (rows['Клики'] / rows['base_Клики'] - 1) * 100.0, np.where(rows['Клики'] > 0, 100.0, 0.0))
    rows['campaign_impression_growth_pct'] = np.where(rows['base_Показы'] > 0, (rows['Показы'] / rows['base_Показы'] - 1) * 100.0, np.where(rows['Показы'] > 0, 100.0, 0.0))
    rows['campaign_cpo'] = np.where(rows['Заказы'] > 0, rows['Расход'] / rows['Заказы'], 0.0)
    rows['campaign_roi_cur'] = np.where(rows['Расход'] > 0, ((rows['Заказы'] * rows['buyout_rate'] * rows['np_unit']) - rows['Расход']) / rows['Расход'], 0.0)
    item_revenue_cur = numeric_series(rows, 'item_revenue_cur', 0.0)
    item_spend_cur = numeric_series(rows, 'item_spend_cur', 0.0)
    item_orders_cur = numeric_series(rows, 'item_orders_cur', 0.0)
    item_order_growth_pct = numeric_series(rows, 'item_order_growth_pct', 0.0)
    base_item_spend_cur = numeric_series(rows, 'base_item_spend_cur', 0.0)
    campaign_orders_cur = numeric_series(rows, 'Заказы', 0.0)
    add_to_cart_conv = numeric_series(rows, 'addToCartConversion', 0.0)
    subj_add_to_cart = numeric_series(rows, 'subj_addToCart', 0.0)
    cart_to_order_conv = numeric_series(rows, 'cartToOrderConversion', 0.0)
    subj_cart_to_order = numeric_series(rows, 'subj_cartToOrder', 0.0)

    rows['blended_drr'] = np.where(item_revenue_cur * rows['buyout_rate'] > 0, item_spend_cur / (item_revenue_cur * rows['buyout_rate']), 0.0)
    rows['total_orders'] = item_orders_cur
    rows['total_revenue'] = item_revenue_cur
    rows['ad_spend'] = item_spend_cur
    rows['ad_orders'] = campaign_orders_cur
    rows['order_growth_pct'] = item_order_growth_pct
    rows['spend_growth_pct'] = np.where(base_item_spend_cur > 0, (item_spend_cur / base_item_spend_cur.replace(0, np.nan) - 1.0) * 100.0, 0.0)
    rows['required_growth_pct'] = rows.apply(lambda r: compute_required_growth(safe_float(r.get('campaign_drr_cur')), safe_float(r.get('spend_growth_pct')), str(r.get('subject_norm',''))), axis=1)
    rows['card_issue'] = ((add_to_cart_conv < 0.5 * subj_add_to_cart) | (cart_to_order_conv < 0.5 * subj_cart_to_order)) if 'subj_addToCart' in rows.columns else False

    if category_plan is not None and not category_plan.empty and 'subject_norm' in category_plan.columns:
        cat_cols = [c for c in ['subject_norm','Факт ВП MTD, ₽','План ВП MTD, ₽','Темп плана ВП, %','Лимит ДРР категории, доля'] if c in category_plan.columns]
        cat_merge = category_plan[cat_cols].drop_duplicates('subject_norm').copy()
        cat_merge = cat_merge.rename(columns={
            'Факт ВП MTD, ₽': 'category_gp_mtd',
            'План ВП MTD, ₽': 'category_gp_plan',
            'Темп плана ВП, %': 'category_plan_attainment_pct',
            'Лимит ДРР категории, доля': 'category_limit_drr_plan',
        })
        rows = rows.merge(cat_merge, on='subject_norm', how='left')
    else:
        rows['category_gp_mtd'] = 0.0
        rows['category_gp_plan'] = 0.0
        rows['category_plan_attainment_pct'] = 100.0
        rows['category_limit_drr_plan'] = np.nan

    if category_diag is not None and not category_diag.empty:
        rows = rows.merge(category_diag[['subject_norm','category_drr_cur','category_gp_cur','category_gp_base','category_orders_cur','category_orders_base','category_orders_growth_pct','category_gp_growth_pct','category_demand_cur','category_demand_base','category_demand_growth_pct','category_limit_drr']], on='subject_norm', how='left')
    else:
        rows['category_drr_cur'] = 0.0
        rows['category_gp_cur'] = 0.0
        rows['category_gp_base'] = 0.0
        rows['category_orders_cur'] = 0.0
        rows['category_orders_base'] = 0.0
        rows['category_orders_growth_pct'] = 0.0
        rows['category_gp_growth_pct'] = 0.0
        rows['category_demand_cur'] = 0.0
        rows['category_demand_base'] = 0.0
        rows['category_demand_growth_pct'] = 0.0
        rows['category_limit_drr'] = rows['subject_norm'].map(get_category_drr_limit)

    rows['category_limit_drr'] = numeric_series(rows, 'category_limit_drr', np.nan).fillna(rows['subject_norm'].map(get_category_drr_limit))
    _category_gp_plan = numeric_series(rows, 'category_gp_plan', 0.0)
    _category_gp_mtd = numeric_series(rows, 'category_gp_mtd', 0.0)
    _category_plan_att_default = pd.Series(
        np.where(_category_gp_plan > 0, _category_gp_mtd / _category_gp_plan.replace(0, np.nan).fillna(1.0) * 100.0, 100.0),
        index=rows.index,
    )
    rows['category_plan_attainment_pct'] = numeric_series(rows, 'category_plan_attainment_pct', np.nan).fillna(_category_plan_att_default)

    category_plan = category_diag.copy() if category_diag is not None else pd.DataFrame(columns=['subject_norm'])
    if category_plan is not None and not category_plan.empty:
        category_plan['Категория'] = category_plan['subject_norm'].map(get_subject_display_name)
        category_plan['Лимит ДРР категории, доля'] = category_plan['category_limit_drr']
        category_plan['Лимит ДРР категории, %'] = category_plan['category_limit_drr'] * 100.0
        category_plan['ДРР категории, доля'] = category_plan['category_drr_cur']
        category_plan['ДРР категории, %'] = category_plan['category_drr_cur'] * 100.0
        category_plan['Факт ВП MTD, ₽'] = category_plan['category_gp_cur']
        category_plan['ВП базовое окно, ₽'] = category_plan['category_gp_base']
        category_plan['Темп категории, %'] = category_plan['category_gp_growth_pct']
        category_plan['Заказы окно'] = category_plan['category_orders_cur']
        category_plan['Заказы база'] = category_plan['category_orders_base']
        category_plan['Рост заказов окна, %'] = category_plan['category_orders_growth_pct']
        category_plan['Спрос окно'] = category_plan['category_demand_cur']
        category_plan['Спрос база'] = category_plan['category_demand_base']
        category_plan['Рост спроса окна, %'] = category_plan['category_demand_growth_pct']
        category_plan['Фиксированный % выкупа'] = category_plan['subject_norm'].map(lambda x: get_subject_buyout_rate(x) * 100.0)
        category_plan['Рабочее окно с'] = window['cur_start']
        category_plan['Рабочее окно по'] = window['cur_end']
        category_plan = category_plan[['Категория','subject_norm','Фиксированный % выкупа','Заказы окно','Заказы база','Рост заказов окна, %','Спрос окно','Спрос база','Рост спроса окна, %','Факт ВП MTD, ₽','ВП базовое окно, ₽','Темп категории, %','ДРР категории, доля','ДРР категории, %','Лимит ДРР категории, доля','Лимит ДРР категории, %','Рабочее окно с','Рабочее окно по']]

    rows = normalize_core_columns(rows)
    if 'supplier_article' not in rows.columns:
        rows['supplier_article'] = rows.get('Артикул продавца', '')
    rows['supplier_article'] = rows['supplier_article'].fillna('').astype(str)
    rows = rows.merge(build_channel_balance(ads_daily, campaigns, master, econ_latest, window, funnel), on='supplier_article', how='left')
    for c in ['cpo_cpc','cpo_cpm','drr_cpc','drr_cpm','gp_after_ads_cpc','gp_after_ads_cpm','orders_cpc','orders_cpm']:
        rows[c] = numeric_series(rows, c, 0.0)
    rows['plan_attainment_pct'] = numeric_series(rows, 'Темп плана ВП, %', 0.0)

    subject_benchmarks = build_subject_benchmarks(rows)
    rows = rows.merge(subject_benchmarks, on=['subject_norm','placement'], how='left')
    rows['capture_imp'] = np.where(rows['demand_week'] > 0, rows['Показы'] / rows['demand_week'], 0.0)
    rows['capture_click'] = np.where(rows['keyword_clicks'] > 0, rows['Клики'] / rows['keyword_clicks'], 0.0)
    bench_capture_imp = numeric_series(rows, 'bench_capture_imp', 0.0)
    rows['eff_index_imp'] = np.where(bench_capture_imp > 0, rows['capture_imp'] / bench_capture_imp.replace(0, np.nan), 1.0)
    bench_capture_click = numeric_series(rows, 'bench_capture_click', 0.0)
    rows['eff_index_click'] = np.where(bench_capture_click > 0, rows['capture_click'] / bench_capture_click.replace(0, np.nan), 1.0)

    limits = rows.apply(lambda r: pd.Series(compute_bid_limits(r, subject_benchmarks), index=['comfort_bid_rub','max_bid_rub','experiment_bid_rub','limit_type']), axis=1)
    rows = pd.concat([rows, limits], axis=1)
    rows['Причина лимита'] = rows.apply(explain_limit_reason, axis=1)

    rows = rows.sort_values(['id_campaign','nmId','placement']).drop_duplicates(['id_campaign','nmId','placement','payment_type'])

    decisions = []
    for _, r in rows.iterrows():
        action, new_bid, reason, rate_limit = determine_action(r, cfg)
        max_bid = safe_float(r.get('max_bid_rub'))
        if max_bid > 0 and new_bid > max_bid:
            new_bid = max_bid
        decisions.append({
            'Дата запуска': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'ID кампании': safe_int(r.get('id_campaign')),
            'Артикул WB': safe_int(r.get('nmId')),
            'Артикул продавца': r.get('supplier_article',''),
            'Товар': r.get('supplier_article',''),
            'Предмет': r.get('subject',''),
            'Плейсмент': r.get('placement',''),
            'Тип кампании': f"{r.get('payment_type','')}_{r.get('placement','')}",
            'Текущая ставка, ₽': round(safe_float(r.get('current_bid_rub')), 2),
            'Комфортная ставка, ₽': round(safe_float(r.get('comfort_bid_rub')), 2) if pd.notna(r.get('comfort_bid_rub')) else None,
            'Максимальная ставка, ₽': round(safe_float(r.get('max_bid_rub')), 2) if pd.notna(r.get('max_bid_rub')) else None,
            'Экспериментальная ставка, ₽': round(safe_float(r.get('experiment_bid_rub')), 2) if pd.notna(r.get('experiment_bid_rub')) else None,
            'Тип лимита': r.get('limit_type',''),
            'Причина лимита': r.get('Причина лимита',''),
            'Статус кампании': r.get('campaign_status',''),
            'Активна для API': bool(r.get('campaign_is_active', False)),
            'Действие': action,
            'Новая ставка, ₽': round(safe_float(new_bid), 2),
            'Причина': reason,
            'Показы': round(safe_float(r.get('Показы')), 0),
            'Клики': round(safe_float(r.get('Клики')), 0),
            'CTR, %': round(safe_float(r.get('ctr_pct')), 2),
            'Заказы РК': round(safe_float(r.get('Заказы')), 2),
            'Выручка РК, ₽': round(safe_float(r.get('Сумма_заказов')), 2),
            'Расход РК, ₽': round(safe_float(r.get('Расход')), 2),
            'ДРР кампании, %': round(safe_float(r.get('campaign_drr_cur')) * 100, 2),
            'ВП кампании текущее окно после рекламы, ₽': round(safe_float(r.get('campaign_gp_cur')), 2),
            'ВП кампании базовое окно после рекламы, ₽': round(safe_float(r.get('campaign_gp_base')), 2),
            'Рост ВП кампании, %': round(safe_float(r.get('campaign_gp_growth_pct')), 2),
            'ROI кампании, %': round(safe_float(r.get('campaign_roi_cur')) * 100, 2),
            'CPO кампании, ₽': round(safe_float(r.get('campaign_cpo')), 2),
            'Все заказы товара': round(safe_float(r.get('total_orders')), 2),
            'Выручка товара, ₽': round(safe_float(r.get('total_revenue')), 2),
            'Общий ДРР товара, %': round(safe_float(r.get('blended_drr')) * 100, 2),
            'ВП товара, % рост': round(safe_float(r.get('item_gp_growth_pct')), 2),
            'План ВП MTD, ₽': round(safe_float(r.get('План ВП MTD, ₽')), 2),
            'Факт ВП MTD, ₽': round(safe_float(r.get('Факт ВП MTD, ₽')), 2),
            'Темп плана ВП, %': round(safe_float(r.get('Темп плана ВП, %')), 2),
            'Причина плана': r.get('Причина плана',''),
            'ДРР категории, %': round(safe_float(r.get('category_drr_cur')) * 100, 2),
            'Лимит ДРР категории, %': round(safe_float(r.get('category_limit_drr')) * 100, 2),
            'Факт ВП категории MTD, ₽': round(safe_float(r.get('category_gp_cur')), 2),
            'Темп категории, %': round(safe_float(r.get('category_gp_growth_pct')), 2),
            'Рост заказов категории, %': round(safe_float(r.get('category_orders_growth_pct')), 2),
            'Рост спроса категории, %': round(safe_float(r.get('category_demand_growth_pct')), 2),
            'Темп плана категории, %': round(safe_float(r.get('category_plan_attainment_pct')), 2),
            'Рост показов кампании, %': round(safe_float(r.get('campaign_impression_growth_pct')), 2),
            'Рост кликов кампании, %': round(safe_float(r.get('campaign_click_growth_pct')), 2),
            'Рост заказов кампании, %': round(safe_float(r.get('campaign_order_growth_pct')), 2),
            'CPO CPC, ₽': round(safe_float(r.get('cpo_cpc')), 2),
            'CPO Полок, ₽': round(safe_float(r.get('cpo_cpm')), 2),
            'ДРР CPC, %': round(safe_float(r.get('drr_cpc')) * 100, 2),
            'ДРР Полок, %': round(safe_float(r.get('drr_cpm')) * 100, 2),
            'ВП CPC, ₽': round(safe_float(r.get('gp_after_ads_cpc')), 2),
            'ВП Полок, ₽': round(safe_float(r.get('gp_after_ads_cpm')), 2),
            'Лучший канал': r.get('better_channel',''),
            'Статус риска': ('Критический' if safe_float(r.get('campaign_drr_cur')) > cfg.campaign_hard_drr else ('Высокий' if safe_float(r.get('campaign_drr_cur')) > cfg.campaign_target_drr or safe_float(r.get('campaign_gp_cur')) <= 0 else 'Низкий')),
            'Потенциал роста': ('Высокий' if r.get('better_channel','') == ('CPC' if str(r.get('payment_type','')).lower() == 'cpc' else 'CPM') and safe_float(r.get('campaign_gp_cur')) > 0 and safe_float(r.get('campaign_drr_cur')) <= cfg.campaign_target_drr else 'Низкий')
        })
    decisions_df = pd.DataFrame(decisions).drop_duplicates(['ID кампании','Артикул WB','Плейсмент'])

    weak = decisions_df[(decisions_df['Действие'].isin(['Снизить','Предел эффективности ставки'])) | (decisions_df.get('Статус риска', '').astype(str).eq('Критический'))].copy() if not decisions_df.empty else pd.DataFrame()
    if not weak.empty:
        weak['Комментарий'] = weak['Причина']
        weak = weak[['Артикул продавца','Артикул WB','ID кампании','Тип кампании','Плейсмент','Действие','Комментарий']].drop_duplicates()
    else:
        weak = pd.DataFrame([{'Комментарий':'Нет слабых позиций'}])

    product_metrics = daily_item.groupby('Артикул продавца', as_index=False).agg(
        Заказы=('Заказы','sum'), ДРР=('ДРР, доля','mean'), Валовая_прибыль_после_рекламы=('Валовая прибыль после рекламы, ₽','sum'), Клики=('Клики','sum'), CPO=('CPO, ₽','mean'), Чистая_прибыль=('Прогнозная чистая прибыль, ₽','sum')
    ) if not daily_item.empty and 'Комментарий' not in daily_item.columns else pd.DataFrame([{'Комментарий':'Нет дневной товарной истории'}])

    campaign_profit = daily_campaign.groupby(['ID кампании','Тип'], as_index=False).agg(
        Заказы=('Заказы_РК','sum'), ДРР=('ДРР кампании, доля','mean'), Валовая_прибыль_после_рекламы=('ВП кампании после рекламы, ₽','sum'), ROI=('ROI кампании','mean'), CPO=('CPO кампании, ₽','mean')
    ) if not daily_campaign.empty and 'Комментарий' not in daily_campaign.columns else pd.DataFrame([{'Комментарий':'Нет дневной кампанийной истории'}])

    effects = decisions_df[decisions_df['Действие'].isin(['Повысить','Снизить']) & (decisions_df['Новая ставка, ₽'] != decisions_df['Текущая ставка, ₽'])].copy()
    effects = effects[['Дата запуска','Артикул продавца','ID кампании','Тип кампании','Текущая ставка, ₽','Новая ставка, ₽','Действие','Причина']] if not effects.empty else pd.DataFrame([{'Комментарий':'Нет изменений ставок'}])

    return {
        'rows': rows,
        'decisions': decisions_df,
        'weak': weak,
        'product_metrics': product_metrics,
        'campaign_profit': campaign_profit,
        'effects': effects,
        'eff_history_sheets': build_efficiency_history(ads_daily, campaigns, aggregate_keyword_daily(keywords), master, bid_history, as_of_date),
        'window': window,
        'daily_item_history': daily_item,
        'daily_campaign_history': daily_campaign,
        'plan_sheet': plan_df,
        'category_plan': category_plan,
    }


def decisions_to_payload(decisions_df: pd.DataFrame) -> Dict[str, Any]:
    if decisions_df.empty:
        return {'bids': []}
    work = decisions_df.copy()
    if 'Активна для API' in work.columns:
        work = work[work['Активна для API'].fillna(False).astype(bool)].copy()
    elif 'Статус кампании' in work.columns:
        work = work[work['Статус кампании'].map(is_active_campaign_status)].copy()
    changed = work[work['Действие'].isin(['Повысить','Снизить']) & (work['Новая ставка, ₽'] != work['Текущая ставка, ₽'])].copy()
    changed = changed.drop_duplicates(['ID кампании','Артикул WB','Плейсмент'])
    grouped = {}
    for _, r in changed.iterrows():
        advert = safe_int(r['ID кампании'])
        nm_id = safe_int(r['Артикул WB'])
        payment_type = 'cpc' if 'cpc' in str(r['Тип кампании']).lower() else 'cpm'
        placement = str(r['Плейсмент'])
        grouped.setdefault((advert, payment_type), []).append({
            'nm_id': nm_id,
            'placement': placement_for_bids_endpoint(placement),
            'bid_kopecks': normalize_bid_for_wb(r['Новая ставка, ₽'], payment_type, placement),
        })
    out = []
    for (advert, payment_type), items in grouped.items():
        uniq = {(it['nm_id'], it['placement']): it for it in items}
        out.append({'advert_id': advert, 'payment_type': payment_type, 'nm_bids': list(uniq.values())})
    return {'bids': out}


def apply_shade_actions(actions_df: pd.DataFrame, api_key: str, dry_run: bool) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    empty_log = pd.DataFrame([{'Комментарий':'Блок оттенков отключён: каждый оттенок теперь ведётся отдельной рекламной кампанией'}])
    empty_actions = pd.DataFrame([{'Комментарий':'Блок оттенков отключён'}])
    empty_tests = pd.DataFrame([{'Комментарий':'Блок оттенков отключён'}])
    return empty_log, empty_actions, empty_tests


def build_history_append(decisions: pd.DataFrame, as_of_date: date) -> pd.DataFrame:
    if decisions.empty:
        return pd.DataFrame()
    hist = decisions.copy()
    hist['Дата запуска'] = now_ts()
    hist['Дата дня'] = as_of_date.isoformat()
    return hist



def save_outputs(provider: BaseProvider, results: Dict[str, Any], run_mode: str, bid_send_log: Optional[pd.DataFrame], shade_apply_log: Optional[pd.DataFrame] = None, history_append: Optional[pd.DataFrame] = None) -> None:
    decisions = results.get('decisions', pd.DataFrame()).copy()
    plan_sheet = results.get('plan_sheet', pd.DataFrame()).copy()
    category_plan = results.get('category_plan', pd.DataFrame()).copy()
    daily_item = results.get('daily_item_history', pd.DataFrame()).copy()
    daily_campaign = results.get('daily_campaign_history', pd.DataFrame()).copy()
    product_metrics = results.get('product_metrics', pd.DataFrame()).copy()
    campaign_profit = results.get('campaign_profit', pd.DataFrame()).copy()
    effects = results.get('effects', pd.DataFrame()).copy()
    weak = results.get('weak', pd.DataFrame()).copy()
    min_bids_df = results.get('min_bids_df', pd.DataFrame()).copy()
    history_append = history_append.copy() if isinstance(history_append, pd.DataFrame) else pd.DataFrame()

    item_cols = ['Дата', 'Товар', 'Артикул WB', 'Артикул продавца', 'Предмет', 'Заказы', 'Сумма_заказов', '% выкупа', 'Расходы_РК',
                 'Валовая прибыль после рекламы, ₽', 'Клики', 'CPO, ₽', 'Прогнозная чистая прибыль, ₽', 'Показы', 'Заказы_РК', 'Выручка_РК',
                 'ДРР, доля', 'Конверсия в корзину, %', 'Конверсия в заказ, %', 'Частота запросов', 'Спрос по ключам', 'Медианная позиция', 'Видимость, %', 'День зрелый']
    campaign_cols = ['Дата', 'ID кампании', 'Артикул WB', 'Артикул продавца', 'Предмет', 'Тип кампании', 'Тип', 'Плейсмент', 'Ставка, ₽', 'Показы', 'Клики',
                     'Заказы_РК', 'Выручка_РК', 'Расходы_РК', 'ДРР кампании, доля', 'ВП кампании после рекламы, ₽', 'ROI кампании',
                     'CPO кампании, ₽', 'Спрос по ключам', 'Частота запросов', 'День зрелый']

    daily_item = trim_to_columns(daily_item, item_cols)
    daily_campaign = trim_to_columns(daily_campaign, campaign_cols)

    summary = pd.DataFrame([{
        'Режим': 'run',
        'Дата формирования': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'Всего рекомендаций': int(len(decisions)),
        'Изменённых ставок': int(len(decisions[decisions['Действие'].isin(['Повысить', 'Снизить']) & (decisions['Новая ставка, ₽'] != decisions['Текущая ставка, ₽'])])) if not decisions.empty else 0,
        'Блоков отправки ставок': 0 if bid_send_log is None or bid_send_log.empty else int(len(bid_send_log)),
        'Текущее окно с': results['window']['cur_start'],
        'Текущее окно по': results['window']['cur_end'],
        'База с': results['window']['base_start'],
        'База по': results['window']['base_end'],
    }])

    old_sheets = {}
    try:
        if provider.file_exists(OUT_SINGLE_REPORT):
            old_sheets = provider.read_excel_all_sheets(OUT_SINGLE_REPORT)
    except Exception:
        old_sheets = {}

    def append_dedup(old_df: pd.DataFrame, new_df: pd.DataFrame, keys: List[str], final_cols: Optional[List[str]] = None) -> pd.DataFrame:
        old_df = old_df.copy() if isinstance(old_df, pd.DataFrame) else pd.DataFrame()
        new_df = new_df.copy() if isinstance(new_df, pd.DataFrame) else pd.DataFrame()
        if final_cols:
            old_df = trim_to_columns(old_df, final_cols)
            new_df = trim_to_columns(new_df, final_cols)
        if new_df.empty:
            base = old_df
        elif old_df.empty:
            base = new_df
        else:
            base = pd.concat([old_df, new_df], ignore_index=True, sort=False)
        real_keys = [k for k in keys if k in base.columns]
        base = base.drop_duplicates(real_keys, keep='last') if real_keys else base
        if final_cols:
            base = trim_to_columns(base, final_cols)
        return base

    decisions_hist = append_dedup(old_sheets.get('История решений день', pd.DataFrame()), history_append, ['Дата запуска', 'ID кампании', 'Артикул WB', 'Плейсмент'])
    item_hist_all = append_dedup(old_sheets.get('История день товар', pd.DataFrame()), daily_item, ['Дата', 'Артикул WB'], item_cols)
    campaign_hist_all = append_dedup(old_sheets.get('История день кампания', pd.DataFrame()), daily_campaign, ['Дата', 'ID кампании', 'Артикул WB'], campaign_cols)
    old_bid_hist = old_sheets.get('История ставок', pd.DataFrame())
    bid_hist_all = append_dedup(old_bid_hist, decisions[decisions['Действие'].isin(['Повысить', 'Снизить']) & (decisions['Новая ставка, ₽'] != decisions['Текущая ставка, ₽'])].copy(), ['ID кампании', 'Артикул WB', 'Плейсмент', 'Дата запуска'])
    archive_all = append_dedup(old_sheets.get('Архив решений', pd.DataFrame()), decisions, ['Дата запуска', 'ID кампании', 'Артикул WB', 'Плейсмент'])

    eff_stub = pd.DataFrame([{'Комментарий': 'Детальная эффективность ставки вынесена в отдельный файл Эффективность_ставки_ежедневно.xlsx'}])

    sheets = {
        'Сводка': summary,
        'Решения': decisions,
        'План': plan_sheet,
        'План категорий': category_plan,
        'Товар день': daily_item,
        'Кампания день': daily_campaign,
        'Товар итог': product_metrics,
        'Кампании прибыль': campaign_profit,
        'Фактически изменённые ставки': bid_send_log if bid_send_log is not None and not bid_send_log.empty else pd.DataFrame([{'Комментарий': 'Нет отправленных блоков ставок'}]),
        'Лимиты ставок': decisions[[c for c in ['Артикул продавца', 'ID кампании', 'Тип кампании', 'Статус кампании', 'Активна для API', 'Текущая ставка, ₽', 'Комфортная ставка, ₽', 'Максимальная ставка, ₽', 'Экспериментальная ставка, ₽', 'Тип лимита', 'Причина лимита'] if c in decisions.columns]].copy() if not decisions.empty else pd.DataFrame(),
        'Минимальные ставки WB': min_bids_df if not min_bids_df.empty else pd.DataFrame([{'Комментарий': 'Нет данных по min bid'}]),
        'Слабые позиции': weak,
        'Эффект изменений': effects,
        'Лог API': pd.DataFrame(API_CALL_LOGS) if API_CALL_LOGS else pd.DataFrame([{'Комментарий': 'Нет вызовов API'}]),
        'История решений день': decisions_hist,
        'История ставок': bid_hist_all,
        'Архив решений': archive_all,
        'Эффективность ставки': eff_stub,
        'История день товар': item_hist_all,
        'История день кампания': campaign_hist_all,
    }

    sheets = {name: normalize_output_df(df) for name, df in sheets.items()}
    provider.write_excel(OUT_SINGLE_REPORT, sheets)
    provider.write_excel(OUT_PREVIEW, sheets)
    provider.write_text(OUT_SUMMARY, json.dumps(summary.iloc[0].to_dict(), ensure_ascii=False, default=str, indent=2))

    eff_sheets = results.get('eff_history_sheets', {})
    if not eff_sheets:
        eff_sheets = {'Комментарий': pd.DataFrame([{'Комментарий': 'Нет данных по эффективности ставки'}])}
    eff_sheets = {name: normalize_output_df(df) for name, df in eff_sheets.items()}
    provider.write_excel(OUT_EFF, eff_sheets)

def run_manager(args: argparse.Namespace) -> None:
    API_CALL_LOGS.clear()
    MIN_BID_ROWS.clear()
    provider = choose_provider(args.local_data_dir)
    as_of_date = datetime.strptime(args.as_of_date, '%Y-%m-%d').date() if args.as_of_date else datetime.now().date()
    cfg = Config()
    results = prepare_metrics(provider, cfg, as_of_date)
    api_key = os.getenv('WB_PROMO_KEY_TOPFACE','').strip()
    results = enrich_with_min_bids(results, api_key)
    decisions = results['decisions'].copy()
    log(f'✅ Всего строк решений: {len(decisions)}')
    changed = decisions[decisions['Действие'].isin(['Повысить','Снизить']) & (decisions['Новая ставка, ₽'] != decisions['Текущая ставка, ₽'])].copy()
    log(f'🔁 Изменённых ставок: {len(changed)}')
    if not changed.empty:
        print(changed[['Товар','Артикул продавца','Предмет','ID кампании','Плейсмент','Текущая ставка, ₽','Новая ставка, ₽','Действие','Причина']].head(30).to_string(index=False), flush=True)
    payload = decisions_to_payload(decisions)
    bid_send_log = send_payload(payload, api_key, dry_run=False)
    log(f"📤 Отправлено блоков в WB: {0 if bid_send_log is None or bid_send_log.empty else len(bid_send_log)}")
    history_append = build_history_append(decisions, as_of_date)
    save_outputs(provider, results, 'run', bid_send_log, None, history_append)




def parse_abc_snapshot_dt(key: str) -> datetime:
    text = str(key or '')
    m = re.search(r'_at_(\d{4}-\d{2}-\d{2})_(\d{2})-(\d{2})', text)
    if m:
        try:
            return datetime.strptime(f"{m.group(1)} {m.group(2)}:{m.group(3)}", "%Y-%m-%d %H:%M")
        except Exception:
            pass
    return datetime.min


def load_latest_abc_reference(provider: BaseProvider) -> Tuple[pd.DataFrame, Dict[str, float]]:
    try:
        keys = provider.list_keys(ABC_PREFIX)
    except Exception:
        return pd.DataFrame(), {}
    abc_keys = [k for k in keys if re.search(r'wb_abc_report_goods__.*\.xlsx$', Path(str(k)).name, flags=re.I)]
    if not abc_keys:
        return pd.DataFrame(), {}
    best_key = max(abc_keys, key=parse_abc_snapshot_dt)
    try:
        df = provider.read_excel(best_key).copy()
    except Exception:
        return pd.DataFrame(), {}
    if df.empty:
        return pd.DataFrame(), {}

    df = df.rename(columns={
        'Артикул WB': 'nmId',
        'Артикул продавца': 'supplier_article',
        'Предмет': 'subject',
        'Ваша категория': 'manager_category',
        'Процент выкупов, %': 'buyout_pct',
        'Заказы': 'orders_cnt',
    })
    if 'nmId' in df.columns:
        df['nmId'] = pd.to_numeric(df['nmId'], errors='coerce')
    else:
        df['nmId'] = np.nan
    df['supplier_article'] = series_or_default(df, 'supplier_article', '').fillna('').astype(str)
    df['subject'] = series_or_default(df, 'subject', '').fillna('').astype(str)
    df['subject_norm'] = df['subject'].map(canonical_subject)
    df['manager_category'] = series_or_default(df, 'manager_category', '').fillna('').astype(str).str.strip()
    df['manager_category'] = df['manager_category'].replace({'': 'Без менеджера'})
    df['buyout_pct'] = pd.to_numeric(df.get('buyout_pct'), errors='coerce')
    df['orders_cnt'] = numeric_series(df, 'orders_cnt', 0.0)
    df['buyout_pct_clipped'] = df['buyout_pct'].clip(lower=0.0, upper=100.0)

    rates: Dict[str, float] = {}
    for subject_norm, grp in df.groupby('subject_norm'):
        if not subject_norm:
            continue
        valid = grp['buyout_pct_clipped'].notna()
        if not valid.any():
            continue
        w = grp.loc[valid, 'orders_cnt'].fillna(0.0)
        x = grp.loc[valid, 'buyout_pct_clipped'].fillna(0.0)
        if float(w.sum()) > 0:
            pct_value = float((x * w).sum() / w.sum())
        else:
            pct_value = float(x.mean())
        rates[subject_norm] = round(clamp(pct_value / 100.0, 0.0, 1.0), 4)

    ref_cols = [c for c in ['nmId', 'supplier_article', 'subject', 'subject_norm', 'manager_category', 'buyout_pct_clipped', 'orders_cnt'] if c in df.columns]
    ref = df[ref_cols].drop_duplicates().copy()
    ref.attrs['source_key'] = best_key
    return ref, rates


def _normalize_funnel_subject_sheet(raw_df: pd.DataFrame) -> pd.DataFrame:
    if raw_df is None or raw_df.empty:
        return pd.DataFrame(columns=['supplier_article', 'nmId', 'subject', 'subject_norm', 'buyout_rate', 'orders_qty', 'orders_sum'])
    df = raw_df.copy()
    first_row = df.iloc[0].astype(str).tolist() if len(df.index) > 0 else []
    if 'Предмет' in first_row and ('Процент выкупа' in first_row or 'Процент выкупа (предыдущий период)' in first_row):
        df.columns = [str(x).strip() for x in df.iloc[0].tolist()]
        df = df.iloc[1:].copy()
    df = df.rename(columns={
        'Артикул продавца': 'supplier_article',
        'Артикул WB': 'nmId',
        'Предмет': 'subject',
        'Процент выкупа': 'buyout_pct',
        'Заказали, шт': 'orders_qty',
        'Заказали на сумму, ₽': 'orders_sum',
        'Выкупили на сумму, ₽': 'buyout_sum',
        'ordersSumRub': 'orders_sum',
        'buyoutsSumRub': 'buyout_sum',
        'ordersCount': 'orders_qty',
        'buyoutPercent': 'buyout_pct',
    })
    if 'subject' not in df.columns:
        return pd.DataFrame(columns=['supplier_article', 'nmId', 'subject', 'subject_norm', 'buyout_rate', 'orders_qty', 'orders_sum'])
    out = df.copy()
    out['supplier_article'] = series_or_default(out, 'supplier_article', '').fillna('').astype(str).str.strip()
    out['nmId'] = pd.to_numeric(out.get('nmId'), errors='coerce')
    out['subject'] = series_or_default(out, 'subject', '').fillna('').astype(str).str.strip()
    out['subject_norm'] = out['subject'].map(canonical_subject)
    direct_rate = to_buyout_rate(out.get('buyout_pct', pd.Series(index=out.index, dtype=float)), default=np.nan)
    if 'orders_sum' in out.columns and 'buyout_sum' in out.columns:
        orders_sum = numeric_series(out, 'orders_sum', 0.0)
        buyout_sum = numeric_series(out, 'buyout_sum', 0.0)
        ratio_sum = pd.Series(np.where(orders_sum > 0, buyout_sum / orders_sum, np.nan), index=out.index)
        direct_rate = direct_rate.where(~direct_rate.isna(), ratio_sum)
    out['buyout_rate'] = pd.to_numeric(direct_rate, errors='coerce').clip(lower=0.0, upper=1.0)
    out['orders_qty'] = numeric_series(out, 'orders_qty', 0.0)
    out['orders_sum'] = numeric_series(out, 'orders_sum', 0.0)
    out = out[out['subject_norm'].ne('') & out['buyout_rate'].notna()].copy()
    return out[['supplier_article', 'nmId', 'subject', 'subject_norm', 'buyout_rate', 'orders_qty', 'orders_sum']]


def load_latest_funnel_subject_reference(provider: BaseProvider) -> Tuple[pd.DataFrame, Dict[str, float]]:
    candidates: List[Tuple[str, pd.DataFrame]] = []

    def _append_candidate(name: str, raw_df: pd.DataFrame) -> None:
        norm = _normalize_funnel_subject_sheet(raw_df)
        if not norm.empty:
            candidates.append((name, norm))

    try:
        sheets = provider.read_excel_all_sheets(FUNNEL_KEY)
        if isinstance(sheets, dict):
            for sheet_name, raw_df in sheets.items():
                if str(sheet_name).strip().lower() in {'товары', 'products'}:
                    _append_candidate(f'{FUNNEL_KEY}::{sheet_name}', raw_df)
                    break
            if not candidates:
                for sheet_name, raw_df in sheets.items():
                    _append_candidate(f'{FUNNEL_KEY}::{sheet_name}', raw_df)
                    if candidates:
                        break
    except Exception:
        pass

    if not candidates and isinstance(provider, LocalProvider):
        local_candidates = sorted(provider.base_dir.glob('*Воронка продаж*.zip')) + sorted(provider.base_dir.glob('*Воронка продаж*.xlsx'))
        for path in local_candidates:
            try:
                if path.suffix.lower() == '.zip':
                    with zipfile.ZipFile(path) as zf:
                        for member in zf.namelist():
                            if member.lower().endswith('.xlsx'):
                                data = io.BytesIO(zf.read(member))
                                xls = pd.ExcelFile(data)
                                for sheet_name in xls.sheet_names:
                                    raw_df = pd.read_excel(data, sheet_name=sheet_name)
                                    _append_candidate(f'{path.name}::{sheet_name}', raw_df)
                                    if candidates:
                                        break
                                if candidates:
                                    break
                else:
                    xls = pd.ExcelFile(path)
                    for sheet_name in xls.sheet_names:
                        raw_df = pd.read_excel(path, sheet_name=sheet_name)
                        _append_candidate(f'{path.name}::{sheet_name}', raw_df)
                        if candidates:
                            break
                if candidates:
                    break
            except Exception:
                continue

    if not candidates:
        return pd.DataFrame(), {}

    source_name, ref = candidates[0]
    rates: Dict[str, float] = {}
    for subject_norm, grp in ref.groupby('subject_norm'):
        if not subject_norm:
            continue
        weights = grp['orders_qty'].fillna(0.0)
        if float(weights.sum()) <= 0:
            weights = grp['orders_sum'].fillna(0.0)
        values = grp['buyout_rate'].fillna(np.nan)
        valid = values.notna()
        if not valid.any():
            continue
        values = values[valid]
        weights = weights[valid]
        if float(weights.sum()) > 0:
            rate_value = float((values * weights).sum() / weights.sum())
        else:
            rate_value = float(values.mean())
        rates[subject_norm] = round(clamp(rate_value, 0.0, 1.0), 4)

    ref = ref.drop_duplicates().copy()
    ref.attrs['source_key'] = source_name
    return ref, rates


def write_binary_provider(provider: BaseProvider, key: str, data: bytes, content_type: str = 'application/octet-stream') -> None:
    if isinstance(provider, S3Provider):
        provider.s3.put_object(Bucket=provider.bucket, Key=key, Body=data, ContentType=content_type)
        return
    if isinstance(provider, LocalProvider):
        path = provider._resolve(key)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(data)
        return
    raise TypeError(f'Unsupported provider type for binary write: {type(provider)}')


def article_natural_key(value: Any) -> List[Any]:
    text = str(value or '').strip()
    parts = re.split(r'(\d+)', text)
    out: List[Any] = []
    for part in parts:
        if not part:
            continue
        out.append(int(part) if part.isdigit() else part.lower())
    return out


def wrap_text_to_width(draw: Any, text: str, font: Any, max_width: int) -> List[str]:
    text = str(text or '').strip()
    if not text:
        return ['']
    words = text.split()
    if not words:
        return ['']
    lines: List[str] = []
    current = words[0]
    for word in words[1:]:
        candidate = f'{current} {word}'
        width = draw.textbbox((0, 0), candidate, font=font)[2]
        if width <= max_width:
            current = candidate
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return lines



def build_manager_pdf_bytes(manager_name: str, manager_df: pd.DataFrame, run_dt_text: str, source_key: str = '') -> bytes:
    """
    Build PDF bytes for manager recommendations.

    Fallback order:
    1) reportlab
    2) matplotlib PdfPages
    3) minimal built-in PDF (ASCII-safe transliterated text)
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import HRFlowable, KeepTogether, Paragraph, SimpleDocTemplate, Spacer

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=A4,
            leftMargin=14 * mm,
            rightMargin=14 * mm,
            topMargin=14 * mm,
            bottomMargin=14 * mm,
            title=f'Рекомендации по ставкам - {manager_name}',
            author='Ассистент WB',
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ManagerTitle',
            parent=styles['Title'],
            fontName='Helvetica-Bold',
            fontSize=16,
            leading=20,
            alignment=TA_LEFT,
            textColor=colors.black,
            spaceAfter=6,
        )
        meta_style = ParagraphStyle(
            'ManagerMeta',
            parent=styles['BodyText'],
            fontName='Helvetica',
            fontSize=9,
            leading=12,
            textColor=colors.HexColor('#444444'),
            spaceAfter=2,
        )
        head_style = ParagraphStyle(
            'ManagerHead',
            parent=styles['Heading4'],
            fontName='Helvetica-Bold',
            fontSize=10.5,
            leading=13,
            textColor=colors.black,
            spaceAfter=4,
        )
        body_style = ParagraphStyle(
            'ManagerBody',
            parent=styles['BodyText'],
            fontName='Helvetica',
            fontSize=9.5,
            leading=12,
            textColor=colors.black,
            spaceAfter=3,
        )
        reason_style = ParagraphStyle(
            'ManagerReason',
            parent=styles['BodyText'],
            fontName='Helvetica',
            fontSize=8.8,
            leading=11,
            textColor=colors.HexColor('#222222'),
            spaceAfter=0,
        )

        def esc(value: Any) -> str:
            text_value = '' if value is None else str(value)
            return (
                text_value.replace('&', '&amp;')
                .replace('<', '&lt;')
                .replace('>', '&gt;')
                .replace('\n', '<br/>')
            )

        story = [
            Paragraph(esc(f'Рекомендации по ставкам - {manager_name}'), title_style),
            Paragraph(esc(f'Дата формирования: {run_dt_text}'), meta_style),
            Paragraph(esc(f'Строк рекомендаций: {len(manager_df)}'), meta_style),
        ]
        if source_key:
            story.append(Paragraph(esc(f'Источник менеджеров/выкупа: {Path(source_key).name}'), meta_style))
        story.extend([Spacer(1, 4), HRFlowable(width='100%', thickness=1, color=colors.black), Spacer(1, 8)])

        for _, row in manager_df.iterrows():
            current_bid = safe_float(row.get('Текущая ставка, ₽'))
            new_bid = safe_float(row.get('Новая ставка, ₽'))
            drr_campaign = safe_float(row.get('ДРР кампании, %'))
            drr_category = safe_float(row.get('ДРР категории, %'))
            spend = safe_float(row.get('Расход РК, ₽'))
            orders = safe_float(row.get('Заказы РК'))
            gp = safe_float(row.get('ВП кампании текущее окно после рекламы, ₽'))

            head = (
                f"Артикул {esc(row.get('Артикул продавца', ''))} | "
                f"WB {safe_int(row.get('Артикул WB'))} | "
                f"ID кампании {safe_int(row.get('ID кампании'))} | "
                f"{esc(row.get('Тип кампании', ''))}"
            )
            rec = (
                f"<b>Рекомендация:</b> {esc(row.get('Действие', ''))} | "
                f"ставка {current_bid:.0f} -> {new_bid:.0f} ₽"
            )
            metrics = (
                f"<b>Цифры:</b> расход {spend:.0f} ₽; "
                f"заказы РК {orders:.0f}; "
                f"ДРР кампании {drr_campaign:.1f}%; "
                f"ВП кампании {gp:.0f} ₽; "
                f"ДРР категории {drr_category:.1f}%"
            )
            reason = f"<b>Почему:</b> {esc(row.get('Причина', ''))}"

            block = [
                Paragraph(head, head_style),
                Paragraph(rec, body_style),
                Paragraph(metrics, body_style),
                Paragraph(reason, reason_style),
                Spacer(1, 6),
                HRFlowable(width='100%', thickness=0.6, color=colors.HexColor('#B5B5B5')),
                Spacer(1, 7),
            ]
            story.append(KeepTogether(block))

        doc.build(story)
        return buf.getvalue()
    except Exception:
        pass

    try:
        import math
        import textwrap
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_pdf import PdfPages

        lines: List[str] = []
        lines.append(f'Рекомендации по ставкам - {manager_name}')
        lines.append(f'Дата формирования: {run_dt_text}')
        lines.append(f'Строк рекомендаций: {len(manager_df)}')
        if source_key:
            lines.append(f'Источник менеджеров/выкупа: {Path(source_key).name}')
        lines.append('')

        for _, row in manager_df.iterrows():
            current_bid = safe_float(row.get('Текущая ставка, ₽'))
            new_bid = safe_float(row.get('Новая ставка, ₽'))
            drr_campaign = safe_float(row.get('ДРР кампании, %'))
            drr_category = safe_float(row.get('ДРР категории, %'))
            spend = safe_float(row.get('Расход РК, ₽'))
            orders = safe_float(row.get('Заказы РК'))
            gp = safe_float(row.get('ВП кампании текущее окно после рекламы, ₽'))

            block = [
                f"Артикул {row.get('Артикул продавца', '')} | WB {safe_int(row.get('Артикул WB'))} | ID кампании {safe_int(row.get('ID кампании'))} | {row.get('Тип кампании', '')}",
                f"Рекомендация: {row.get('Действие', '')} | ставка {current_bid:.0f} -> {new_bid:.0f} ₽",
                f"Цифры: расход {spend:.0f} ₽; заказы РК {orders:.0f}; ДРР кампании {drr_campaign:.1f}%; ВП кампании {gp:.0f} ₽; ДРР категории {drr_category:.1f}%",
                f"Почему: {row.get('Причина', '')}",
                "-" * 110,
            ]
            for b in block:
                wrapped = textwrap.wrap(str(b), width=115, break_long_words=False, break_on_hyphens=False) or ['']
                lines.extend(wrapped)

        line_height = 0.022
        usable_lines = 42
        total_pages = max(1, math.ceil(len(lines) / usable_lines))
        buf = io.BytesIO()
        with PdfPages(buf) as pdf:
            for page_idx in range(total_pages):
                page_lines = lines[page_idx * usable_lines:(page_idx + 1) * usable_lines]
                fig = plt.figure(figsize=(8.27, 11.69))
                ax = fig.add_axes([0, 0, 1, 1])
                ax.axis('off')
                y = 0.97
                for i, line in enumerate(page_lines):
                    fontsize = 12 if (page_idx == 0 and i == 0) else 8.5
                    weight = 'bold' if (page_idx == 0 and i == 0) else 'normal'
                    ax.text(0.04, y, str(line), ha='left', va='top', fontsize=fontsize, fontweight=weight, family='DejaVu Sans')
                    y -= line_height
                pdf.savefig(fig)
                plt.close(fig)
        return buf.getvalue()
    except Exception:
        pass

    def _translit(value: Any) -> str:
        mapping = {
            'А':'A','Б':'B','В':'V','Г':'G','Д':'D','Е':'E','Ё':'E','Ж':'Zh','З':'Z','И':'I','Й':'Y','К':'K','Л':'L','М':'M','Н':'N','О':'O','П':'P','Р':'R','С':'S','Т':'T','У':'U','Ф':'F','Х':'Kh','Ц':'Ts','Ч':'Ch','Ш':'Sh','Щ':'Sch','Ъ':'','Ы':'Y','Ь':'','Э':'E','Ю':'Yu','Я':'Ya',
            'а':'a','б':'b','в':'v','г':'g','д':'d','е':'e','ё':'e','ж':'zh','з':'z','и':'i','й':'y','к':'k','л':'l','м':'m','н':'n','о':'o','п':'p','р':'r','с':'s','т':'t','у':'u','ф':'f','х':'kh','ц':'ts','ч':'ch','ш':'sh','щ':'sch','ъ':'','ы':'y','ь':'','э':'e','ю':'yu','я':'ya',
            '₽':'RUB','—':'-','–':'-','№':'No '
        }
        s = '' if value is None else str(value)
        return ''.join(mapping.get(ch, ch if ord(ch) < 128 else '?') for ch in s)

    def _pdf_escape(text_value: str) -> str:
        return text_value.replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')

    lines = [
        _translit(f'Rekomendatsii po stavkam - {manager_name}'),
        _translit(f'Data formirovaniya: {run_dt_text}'),
        _translit(f'Strok rekomendatsiy: {len(manager_df)}'),
    ]
    if source_key:
        lines.append(_translit(f'Istochnik menedzherov/vykupa: {Path(source_key).name}'))
    lines.append('')

    import textwrap
    for _, row in manager_df.iterrows():
        current_bid = safe_float(row.get('Текущая ставка, ₽'))
        new_bid = safe_float(row.get('Новая ставка, ₽'))
        drr_campaign = safe_float(row.get('ДРР кампании, %'))
        drr_category = safe_float(row.get('ДРР категории, %'))
        spend = safe_float(row.get('Расход РК, ₽'))
        orders = safe_float(row.get('Заказы РК'))
        gp = safe_float(row.get('ВП кампании текущее окно после рекламы, ₽'))
        block = [
            _translit(f"Artikul {row.get('Артикул продавца', '')} | WB {safe_int(row.get('Артикул WB'))} | ID kampanii {safe_int(row.get('ID кампании'))} | {row.get('Тип кампании', '')}"),
            _translit(f"Rekomendatsiya: {row.get('Действие', '')} | stavka {current_bid:.0f} -> {new_bid:.0f} RUB"),
            _translit(f"Tsifry: rashod {spend:.0f} RUB; zakazy RK {orders:.0f}; DRR kampanii {drr_campaign:.1f}%; VP kampanii {gp:.0f} RUB; DRR kategorii {drr_category:.1f}%"),
            _translit(f"Pochemu: {row.get('Причина', '')}"),
            "-" * 110,
        ]
        for b in block:
            lines.extend(textwrap.wrap(str(b), width=110, break_long_words=False, break_on_hyphens=False) or [''])

    lines_per_page = 46
    page_width = 595
    page_height = 842
    content = []
    page_objects = []
    font_obj_num = 3
    obj_num = 4

    pages = [lines[i:i+lines_per_page] for i in range(0, len(lines), lines_per_page)] or [[]]
    for page_lines in pages:
        stream_lines = ['BT', '/F1 9 Tf', '14 TL', '40 800 Td']
        for i, line in enumerate(page_lines):
            if i > 0:
                stream_lines.append('T*')
            stream_lines.append(f'({_pdf_escape(line)}) Tj')
        stream_lines.append('ET')
        stream = '\n'.join(stream_lines).encode('latin-1', errors='replace')
        content_obj = obj_num
        page_obj = obj_num + 1
        content.append((content_obj, f"<< /Length {len(stream)} >>\nstream\n".encode('latin-1') + stream + b"\nendstream"))
        page_objects.append(page_obj)
        content.append((page_obj, f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_width} {page_height}] /Contents {content_obj} 0 R /Resources << /Font << /F1 {font_obj_num} 0 R >> >> >>".encode('latin-1')))
        obj_num += 2

    objects = [
        (1, b"<< /Type /Catalog /Pages 2 0 R >>"),
        (2, f"<< /Type /Pages /Kids [{' '.join(f'{n} 0 R' for n in page_objects)}] /Count {len(page_objects)} >>".encode('latin-1')),
        (3, b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>"),
    ] + content

    pdf = io.BytesIO()
    pdf.write(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = {0: 0}
    for num, body in objects:
        offsets[num] = pdf.tell()
        pdf.write(f"{num} 0 obj\n".encode('latin-1'))
        pdf.write(body)
        pdf.write(b"\nendobj\n")
    xref_pos = pdf.tell()
    max_obj = max(offsets)
    pdf.write(f"xref\n0 {max_obj + 1}\n".encode('latin-1'))
    pdf.write(b"0000000000 65535 f \n")
    for i in range(1, max_obj + 1):
        pdf.write(f"{offsets.get(i, 0):010d} 00000 n \n".encode('latin-1'))
    pdf.write(f"trailer\n<< /Size {max_obj + 1} /Root 1 0 R >>\nstartxref\n{xref_pos}\n%%EOF".encode('latin-1'))
    return pdf.getvalue()


def build_manager_recommendations(decisions_df: pd.DataFrame, abc_ref: pd.DataFrame) -> pd.DataFrame:
    work = decisions_df.copy()
    if work.empty:
        return work
    if abc_ref is not None and not abc_ref.empty:
        manager_map = abc_ref[['nmId', 'manager_category']].dropna().drop_duplicates('nmId').copy()
        manager_map = manager_map.rename(columns={'nmId': 'Артикул WB', 'manager_category': 'Менеджер'})
        work = work.merge(manager_map, on='Артикул WB', how='left')
    if 'Менеджер' not in work.columns:
        work['Менеджер'] = 'Без менеджера'
    work['Менеджер'] = work['Менеджер'].fillna('Без менеджера').astype(str).str.strip().replace({'': 'Без менеджера'})
    work['_article_sort'] = work['Артикул продавца'].map(lambda x: '|'.join(f"{p:010d}" if isinstance(p, int) else str(p) for p in article_natural_key(x)))
    work = work.sort_values(['Менеджер', '_article_sort', 'ID кампании', 'Тип кампании'], kind='stable').drop(columns=['_article_sort'])
    return work


def save_manager_recommendation_pdfs(provider: BaseProvider, decisions_df: pd.DataFrame, abc_ref: pd.DataFrame, run_dt_text: str) -> pd.DataFrame:
    work = build_manager_recommendations(decisions_df, abc_ref)
    if work.empty:
        return pd.DataFrame([{'Менеджер': 'Нет данных', 'Файл': '', 'Строк рекомендаций': 0}])

    source_key = ''
    if abc_ref is not None and hasattr(abc_ref, 'attrs'):
        source_key = str(abc_ref.attrs.get('source_key') or '')

    index_rows: List[Dict[str, Any]] = []
    base_prefix = SERVICE_ROOT + 'Рекомендации_менеджерам/'
    for manager_name, grp in work.groupby('Менеджер', dropna=False):
        manager_name = str(manager_name or 'Без менеджера').strip() or 'Без менеджера'
        pdf_bytes = build_manager_pdf_bytes(manager_name, grp, run_dt_text, source_key)
        safe_name = re.sub(r'[^0-9A-Za-zА-Яа-я_-]+', '_', manager_name).strip('_') or 'manager'
        key = base_prefix + f'Рекомендации_{safe_name}.pdf'
        write_binary_provider(provider, key, pdf_bytes, content_type='application/pdf')
        index_rows.append({
            'Менеджер': manager_name,
            'Файл': key,
            'Строк рекомендаций': int(len(grp)),
        })
    return pd.DataFrame(index_rows).sort_values(['Менеджер']).reset_index(drop=True)


def save_outputs(provider: BaseProvider, results: Dict[str, Any], run_mode: str, bid_send_log: Optional[pd.DataFrame], shade_apply_log: Optional[pd.DataFrame] = None, history_append: Optional[pd.DataFrame] = None) -> None:
    decisions = results.get('decisions', pd.DataFrame()).copy()
    plan_sheet = results.get('plan_sheet', pd.DataFrame()).copy()
    category_plan = results.get('category_plan', pd.DataFrame()).copy()
    daily_item = results.get('daily_item_history', pd.DataFrame()).copy()
    daily_campaign = results.get('daily_campaign_history', pd.DataFrame()).copy()
    product_metrics = results.get('product_metrics', pd.DataFrame()).copy()
    campaign_profit = results.get('campaign_profit', pd.DataFrame()).copy()
    effects = results.get('effects', pd.DataFrame()).copy()
    weak = results.get('weak', pd.DataFrame()).copy()
    min_bids_df = results.get('min_bids_df', pd.DataFrame()).copy()
    history_append = history_append.copy() if isinstance(history_append, pd.DataFrame) else pd.DataFrame()

    abc_ref = globals().get('ABC_REFERENCE_DF', pd.DataFrame())
    if abc_ref is None or not isinstance(abc_ref, pd.DataFrame):
        abc_ref = pd.DataFrame()
    manager_index = save_manager_recommendation_pdfs(provider, decisions, abc_ref, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

    if not decisions.empty:
        decisions = build_manager_recommendations(decisions, abc_ref)

    item_cols = ['Дата', 'Товар', 'Артикул WB', 'Артикул продавца', 'Предмет', 'Заказы', 'Сумма_заказов', '% выкупа', 'Расходы_РК',
                 'Валовая прибыль после рекламы, ₽', 'Клики', 'CPO, ₽', 'Прогнозная чистая прибыль, ₽', 'Показы', 'Заказы_РК', 'Выручка_РК',
                 'ДРР, доля', 'Конверсия в корзину, %', 'Конверсия в заказ, %', 'Частота запросов', 'Спрос по ключам', 'Медианная позиция', 'Видимость, %', 'День зрелый']
    campaign_cols = ['Дата', 'ID кампании', 'Артикул WB', 'Артикул продавца', 'Предмет', 'Тип кампании', 'Тип', 'Плейсмент', 'Ставка, ₽', 'Показы', 'Клики',
                     'Заказы_РК', 'Выручка_РК', 'Расходы_РК', 'ДРР кампании, доля', 'ВП кампании после рекламы, ₽', 'ROI кампании',
                     'CPO кампании, ₽', 'Спрос по ключам', 'Частота запросов', 'День зрелый']

    daily_item = trim_to_columns(daily_item, item_cols)
    daily_campaign = trim_to_columns(daily_campaign, campaign_cols)

    summary = pd.DataFrame([{
        'Режим': 'run',
        'Дата формирования': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'Всего рекомендаций': int(len(decisions)),
        'Изменённых ставок': int(len(decisions[decisions['Действие'].isin(['Повысить', 'Снизить']) & (decisions['Новая ставка, ₽'] != decisions['Текущая ставка, ₽'])])) if not decisions.empty else 0,
        'Блоков отправки ставок': 0 if bid_send_log is None or bid_send_log.empty else int(len(bid_send_log)),
        'PDF менеджерам': int(len(manager_index)) if not manager_index.empty else 0,
        'Текущее окно с': results['window']['cur_start'],
        'Текущее окно по': results['window']['cur_end'],
        'База с': results['window']['base_start'],
        'База по': results['window']['base_end'],
    }])

    old_sheets = {}
    try:
        if provider.file_exists(OUT_SINGLE_REPORT):
            old_sheets = provider.read_excel_all_sheets(OUT_SINGLE_REPORT)
    except Exception:
        old_sheets = {}

    def append_dedup(old_df: pd.DataFrame, new_df: pd.DataFrame, keys: List[str], final_cols: Optional[List[str]] = None) -> pd.DataFrame:
        old_df = old_df.copy() if isinstance(old_df, pd.DataFrame) else pd.DataFrame()
        new_df = new_df.copy() if isinstance(new_df, pd.DataFrame) else pd.DataFrame()
        if final_cols:
            old_df = trim_to_columns(old_df, final_cols)
            new_df = trim_to_columns(new_df, final_cols)
        if new_df.empty:
            base = old_df
        elif old_df.empty:
            base = new_df
        else:
            base = pd.concat([old_df, new_df], ignore_index=True, sort=False)
        real_keys = [k for k in keys if k in base.columns]
        base = base.drop_duplicates(real_keys, keep='last') if real_keys else base
        if final_cols:
            base = trim_to_columns(base, final_cols)
        return base

    decisions_hist = append_dedup(old_sheets.get('История решений день', pd.DataFrame()), history_append, ['Дата запуска', 'ID кампании', 'Артикул WB', 'Плейсмент'])
    item_hist_all = append_dedup(old_sheets.get('История день товар', pd.DataFrame()), daily_item, ['Дата', 'Артикул WB'], item_cols)
    campaign_hist_all = append_dedup(old_sheets.get('История день кампания', pd.DataFrame()), daily_campaign, ['Дата', 'ID кампании', 'Артикул WB'], campaign_cols)
    old_bid_hist = old_sheets.get('История ставок', pd.DataFrame())
    bid_hist_all = append_dedup(old_bid_hist, decisions[decisions['Действие'].isin(['Повысить', 'Снизить']) & (decisions['Новая ставка, ₽'] != decisions['Текущая ставка, ₽'])].copy(), ['ID кампании', 'Артикул WB', 'Плейсмент', 'Дата запуска'])
    archive_all = append_dedup(old_sheets.get('Архив решений', pd.DataFrame()), decisions, ['Дата запуска', 'ID кампании', 'Артикул WB', 'Плейсмент'])

    eff_stub = pd.DataFrame([{'Комментарий': 'Детальная эффективность ставки вынесена в отдельный файл Эффективность_ставки_ежедневно.xlsx'}])

    sheets = {
        'Сводка': summary,
        'Решения': decisions,
        'План': plan_sheet,
        'План категорий': category_plan,
        'Товар день': daily_item,
        'Кампания день': daily_campaign,
        'Товар итог': product_metrics,
        'Кампании прибыль': campaign_profit,
        'Фактически изменённые ставки': bid_send_log if bid_send_log is not None and not bid_send_log.empty else pd.DataFrame([{'Комментарий': 'Нет отправленных блоков ставок'}]),
        'Лимиты ставок': decisions[[c for c in ['Менеджер', 'Артикул продавца', 'ID кампании', 'Тип кампании', 'Статус кампании', 'Активна для API', 'Текущая ставка, ₽', 'Комфортная ставка, ₽', 'Максимальная ставка, ₽', 'Экспериментальная ставка, ₽', 'Тип лимита', 'Причина лимита'] if c in decisions.columns]].copy() if not decisions.empty else pd.DataFrame(),
        'Минимальные ставки WB': min_bids_df if not min_bids_df.empty else pd.DataFrame([{'Комментарий': 'Нет данных по min bid'}]),
        'Слабые позиции': weak,
        'Эффект изменений': effects,
        'Рекомендации менеджерам': manager_index if not manager_index.empty else pd.DataFrame([{'Комментарий': 'PDF не сформированы'}]),
        'Лог API': pd.DataFrame(API_CALL_LOGS) if API_CALL_LOGS else pd.DataFrame([{'Комментарий': 'Нет вызовов API'}]),
        'История решений день': decisions_hist,
        'История ставок': bid_hist_all,
        'Архив решений': archive_all,
        'Эффективность ставки': eff_stub,
        'История день товар': item_hist_all,
        'История день кампания': campaign_hist_all,
    }

    sheets = {name: normalize_output_df(df) for name, df in sheets.items()}
    provider.write_excel(OUT_SINGLE_REPORT, sheets)
    provider.write_excel(OUT_PREVIEW, sheets)
    provider.write_text(OUT_SUMMARY, json.dumps(summary.iloc[0].to_dict(), ensure_ascii=False, default=str, indent=2))

    eff_sheets = results.get('eff_history_sheets', {})
    if not eff_sheets:
        eff_sheets = {'Комментарий': pd.DataFrame([{'Комментарий': 'Нет данных по эффективности ставки'}])}
    eff_sheets = {name: normalize_output_df(df) for name, df in eff_sheets.items()}
    provider.write_excel(OUT_EFF, eff_sheets)


def run_manager(args: argparse.Namespace) -> None:
    API_CALL_LOGS.clear()
    MIN_BID_ROWS.clear()
    provider = choose_provider(args.local_data_dir)
    abc_ref, abc_rates = load_latest_abc_reference(provider)
    funnel_ref, funnel_rates = load_latest_funnel_subject_reference(provider)
    globals()['ABC_REFERENCE_DF'] = abc_ref
    globals()['FUNNEL_SUBJECT_REFERENCE_DF'] = funnel_ref

    merged_rates = dict(SUBJECT_FIXED_BUYOUT_RATES)
    if abc_rates:
        merged_rates.update(abc_rates)
    if funnel_rates:
        merged_rates.update(funnel_rates)
    globals()['SUBJECT_FIXED_BUYOUT_RATES'] = merged_rates

    if merged_rates:
        target_rate_text = ', '.join(
            f"{get_subject_display_name(k)}={v*100:.1f}%"
            for k, v in sorted(merged_rates.items())
            if k in TARGET_SUBJECTS
        )
        all_subjects_loaded = len([k for k in merged_rates.keys() if str(k).strip()])
        source_parts = []
        if abc_rates:
            source_parts.append(f'ABC={len(abc_rates)}')
        if funnel_rates:
            source_parts.append(f'Воронка={len(funnel_rates)}')
        source_text = ', '.join(source_parts) if source_parts else 'fallback'
        log(f'📦 Выкуп по всем предметам загружен: {all_subjects_loaded} предметов ({source_text})')
        if target_rate_text:
            log(f'📦 Целевые категории: {target_rate_text}')
    as_of_date = datetime.strptime(args.as_of_date, '%Y-%m-%d').date() if args.as_of_date else datetime.now().date()
    cfg = Config()
    results = prepare_metrics(provider, cfg, as_of_date)
    api_key = os.getenv('WB_PROMO_KEY_TOPFACE','').strip()
    results = enrich_with_min_bids(results, api_key)
    decisions = results['decisions'].copy()
    log(f'✅ Всего строк решений: {len(decisions)}')
    changed = decisions[decisions['Действие'].isin(['Повысить','Снизить']) & (decisions['Новая ставка, ₽'] != decisions['Текущая ставка, ₽'])].copy()
    log(f'🔁 Изменённых ставок: {len(changed)}')
    if not changed.empty:
        print(changed[['Товар','Артикул продавца','Предмет','ID кампании','Плейсмент','Текущая ставка, ₽','Новая ставка, ₽','Действие','Причина']].head(30).to_string(index=False), flush=True)
    payload = decisions_to_payload(decisions)
    bid_send_log = send_payload(payload, api_key, dry_run=False)
    log(f"📤 Отправлено блоков в WB: {0 if bid_send_log is None or bid_send_log.empty else len(bid_send_log)}")
    history_append = build_history_append(decisions, as_of_date)
    save_outputs(provider, results, 'run', bid_send_log, None, history_append)



# ===================== TRAFFIC SHARE / POST-CHECK OVERRIDES =====================

_BASE_prepare_metrics = prepare_metrics
_BASE_save_outputs = save_outputs


def _read_bid_history_any(provider: BaseProvider) -> pd.DataFrame:
    """Read bid history from dedicated file or from workbook sheet, robustly."""
    hist = pd.DataFrame()
    try:
        hist = load_bid_history(provider)
    except Exception:
        hist = pd.DataFrame()
    if hist is not None and not hist.empty:
        return hist

    for candidate in [OUT_SINGLE_REPORT, OUT_PREVIEW]:
        try:
            if provider.file_exists(candidate):
                sheets = provider.read_excel_all_sheets(candidate)
                for sh in ['История ставок', 'История_ставок']:
                    if sh in sheets and not sheets[sh].empty:
                        raw = sheets[sh].copy()
                        raw = raw.rename(columns={'Дата запуска': 'run_ts', 'ID кампании': 'id_campaign', 'Артикул WB': 'nmId', 'Тип кампании': 'campaign_type'})
                        raw['run_ts'] = pd.to_datetime(raw.get('run_ts'), errors='coerce')
                        raw['date'] = raw['run_ts'].dt.normalize().astype('datetime64[ns]')
                        if 'Новая ставка, ₽' in raw.columns:
                            raw['bid_rub'] = pd.to_numeric(raw.get('Новая ставка, ₽'), errors='coerce').fillna(0.0)
                        else:
                            search_col = pd.to_numeric(raw.get('Ставка поиск, коп', 0), errors='coerce') if 'Ставка поиск, коп' in raw.columns else pd.Series(0, index=raw.index, dtype=float)
                            reco_col = pd.to_numeric(raw.get('Ставка рекомендации, коп', 0), errors='coerce') if 'Ставка рекомендации, коп' in raw.columns else pd.Series(0, index=raw.index, dtype=float)
                            bid_kop = search_col.where(search_col.fillna(0) > 0, reco_col)
                            raw['bid_rub'] = (bid_kop.fillna(0) / 100.0).astype(float)
                        raw['id_campaign'] = pd.to_numeric(raw.get('id_campaign'), errors='coerce')
                        raw['nmId'] = pd.to_numeric(raw.get('nmId'), errors='coerce')
                        raw = raw.dropna(subset=['run_ts', 'date', 'id_campaign', 'nmId']).copy()
                        if raw.empty:
                            continue
                        raw['id_campaign'] = raw['id_campaign'].astype('int64')
                        raw['nmId'] = raw['nmId'].astype('int64')
                        return raw
        except Exception:
            continue
    return pd.DataFrame()


def _keyword_daily_unique(keywords: pd.DataFrame) -> pd.DataFrame:
    """Deduplicate WB keyword rows because API returns same query in 3 filters."""
    if keywords is None or keywords.empty:
        return pd.DataFrame(columns=[
            'date', 'nmId', 'supplier_article', 'subject_norm', 'query_text', 'query_freq',
            'keyword_clicks', 'keyword_orders', 'visibility_pct', 'median_position', 'query_filters'
        ])
    df = keywords.copy()
    rename_map = {
        'Поисковый запрос': 'query_text',
        'Переходы в карточку': 'keyword_clicks',
        'Заказы': 'keyword_orders',
        'Частота запросов': 'query_freq',
        'Видимость %': 'visibility_pct',
        'Медианная позиция': 'median_position',
        'Фильтр': 'query_filter',
    }
    for src, dst in rename_map.items():
        if src in df.columns and dst not in df.columns:
            df[dst] = df[src]
    for c, default in [
        ('query_text', ''),
        ('query_filter', ''),
        ('supplier_article', ''),
        ('subject_norm', ''),
        ('query_freq', 0.0),
        ('keyword_clicks', 0.0),
        ('keyword_orders', 0.0),
        ('visibility_pct', 0.0),
        ('median_position', 0.0),
    ]:
        if c not in df.columns:
            df[c] = default
    if 'nmId' not in df.columns:
        df['nmId'] = pd.to_numeric(df.get('Артикул WB'), errors='coerce')
    df['date'] = pd.to_datetime(df.get('date'), errors='coerce').dt.date
    df['nmId'] = pd.to_numeric(df.get('nmId'), errors='coerce')
    df['supplier_article'] = df['supplier_article'].fillna('').astype(str)
    df['subject_norm'] = df['subject_norm'].fillna('').astype(str).map(canonical_subject)
    df['query_text'] = df['query_text'].fillna('').astype(str).str.strip()
    df['query_filter'] = df['query_filter'].fillna('').astype(str).str.strip().str.lower()
    df = df[df['date'].notna() & df['nmId'].notna() & df['query_text'].ne('')].copy()
    if df.empty:
        return pd.DataFrame(columns=[
            'date', 'nmId', 'supplier_article', 'subject_norm', 'query_text', 'query_freq',
            'keyword_clicks', 'keyword_orders', 'visibility_pct', 'median_position', 'query_filters'
        ])
    for c in ['query_freq', 'keyword_clicks', 'keyword_orders', 'visibility_pct', 'median_position']:
        df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0.0)
    grouped = df.groupby(['date', 'nmId', 'supplier_article', 'subject_norm', 'query_text'], as_index=False).agg(
        query_freq=('query_freq', 'max'),
        keyword_clicks=('keyword_clicks', 'max'),
        keyword_orders=('keyword_orders', 'max'),
        visibility_pct=('visibility_pct', 'max'),
        median_position=('median_position', 'min'),
        query_filters=('query_filter', lambda s: '|'.join(sorted({x for x in s if x}))),
    )
    grouped['keyword_click_share_pct'] = np.where(
        grouped['query_freq'] > 0,
        grouped['keyword_clicks'] / grouped['query_freq'] * 100.0,
        0.0,
    )
    return grouped


def _classify_primary_queries(kw_daily: pd.DataFrame, as_of_date: date) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Split queries into primary (80% orders / clicks) and secondary."""
    if kw_daily.empty:
        empty = pd.DataFrame(columns=['nmId', 'supplier_article', 'query_text', 'query_class', 'query_rank', 'metric_share_pct'])
        top20 = pd.DataFrame(columns=['nmId', 'supplier_article', 'article_metric'])
        return empty, top20

    hist = pd.DataFrame()
    for days in [14, 28]:
        tmp = kw_daily[(kw_daily['date'] >= (as_of_date - timedelta(days=days - 1))) & (kw_daily['date'] <= as_of_date)].copy()
        if not tmp.empty:
            hist = tmp
            break
    if hist.empty:
        hist = kw_daily.copy()

    sku_metric = hist.groupby(['nmId', 'supplier_article'], as_index=False).agg(
        article_orders=('keyword_orders', 'sum'),
        article_clicks=('keyword_clicks', 'sum'),
        article_freq=('query_freq', 'sum'),
    )
    sku_metric['article_metric'] = np.where(
        sku_metric['article_orders'] > 0,
        sku_metric['article_orders'],
        np.where(sku_metric['article_clicks'] > 0, sku_metric['article_clicks'], sku_metric['article_freq'])
    )
    top20 = sku_metric.sort_values(['article_metric', 'article_orders', 'article_clicks'], ascending=[False, False, False]).head(20)[['nmId', 'supplier_article', 'article_metric']]

    q = hist.groupby(['nmId', 'supplier_article', 'query_text'], as_index=False).agg(
        query_orders=('keyword_orders', 'sum'),
        query_clicks=('keyword_clicks', 'sum'),
        query_freq=('query_freq', 'sum'),
    )
    q['metric_value'] = np.where(
        q['query_orders'] > 0,
        q['query_orders'],
        np.where(q['query_clicks'] > 0, q['query_clicks'], q['query_freq'])
    )

    out_parts = []
    for (nm_id, supplier_article), g in q.groupby(['nmId', 'supplier_article']):
        gg = g.sort_values(['metric_value', 'query_orders', 'query_clicks', 'query_freq', 'query_text'], ascending=[False, False, False, False, True]).copy()
        total_metric = safe_float(gg['metric_value'].sum())
        if total_metric > 0:
            gg['metric_share_pct'] = gg['metric_value'] / total_metric * 100.0
            gg['cum_share_pct'] = gg['metric_share_pct'].cumsum()
            gg['query_rank'] = np.arange(1, len(gg) + 1)
            gg['query_class'] = np.where((gg['cum_share_pct'] <= 80.0) | (gg['query_rank'] == 1), 'primary', 'secondary')
        else:
            gg['query_rank'] = np.arange(1, len(gg) + 1)
            gg['metric_share_pct'] = 0.0
            gg['query_class'] = np.where(gg['query_rank'] <= 3, 'primary', 'secondary')
        out_parts.append(gg[['nmId', 'supplier_article', 'query_text', 'query_class', 'query_rank', 'metric_share_pct']])
    class_map = pd.concat(out_parts, ignore_index=True) if out_parts else pd.DataFrame(columns=['nmId', 'supplier_article', 'query_text', 'query_class', 'query_rank', 'metric_share_pct'])
    return class_map, top20


def _build_top20_query_history(kw_daily: pd.DataFrame, class_map: pd.DataFrame, top20: pd.DataFrame) -> pd.DataFrame:
    if kw_daily.empty or class_map.empty or top20.empty:
        return pd.DataFrame([{'Комментарий': 'Нет данных по ключевым запросам для топ-20 артикулов'}])
    df = kw_daily.merge(class_map, on=['nmId', 'supplier_article', 'query_text'], how='left')
    df['query_class'] = df['query_class'].fillna('secondary')
    df = df.merge(top20[['nmId', 'supplier_article']], on=['nmId', 'supplier_article'], how='inner')
    if df.empty:
        return pd.DataFrame([{'Комментарий': 'Нет данных по ключевым запросам для топ-20 артикулов'}])
    df['Доля кликов от спроса, %'] = np.where(df['query_freq'] > 0, df['keyword_clicks'] / df['query_freq'] * 100.0, 0.0)
    df = df.rename(columns={
        'date': 'Дата',
        'nmId': 'Артикул WB',
        'supplier_article': 'Артикул продавца',
        'query_text': 'Поисковый запрос',
        'query_class': 'Тип запроса',
        'query_freq': 'Частотность запроса',
        'keyword_clicks': 'Переходы в карточку',
        'keyword_orders': 'Заказы по запросу',
        'visibility_pct': 'Видимость, %',
        'median_position': 'Медианная позиция',
        'metric_share_pct': 'Доля вклада запроса, %',
    })
    keep_cols = [
        'Дата', 'Артикул WB', 'Артикул продавца', 'Поисковый запрос', 'Тип запроса', 'Частотность запроса',
        'Переходы в карточку', 'Заказы по запросу', 'Видимость, %', 'Медианная позиция',
        'Доля кликов от спроса, %', 'Доля вклада запроса, %'
    ]
    return df[keep_cols].sort_values(['Артикул продавца', 'Дата', 'Тип запроса', 'Заказы по запросу', 'Частотность запроса'], ascending=[True, True, True, False, False])


def _weighted_mean(values: pd.Series, weights: pd.Series) -> float:
    v = pd.to_numeric(values, errors='coerce').fillna(0.0)
    w = pd.to_numeric(weights, errors='coerce').fillna(0.0)
    s = w.sum()
    if s <= 0:
        return safe_float(v.mean()) if len(v) else 0.0
    return safe_float((v * w).sum() / s)


def _query_group_snapshot(kw_daily: pd.DataFrame, class_map: pd.DataFrame, nm_id: int, supplier_article: str, start_date: date, end_date: date) -> Dict[str, float]:
    snap = {
        'primary_freq': 0.0, 'primary_clicks': 0.0, 'primary_orders': 0.0, 'primary_click_share_pct': 0.0,
        'primary_visibility_pct': 0.0, 'primary_position': 0.0,
        'secondary_freq': 0.0, 'secondary_clicks': 0.0, 'secondary_orders': 0.0, 'secondary_click_share_pct': 0.0,
        'secondary_visibility_pct': 0.0, 'secondary_position': 0.0,
    }
    if kw_daily.empty or class_map.empty or nm_id <= 0:
        return snap
    df = kw_daily[(kw_daily['date'] >= start_date) & (kw_daily['date'] <= end_date) & (pd.to_numeric(kw_daily['nmId'], errors='coerce') == nm_id)].copy()
    if supplier_article:
        df = df[df['supplier_article'].astype(str) == str(supplier_article)].copy()
    if df.empty:
        return snap
    df = df.merge(class_map, on=['nmId', 'supplier_article', 'query_text'], how='left')
    df['query_class'] = df['query_class'].fillna('secondary')
    out = {}
    for qclass in ['primary', 'secondary']:
        g = df[df['query_class'] == qclass].copy()
        if g.empty:
            out.update({
                f'{qclass}_freq': 0.0, f'{qclass}_clicks': 0.0, f'{qclass}_orders': 0.0,
                f'{qclass}_click_share_pct': 0.0, f'{qclass}_visibility_pct': 0.0, f'{qclass}_position': 0.0,
            })
            continue
        freq = safe_float(g['query_freq'].sum())
        clicks = safe_float(g['keyword_clicks'].sum())
        orders = safe_float(g['keyword_orders'].sum())
        out.update({
            f'{qclass}_freq': freq,
            f'{qclass}_clicks': clicks,
            f'{qclass}_orders': orders,
            f'{qclass}_click_share_pct': (clicks / freq * 100.0) if freq > 0 else 0.0,
            f'{qclass}_visibility_pct': _weighted_mean(g['visibility_pct'], g['query_freq']),
            f'{qclass}_position': _weighted_mean(g['median_position'].replace(0, np.nan).fillna(0.0), g['query_freq']),
        })
    snap.update(out)
    return snap


def _ad_snapshot(ads_daily: pd.DataFrame, campaigns: pd.DataFrame, econ_latest: pd.DataFrame, campaign_id: int, nm_id: int, start_date: date, end_date: date) -> Dict[str, float]:
    cols = {
        'impressions': 0.0, 'clicks': 0.0, 'orders': 0.0, 'spend': 0.0, 'revenue': 0.0,
        'ctr_pct': 0.0, 'cpo': 0.0, 'drr_pct': 0.0, 'gp_after_ads': 0.0
    }
    if ads_daily.empty or campaign_id <= 0 or nm_id <= 0:
        return cols
    df = ads_daily[(ads_daily['date'] >= start_date) & (ads_daily['date'] <= end_date)].copy()
    if df.empty:
        return cols
    df['id_campaign'] = pd.to_numeric(df.get('id_campaign'), errors='coerce')
    df['nmId'] = pd.to_numeric(df.get('nmId'), errors='coerce')
    df = df[(df['id_campaign'] == campaign_id) & (df['nmId'] == nm_id)].copy()
    if df.empty:
        return cols
    gp_map = latest_econ_rows(econ_latest, ['nmId', 'gp_realized']) if isinstance(econ_latest, pd.DataFrame) and not econ_latest.empty else pd.DataFrame(columns=['nmId', 'gp_realized'])
    gp_realized = 0.0
    if not gp_map.empty and 'gp_realized' in gp_map.columns:
        gp_match = gp_map[pd.to_numeric(gp_map['nmId'], errors='coerce') == nm_id]
        if not gp_match.empty:
            gp_realized = safe_float(gp_match['gp_realized'].iloc[0])
    impressions = safe_float(pd.to_numeric(df.get('Показы'), errors='coerce').fillna(0.0).sum())
    clicks = safe_float(pd.to_numeric(df.get('Клики'), errors='coerce').fillna(0.0).sum())
    orders = safe_float(pd.to_numeric(df.get('Заказы'), errors='coerce').fillna(0.0).sum())
    spend = safe_float(pd.to_numeric(df.get('Расход'), errors='coerce').fillna(0.0).sum())
    revenue = safe_float(pd.to_numeric(df.get('Сумма заказов'), errors='coerce').fillna(0.0).sum())
    cols.update({
        'impressions': impressions,
        'clicks': clicks,
        'orders': orders,
        'spend': spend,
        'revenue': revenue,
        'ctr_pct': (clicks / impressions * 100.0) if impressions > 0 else 0.0,
        'cpo': (spend / orders) if orders > 0 else 0.0,
        'drr_pct': (spend / revenue * 100.0) if revenue > 0 else 0.0,
        'gp_after_ads': orders * gp_realized - spend,
    })
    return cols


def _build_bid_change_events(hist: pd.DataFrame) -> pd.DataFrame:
    if hist is None or hist.empty:
        return pd.DataFrame(columns=['run_ts', 'change_date', 'id_campaign', 'nmId', 'supplier_article', 'campaign_type', 'old_bid', 'new_bid', 'direction'])
    df = hist.copy()
    if 'Старая ставка, ₽' in df.columns and 'Новая ставка, ₽' in df.columns:
        df['old_bid'] = pd.to_numeric(df.get('Старая ставка, ₽'), errors='coerce')
        df['new_bid'] = pd.to_numeric(df.get('Новая ставка, ₽'), errors='coerce')
        df['direction'] = np.where(df['new_bid'] > df['old_bid'], 'up', np.where(df['new_bid'] < df['old_bid'], 'down', 'hold'))
        df['campaign_type'] = df.get('campaign_type', df.get('Тип кампании', ''))
        df['supplier_article'] = df.get('Артикул продавца', df.get('supplier_article', '')).fillna('').astype(str)
        events = df[df['direction'].isin(['up', 'down'])].copy()
        if events.empty:
            return pd.DataFrame(columns=['run_ts', 'change_date', 'id_campaign', 'nmId', 'supplier_article', 'campaign_type', 'old_bid', 'new_bid', 'direction'])
        events['change_date'] = pd.to_datetime(events.get('run_ts'), errors='coerce').dt.date
        events = events.dropna(subset=['change_date']).copy()
        return events[['run_ts', 'change_date', 'id_campaign', 'nmId', 'supplier_article', 'campaign_type', 'old_bid', 'new_bid', 'direction']].drop_duplicates()
    df['supplier_article'] = df.get('supplier_article', '').fillna('').astype(str)
    df = df.sort_values(['id_campaign', 'nmId', 'run_ts'])
    parts = []
    for (cid, nm), g in df.groupby(['id_campaign', 'nmId']):
        gg = g.sort_values('run_ts').copy()
        gg['old_bid'] = gg['bid_rub'].shift(1)
        gg['new_bid'] = gg['bid_rub']
        gg['direction'] = np.where(gg['new_bid'] > gg['old_bid'], 'up', np.where(gg['new_bid'] < gg['old_bid'], 'down', 'hold'))
        parts.append(gg)
    events = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()
    if events.empty:
        return pd.DataFrame(columns=['run_ts', 'change_date', 'id_campaign', 'nmId', 'supplier_article', 'campaign_type', 'old_bid', 'new_bid', 'direction'])
    events = events[events['direction'].isin(['up', 'down'])].copy()
    events['change_date'] = pd.to_datetime(events.get('run_ts'), errors='coerce').dt.date
    return events[['run_ts', 'change_date', 'id_campaign', 'nmId', 'supplier_article', 'campaign_type', 'old_bid', 'new_bid', 'direction']].dropna(subset=['change_date'])


def _evaluate_bid_change_effects(events: pd.DataFrame, ads_daily: pd.DataFrame, campaigns: pd.DataFrame, econ_latest: pd.DataFrame, kw_daily: pd.DataFrame, class_map: pd.DataFrame) -> pd.DataFrame:
    if events is None or events.empty:
        return pd.DataFrame([{'Комментарий': 'Нет истории изменений ставок для оценки эффекта'}])
    max_ads_date = pd.to_datetime(ads_daily.get('date'), errors='coerce').dt.date.max() if isinstance(ads_daily, pd.DataFrame) and not ads_daily.empty else None
    out_rows = []
    for _, ev in events.sort_values('run_ts').iterrows():
        cid = safe_int(ev.get('id_campaign'))
        nm = safe_int(ev.get('nmId'))
        sa = str(ev.get('supplier_article') or '')
        ch_date = ev.get('change_date')
        if cid <= 0 or nm <= 0 or not isinstance(ch_date, date):
            continue
        pre_start = ch_date - timedelta(days=3)
        pre_end = ch_date - timedelta(days=1)
        post1_start = ch_date + timedelta(days=1)
        post1_end = min(ch_date + timedelta(days=1), max_ads_date) if max_ads_date else ch_date + timedelta(days=1)
        post3_start = ch_date + timedelta(days=1)
        post3_end = min(ch_date + timedelta(days=3), max_ads_date) if max_ads_date else ch_date + timedelta(days=3)
        if max_ads_date and post3_start > max_ads_date:
            continue

        ad_pre = _ad_snapshot(ads_daily, campaigns, econ_latest, cid, nm, pre_start, pre_end)
        ad_post1 = _ad_snapshot(ads_daily, campaigns, econ_latest, cid, nm, post1_start, post1_end)
        ad_post3 = _ad_snapshot(ads_daily, campaigns, econ_latest, cid, nm, post3_start, post3_end)
        q_pre = _query_group_snapshot(kw_daily, class_map, nm, sa, pre_start, pre_end)
        q_post1 = _query_group_snapshot(kw_daily, class_map, nm, sa, post1_start, post1_end)
        q_post3 = _query_group_snapshot(kw_daily, class_map, nm, sa, post3_start, post3_end)

        primary_share_pre = safe_float(q_pre.get('primary_click_share_pct'))
        primary_share_post1 = safe_float(q_post1.get('primary_click_share_pct'))
        primary_share_post3 = safe_float(q_post3.get('primary_click_share_pct'))
        primary_vis_pre = safe_float(q_pre.get('primary_visibility_pct'))
        primary_vis_post1 = safe_float(q_post1.get('primary_visibility_pct'))
        primary_vis_post3 = safe_float(q_post3.get('primary_visibility_pct'))
        primary_pos_pre = safe_float(q_pre.get('primary_position'))
        primary_pos_post1 = safe_float(q_post1.get('primary_position'))
        primary_pos_post3 = safe_float(q_post3.get('primary_position'))
        market_growth_primary_pct = growth_pct(safe_float(q_post3.get('primary_freq')), safe_float(q_pre.get('primary_freq')))

        if ev.get('direction') == 'up':
            if (primary_share_post1 <= primary_share_pre * 1.01) and (primary_vis_post1 <= primary_vis_pre + 1.0) and (safe_float(ad_post1.get('clicks')) <= safe_float(ad_pre.get('clicks')) * 1.03):
                effect = 'рост не дал долю'
                verdict = 'неуспешно'
            elif (safe_float(ad_post3.get('orders')) <= safe_float(ad_pre.get('orders')) * 1.02) and (safe_float(ad_post3.get('gp_after_ads')) <= safe_float(ad_pre.get('gp_after_ads')) * 1.02) and (primary_share_post3 <= primary_share_pre * 1.02):
                effect = 'рост не привёл к заказам/ВП'
                verdict = 'неуспешно'
            elif (primary_share_post3 > primary_share_pre * 1.05) or (primary_vis_post3 > primary_vis_pre + 3.0) or (primary_pos_post3 > 0 and primary_pos_pre > 0 and primary_pos_post3 < primary_pos_pre - 0.5):
                effect = 'рост оправдан'
                verdict = 'успешно'
            else:
                effect = 'недостаточно данных'
                verdict = 'ожидание'
        else:
            drr_pre = safe_float(ad_pre.get('drr_pct'))
            drr_post3 = safe_float(ad_post3.get('drr_pct'))
            if (primary_share_post3 < primary_share_pre * 0.90) and (safe_float(ad_post3.get('orders')) < safe_float(ad_pre.get('orders')) * 0.90) and (safe_float(ad_post3.get('gp_after_ads')) < safe_float(ad_pre.get('gp_after_ads')) * 0.90) and (drr_post3 >= drr_pre - 0.5):
                effect = 'снижение вредно — потеряли ключевую долю'
                verdict = 'неуспешно'
            elif (primary_share_post3 >= primary_share_pre * 0.95) and ((drr_post3 < drr_pre - 0.3) or ((safe_float(ad_pre.get('cpo')) > 0) and (safe_float(ad_post3.get('cpo')) < safe_float(ad_pre.get('cpo')) * 0.95))):
                effect = 'снижение оправдано'
                verdict = 'успешно'
            else:
                effect = 'недостаточно данных'
                verdict = 'ожидание'

        out_rows.append({
            'Дата изменения': ch_date,
            'ID кампании': cid,
            'Артикул WB': nm,
            'Артикул продавца': sa,
            'Тип кампании': ev.get('campaign_type', ''),
            'Направление': ev.get('direction', ''),
            'Старая ставка, ₽': round(safe_float(ev.get('old_bid')), 2),
            'Новая ставка, ₽': round(safe_float(ev.get('new_bid')), 2),
            'Показы до': round(safe_float(ad_pre.get('impressions')), 0),
            'Показы D+1': round(safe_float(ad_post1.get('impressions')), 0),
            'Показы D+3': round(safe_float(ad_post3.get('impressions')), 0),
            'Клики до': round(safe_float(ad_pre.get('clicks')), 0),
            'Клики D+1': round(safe_float(ad_post1.get('clicks')), 0),
            'Клики D+3': round(safe_float(ad_post3.get('clicks')), 0),
            'Заказы до': round(safe_float(ad_pre.get('orders')), 2),
            'Заказы D+3': round(safe_float(ad_post3.get('orders')), 2),
            'ВП до, ₽': round(safe_float(ad_pre.get('gp_after_ads')), 2),
            'ВП D+3, ₽': round(safe_float(ad_post3.get('gp_after_ads')), 2),
            'ДРР до, %': round(safe_float(ad_pre.get('drr_pct')), 2),
            'ДРР D+3, %': round(safe_float(ad_post3.get('drr_pct')), 2),
            'Ключевая доля кликов до, %': round(primary_share_pre, 2),
            'Ключевая доля кликов D+1, %': round(primary_share_post1, 2),
            'Ключевая доля кликов D+3, %': round(primary_share_post3, 2),
            'Ключевая видимость до, %': round(primary_vis_pre, 2),
            'Ключевая видимость D+1, %': round(primary_vis_post1, 2),
            'Ключевая видимость D+3, %': round(primary_vis_post3, 2),
            'Ключевая позиция до': round(primary_pos_pre, 2),
            'Ключевая позиция D+1': round(primary_pos_post1, 2),
            'Ключевая позиция D+3': round(primary_pos_post3, 2),
            'Рост частотности ключевых запросов, %': round(market_growth_primary_pct, 2),
            'Вывод': effect,
            'Статус': verdict,
        })
    if not out_rows:
        return pd.DataFrame([{'Комментарий': 'Нет валидных событий изменения ставки для оценки'}])
    out = pd.DataFrame(out_rows).sort_values(['Дата изменения', 'ID кампании', 'Артикул WB'], ascending=[False, True, True])
    return out


def _apply_decision_guardrails(decisions: pd.DataFrame, effects: pd.DataFrame) -> pd.DataFrame:
    if decisions is None or decisions.empty:
        return decisions
    df = decisions.copy()
    for c in ['Максимальная ставка, ₽', 'ДРР кампании, %', 'ВП кампании текущее окно после рекламы, ₽', 'ДРР категории, %', 'CPO кампании, ₽']:
        if c not in df.columns:
            df[c] = 0.0
        df[c] = pd.to_numeric(df[c], errors='coerce')
    if 'Минимальная ставка WB, ₽' in df.columns:
        df['Минимальная ставка WB, ₽'] = pd.to_numeric(df.get('Минимальная ставка WB, ₽'), errors='coerce')
    else:
        df['Минимальная ставка WB, ₽'] = np.nan

    if isinstance(effects, pd.DataFrame) and not effects.empty and 'Дата изменения' in effects.columns and 'ID кампании' in effects.columns and 'Артикул WB' in effects.columns:
        eff = effects.copy()
        eff['Дата изменения'] = pd.to_datetime(eff['Дата изменения'], errors='coerce')
        eff = eff.sort_values('Дата изменения').drop_duplicates(subset=['ID кампании', 'Артикул WB'], keep='last')
        keep = ['ID кампании', 'Артикул WB', 'Вывод', 'Статус', 'Ключевая доля кликов до, %', 'Ключевая доля кликов D+3, %']
        eff = eff[[c for c in keep if c in eff.columns]]
        df = df.merge(eff, on=['ID кампании', 'Артикул WB'], how='left')
    else:
        df['Вывод'] = ''
        df['Статус'] = ''

    no_max_raise = df['Действие'].astype(str).isin(['Повысить', 'Тест роста']) & (df['Максимальная ставка, ₽'].fillna(0) <= 0)
    df.loc[no_max_raise, 'Действие'] = 'Без изменений'
    df.loc[no_max_raise, 'Новая ставка, ₽'] = df.loc[no_max_raise, 'Текущая ставка, ₽']
    df.loc[no_max_raise, 'Причина'] = df.loc[no_max_raise, 'Причина'].astype(str) + ' | Рост запрещён: не рассчитана максимальная ставка'

    crisis = pd.to_numeric(df.get('ДРР категории, %'), errors='coerce').fillna(0.0) > 10.0
    expensive_campaign = pd.to_numeric(df.get('ДРР кампании, %'), errors='coerce').fillna(0.0) > 10.0
    negative_gp = pd.to_numeric(df.get('ВП кампании текущее окно после рекламы, ₽'), errors='coerce').fillna(0.0) < 0.0
    current_bid = pd.to_numeric(df.get('Текущая ставка, ₽'), errors='coerce').fillna(0.0)
    min_bid = pd.to_numeric(df.get('Минимальная ставка WB, ₽'), errors='coerce')
    type_series = df.get('Тип оплаты', df.get('Тип кампании', '')).astype(str).str.lower()
    default_floor = np.where(type_series.str.contains('cpc'), 4.0, 80.0)
    floor = np.where(min_bid.fillna(0) > 0, min_bid.fillna(0), default_floor)
    bid_step = np.where(type_series.str.contains('cpc'), 1.0, 6.0)

    force_down = crisis & expensive_campaign & (negative_gp | df['Действие'].astype(str).eq('Без изменений')) & (current_bid > floor + 0.009)
    if force_down.any():
        new_bid = np.maximum(current_bid - bid_step, floor)
        df.loc[force_down, 'Действие'] = 'Снизить'
        df.loc[force_down, 'Новая ставка, ₽'] = new_bid[force_down]
        df.loc[force_down, 'Причина'] = df.loc[force_down, 'Причина'].astype(str) + ' | Антикризис: ДРР категории > 10%, дорогую кампанию сушим'

    bad_recent_raise = df['Действие'].astype(str).isin(['Повысить', 'Тест роста']) & df['Вывод'].astype(str).isin(['рост не дал долю', 'рост не привёл к заказам/ВП'])
    df.loc[bad_recent_raise, 'Действие'] = 'Без изменений'
    df.loc[bad_recent_raise, 'Новая ставка, ₽'] = df.loc[bad_recent_raise, 'Текущая ставка, ₽']
    df.loc[bad_recent_raise, 'Причина'] = df.loc[bad_recent_raise, 'Причина'].astype(str) + ' | Повторный рост запрещён: прошлое повышение не улучшило долю/результат'

    bad_recent_cut = df['Действие'].astype(str).eq('Снизить') & df['Вывод'].astype(str).eq('снижение вредно — потеряли ключевую долю')
    df.loc[bad_recent_cut, 'Действие'] = 'Без изменений'
    df.loc[bad_recent_cut, 'Новая ставка, ₽'] = df.loc[bad_recent_cut, 'Текущая ставка, ₽']
    df.loc[bad_recent_cut, 'Причина'] = df.loc[bad_recent_cut, 'Причина'].astype(str) + ' | Снижение не повторяем: прошлое снижение было вредным'

    if 'CPO кампании, ₽' in df.columns:
        df['_block_key'] = df.get('Товар', '').astype(str)
        block_cpo = df.groupby('_block_key', as_index=False).agg(block_cpo_min=('CPO кампании, ₽', 'min'), block_cpo_avg=('CPO кампании, ₽', 'mean'))
        df = df.merge(block_cpo, on='_block_key', how='left')
        risky_raise = crisis & df['Действие'].astype(str).isin(['Повысить', 'Тест роста']) & ((df['ВП кампании текущее окно после рекламы, ₽'].fillna(0) <= 0) | (df['CPO кампании, ₽'].fillna(0) > df['block_cpo_avg'].fillna(np.inf)))
        df.loc[risky_raise, 'Действие'] = 'Без изменений'
        df.loc[risky_raise, 'Новая ставка, ₽'] = df.loc[risky_raise, 'Текущая ставка, ₽']
        df.loc[risky_raise, 'Причина'] = df.loc[risky_raise, 'Причина'].astype(str) + ' | В кризис растим только прибыльные кампании с CPO не хуже среднего по блоку'
        df = df.drop(columns=['_block_key', 'block_cpo_min', 'block_cpo_avg'], errors='ignore')

    return df


def _build_traffic_share_summary(kw_daily: pd.DataFrame, class_map: pd.DataFrame, top20: pd.DataFrame, as_of_date: date) -> pd.DataFrame:
    if kw_daily.empty or class_map.empty or top20.empty:
        return pd.DataFrame([{'Комментарий': 'Нет данных для сводки по доле трафика'}])
    cur_start = as_of_date - timedelta(days=6)
    prev_start = as_of_date - timedelta(days=13)
    prev_end = as_of_date - timedelta(days=7)
    df = kw_daily.merge(class_map, on=['nmId', 'supplier_article', 'query_text'], how='left')
    df['query_class'] = df['query_class'].fillna('secondary')
    df = df.merge(top20[['nmId', 'supplier_article']], on=['nmId', 'supplier_article'], how='inner')
    out_rows = []
    for (nm, sa), g in df.groupby(['nmId', 'supplier_article']):
        g_cur = g[(g['date'] >= cur_start) & (g['date'] <= as_of_date)].copy()
        g_prev = g[(g['date'] >= prev_start) & (g['date'] <= prev_end)].copy()
        if g_cur.empty and g_prev.empty:
            continue
        def _agg(x, cls):
            y = x[x['query_class'] == cls].copy()
            freq = safe_float(y['query_freq'].sum())
            clicks = safe_float(y['keyword_clicks'].sum())
            orders = safe_float(y['keyword_orders'].sum())
            vis = _weighted_mean(y['visibility_pct'], y['query_freq']) if not y.empty else 0.0
            pos = _weighted_mean(y['median_position'], y['query_freq']) if not y.empty else 0.0
            return freq, clicks, orders, (clicks / freq * 100.0) if freq > 0 else 0.0, vis, pos
        cur_p = _agg(g_cur, 'primary')
        prev_p = _agg(g_prev, 'primary')
        cur_s = _agg(g_cur, 'secondary')
        prev_s = _agg(g_prev, 'secondary')
        out_rows.append({
            'Артикул WB': safe_int(nm),
            'Артикул продавца': sa,
            'Ключевая частотность текущая': round(cur_p[0], 0),
            'Ключевая частотность база': round(prev_p[0], 0),
            'Ключевая доля кликов текущая, %': round(cur_p[3], 2),
            'Ключевая доля кликов база, %': round(prev_p[3], 2),
            'Ключевая видимость текущая, %': round(cur_p[4], 2),
            'Ключевая видимость база, %': round(prev_p[4], 2),
            'Ключевая позиция текущая': round(cur_p[5], 2),
            'Ключевая позиция база': round(prev_p[5], 2),
            'Вторичная доля кликов текущая, %': round(cur_s[3], 2),
            'Вторичная доля кликов база, %': round(prev_s[3], 2),
            'Рост ключевой частотности, %': round(growth_pct(cur_p[0], prev_p[0]), 2),
            'Рост ключевой доли кликов, %': round(growth_pct(cur_p[3], prev_p[3]), 2) if prev_p[3] > 0 else round(cur_p[3], 2),
            'Рост заказов по ключевым, %': round(growth_pct(cur_p[2], prev_p[2]), 2),
        })
    if not out_rows:
        return pd.DataFrame([{'Комментарий': 'Нет данных для сводки по доле трафика'}])
    return pd.DataFrame(out_rows).sort_values(['Рост заказов по ключевым, %', 'Рост ключевой доли кликов, %'], ascending=[False, False])


def _load_pause_history_any(provider: BaseProvider) -> pd.DataFrame:
    for candidate in [OUT_SINGLE_REPORT, OUT_PREVIEW]:
        try:
            if provider.file_exists(candidate):
                sheets = provider.read_excel_all_sheets(candidate)
                for sh in ['История пауз', 'История_пауз']:
                    if sh in sheets:
                        return sheets[sh].copy()
        except Exception:
            continue
    return pd.DataFrame(columns=['Дата', 'ID кампании', 'Артикул продавца', 'Статус', 'Причина'])


def _build_pause_candidates(decisions: pd.DataFrame) -> pd.DataFrame:
    if decisions is None or decisions.empty:
        return pd.DataFrame([{'Комментарий': 'Нет данных для кандидатов на паузу'}])
    df = decisions.copy()
    df['campaign_gp'] = pd.to_numeric(df.get('ВП кампании текущее окно после рекламы, ₽'), errors='coerce').fillna(0.0)
    df['campaign_drr'] = pd.to_numeric(df.get('ДРР кампании, %'), errors='coerce').fillna(0.0)
    df['current_bid'] = pd.to_numeric(df.get('Текущая ставка, ₽'), errors='coerce').fillna(0.0)
    df['min_bid'] = pd.to_numeric(df.get('Минимальная ставка WB, ₽'), errors='coerce').fillna(0.0)
    df['Показы'] = pd.to_numeric(df.get('Показы'), errors='coerce').fillna(0.0)
    df['block_key'] = df.get('Товар', '').astype(str)
    is_brush = df.get('Предмет', '').astype(str).str.lower().eq('кисти косметические') | df.get('Артикул продавца', '').astype(str).str.startswith('901')
    df['CPO кампании, ₽'] = pd.to_numeric(df.get('CPO кампании, ₽'), errors='coerce').fillna(np.inf)
    df['rank_in_block'] = df.groupby('block_key')['CPO кампании, ₽'].rank(method='dense', ascending=True)
    candidate = (~is_brush) & (df['rank_in_block'] > 1) & (df['campaign_gp'] < 0) & (df['campaign_drr'] >= 18.0) & (df['current_bid'] <= df['min_bid'].replace(0, np.nan).fillna(df['current_bid'])) & (df['Показы'] >= 10000)
    out = df.loc[candidate, ['ID кампании', 'Артикул продавца', 'Предмет', 'Товар', 'Плейсмент', 'Текущая ставка, ₽', 'Минимальная ставка WB, ₽', 'Показы', 'ДРР кампании, %', 'ВП кампании текущее окно после рекламы, ₽']].copy()
    if out.empty:
        return pd.DataFrame([{'Комментарий': 'Нет кандидатов на паузу по текущим правилам'}])
    out['Причина'] = 'Альтернативная кампания: отрицательная ВП на минимальной ставке, ДРР >= 18%, высокий объём показов'
    out['Рекомендация'] = 'Кандидат на паузу'
    return out.sort_values(['Предмет', 'Товар', 'ДРР кампании, %'], ascending=[True, True, False])


def prepare_metrics(provider: BaseProvider, cfg: Config, as_of_date: date) -> Dict[str, Any]:
    results = _BASE_prepare_metrics(provider, cfg, as_of_date)
    try:
        ads_daily, campaigns = load_ads(provider)
        keywords = load_keywords(provider)
        econ = load_economics(provider)
        econ_latest = latest_econ_rows(econ, ['nmId', 'supplier_article', 'gp_realized']) if not econ.empty else pd.DataFrame(columns=['nmId', 'supplier_article', 'gp_realized'])
        kw_daily = _keyword_daily_unique(keywords)
        class_map, top20 = _classify_primary_queries(kw_daily, as_of_date)
        bid_hist = _read_bid_history_any(provider)
        events = _build_bid_change_events(bid_hist)
        effects = _evaluate_bid_change_effects(events, ads_daily, campaigns, econ_latest, kw_daily, class_map)
        top20_hist = _build_top20_query_history(kw_daily, class_map, top20)
        traffic_share = _build_traffic_share_summary(kw_daily, class_map, top20, as_of_date)
        decisions = _apply_decision_guardrails(results.get('decisions', pd.DataFrame()), effects)
        results['decisions'] = decisions
        results['История_изменений_ставок'] = normalize_output_df(events.rename(columns={
            'run_ts': 'Дата запуска', 'change_date': 'Дата изменения', 'id_campaign': 'ID кампании',
            'nmId': 'Артикул WB', 'supplier_article': 'Артикул продавца', 'campaign_type': 'Тип кампании',
            'old_bid': 'Старая ставка, ₽', 'new_bid': 'Новая ставка, ₽', 'direction': 'Направление'
        }) if not events.empty else pd.DataFrame([{'Комментарий': 'Нет истории изменений ставок'}]))
        results['Эффект_изменения_ставки'] = normalize_output_df(effects)
        results['История_ключевых_запросов'] = normalize_output_df(top20_hist)
        results['Сводка_по_доле_трафика'] = normalize_output_df(traffic_share)
        results['Кандидаты_на_паузу'] = normalize_output_df(_build_pause_candidates(decisions))
        results['История_пауз'] = normalize_output_df(_load_pause_history_any(provider))
    except Exception as e:
        msg = str(e)
        results['Эффект_изменения_ставки'] = pd.DataFrame([{'Комментарий': f'Не удалось построить post-check изменения ставок: {msg}'}])
        results['История_ключевых_запросов'] = pd.DataFrame([{'Комментарий': f'Не удалось построить историю ключевых запросов: {msg}'}])
        results['Сводка_по_доле_трафика'] = pd.DataFrame([{'Комментарий': f'Не удалось построить сводку доли трафика: {msg}'}])
        results['Кандидаты_на_паузу'] = pd.DataFrame([{'Комментарий': f'Не удалось построить кандидатов на паузу: {msg}'}])
        results['История_пауз'] = pd.DataFrame([{'Комментарий': 'История пауз пока пуста'}])
    return results


def save_outputs(provider: BaseProvider, results: Dict[str, Any], run_mode: str, bid_send_log: Optional[pd.DataFrame], shade_apply_log: Optional[pd.DataFrame], history_append: pd.DataFrame) -> None:
    _BASE_save_outputs(provider, results, run_mode, bid_send_log, shade_apply_log, history_append)
    extra_sheet_names = [
        'История_изменений_ставок',
        'Эффект_изменения_ставки',
        'История_ключевых_запросов',
        'Сводка_по_доле_трафика',
        'Кандидаты_на_паузу',
        'История_пауз',
    ]
    extra = {name: normalize_output_df(results.get(name, pd.DataFrame([{'Комментарий': 'Нет данных'}]))) for name in extra_sheet_names}
    try:
        sheets = provider.read_excel_all_sheets(OUT_SINGLE_REPORT) if provider.file_exists(OUT_SINGLE_REPORT) else {}
    except Exception:
        sheets = {}
    sheets.update(extra)
    provider.write_excel(OUT_SINGLE_REPORT, sheets)
    try:
        provider.write_excel(OUT_PREVIEW, sheets)
    except Exception:
        pass
    try:
        hist_sheet = sheets.get('История ставок', sheets.get('История_ставок', pd.DataFrame()))
        if hist_sheet is not None and not hist_sheet.empty:
            provider.write_excel(OUT_BID_HISTORY, {'История ставок': hist_sheet})
    except Exception:
        pass

def main() -> None:
    args = build_parser().parse_args()
    args.mode = 'run'
    run_manager(args)

if __name__ == '__main__':
    main()
