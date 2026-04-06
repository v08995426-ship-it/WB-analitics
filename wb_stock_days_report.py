#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import io
import math
import os
import re
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, Iterable, Optional, Sequence

import boto3
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

STORE_NAME = "TOPFACE"
WB_STOCKS_PREFIX = f"Отчёты/Остатки/{STORE_NAME}/Недельные/"
WB_ORDERS_PREFIX = f"Отчёты/Заказы/{STORE_NAME}/Недельные/"
ARTICLE_MAP_KEY = "Отчёты/Остатки/1С/Артикулы 1с.xlsx"
STOCKS_1C_KEY = "Отчёты/Остатки/1С/Остатки 1С.xlsx"
RRC_KEY = f"Отчёты/Финансовые показатели/{STORE_NAME}/РРЦ.xlsx"
INBOUND_PREFIX = "Отчёты/Остатки/1С/"
ABC_NAME_FRAGMENT = "abc_report_goods"
OUT_DIR = "output"

SHEET_CRITICAL = "Критично <14 дней"
SHEET_CALC = "Расчёт"
SHEET_DEAD = "Dead_Stock"
SHEET_MONITOR = "Мониторинг остатков"

FONT_NAME = "Calibri"
FONT_SIZE = 14

FILL_HEADER = PatternFill("solid", fgColor="1F4E78")
FILL_LIGHT_GREEN = PatternFill("solid", fgColor="CCFFCC")
FILL_BLACK = PatternFill("solid", fgColor="000000")
FILL_ORANGE = PatternFill("solid", fgColor="FCE4D6")
FILL_BLUE_ROW = PatternFill("solid", fgColor="DDEBF7")

BORDER_THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

REDISTRIBUTION_SHEET = "Перераспределения"
REDISTRIBUTION_WAREHOUSES_SHEET = "Склады"
REDISTRIBUTION_INSTRUCTION_SHEET = "Инструкция"

MOSCOW_CLUSTER_GROUP = "__MOSCOW_CLUSTER__"
MOSCOW_CLUSTER_WEIGHTS: dict[str, float] = {
    "Коледино": 0.5,
    "Электросталь": 0.5,
}
CENTRAL_HUBS: tuple[str, ...] = ("Коледино", "Электросталь", "Белые Столбы")

WAREHOUSE_ALIASES_REDISTRIBUTION: dict[str, str] = {
    "Москва": "Коледино",
    "Самара (Новосемейкино)": "Новосемейкино",
    "Самара Новосемейкино": "Новосемейкино",
    "Санкт-Петербург Уткина Заводь": "СПБ Шушары",
    "СПб Уткина Заводь": "СПБ Шушары",
    "СПБ Уткина Заводь": "СПБ Шушары",
    "Санкт Петербург Уткина Заводь": "СПБ Шушары",
    "Владимир": "Владимир Воршинское",
    "Рязань": "Рязань (Тюшевское)",
    "Екатеринбург Перспективная 14": "Екатеринбург - Перспективная 14",
    "Екатеринбург - Перспективный 12": "Екатеринбург - Перспективная 14",
    "Екатеринбург - Перспективный 12": "Екатеринбург - Перспективная 14",
    "Екатеринбург - Перспективный 14": "Екатеринбург - Перспективная 14",
    "Екатеринбург - Перспективный 14г": "Екатеринбург - Перспективная 14",
    "Екатеринбург - Испытателей 14г": "Екатеринбург - Испытателей 14г",
    "Владимир Воршинское": "Владимир Воршинское",
}

WAREHOUSE_ZONE: dict[str, str] = {
    "Коледино": "Центр",
    "Электросталь": "Центр",
    "Белые Столбы": "Центр",
    "Тула": "Центр",
    "Рязань (Тюшевское)": "Центр",
    "Котовск": "Центр",
    "Владимир Воршинское": "Центр",
    "Пенза": "Центр",
    "СПБ Шушары": "Северо-Запад",
    "Краснодар": "Юг",
    "Невинномысск": "Юг",
    "Волгоград": "Юг",
    "Казань": "Поволжье",
    "Новосемейкино": "Поволжье",
    "Сарапул": "Поволжье",
    "Екатеринбург - Испытателей 14г": "Урал",
    "Екатеринбург - Перспективная 14": "Урал",
    "Новосибирск": "Сибирь",
}


MP_STOCK_FORMULA_WEIGHTS: dict[str, float] = {
    "Адресный склад": 1.0,
    'Оптовый склад Луганск- ООО "Хайлер"': 0.5,
    'Основной склад - ООО "Хайлер"': 0.5,
}

MANAGER_OVERRIDES_BY_ARTICLE_1C: dict[str, str] = {
    "PT901.F25": "",
    "PT901.F26": "",
    "PT901.F27": "",
    "PT901.F28": "",
    "PT901.SET-1": "",
    "PT810.001": "Игорь",
    "PT811.001": "Игорь",
    "PT554.007K": "Игорь",
    "PT567.001K": "Юлия",
}



@dataclass
class Config:
    bucket: str
    access_key: str
    secret_key: str
    endpoint_url: str
    region_name: str
    telegram_bot_token: str
    telegram_chat_id: str
    stop_articles_raw: str
    force_send: bool
    run_date: date
    redistribution_template_key: str
    redistribution_template_local: str
    send_redistribution_always: bool
    redistribution_days: int
    redistribution_target_days: int


class S3Storage:
    def __init__(self, cfg: Config) -> None:
        self.bucket = cfg.bucket
        self.client = boto3.client(
            "s3",
            endpoint_url=cfg.endpoint_url,
            aws_access_key_id=cfg.access_key,
            aws_secret_access_key=cfg.secret_key,
            region_name=cfg.region_name,
        )

    def list_keys(self, prefix: str) -> list[str]:
        keys: list[str] = []
        token = None
        while True:
            kwargs = {"Bucket": self.bucket, "Prefix": prefix, "MaxKeys": 1000}
            if token:
                kwargs["ContinuationToken"] = token
            resp = self.client.list_objects_v2(**kwargs)
            for item in resp.get("Contents", []):
                key = item["Key"]
                if not key.endswith("/"):
                    keys.append(key)
            if not resp.get("IsTruncated"):
                break
            token = resp.get("NextContinuationToken")
        return keys

    def read_excel(self, key: str, **kwargs) -> pd.DataFrame:
        obj = self.client.get_object(Bucket=self.bucket, Key=key)
        return pd.read_excel(io.BytesIO(obj["Body"].read()), **kwargs)


def log(msg: str) -> None:
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}", flush=True)


def normalize_text(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, float) and float(value).is_integer():
        return str(int(value)).strip()
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".")[0]
    return text


def normalize_key(value: object) -> str:
    return normalize_text(value).upper()


def safe_float(value: object) -> float:
    if value is None or pd.isna(value):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "").replace(",", ".")
    if not text:
        return 0.0
    try:
        return float(text)
    except Exception:
        return 0.0


def round_int(value: object) -> int:
    return int(round(safe_float(value)))


def ceil_int(value: object) -> int:
    return int(math.ceil(safe_float(value)))


def choose_existing_column(df: pd.DataFrame, candidates: Iterable[str], label: str) -> str:
    mapping = {str(c).strip().lower(): c for c in df.columns}
    for candidate in candidates:
        real = mapping.get(candidate.strip().lower())
        if real is not None:
            return real
    raise KeyError(f"Не найдена колонка для '{label}'. Доступные колонки: {list(df.columns)}")


def try_choose_column(df: pd.DataFrame, candidates: Iterable[str]) -> Optional[str]:
    mapping = {str(c).strip().lower(): c for c in df.columns}
    for candidate in candidates:
        real = mapping.get(candidate.strip().lower())
        if real is not None:
            return real
    return None


def parse_stop_articles(raw: str) -> set[str]:
    if not raw:
        return set()
    text = raw.replace("\r", "\n").replace(";", "\n").replace(",", "\n")
    return {normalize_key(x) for x in text.split("\n") if normalize_text(x)}


def parse_iso_week_from_key(key: str) -> tuple[int, int]:
    m = re.search(r"_(\d{4})-W(\d{2})\.xlsx$", key, flags=re.IGNORECASE)
    if not m:
        return (0, 0)
    return int(m.group(1)), int(m.group(2))


def latest_weekly_key(keys: list[str]) -> str:
    xlsx = [k for k in keys if k.lower().endswith(".xlsx")]
    if not xlsx:
        raise FileNotFoundError("Не найдены weekly xlsx файлы")
    return sorted(xlsx, key=parse_iso_week_from_key)[-1]


def latest_n_weekly_keys(keys: list[str], n: int) -> list[str]:
    xlsx = [k for k in keys if k.lower().endswith(".xlsx")]
    return sorted(xlsx, key=parse_iso_week_from_key)[-n:]


def get_config() -> Config:
    bucket = (os.getenv("YC_BUCKET_NAME") or os.getenv("CLOUD_RU_BUCKET") or os.getenv("WB_S3_BUCKET") or "").strip()
    access_key = (os.getenv("YC_ACCESS_KEY_ID") or os.getenv("CLOUD_RU_ACCESS_KEY") or os.getenv("WB_S3_ACCESS_KEY") or "").strip()
    secret_key = (os.getenv("YC_SECRET_ACCESS_KEY") or os.getenv("CLOUD_RU_SECRET_KEY") or os.getenv("WB_S3_SECRET_KEY") or "").strip()
    endpoint_url = (os.getenv("YC_ENDPOINT_URL") or os.getenv("WB_S3_ENDPOINT") or "https://storage.yandexcloud.net").strip()
    region_name = (os.getenv("WB_S3_REGION") or "ru-central1").strip()
    if not bucket or not access_key or not secret_key:
        raise ValueError("Не заданы параметры Object Storage")
    return Config(
        bucket=bucket,
        access_key=access_key,
        secret_key=secret_key,
        endpoint_url=endpoint_url,
        region_name=region_name,
        telegram_bot_token=(os.getenv("TELEGRAM_BOT_TOKEN") or "").strip(),
        telegram_chat_id=(os.getenv("TELEGRAM_CHAT_ID") or "").strip(),
        stop_articles_raw=os.getenv("WB_STOP_LIST_KEY", ""),
        force_send=(os.getenv("WB_FORCE_SEND", "false").strip().lower() == "true"),
        run_date=date.today(),
        redistribution_template_key=(os.getenv("WB_REDISTRIBUTION_TEMPLATE_KEY") or "").strip(),
        redistribution_template_local=(os.getenv("WB_REDISTRIBUTION_TEMPLATE_LOCAL") or "").strip(),
        send_redistribution_always=(os.getenv("WB_SEND_REDISTRIBUTION_ALWAYS", "false").strip().lower() == "true"),
        redistribution_days=max(int(os.getenv("WB_REDISTRIBUTION_LOOKBACK_DAYS", "14") or 14), 1),
        redistribution_target_days=max(int(os.getenv("WB_REDISTRIBUTION_TARGET_DAYS", "21") or 21), 1),
    )


def should_send_report(cfg: Config) -> bool:
    if cfg.force_send:
        return True
    return cfg.run_date.weekday() in (0, 4)


def should_send_redistribution(cfg: Config) -> bool:
    if cfg.force_send or cfg.send_redistribution_always:
        return True
    return cfg.run_date.weekday() == 0


def load_article_map(storage: S3Storage) -> dict[str, str]:
    df = storage.read_excel(ARTICLE_MAP_KEY)
    wb_col = df.columns[0]
    article_col = df.columns[2]
    temp = df[[wb_col, article_col]].copy()
    temp.columns = ["Артикул WB", "Артикул 1С"]
    temp["Артикул WB"] = temp["Артикул WB"].map(normalize_key)
    temp["Артикул 1С"] = temp["Артикул 1С"].map(normalize_text)
    temp = temp[(temp["Артикул WB"] != "") & (temp["Артикул 1С"] != "")]
    temp = temp.drop_duplicates(subset=["Артикул WB"], keep="first")
    mapping = dict(zip(temp["Артикул WB"], temp["Артикул 1С"]))
    log(f"Загружено соответствий WB -> 1С: {len(mapping)}")
    return mapping



def load_stocks_1c(storage: S3Storage) -> pd.DataFrame:
    df = storage.read_excel(STOCKS_1C_KEY)
    article_col = choose_existing_column(df, ["Артикул", "АРТ", "Артикул 1С"], "Артикул 1С")

    legacy_stock_col = try_choose_column(df, ["Остатки МП", "Остатки МП (Липецк), шт", "Остатки МП(Липецк), шт"])
    if legacy_stock_col is not None:
        stock_series = pd.to_numeric(df[legacy_stock_col], errors="coerce").fillna(0)
        log(f"1С остатки МП взяты из legacy-колонки: {legacy_stock_col}")
    else:
        selected_columns: list[tuple[str, float]] = []
        missing_columns: list[str] = []
        for column_name, weight in MP_STOCK_FORMULA_WEIGHTS.items():
            resolved = try_choose_column(df, [column_name])
            if resolved is None:
                missing_columns.append(column_name)
                continue
            selected_columns.append((resolved, weight))

        if not selected_columns:
            raise KeyError(
                "Не найдены колонки для расчёта 'Остатки МП' по формуле. "
                f"Доступные колонки: {list(df.columns)}"
            )

        stock_series = pd.Series(0.0, index=df.index)
        for col_name, weight in selected_columns:
            stock_series = stock_series.add(
                pd.to_numeric(df[col_name], errors="coerce").fillna(0) * float(weight),
                fill_value=0,
            )
        log(
            "1С остатки МП собраны по формуле: "
            + " + ".join([f"{col}*{weight:g}" for col, weight in selected_columns])
            + (f" | отсутствуют колонки: {missing_columns}" if missing_columns else "")
        )

    temp = pd.DataFrame({
        "Артикул 1С": df[article_col].map(normalize_text),
        "Остатки МП (Липецк), шт": stock_series.map(ceil_int),
    })
    temp = temp[temp["Артикул 1С"] != ""]
    temp = temp.groupby("Артикул 1С", as_index=False, dropna=False)["Остатки МП (Липецк), шт"].sum()
    return temp

def load_rrc(storage: S3Storage) -> pd.DataFrame:
    df = storage.read_excel(RRC_KEY)
    article_col = df.columns[0]
    rrc_col = df.columns[3]
    temp = pd.DataFrame({
        "Артикул 1С": df[article_col].map(normalize_text),
        "РРЦ": df[rrc_col].map(round_int),
    })
    temp = temp[temp["Артикул 1С"] != ""]
    return temp.drop_duplicates(subset=["Артикул 1С"], keep="first")


def load_abc_managers(storage: S3Storage) -> pd.DataFrame:
    try:
        keys = [k for k in storage.list_keys("") if k.lower().endswith(".xlsx") and ABC_NAME_FRAGMENT in os.path.basename(k).lower()]
        if not keys:
            return pd.DataFrame(columns=["Артикул WB", "Артикул WB продавца", "Менеджер"])
        key = sorted(keys)[-1]
        log(f"Берём ABC-отчёт: {key}")
        df = storage.read_excel(key)
        wb_col = choose_existing_column(df, ["Артикул WB"], "Артикул WB в ABC")
        seller_col = choose_existing_column(df, ["Артикул продавца"], "Артикул продавца в ABC")
        mgr_col = choose_existing_column(df, ["Ваша категория"], "Ваша категория в ABC")
        temp = pd.DataFrame({
            "Артикул WB": df[wb_col].map(normalize_key),
            "Артикул WB продавца": df[seller_col].map(normalize_text),
            "Менеджер": df[mgr_col].map(normalize_text),
        })
        temp = temp[temp["Менеджер"] != ""]
        return temp.drop_duplicates(subset=["Артикул WB", "Артикул WB продавца"], keep="first")
    except Exception as exc:
        log(f"ABC-отчёт не загружен: {exc}")
        return pd.DataFrame(columns=["Артикул WB", "Артикул WB продавца", "Менеджер"])


def load_latest_wb_stocks(storage: S3Storage) -> tuple[pd.DataFrame, str]:
    latest_key = latest_weekly_key(storage.list_keys(WB_STOCKS_PREFIX))
    log(f"Берём остатки WB из файла: {latest_key}")
    df = storage.read_excel(latest_key)

    sample_col = choose_existing_column(df, ["Дата сбора", "Дата запроса"], "дата среза")
    df["_sample_dt"] = pd.to_datetime(df[sample_col], errors="coerce")
    latest_dt = df["_sample_dt"].max()
    if pd.notna(latest_dt):
        df = df[df["_sample_dt"] == latest_dt].copy()

    wb_col = choose_existing_column(df, ["Артикул WB", "nmId"], "Артикул WB")
    seller_col = choose_existing_column(df, ["Артикул продавца"], "Артикул продавца")
    stock_col = choose_existing_column(df, ["Доступно для продажи", "Полное количество", "Количество", "Доступно", "Остаток", "Остатки"], "остатка WB")

    temp = pd.DataFrame({
        "Артикул WB": df[wb_col].map(normalize_key),
        "Артикул WB продавца": df[seller_col].map(normalize_text),
        "Остаток WB, шт": df[stock_col].map(round_int),
    })
    temp = temp[(temp["Артикул WB"] != "") | (temp["Артикул WB продавца"] != "")]
    temp = temp.groupby(["Артикул WB", "Артикул WB продавца"], as_index=False)["Остаток WB, шт"].sum()
    return temp, latest_key


def load_orders_metrics(storage: S3Storage) -> tuple[pd.DataFrame, list[str]]:
    keys = latest_n_weekly_keys(storage.list_keys(WB_ORDERS_PREFIX), 10)
    log(f"Берём заказы WB из файлов: {keys}")
    frames: list[pd.DataFrame] = []
    for key in keys:
        df = storage.read_excel(key)
        frames.append(df)
    orders = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    if orders.empty:
        return pd.DataFrame(columns=[
            "Артикул WB", "Артикул WB продавца", "Продажи 7 дней, шт", "Продажи 60 дней, шт",
            "Среднесуточные продажи 7д", "Среднесуточные продажи 60д", "Цена покупателя"
        ]), keys

    wb_col = choose_existing_column(orders, ["nmId", "Артикул WB"], "Артикул WB в заказах")
    seller_col = choose_existing_column(orders, ["supplierArticle", "Артикул продавца"], "Артикул продавца в заказах")
    date_col = choose_existing_column(orders, ["date", "Дата", "Дата заказа", "lastChangeDate", "Дата продажи"], "дата в заказах")

    work = pd.DataFrame({
        "Артикул WB": orders[wb_col].map(normalize_key),
        "Артикул WB продавца": orders[seller_col].map(normalize_text),
        "dt": pd.to_datetime(orders[date_col], errors="coerce").dt.normalize(),
    })
    if "finishedPrice" in orders.columns:
        work["finishedPrice"] = orders["finishedPrice"].map(safe_float)
    else:
        work["finishedPrice"] = 0.0

    work = work[((work["Артикул WB"] != "") | (work["Артикул WB продавца"] != "")) & work["dt"].notna()].copy()
    if work.empty:
        return pd.DataFrame(columns=[
            "Артикул WB", "Артикул WB продавца", "Продажи 7 дней, шт", "Продажи 60 дней, шт",
            "Среднесуточные продажи 7д", "Среднесуточные продажи 60д", "Цена покупателя"
        ]), keys

    max_dt = work["dt"].max()
    start_7 = max_dt - pd.Timedelta(days=6)
    start_60 = max_dt - pd.Timedelta(days=59)
    group_cols = ["Артикул WB", "Артикул WB продавца"]

    sales_7 = work[work["dt"] >= start_7].groupby(group_cols).size().rename("sales_7d")
    sales_60 = work[work["dt"] >= start_60].groupby(group_cols).size().rename("sales_60d")
    metrics = pd.concat([sales_7, sales_60], axis=1).fillna(0).reset_index()
    metrics["sales_7d"] = metrics["sales_7d"].astype(int)
    metrics["sales_60d"] = metrics["sales_60d"].astype(int)
    metrics["avg_daily_sales_7d"] = metrics["sales_7d"] / 7.0
    metrics["avg_daily_sales_60d"] = metrics["sales_60d"] / 60.0

    price_last = work[work["dt"] == max_dt].groupby(group_cols)["finishedPrice"].mean().rename("Цена покупателя").reset_index()
    price_last["Цена покупателя"] = price_last["Цена покупателя"].map(round_int)
    metrics = metrics.merge(price_last, on=group_cols, how="left")
    return metrics, keys



def parse_inbound_base_date(filename: str) -> Optional[date]:
    for pattern, dt_format in (
        (r"(\d{2}-\d{2}-\d{4})", "%d-%m-%Y"),
        (r"(\d{2}-\d{2}-\d{2})(?!\d)", "%d-%m-%y"),
    ):
        m = re.search(pattern, filename)
        if m:
            try:
                return datetime.strptime(m.group(1), dt_format).date()
            except ValueError:
                continue
    return None


def detect_inbound_mp_column(df: pd.DataFrame) -> tuple[Optional[str], int]:
    direct = try_choose_column(df, ["Заказ МП", "ЗаказМП"])
    if direct is not None:
        return direct, 0

    target = "ЗАКАЗМП"
    scan_rows = min(len(df), 5)
    for row_idx in range(scan_rows):
        for col in df.columns:
            cell_value = normalize_text(df.iloc[row_idx][col]).upper().replace(" ", "")
            if cell_value == target:
                return str(col), row_idx + 1
    return None, 0


def load_inbound(storage: S3Storage, run_date: date) -> pd.DataFrame:
    keys = [
        k for k in storage.list_keys(INBOUND_PREFIX)
        if k.lower().endswith(".xlsx") and "в пути" in os.path.basename(k).lower()
    ]
    frames: list[pd.DataFrame] = []

    for key in keys:
        fname = os.path.basename(key)
        base_date = parse_inbound_base_date(fname)
        if base_date is None:
            log(f"Файл 'В пути' пропущен: не смогли распознать дату в имени {fname}")
            continue

        arrival_date = base_date + timedelta(days=14)
        df = storage.read_excel(key)
        if df.empty:
            continue

        code_col = try_choose_column(df, ["CODES", "Артикул 1С", "Артикул"])
        if code_col is None:
            log(f"Файл 'В пути' пропущен: не найдена колонка артикула в {fname}")
            continue

        qty_col, data_start_row = detect_inbound_mp_column(df)
        if qty_col is None:
            log(f"Файл 'В пути' пропущен: не найдена колонка 'Заказ МП' в {fname}")
            continue

        temp = pd.DataFrame({
            "Артикул 1С": df[code_col].map(normalize_text),
            "qty_raw": df[qty_col],
        })
        if data_start_row > 0:
            temp = temp.iloc[data_start_row:].copy()

        temp = temp[
            (~temp["Артикул 1С"].str.upper().isin({"CODES", "КОДЫ"}))
            & (temp["Артикул 1С"] != "")
        ].copy()
        temp["Партия в пути, шт"] = pd.to_numeric(temp["qty_raw"], errors="coerce").fillna(0).map(round_int)
        temp = temp[temp["Партия в пути, шт"] > 0].copy()
        if temp.empty:
            continue

        temp = temp[["Артикул 1С", "Партия в пути, шт"]].copy()
        temp["Дата поступления"] = pd.to_datetime(arrival_date)
        temp["Дней до поступления"] = max((arrival_date - run_date).days, 0)
        temp["Файл в пути"] = fname
        frames.append(temp)

    if not frames:
        return pd.DataFrame(
            columns=[
                "Артикул 1С",
                "Товары в пути, шт",
                "Ближайшее поступление, шт",
                "Дата поступления",
                "Дней до поступления",
                "Партий в пути, шт",
            ]
        )

    all_inbound = pd.concat(frames, ignore_index=True)
    all_inbound["Дата поступления"] = pd.to_datetime(all_inbound["Дата поступления"], errors="coerce")
    all_inbound = all_inbound.sort_values(["Артикул 1С", "Дата поступления", "Файл в пути"]).reset_index(drop=True)

    total_qty = (
        all_inbound.groupby("Артикул 1С", as_index=False)["Партия в пути, шт"]
        .sum()
        .rename(columns={"Партия в пути, шт": "Товары в пути, шт"})
    )

    nearest_rows = []
    for article, part in all_inbound.groupby("Артикул 1С", dropna=False):
        nearest_date = part["Дата поступления"].min()
        nearest_part = part[part["Дата поступления"] == nearest_date]
        nearest_rows.append({
            "Артикул 1С": article,
            "Ближайшее поступление, шт": int(nearest_part["Партия в пути, шт"].sum()),
            "Дата поступления": nearest_date,
            "Дней до поступления": int(nearest_part["Дней до поступления"].min()),
            "Партий в пути, шт": int(part["Дата поступления"].nunique()),
        })

    nearest_df = pd.DataFrame(nearest_rows)
    return total_qty.merge(nearest_df, on="Артикул 1С", how="left")

def load_current_month_zero_days(storage: S3Storage, zero_articles: set[str], avg7_map: dict[str, float], run_date: date) -> dict[str, int]:
    if not zero_articles:
        return {}
    month_start = run_date.replace(day=1)
    rows: list[pd.DataFrame] = []
    for key in sorted(storage.list_keys(WB_STOCKS_PREFIX), key=parse_iso_week_from_key):
        if not key.lower().endswith(".xlsx"):
            continue
        try:
            df = storage.read_excel(key)
        except Exception:
            continue
        wb_col = choose_existing_column(df, ["Артикул WB", "nmId"], "Артикул WB")
        stock_col = choose_existing_column(df, ["Доступно для продажи", "Полное количество"], "остаток WB")
        sample_col = choose_existing_column(df, ["Дата сбора", "Дата запроса"], "дата среза")
        temp = pd.DataFrame({
            "Артикул WB": df[wb_col].map(normalize_key),
            "stock_wb": df[stock_col].map(safe_float),
            "sample_dt": pd.to_datetime(df[sample_col], errors="coerce").dt.normalize(),
        })
        temp = temp[(temp["Артикул WB"].isin(zero_articles)) & temp["sample_dt"].notna()]
        temp = temp[temp["sample_dt"].dt.date >= month_start]
        if temp.empty:
            continue
        temp = temp.groupby(["Артикул WB", "sample_dt"], as_index=False)["stock_wb"].sum()
        rows.append(temp)
    if not rows:
        return {}
    month_df = pd.concat(rows, ignore_index=True)

    def is_zero_like(row: pd.Series) -> bool:
        threshold = 0.5 * float(avg7_map.get(row["Артикул WB"], 0.0) or 0.0)
        return float(row["stock_wb"]) <= threshold

    month_df["is_zero_like"] = month_df.apply(is_zero_like, axis=1)
    return {k: int(v) for k, v in month_df.groupby("Артикул WB")["is_zero_like"].sum().to_dict().items()}


def compute_coef_rrc(price: int, rrc: int) -> str:
    if rrc <= 0 or price <= 0:
        return ""
    return f"{price / rrc:.2f}".replace(".", ",") + "_РРЦ"



def build_report_dataframe(
    wb_stocks: pd.DataFrame,
    sales: pd.DataFrame,
    article_map: dict[str, str],
    stocks_1c: pd.DataFrame,
    stop_articles: set[str],
    rrc_df: pd.DataFrame,
    inbound_df: pd.DataFrame,
    zero_days_map: dict[str, int],
    abc_df: pd.DataFrame,
) -> pd.DataFrame:
    df = wb_stocks.merge(sales, on=["Артикул WB", "Артикул WB продавца"], how="left")
    for col, default in {
        "sales_7d": 0,
        "sales_60d": 0,
        "avg_daily_sales_7d": 0.0,
        "avg_daily_sales_60d": 0.0,
        "Цена покупателя": 0,
    }.items():
        if col not in df.columns:
            df[col] = default
        df[col] = df[col].fillna(default)

    df["Артикул 1С"] = df["Артикул WB"].map(article_map)
    missing = df["Артикул 1С"].isna() | (df["Артикул 1С"].astype(str).str.strip() == "")
    df.loc[missing, "Артикул 1С"] = df.loc[missing, "Артикул WB продавца"]
    df["Артикул 1С"] = df["Артикул 1С"].map(normalize_text)
    df = df[(df["Артикул 1С"] != "") & (~df["Артикул 1С"].str.startswith("PT104", na=False))].copy()

    df = df.merge(stocks_1c, on="Артикул 1С", how="left")
    df["Остатки МП (Липецк), шт"] = df["Остатки МП (Липецк), шт"].fillna(0).map(ceil_int)

    df = df.merge(inbound_df, on="Артикул 1С", how="left")
    for col in ["Товары в пути, шт", "Ближайшее поступление, шт", "Партий в пути, шт"]:
        if col not in df.columns:
            df[col] = 0
        df[col] = df[col].fillna(0).map(round_int)
    df["Дней до поступления"] = pd.to_numeric(df.get("Дней до поступления"), errors="coerce")
    df["Дата поступления"] = pd.to_datetime(df.get("Дата поступления"), errors="coerce")
    no_inbound_mask = df["Товары в пути, шт"] <= 0
    df.loc[no_inbound_mask, "Дней до поступления"] = pd.NA
    df.loc[no_inbound_mask, "Дата поступления"] = pd.NaT
    df.loc[no_inbound_mask, "Ближайшее поступление, шт"] = 0
    df.loc[no_inbound_mask, "Партий в пути, шт"] = 0

    df = df.merge(abc_df, on=["Артикул WB", "Артикул WB продавца"], how="left")
    if "Менеджер" not in df.columns:
        df["Менеджер"] = ""
    df["Менеджер"] = df["Менеджер"].fillna("")
    df["Менеджер"] = df.apply(
        lambda r: MANAGER_OVERRIDES_BY_ARTICLE_1C.get(normalize_text(r.get("Артикул 1С")), normalize_text(r.get("Менеджер"))),
        axis=1,
    )

    df["Продажи 7 дней, шт"] = df["sales_7d"].map(round_int)
    df["Продажи 60 дней, шт"] = df["sales_60d"].map(round_int)
    df["Среднесуточные продажи 7д"] = df["avg_daily_sales_7d"].map(safe_float)
    df["Среднесуточные продажи 60д"] = df["avg_daily_sales_60d"].map(safe_float)

    def daily_demand(row: pd.Series) -> float:
        stock = safe_float(row["Остаток WB, шт"])
        avg7 = safe_float(row["Среднесуточные продажи 7д"])
        avg60 = safe_float(row["Среднесуточные продажи 60д"])
        if stock <= 0 or avg7 <= 0:
            return avg60
        return avg7

    df["Расчётный спрос в день, шт"] = df.apply(daily_demand, axis=1)
    df["WB хватит, дней"] = df.apply(
        lambda r: safe_float(r["Остаток WB, шт"]) / safe_float(r["Расчётный спрос в день, шт"])
        if safe_float(r["Расчётный спрос в день, шт"]) > 0 else 0.0,
        axis=1,
    )
    df["WB + Липецк, дней"] = df.apply(
        lambda r: (safe_float(r["Остаток WB, шт"]) + safe_float(r["Остатки МП (Липецк), шт"])) / safe_float(r["Расчётный спрос в день, шт"])
        if safe_float(r["Расчётный спрос в день, шт"]) > 0 else 0.0,
        axis=1,
    )
    df["После ближайшего поступления, дней"] = df.apply(
        lambda r: (
            safe_float(r["Остаток WB, шт"]) + safe_float(r["Остатки МП (Липецк), шт"]) + safe_float(r["Ближайшее поступление, шт"])
        ) / safe_float(r["Расчётный спрос в день, шт"])
        if safe_float(r["Расчётный спрос в день, шт"]) > 0 else 0.0,
        axis=1,
    )
    df["WB + Липецк + в пути, дней"] = df.apply(
        lambda r: (
            safe_float(r["Остаток WB, шт"]) + safe_float(r["Остатки МП (Липецк), шт"]) + safe_float(r["Товары в пути, шт"])
        ) / safe_float(r["Расчётный спрос в день, шт"])
        if safe_float(r["Расчётный спрос в день, шт"]) > 0 else 0.0,
        axis=1,
    )

    def enough_to_arrival(row: pd.Series) -> str:
        if pd.isna(row["Дней до поступления"]):
            return ""
        if safe_float(row["Расчётный спрос в день, шт"]) <= 0:
            return "Да"
        current_cover = (
            safe_float(row["Остаток WB, шт"]) + safe_float(row["Остатки МП (Липецк), шт"])
        ) / safe_float(row["Расчётный спрос в день, шт"])
        return "Да" if current_cover >= safe_float(row["Дней до поступления"]) else "Нет"

    df["Хватит до поступления"] = df.apply(enough_to_arrival, axis=1)
    df["Out of stock, days"] = df["WB + Липецк + в пути, дней"].map(lambda x: round_int(max(60 - safe_float(x), 0)))
    df["Хватит на 60 дней"] = df["WB + Липецк + в пути, дней"].map(
        lambda x: "Да" if safe_float(x) >= 60 else f"Дефицит {round_int(60 - safe_float(x))} дн."
    )

    df["Дней без остатка WB в текущем месяце"] = df["Артикул WB"].map(zero_days_map).fillna(0).astype(int)
    df.loc[df["Остаток WB, шт"] > 0, "Дней без остатка WB в текущем месяце"] = 0
    df["Delist"] = df["Артикул 1С"].map(lambda x: "Delist" if normalize_key(x) in stop_articles else "")

    df = df.merge(rrc_df, on="Артикул 1С", how="left")
    df["РРЦ"] = df["РРЦ"].fillna(0).map(round_int)
    df["Цена покупателя"] = df["Цена покупателя"].fillna(0).map(round_int)
    df["Коэффициент"] = df.apply(
        lambda r: compute_coef_rrc(round_int(r["Цена покупателя"]), round_int(r["РРЦ"])),
        axis=1,
    )

    df = df[df["Продажи 60 дней, шт"] >= 20].copy()

    for col in [
        "Среднесуточные продажи 7д",
        "Среднесуточные продажи 60д",
        "Расчётный спрос в день, шт",
        "WB хватит, дней",
        "WB + Липецк, дней",
        "После ближайшего поступления, дней",
        "WB + Липецк + в пути, дней",
    ]:
        df[col] = df[col].map(round_int)

    return df.sort_values(
        by="Артикул 1С",
        key=lambda s: s.map(lambda x: [int(t) if t.isdigit() else t.lower() for t in re.split(r"(\d+)", str(x))]),
    ).reset_index(drop=True)


def split_sheets(report_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    crit_mask = (
        (report_df["Остаток WB, шт"] <= 0)
        | (report_df["WB + Липецк, дней"] < 14)
        | ((report_df["Товары в пути, шт"] > 0) & (report_df["Хватит до поступления"] == "Нет"))
    )
    critical = report_df[crit_mask].copy()
    critical["Комментарий"] = critical.apply(
        lambda r: "Не хватает до поставки" if (safe_float(r["Товары в пути, шт"]) > 0 and r["Хватит до поступления"] == "Нет") else "",
        axis=1,
    )
    critical = critical[[
        "Артикул 1С", "Продажи 60 дней, шт", "WB хватит, дней", "Out of stock, days", "WB + Липецк, дней",
        "Товары в пути, шт", "Ближайшее поступление, шт", "Дней до поступления",
        "Остаток WB, шт", "Остатки МП (Липецк), шт", "Дней без остатка WB в текущем месяце",
        "Комментарий", "Менеджер", "Delist",
    ]].copy()

    calc = report_df[[
        "Артикул 1С", "Менеджер", "Артикул WB", "Артикул WB продавца", "Остаток WB, шт",
        "Остатки МП (Липецк), шт", "Товары в пути, шт", "Ближайшее поступление, шт",
        "Дата поступления", "Дней до поступления", "Партий в пути, шт",
        "Продажи 7 дней, шт", "Продажи 60 дней, шт", "Среднесуточные продажи 7д", "Среднесуточные продажи 60д",
        "Расчётный спрос в день, шт", "WB хватит, дней", "WB + Липецк, дней",
        "После ближайшего поступления, дней", "WB + Липецк + в пути, дней",
        "Хватит до поступления", "Out of stock, days", "Хватит на 60 дней",
        "Дней без остатка WB в текущем месяце", "Цена покупателя", "РРЦ", "Коэффициент", "Delist",
    ]].copy()

    dead = report_df[report_df["WB + Липецк + в пути, дней"] > 120].copy()
    dead = dead[[
        "Артикул 1С", "Менеджер", "WB хватит, дней", "WB + Липецк, дней",
        "После ближайшего поступления, дней", "WB + Липецк + в пути, дней",
        "Остаток WB, шт", "Остатки МП (Липецк), шт", "Товары в пути, шт", "Ближайшее поступление, шт",
        "Продажи 60 дней, шт", "Цена покупателя", "РРЦ", "Коэффициент", "Delist",
    ]].copy()

    monitor = report_df[report_df["Delist"] != "Delist"].copy()
    monitor = monitor[[
        "Артикул 1С", "Продажи 60 дней, шт", "Out of stock, days", "Хватит на 60 дней",
        "WB + Липецк, дней", "После ближайшего поступления, дней", "WB + Липецк + в пути, дней",
        "Товары в пути, шт", "Ближайшее поступление, шт", "Хватит до поступления",
        "Остаток WB, шт", "Остатки МП (Липецк), шт", "Дней без остатка WB в текущем месяце", "Менеджер",
    ]].copy()
    return critical, calc, dead, monitor

def auto_fit_columns(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            text = "" if cell.value is None else str(cell.value)
            max_len = max((len(part) for part in text.split("\n")), default=0)
            widths[cell.column] = max(widths.get(cell.column, 0), max_len)
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 3, 16), 42)


def style_sheet(ws, monitor: bool = False, dead_days_col: Optional[int] = None) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_THIN
            cell.font = Font(name=FONT_NAME, size=FONT_SIZE, color="000000")
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color="FFFFFF")
    for row in ws.iter_rows(min_row=2):
        row[0].alignment = ALIGN_LEFT
    auto_fit_columns(ws)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    headers = [c.value for c in ws[1]]

    if ws.title == SHEET_CRITICAL:
        wb_days_idx = headers.index("WB хватит, дней") + 1 if "WB хватит, дней" in headers else None
        comment_idx = headers.index("Комментарий") + 1 if "Комментарий" in headers else None
        for r in range(2, ws.max_row + 1):
            if wb_days_idx is not None and safe_float(ws.cell(r, wb_days_idx).value) == 0:
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).fill = FILL_ORANGE
            if comment_idx is not None and str(ws.cell(r, comment_idx).value or "").strip() == "Не хватает до поставки":
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).fill = FILL_BLUE_ROW

    if ws.title == SHEET_DEAD and dead_days_col is not None:
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(r, dead_days_col)
            if safe_float(cell.value) > 180:
                cell.fill = FILL_BLACK
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color="FFFFFF")

    if monitor and "Хватит на 60 дней" in headers:
        idx = headers.index("Хватит на 60 дней") + 1
        for r in range(2, ws.max_row + 1):
            value = str(ws.cell(r, idx).value or "")
            if value.startswith("Дефицит"):
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).fill = FILL_ORANGE


def save_report(report_path: Path, critical: pd.DataFrame, calc: pd.DataFrame, dead: pd.DataFrame, monitor: pd.DataFrame) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        critical.to_excel(writer, sheet_name=SHEET_CRITICAL, index=False)
        calc.to_excel(writer, sheet_name=SHEET_CALC, index=False)
        dead.to_excel(writer, sheet_name=SHEET_DEAD, index=False)
        monitor.to_excel(writer, sheet_name=SHEET_MONITOR, index=False)
    wb = load_workbook(report_path)
    style_sheet(wb[SHEET_CRITICAL])
    style_sheet(wb[SHEET_CALC])
    dead_headers = [c.value for c in wb[SHEET_DEAD][1]]
    dead_days_col = dead_headers.index("WB + Липецк + в пути, дней") + 1 if "WB + Липецк + в пути, дней" in dead_headers else None
    style_sheet(wb[SHEET_DEAD], dead_days_col=dead_days_col)
    style_sheet(wb[SHEET_MONITOR], monitor=True)
    wb.save(report_path)


def send_document_to_telegram(cfg: Config, path: Path, caption: str) -> None:
    if not cfg.telegram_bot_token or not cfg.telegram_chat_id:
        log("Telegram env не заданы — отправку пропускаем")
        return
    url = f"https://api.telegram.org/bot{cfg.telegram_bot_token}/sendDocument"
    with open(path, "rb") as f:
        resp = requests.post(
            url,
            data={"chat_id": cfg.telegram_chat_id, "caption": caption[:1024]},
            files={"document": (path.name, f)},
            timeout=120,
        )
    resp.raise_for_status()
    log(f"Файл отправлен в Telegram: {path.name}")


def send_to_telegram(cfg: Config, path: Path, critical_count: int, dead_count: int) -> None:
    caption = f"📦 Отчёт по остаткам WB {STORE_NAME}\nКритично: {critical_count}\nDead_Stock: {dead_count}"
    send_document_to_telegram(cfg, path, caption)



def normalize_warehouse_redist(value: object) -> str:
    raw = normalize_text(value)
    return WAREHOUSE_ALIASES_REDISTRIBUTION.get(raw, raw)


def build_region_to_warehouse_group() -> dict[str, str]:
    mapping: dict[str, str] = {}

    def add(group: str, regions: Sequence[str]) -> None:
        for region in regions:
            mapping[region] = group

    add(MOSCOW_CLUSTER_GROUP, ["Москва", "Московская область"])
    add("Краснодар", [
        "Краснодарский край", "Ростовская область", "Республика Крым", "Севастополь",
        "Республика Адыгея", "федеральная территория Сириус",
    ])
    add("СПБ Шушары", ["Санкт-Петербург", "Ленинградская область", "Новгородская область", "Республика Карелия"])
    add("Невинномысск", [
        "Ставропольский край", "Республика Дагестан", "Чеченская Республика",
        "Республика Северная Осетия — Алания", "Кабардино-Балкарская Республика",
        "Карачаево-Черкесская Республика", "Республика Ингушетия", "Республика Калмыкия",
    ])
    add("Казань", ["Республика Татарстан", "Ульяновская область", "Кировская область", "Чувашская Республика", "Республика Коми", "Республика Марий Эл"])
    add("Владимир Воршинское", ["Нижегородская область", "Владимирская область", "Ярославская область", "Ивановская область", "Костромская область"])
    add("Екатеринбург - Перспективная 14", [
        "Свердловская область", "Иркутская область", "Красноярский край", "Челябинская область",
        "Новосибирская область", "Кемеровская область", "Ханты-Мансийский автономный округ",
        "Тюменская область", "Алтайский край", "Омская область", "Томская область",
        "Республика Саха (Якутия)", "Республика Бурятия", "Забайкальский край", "Амурская область",
        "Ямало-Ненецкий автономный округ", "Курганская область", "Республика Алтай",
    ])
    add("Новосемейкино", ["Самарская область", "Оренбургская область"])
    add("Сарапул", ["Республика Башкортостан", "Пермский край", "Удмуртская Республика", "Республика Хакасия"])
    add("Воронеж", ["Воронежская область"])
    add("Тула", ["Тульская область", "Белгородская область", "Курская область", "Брянская область", "Орловская область"])
    add("Волгоград", ["Саратовская область", "Волгоградская область", "Астраханская область"])
    add("Котовск", ["Липецкая область", "Тамбовская область", "Республика Мордовия"])
    add("Пенза", ["Пензенская область"])
    add("Рязань (Тюшевское)", ["Рязанская область"])
    add("Новосибирск", ["Республика Тыва"])

    add(MOSCOW_CLUSTER_GROUP, [
        "Приморский край", "Калужская область", "Вологодская область", "Архангельская область",
        "Тверская область", "Мурманская область", "Смоленская область", "Калининградская область",
        "Хабаровский край", "Сахалинская область", "Псковская область", "Камчатский край",
        "Магаданская область", "Еврейская автономная область", "Ненецкий автономный округ",
        "Чукотский автономный округ",
    ])
    return mapping


REGION_TO_WAREHOUSE_GROUP = build_region_to_warehouse_group()


def read_allowed_template_warehouses(template_path: Path) -> list[str]:
    wb = load_workbook(template_path, data_only=False)
    if REDISTRIBUTION_WAREHOUSES_SHEET not in wb.sheetnames:
        raise KeyError(f"В шаблоне нет листа '{REDISTRIBUTION_WAREHOUSES_SHEET}'")
    ws = wb[REDISTRIBUTION_WAREHOUSES_SHEET]
    warehouses: list[str] = []
    for row in range(1, ws.max_row + 1):
        value = normalize_warehouse_redist(ws.cell(row, 1).value)
        if value:
            warehouses.append(value)
    wb.close()
    unique: list[str] = []
    seen: set[str] = set()
    for wh in warehouses:
        if wh not in seen:
            unique.append(wh)
            seen.add(wh)
    return unique


def resolve_redistribution_template(cfg: Config, storage: S3Storage) -> Path:
    candidates: list[Path] = []
    if cfg.redistribution_template_local:
        candidates.append(Path(cfg.redistribution_template_local))
    candidates.extend([
        Path("/mnt/data/Перераспределения.xlsx"),
        Path("/mnt/data/Перераспределения (6).xlsx"),
    ])
    for candidate in candidates:
        if candidate.exists():
            log(f"Шаблон перераспределения взят локально: {candidate}")
            return candidate

    if cfg.redistribution_template_key:
        target = Path(OUT_DIR) / Path(cfg.redistribution_template_key).name
        target.parent.mkdir(parents=True, exist_ok=True)
        obj = storage.client.get_object(Bucket=storage.bucket, Key=cfg.redistribution_template_key)
        target.write_bytes(obj["Body"].read())
        log(f"Шаблон перераспределения скачан из S3: {cfg.redistribution_template_key}")
        return target

    raise FileNotFoundError(
        "Не найден шаблон перераспределения. "
        "Укажи WB_REDISTRIBUTION_TEMPLATE_LOCAL или WB_REDISTRIBUTION_TEMPLATE_KEY."
    )


def load_orders_for_redistribution(storage: S3Storage, lookback_days: int) -> tuple[pd.DataFrame, list[str]]:
    keys = latest_n_weekly_keys(storage.list_keys(WB_ORDERS_PREFIX), 4)
    log(f"Для перераспределения берём заказы из файлов: {keys}")
    frames: list[pd.DataFrame] = []
    for key in keys:
        df = storage.read_excel(key)
        frames.append(df)
    orders = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    if orders.empty:
        return pd.DataFrame(columns=["Артикул WB", "Артикул WB продавца", "dt", "regionName", "warehouseName", "qty"]), keys

    wb_col = choose_existing_column(orders, ["nmId", "Артикул WB"], "Артикул WB в заказах")
    seller_col = choose_existing_column(orders, ["supplierArticle", "Артикул продавца"], "Артикул продавца в заказах")
    date_col = choose_existing_column(orders, ["date", "Дата", "Дата заказа", "lastChangeDate", "Дата продажи"], "дата в заказах")
    region_col = choose_existing_column(orders, ["regionName", "Регион", "Регион покупателя"], "регион в заказах")
    wh_col = try_choose_column(orders, ["warehouseName", "Склад", "Склад заказа"])

    work = pd.DataFrame({
        "Артикул WB": orders[wb_col].map(normalize_key),
        "Артикул WB продавца": orders[seller_col].map(normalize_text),
        "dt": pd.to_datetime(orders[date_col], errors="coerce").dt.normalize(),
        "regionName": orders[region_col].map(normalize_text),
        "warehouseName": orders[wh_col].map(normalize_warehouse_redist) if wh_col else "",
        "qty": 1.0,
    })
    work = work[(work["Артикул WB"] != "") & work["dt"].notna()].copy()
    if work.empty:
        return work, keys

    max_dt = work["dt"].max()
    start_dt = max_dt - pd.Timedelta(days=lookback_days - 1)
    work = work[(work["dt"] >= start_dt) & (work["dt"] <= max_dt)].copy()
    return work, keys


def load_latest_warehouse_stocks_for_redistribution(storage: S3Storage, allowed_warehouses: Sequence[str]) -> tuple[pd.DataFrame, str]:
    latest_key = latest_weekly_key(storage.list_keys(WB_STOCKS_PREFIX))
    df = storage.read_excel(latest_key)
    sample_col = choose_existing_column(df, ["Дата сбора", "Дата запроса"], "дата среза по складам")
    df["_sample_dt"] = pd.to_datetime(df[sample_col], errors="coerce")
    latest_dt = df["_sample_dt"].max()
    if pd.notna(latest_dt):
        df = df[df["_sample_dt"] == latest_dt].copy()

    wb_col = choose_existing_column(df, ["Артикул WB", "nmId"], "Артикул WB")
    seller_col = choose_existing_column(df, ["Артикул продавца"], "Артикул продавца")
    wh_col = choose_existing_column(df, ["Склад", "warehouseName"], "склад")
    qty_col = try_choose_column(df, ["Доступно для продажи", "Полное количество", "Количество", "Доступно", "Остаток", "Остатки"])
    if qty_col is None:
        raise KeyError("В остатках не найдена колонка с количеством товара по складам")

    temp = pd.DataFrame({
        "Артикул WB": df[wb_col].map(normalize_key),
        "Артикул WB продавца": df[seller_col].map(normalize_text),
        "Склад": df[wh_col].map(normalize_warehouse_redist),
        "Остаток склада, шт": df[qty_col].map(round_int),
    })
    temp = temp[(temp["Артикул WB"] != "") & (temp["Склад"] != "")]
    if allowed_warehouses:
        temp = temp[temp["Склад"].isin(set(allowed_warehouses))].copy()
    temp = temp.groupby(["Артикул WB", "Артикул WB продавца", "Склад"], as_index=False)["Остаток склада, шт"].sum()
    return temp, latest_key


def build_sales_by_warehouse(
    orders_df: pd.DataFrame,
    lookback_days: int,
    allowed_warehouses: Sequence[str],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    rows: list[dict[str, object]] = []
    unmapped_rows: list[dict[str, object]] = []
    allowed_set = set(allowed_warehouses)

    for _, row in orders_df.iterrows():
        region = normalize_text(row.get("regionName"))
        article_wb = normalize_key(row.get("Артикул WB"))
        seller_article = normalize_text(row.get("Артикул WB продавца"))
        group = REGION_TO_WAREHOUSE_GROUP.get(region)
        fallback_wh = normalize_warehouse_redist(row.get("warehouseName"))

        if not group and fallback_wh in allowed_set:
            group = fallback_wh

        if not group:
            unmapped_rows.append({
                "Артикул WB": article_wb,
                "Артикул WB продавца": seller_article,
                "Регион": region,
                "Склад из заказа": fallback_wh,
                "Количество заказов": 1,
            })
            continue

        if group == MOSCOW_CLUSTER_GROUP:
            for warehouse, share in MOSCOW_CLUSTER_WEIGHTS.items():
                if warehouse in allowed_set:
                    rows.append({
                        "Артикул WB": article_wb,
                        "Артикул WB продавца": seller_article,
                        "Склад": warehouse,
                        "Продажи 14 дней, шт": float(row.get("qty", 0)) * share,
                    })
        else:
            warehouse = normalize_warehouse_redist(group)
            if warehouse in allowed_set:
                rows.append({
                    "Артикул WB": article_wb,
                    "Артикул WB продавца": seller_article,
                    "Склад": warehouse,
                    "Продажи 14 дней, шт": float(row.get("qty", 0)),
                })

    sales_df = pd.DataFrame(rows)
    if sales_df.empty:
        sales_df = pd.DataFrame(columns=["Артикул WB", "Артикул WB продавца", "Склад", "Продажи 14 дней, шт"])
    else:
        sales_df = sales_df.groupby(["Артикул WB", "Артикул WB продавца", "Склад"], as_index=False)["Продажи 14 дней, шт"].sum()

    sales_df["Среднесуточные продажи 14д"] = sales_df["Продажи 14 дней, шт"].map(lambda x: safe_float(x) / max(lookback_days, 1))

    unmapped_df = pd.DataFrame(unmapped_rows)
    if not unmapped_df.empty:
        unmapped_df = unmapped_df.groupby(["Артикул WB", "Артикул WB продавца", "Регион", "Склад из заказа"], as_index=False)["Количество заказов"].sum()
    return sales_df, unmapped_df


def build_warehouse_balance(
    sales_df: pd.DataFrame,
    stocks_df: pd.DataFrame,
    article_map: dict[str, str],
    lookback_days: int,
    target_days: int,
) -> pd.DataFrame:
    keys = pd.concat(
        [
            sales_df[["Артикул WB", "Артикул WB продавца", "Склад"]] if not sales_df.empty else pd.DataFrame(columns=["Артикул WB", "Артикул WB продавца", "Склад"]),
            stocks_df[["Артикул WB", "Артикул WB продавца", "Склад"]] if not stocks_df.empty else pd.DataFrame(columns=["Артикул WB", "Артикул WB продавца", "Склад"]),
        ],
        ignore_index=True,
    ).drop_duplicates()

    balance = keys.merge(sales_df, on=["Артикул WB", "Артикул WB продавца", "Склад"], how="left")
    balance = balance.merge(stocks_df, on=["Артикул WB", "Артикул WB продавца", "Склад"], how="left")
    if balance.empty:
        return pd.DataFrame(columns=[
            "Артикул WB", "Артикул WB продавца", "Артикул 1С", "Склад",
            "Продажи 14 дней, шт", "Среднесуточные продажи 14д", f"Целевой запас {target_days} дн., шт",
            "Остаток склада, шт", "Баланс к целевому запасу, шт", "Излишек, шт", "Дефицит, шт", "Статус"
        ])

    balance["Продажи 14 дней, шт"] = balance["Продажи 14 дней, шт"].fillna(0.0)
    balance["Среднесуточные продажи 14д"] = balance["Среднесуточные продажи 14д"].fillna(0.0)
    balance["Остаток склада, шт"] = balance["Остаток склада, шт"].fillna(0).map(round_int)
    target_col = f"Целевой запас {target_days} дн., шт"
    balance[target_col] = balance["Среднесуточные продажи 14д"].map(lambda x: ceil_int(safe_float(x) * target_days))
    balance["Баланс к целевому запасу, шт"] = balance["Остаток склада, шт"] - balance[target_col]
    balance["Излишек, шт"] = balance["Баланс к целевому запасу, шт"].map(lambda x: max(round_int(x), 0))
    balance["Дефицит, шт"] = balance["Баланс к целевому запасу, шт"].map(lambda x: max(-round_int(x), 0))
    balance["Статус"] = balance.apply(
        lambda r: "Излишек" if safe_float(r["Излишек, шт"]) > 0 else ("Дефицит" if safe_float(r["Дефицит, шт"]) > 0 else "Норма"),
        axis=1,
    )
    balance["Артикул 1С"] = balance["Артикул WB"].map(article_map).fillna("")
    balance = balance[
        (balance["Продажи 14 дней, шт"] > 0)
        | (balance["Остаток склада, шт"] > 0)
        | (balance["Излишек, шт"] > 0)
        | (balance["Дефицит, шт"] > 0)
    ].copy()
    return balance.sort_values(["Артикул WB", "Склад"]).reset_index(drop=True)


def donor_rank_for_recipient(donor: str, recipient: str, donor_surplus: int) -> tuple[int, int, str]:
    donor_zone = WAREHOUSE_ZONE.get(donor, "")
    recipient_zone = WAREHOUSE_ZONE.get(recipient, "")
    if donor == recipient:
        return (99, 0, donor)
    if donor_zone and donor_zone == recipient_zone:
        return (0, -donor_surplus, donor)
    if donor in CENTRAL_HUBS:
        return (1, -donor_surplus, donor)
    if recipient in CENTRAL_HUBS:
        return (2, -donor_surplus, donor)
    return (3, -donor_surplus, donor)


def build_transfer_plan(balance_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    plan_rows: list[dict[str, object]] = []
    unresolved_rows: list[dict[str, object]] = []
    route_rows: list[dict[str, object]] = []

    if balance_df.empty:
        empty_cols = ["Артикул WB", "Артикул WB продавца", "Артикул 1С", "Склад откуда", "Склад куда", "Количество"]
        return pd.DataFrame(columns=empty_cols), pd.DataFrame(columns=empty_cols), pd.DataFrame(columns=["Склад откуда", "Склад куда", "Приоритет"])

    warehouses = sorted(balance_df["Склад"].dropna().astype(str).unique())
    for recipient in warehouses:
        for donor in warehouses:
            if donor == recipient:
                continue
            rank = donor_rank_for_recipient(donor, recipient, 0)[0]
            route_rows.append({"Склад откуда": donor, "Склад куда": recipient, "Приоритет": rank})

    for (article_wb, seller_article, article_1c), part in balance_df.groupby(["Артикул WB", "Артикул WB продавца", "Артикул 1С"], dropna=False):
        donors = {
            str(row["Склад"]): int(row["Излишек, шт"])
            for _, row in part.iterrows()
            if round_int(row["Излишек, шт"]) > 0
        }
        recipients = [
            {
                "warehouse": str(row["Склад"]),
                "deficit": int(row["Дефицит, шт"]),
            }
            for _, row in part.iterrows()
            if round_int(row["Дефицит, шт"]) > 0
        ]
        recipients.sort(key=lambda x: (-x["deficit"], x["warehouse"]))

        for recipient in recipients:
            need = recipient["deficit"]
            if need <= 0:
                continue

            donor_candidates = sorted(
                [
                    donor for donor, surplus in donors.items()
                    if surplus > 0 and donor != recipient["warehouse"]
                ],
                key=lambda donor: donor_rank_for_recipient(donor, recipient["warehouse"], donors[donor]),
            )

            for donor in donor_candidates:
                if need <= 0:
                    break
                available = donors.get(donor, 0)
                if available <= 0:
                    continue
                qty = min(available, need)
                plan_rows.append({
                    "Артикул WB": article_wb,
                    "Артикул WB продавца": seller_article,
                    "Артикул 1С": article_1c,
                    "Склад откуда": donor,
                    "Склад куда": recipient["warehouse"],
                    "Количество": qty,
                })
                donors[donor] = available - qty
                need -= qty

            if need > 0:
                unresolved_rows.append({
                    "Артикул WB": article_wb,
                    "Артикул WB продавца": seller_article,
                    "Артикул 1С": article_1c,
                    "Склад откуда": "",
                    "Склад куда": recipient["warehouse"],
                    "Количество": need,
                })

    plan_df = pd.DataFrame(plan_rows)
    if not plan_df.empty:
        plan_df = plan_df.groupby(
            ["Артикул WB", "Артикул WB продавца", "Артикул 1С", "Склад откуда", "Склад куда"],
            as_index=False,
        )["Количество"].sum()
        plan_df = plan_df.sort_values(["Артикул WB", "Склад куда", "Склад откуда"]).reset_index(drop=True)
    else:
        plan_df = pd.DataFrame(columns=["Артикул WB", "Артикул WB продавца", "Артикул 1С", "Склад откуда", "Склад куда", "Количество"])

    unresolved_df = pd.DataFrame(unresolved_rows)
    if unresolved_df.empty:
        unresolved_df = pd.DataFrame(columns=["Артикул WB", "Артикул WB продавца", "Артикул 1С", "Склад откуда", "Склад куда", "Количество"])

    routes_df = pd.DataFrame(route_rows).sort_values(["Склад куда", "Приоритет", "Склад откуда"]).reset_index(drop=True)
    return plan_df, unresolved_df, routes_df



def _resolve_sheet_xml_path(xlsx_path: Path, sheet_name: str) -> str:
    ns_main = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    ns_rel_attr = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ns_pkg_rel = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}

    with zipfile.ZipFile(xlsx_path, "r") as zf:
        workbook_root = ET.fromstring(zf.read("xl/workbook.xml"))
        rel_id = ""
        for sheet in workbook_root.findall("main:sheets/main:sheet", ns_main):
            if sheet.attrib.get("name") == sheet_name:
                rel_id = sheet.attrib.get(f"{{{ns_rel_attr}}}id", "")
                break
        if not rel_id:
            raise KeyError(f"Не найден лист '{sheet_name}' в шаблоне")

        rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        target = ""
        for rel in rels_root.findall("rel:Relationship", ns_pkg_rel):
            if rel.attrib.get("Id") == rel_id:
                target = rel.attrib.get("Target", "")
                break
        if not target:
            raise KeyError(f"Не найден xml-путь для листа '{sheet_name}'")

    target = target.lstrip("/")
    return target if target.startswith("xl/") else f"xl/{target}"


def _xml_cell_number(ref: str, value: int) -> ET.Element:
    cell = ET.Element("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c", {"r": ref})
    v = ET.SubElement(cell, "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v")
    v.text = str(value)
    return cell


def _xml_cell_text(ref: str, value: str) -> ET.Element:
    cell = ET.Element(
        "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c",
        {"r": ref, "t": "inlineStr"},
    )
    is_el = ET.SubElement(cell, "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}is")
    t_el = ET.SubElement(is_el, "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")
    t_el.text = value
    return cell


def fill_redistribution_template(template_path: Path, output_path: Path, plan_df: pd.DataFrame) -> None:
    ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns_x14ac = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"
    ns_rel = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ns_x14 = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
    ns_xm = "http://schemas.microsoft.com/office/excel/2006/main"
    ns_xr = "http://schemas.microsoft.com/office/spreadsheetml/2014/revision"

    ET.register_namespace("", ns_main)
    ET.register_namespace("x14ac", ns_x14ac)
    ET.register_namespace("r", ns_rel)
    ET.register_namespace("x14", ns_x14)
    ET.register_namespace("xm", ns_xm)
    ET.register_namespace("xr", ns_xr)

    sheet_xml_path = _resolve_sheet_xml_path(template_path, REDISTRIBUTION_SHEET)

    with zipfile.ZipFile(template_path, "r") as zf:
        file_map = {name: zf.read(name) for name in zf.namelist()}

    root = ET.fromstring(file_map[sheet_xml_path])
    sheet_data = root.find(f"{{{ns_main}}}sheetData")
    if sheet_data is None:
        raise ValueError("В шаблоне не найден sheetData")

    for row in list(sheet_data):
        row_num = round_int(row.attrib.get("r"))
        if row_num >= 2:
            sheet_data.remove(row)

    for idx, (_, row) in enumerate(plan_df.iterrows(), start=2):
        row_el = ET.Element(
            f"{{{ns_main}}}row",
            {"r": str(idx), "spans": "1:5", f"{{{ns_x14ac}}}dyDescent": "0.25"},
        )
        row_el.append(_xml_cell_number(f"A{idx}", round_int(row.get("Артикул WB"))))
        row_el.append(_xml_cell_text(f"B{idx}", normalize_warehouse_redist(row.get("Склад откуда"))))
        row_el.append(_xml_cell_text(f"C{idx}", normalize_warehouse_redist(row.get("Склад куда"))))
        row_el.append(_xml_cell_number(f"D{idx}", round_int(row.get("Количество"))))
        sheet_data.append(row_el)

    dimension = root.find(f"{{{ns_main}}}dimension")
    if dimension is not None:
        last_row = max(len(plan_df) + 1, 1)
        dimension.set("ref", f"A1:E{last_row}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    file_map[sheet_xml_path] = ET.tostring(root, encoding="utf-8", xml_declaration=False)

    with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for name, data in file_map.items():
            zf.writestr(name, data)


def save_redistribution_workbook(
    path: Path,
    sales_df: pd.DataFrame,
    balance_df: pd.DataFrame,
    plan_df: pd.DataFrame,
    unresolved_df: pd.DataFrame,
    routes_df: pd.DataFrame,
    unmapped_regions_df: pd.DataFrame,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        sales_df.to_excel(writer, sheet_name="Продажи_14д_по_складам", index=False)
        balance_df.to_excel(writer, sheet_name="Баланс_21д", index=False)
        plan_df.to_excel(writer, sheet_name="План_перераспределения", index=False)
        unresolved_df.to_excel(writer, sheet_name="Нехватка_без_покрытия", index=False)
        routes_df.to_excel(writer, sheet_name="Словарь_маршрутов", index=False)
        unmapped_regions_df.to_excel(writer, sheet_name="Не_смогли_сопоставить", index=False)

    wb = load_workbook(path)
    for sheet in wb.sheetnames:
        style_sheet(wb[sheet], monitor=(sheet == "Баланс_21д"))
    wb.save(path)
    wb.close()


def create_redistribution_outputs(storage: S3Storage, cfg: Config, article_map: dict[str, str]) -> tuple[Path, Path]:
    template_path = resolve_redistribution_template(cfg, storage)
    allowed_warehouses = read_allowed_template_warehouses(template_path)
    log(f"Разрешённые склады из шаблона: {', '.join(allowed_warehouses)}")

    orders_df, order_sources = load_orders_for_redistribution(storage, cfg.redistribution_days)
    stocks_df, stock_source = load_latest_warehouse_stocks_for_redistribution(storage, allowed_warehouses)
    sales_by_wh_df, unmapped_regions_df = build_sales_by_warehouse(orders_df, cfg.redistribution_days, allowed_warehouses)
    balance_df = build_warehouse_balance(
        sales_df=sales_by_wh_df,
        stocks_df=stocks_df,
        article_map=article_map,
        lookback_days=cfg.redistribution_days,
        target_days=cfg.redistribution_target_days,
    )
    plan_df, unresolved_df, routes_df = build_transfer_plan(balance_df)

    calc_path = Path(OUT_DIR) / f"Перераспределение_WB_{STORE_NAME}_{cfg.run_date.strftime('%Y%m%d')}.xlsx"
    template_out_path = Path(OUT_DIR) / f"Шаблон_перераспределения_WB_{STORE_NAME}_{cfg.run_date.strftime('%Y%m%d')}.xlsx"

    save_redistribution_workbook(
        path=calc_path,
        sales_df=sales_by_wh_df,
        balance_df=balance_df,
        plan_df=plan_df,
        unresolved_df=unresolved_df,
        routes_df=routes_df,
        unmapped_regions_df=unmapped_regions_df,
    )
    fill_redistribution_template(template_path, template_out_path, plan_df)

    log(f"Файл расчёта перераспределения сохранён: {calc_path}")
    log(f"Заполненный шаблон перераспределения сохранён: {template_out_path}")
    log(f"Источники перераспределения | остатки: {stock_source}")
    log(f"Источники перераспределения | заказы: {', '.join(order_sources)}")
    if not unmapped_regions_df.empty:
        log(f"Регионов без сопоставления: {len(unmapped_regions_df)} строк (см. лист 'Не_смогли_сопоставить')")
    return calc_path, template_out_path

def run() -> Path:
    cfg = get_config()
    storage = S3Storage(cfg)
    stop_articles = parse_stop_articles(cfg.stop_articles_raw)

    wb_stocks, stock_source = load_latest_wb_stocks(storage)
    sales_df, order_sources = load_orders_metrics(storage)
    article_map = load_article_map(storage)
    stocks_1c = load_stocks_1c(storage)
    rrc_df = load_rrc(storage)
    inbound_df = load_inbound(storage, cfg.run_date)
    abc_df = load_abc_managers(storage)

    avg7_map: dict[str, float] = {}
    for _, row in sales_df.iterrows():
        wb_key = normalize_key(row.get("Артикул WB"))
        avg7_map[wb_key] = safe_float(row.get("avg_daily_sales_7d"))

    current_zero_articles = set(wb_stocks.loc[wb_stocks["Остаток WB, шт"] <= 0, "Артикул WB"].tolist())
    zero_days_map = load_current_month_zero_days(storage, current_zero_articles, avg7_map, cfg.run_date)

    report_df = build_report_dataframe(
        wb_stocks=wb_stocks,
        sales=sales_df,
        article_map=article_map,
        stocks_1c=stocks_1c,
        stop_articles=stop_articles,
        rrc_df=rrc_df,
        inbound_df=inbound_df,
        zero_days_map=zero_days_map,
        abc_df=abc_df,
    )

    critical, calc, dead, monitor = split_sheets(report_df)
    report_path = Path(OUT_DIR) / f"Отчёт_дни_остатка_WB_{STORE_NAME}_{cfg.run_date.strftime('%Y%m%d')}.xlsx"
    save_report(report_path, critical, calc, dead, monitor)

    log(f"Отчёт сохранён: {report_path}")
    log(f"Источник остатков: {stock_source}")
    log(f"Источники заказов: {', '.join(order_sources)}")

    redistribution_calc_path, redistribution_template_path = create_redistribution_outputs(
        storage=storage,
        cfg=cfg,
        article_map=article_map,
    )

    if should_send_report(cfg):
        send_to_telegram(cfg, report_path, len(critical), len(dead))
    else:
        log("Отправка отчёта по дням остатка в Telegram пропущена по расписанию")

    if should_send_redistribution(cfg):
        send_document_to_telegram(
            cfg,
            redistribution_calc_path,
            f"🚚 Перераспределение WB {STORE_NAME}\nРасчёт излишков/дефицита на {cfg.redistribution_target_days} дней",
        )
        send_document_to_telegram(
            cfg,
            redistribution_template_path,
            f"🚚 Шаблон перераспределения WB {STORE_NAME}\nЗаполнено автоматически по расчёту за {cfg.redistribution_days} дней",
        )
    else:
        log("Отправка файлов перераспределения в Telegram пропущена по расписанию")

    return report_path


if __name__ == "__main__":
    run()
