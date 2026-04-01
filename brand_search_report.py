
from __future__ import annotations

import io
import json
import math
import os
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Optional

import boto3
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# =========================
# Конфиг
# =========================

DEFAULT_BRAND_VARIANTS = [
    "topface",
    "top face",
    "top-face",
    "топфейс",
    "топ фейс",
    "топ-фейс",
    "топфеис",
    "топ феис",
    "топфэйс",
    "топ фэйс",
    "topfase",
    "top fase",
]

DEFAULT_YANDEX_PHRASES = [
    "topface",
    "топфейс",
    "top face",
    "топ фейс",
]

FONT_NAME = "Calibri"
FONT_SIZE = 11
FILL_HEADER = PatternFill("solid", fgColor="D9EAF7")
BORDER_THIN = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)


@dataclass
class Config:
    bucket: str
    access_key: str
    secret_key: str
    endpoint_url: str
    region_name: str

    telegram_bot_token: str
    telegram_chat_id: str
    force_send: bool
    run_date: date

    store_name: str
    wb_keywords_prefix: str
    output_prefix: str

    yandex_api_key: str
    yandex_folder_id: str
    yandex_region_id: Optional[str]
    yandex_top_url: str
    yandex_dynamics_url: str

    brand_variants: list[str]
    yandex_phrases: list[str]
    wb_weeks_to_compare: int
    wb_use_only_orders_filter: bool


class S3Storage:
    def __init__(self, cfg: Config) -> None:
        self.bucket = cfg.bucket
        self.client = boto3.client(
            "s3",
            endpoint_url=cfg.endpoint_url,
            aws_access_key_id=cfg.access_key,
            aws_secret_access_key=cfg.secret_key,
            region_name=cfg.region_name,
        )

    def list_keys(self, prefix: str) -> list[str]:
        keys: list[str] = []
        token = None
        while True:
            kwargs = {"Bucket": self.bucket, "Prefix": prefix, "MaxKeys": 1000}
            if token:
                kwargs["ContinuationToken"] = token
            resp = self.client.list_objects_v2(**kwargs)
            for item in resp.get("Contents", []):
                key = item["Key"]
                if not key.endswith("/"):
                    keys.append(key)
            if not resp.get("IsTruncated"):
                break
            token = resp.get("NextContinuationToken")
        return keys

    def read_excel(self, key: str, **kwargs) -> pd.DataFrame:
        obj = self.client.get_object(Bucket=self.bucket, Key=key)
        return pd.read_excel(io.BytesIO(obj["Body"].read()), **kwargs)

    def write_bytes(self, key: str, data: bytes, content_type: str = "application/octet-stream") -> None:
        self.client.put_object(
            Bucket=self.bucket,
            Key=key,
            Body=data,
            ContentType=content_type,
        )


def log(message: str) -> None:
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {message}", flush=True)


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, float) and float(value).is_integer():
        return str(int(value)).strip()
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".")[0]
    return text


def safe_float(value: object) -> float:
    if value is None:
        return 0.0
    try:
        if pd.isna(value):
            return 0.0
    except Exception:
        pass
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "").replace(",", ".")
    if not text:
        return 0.0
    try:
        return float(text)
    except Exception:
        return 0.0


def parse_list_env(name: str, default_values: list[str]) -> list[str]:
    raw = (os.getenv(name) or "").strip()
    if not raw:
        return default_values[:]
    items = []
    for part in re.split(r"[;\n,|]+", raw):
        part = part.strip()
        if part:
            items.append(part)
    return items or default_values[:]


def parse_iso_week_from_key(key: str) -> tuple[int, int]:
    m = re.search(r"_(\d{4})-W(\d{2})\.xlsx$", key, flags=re.IGNORECASE)
    if not m:
        return (0, 0)
    return int(m.group(1)), int(m.group(2))


def latest_n_weekly_keys(keys: list[str], n: int) -> list[str]:
    xlsx = [k for k in keys if k.lower().endswith(".xlsx")]
    return sorted(xlsx, key=parse_iso_week_from_key)[-n:]


def should_send_report(cfg: Config) -> bool:
    if cfg.force_send:
        return True
    return cfg.run_date.weekday() == 0  # только понедельник


def get_config() -> Config:
    bucket = (os.getenv("YC_BUCKET_NAME") or os.getenv("CLOUD_RU_BUCKET") or os.getenv("WB_S3_BUCKET") or "").strip()
    access_key = (os.getenv("YC_ACCESS_KEY_ID") or os.getenv("CLOUD_RU_ACCESS_KEY") or os.getenv("WB_S3_ACCESS_KEY") or "").strip()
    secret_key = (os.getenv("YC_SECRET_ACCESS_KEY") or os.getenv("CLOUD_RU_SECRET_KEY") or os.getenv("WB_S3_SECRET_KEY") or "").strip()
    endpoint_url = (os.getenv("YC_ENDPOINT_URL") or os.getenv("WB_S3_ENDPOINT") or "https://storage.yandexcloud.net").strip()
    region_name = (os.getenv("WB_S3_REGION") or "ru-central1").strip()

    if not bucket or not access_key or not secret_key:
        raise ValueError("Не заданы параметры Object Storage")

    store_name = (os.getenv("STORE_NAME") or "TOPFACE").strip()
    wb_keywords_prefix = (
        os.getenv("WB_KEYWORDS_PREFIX")
        or f"Отчёты/Поисковые запросы/{store_name}/Недельные/"
    ).strip()
    output_prefix = (
        os.getenv("BRAND_REPORT_OUTPUT_PREFIX")
        or f"Отчёты/Поисковые запросы/Брендовый отчет/{store_name}/"
    ).strip()

    return Config(
        bucket=bucket,
        access_key=access_key,
        secret_key=secret_key,
        endpoint_url=endpoint_url,
        region_name=region_name,
        telegram_bot_token=(os.getenv("TELEGRAM_BOT_TOKEN") or "").strip(),
        telegram_chat_id=(os.getenv("TELEGRAM_CHAT_ID") or "").strip(),
        force_send=(os.getenv("WB_FORCE_SEND", "false").strip().lower() == "true"),
        run_date=date.today(),
        store_name=store_name,
        wb_keywords_prefix=wb_keywords_prefix,
        output_prefix=output_prefix.rstrip("/") + "/",
        yandex_api_key=(os.getenv("YANDEX_API_KEY") or "").strip(),
        yandex_folder_id=(os.getenv("YANDEX_FOLDER_ID") or "").strip(),
        yandex_region_id=(os.getenv("YANDEX_REGION_ID") or "").strip() or None,
        yandex_top_url=(os.getenv("YANDEX_WORDSTAT_TOP_URL") or "https://searchapi.api.cloud.yandex.net/v2/wordstat/topRequests").strip(),
        yandex_dynamics_url=(os.getenv("YANDEX_WORDSTAT_DYNAMICS_URL") or "https://searchapi.api.cloud.yandex.net/v2/wordstat/dynamics").strip(),
        brand_variants=parse_list_env("BRAND_VARIANTS", DEFAULT_BRAND_VARIANTS),
        yandex_phrases=parse_list_env("YANDEX_PHRASES", DEFAULT_YANDEX_PHRASES),
        wb_weeks_to_compare=int((os.getenv("WB_WEEKS_TO_COMPARE") or "8").strip()),
        wb_use_only_orders_filter=(os.getenv("WB_USE_ONLY_ORDERS_FILTER", "true").strip().lower() != "false"),
    )


# =========================
# WB
# =========================

def build_brand_regex(variants: Iterable[str]) -> re.Pattern[str]:
    escaped = [re.escape(v.strip().lower()) for v in variants if str(v).strip()]
    if not escaped:
        raise ValueError("Список вариантов бренда пуст")
    return re.compile("|".join(sorted(set(escaped), key=len, reverse=True)), flags=re.IGNORECASE)


def extract_detected_variants(df: pd.DataFrame, regex: re.Pattern[str]) -> list[str]:
    variants = set()
    for value in df["Поисковый запрос"].dropna().astype(str).str.lower():
        for match in regex.finditer(value):
            variants.add(match.group(0))
    return sorted(variants)


def prepare_wb_week(df: pd.DataFrame, week_key: str, regex: re.Pattern[str], only_orders_filter: bool) -> pd.DataFrame:
    temp = df.copy()
    required_cols = [
        "Дата", "Поисковый запрос", "Фильтр", "Артикул WB", "Артикул продавца",
        "Бренд", "Частота запросов", "Частота за неделю", "Переходы в карточку",
        "Добавления в корзину", "Заказы", "Видимость %"
    ]
    for col in required_cols:
        if col not in temp.columns:
            temp[col] = None

    temp["Поисковый запрос"] = temp["Поисковый запрос"].map(normalize_text)
    temp["query_lower"] = temp["Поисковый запрос"].str.lower()
    temp = temp[temp["query_lower"].str.contains(regex, na=False)]
    if only_orders_filter and "Фильтр" in temp.columns:
        temp = temp[temp["Фильтр"].astype(str).str.lower() == "orders"]

    if temp.empty:
        return temp

    temp["Дата_dt"] = pd.to_datetime(temp["Дата"], errors="coerce")
    last_snapshot = temp["Дата_dt"].max()
    if pd.notna(last_snapshot):
        temp = temp[temp["Дата_dt"] == last_snapshot]

    numeric_cols = ["Частота запросов", "Частота за неделю", "Переходы в карточку", "Добавления в корзину", "Заказы", "Видимость %"]
    for col in numeric_cols:
        temp[col] = temp[col].map(safe_float)

    temp["week_key"] = week_key
    # Внутри недели на уровне query + article берём max, чтобы не дублировать случайные повторы
    grouped = (
        temp.groupby(["week_key", "Поисковый запрос", "Артикул WB", "Артикул продавца"], dropna=False, as_index=False)
        .agg({
            "Бренд": "first",
            "Частота запросов": "max",
            "Частота за неделю": "max",
            "Переходы в карточку": "max",
            "Добавления в корзину": "max",
            "Заказы": "max",
            "Видимость %": "max",
        })
    )
    return grouped


def load_wb_brand_data(storage: S3Storage, cfg: Config, regex: re.Pattern[str]) -> tuple[pd.DataFrame, list[str]]:
    keys = storage.list_keys(cfg.wb_keywords_prefix)
    week_keys = latest_n_weekly_keys(keys, cfg.wb_weeks_to_compare)
    if not week_keys:
        raise FileNotFoundError(f"Не найдены weekly xlsx по префиксу: {cfg.wb_keywords_prefix}")

    frames: list[pd.DataFrame] = []
    for key in week_keys:
        log(f"WB: читаю {key}")
        df = storage.read_excel(key)
        week_df = prepare_wb_week(
            df=df,
            week_key=Path(key).name.replace(".xlsx", ""),
            regex=regex,
            only_orders_filter=cfg.wb_use_only_orders_filter,
        )
        if not week_df.empty:
            frames.append(week_df)

    result = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    return result, week_keys


def build_wb_weekly_summary(wb_brand_df: pd.DataFrame) -> pd.DataFrame:
    if wb_brand_df.empty:
        return pd.DataFrame(columns=[
            "Неделя", "Уникальных брендовых запросов", "Уникальных артикулов",
            "Сумма частотности", "Сумма частотности за неделю", "Переходы в карточку",
            "Добавления в корзину", "Заказы"
        ])

    summary = (
        wb_brand_df.groupby("week_key", as_index=False)
        .agg({
            "Поисковый запрос": pd.Series.nunique,
            "Артикул WB": pd.Series.nunique,
            "Частота запросов": "sum",
            "Частота за неделю": "sum",
            "Переходы в карточку": "sum",
            "Добавления в корзину": "sum",
            "Заказы": "sum",
        })
        .rename(columns={
            "week_key": "Неделя",
            "Поисковый запрос": "Уникальных брендовых запросов",
            "Артикул WB": "Уникальных артикулов",
            "Частота запросов": "Сумма частотности",
            "Частота за неделю": "Сумма частотности за неделю",
        })
        .sort_values("Неделя")
        .reset_index(drop=True)
    )
    return summary


def build_wb_compare_tables(wb_brand_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if wb_brand_df.empty:
        empty = pd.DataFrame()
        return empty, empty, empty

    weeks = sorted(wb_brand_df["week_key"].dropna().astype(str).unique())
    if len(weeks) < 2:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    prev_week = weeks[-2]
    curr_week = weeks[-1]

    prev_df = wb_brand_df[wb_brand_df["week_key"] == prev_week].copy()
    curr_df = wb_brand_df[wb_brand_df["week_key"] == curr_week].copy()

    prev_agg = (
        prev_df.groupby("Поисковый запрос", as_index=False)
        .agg({
            "Частота запросов": "sum",
            "Частота за неделю": "sum",
            "Переходы в карточку": "sum",
            "Добавления в корзину": "sum",
            "Заказы": "sum",
        })
        .rename(columns=lambda c: f"{c} ({prev_week})" if c != "Поисковый запрос" else c)
    )

    curr_agg = (
        curr_df.groupby("Поисковый запрос", as_index=False)
        .agg({
            "Частота запросов": "sum",
            "Частота за неделю": "sum",
            "Переходы в карточку": "sum",
            "Добавления в корзину": "sum",
            "Заказы": "sum",
        })
        .rename(columns=lambda c: f"{c} ({curr_week})" if c != "Поисковый запрос" else c)
    )

    merged = prev_agg.merge(curr_agg, on="Поисковый запрос", how="outer").fillna(0)
    merged["Δ Частота запросов"] = merged[f"Частота запросов ({curr_week})"] - merged[f"Частота запросов ({prev_week})"]
    merged["Δ Частота за неделю"] = merged[f"Частота за неделю ({curr_week})"] - merged[f"Частота за неделю ({prev_week})"]
    merged["Δ Переходы"] = merged[f"Переходы в карточку ({curr_week})"] - merged[f"Переходы в карточку ({prev_week})"]
    merged["Δ Корзина"] = merged[f"Добавления в корзину ({curr_week})"] - merged[f"Добавления в корзину ({prev_week})"]
    merged["Δ Заказы"] = merged[f"Заказы ({curr_week})"] - merged[f"Заказы ({prev_week})"]
    merged["Статус"] = "Без изменений"
    merged.loc[
        (merged[f"Частота за неделю ({prev_week})"] == 0) & (merged[f"Частота за неделю ({curr_week})"] > 0),
        "Статус"
    ] = "Новый запрос"
    merged.loc[
        (merged[f"Частота за неделю ({prev_week})"] > 0) & (merged[f"Частота за неделю ({curr_week})"] == 0),
        "Статус"
    ] = "Исчез запрос"

    top_growth = merged.sort_values(["Δ Частота за неделю", "Δ Заказы"], ascending=[False, False]).head(50).reset_index(drop=True)
    top_decline = merged.sort_values(["Δ Частота за неделю", "Δ Заказы"], ascending=[True, True]).head(50).reset_index(drop=True)
    merged = merged.sort_values(["Δ Частота за неделю", "Δ Заказы"], ascending=[False, False]).reset_index(drop=True)
    return merged, top_growth, top_decline


# =========================
# Yandex Wordstat
# =========================

def yandex_headers(cfg: Config) -> dict[str, str]:
    if not cfg.yandex_api_key:
        raise ValueError("Не задан YANDEX_API_KEY")
    return {
        "Authorization": f"Api-Key {cfg.yandex_api_key}",
        "Content-Type": "application/json",
    }


def yandex_post(url: str, payload: dict[str, Any], cfg: Config) -> dict[str, Any]:
    resp = requests.post(url, headers=yandex_headers(cfg), json=payload, timeout=120)
    try:
        data = resp.json()
    except Exception:
        data = {"raw_text": resp.text}
    if resp.status_code >= 400:
        raise RuntimeError(f"Yandex API {resp.status_code}: {json.dumps(data, ensure_ascii=False)[:1000]}")
    return data


def build_yandex_payload(
    cfg: Config,
    phrase: str,
    *,
    for_dynamics: bool = False,
    dynamics_period: str = "PERIOD_WEEKLY",
) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "folderId": cfg.yandex_folder_id,
        "phrase": phrase,
    }
    if cfg.yandex_region_id:
        payload["regions"] = [str(cfg.yandex_region_id)]
    if for_dynamics:
        payload["period"] = dynamics_period
    else:
        payload["numPhrases"] = 100
    return payload


def try_yandex_dynamics(cfg: Config, phrase: str) -> dict[str, Any]:
    last_error: Optional[Exception] = None
    period_candidates = [
        "PERIOD_WEEKLY",
        "PERIOD_WEEKLY",
        "WEEK",
    ]
    for period_value in period_candidates:
        payload = build_yandex_payload(cfg, phrase, for_dynamics=True, dynamics_period=period_value)
        try:
            return yandex_post(cfg.yandex_dynamics_url, payload, cfg)
        except Exception as exc:
            last_error = exc
            log(f"Yandex Wordstat: GetDynamics не принял period='{period_value}' для фразы '{phrase}'")
            continue
    if last_error:
        raise last_error
    raise RuntimeError("Не удалось выполнить GetDynamics")

def parse_possible_date(value: object) -> Optional[pd.Timestamp]:
    text = normalize_text(value)
    if not text:
        return None
    dt = pd.to_datetime(text, errors="coerce")
    if pd.isna(dt):
        return None
    return pd.Timestamp(dt).normalize()

def recursive_find_dicts(obj: Any) -> list[dict[str, Any]]:
    found: list[dict[str, Any]] = []
    if isinstance(obj, dict):
        found.append(obj)
        for value in obj.values():
            found.extend(recursive_find_dicts(value))
    elif isinstance(obj, list):
        for item in obj:
            found.extend(recursive_find_dicts(item))
    return found


def extract_yandex_top_records(response_json: dict[str, Any], source_phrase: str) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for item in recursive_find_dicts(response_json):
        query = (
            item.get("query")
            or item.get("phrase")
            or item.get("text")
            or item.get("keyword")
            or item.get("searchText")
        )
        if not query:
            continue

        rows.append({
            "Источник фраза": source_phrase,
            "Запрос": normalize_text(query),
            "Показы": safe_float(
                item.get("shows")
                or item.get("showsCount")
                or item.get("count")
                or item.get("freq")
                or item.get("value")
                or item.get("searches")
                or item.get("frequency")
            ),
            "Тип": normalize_text(item.get("type") or item.get("kind") or item.get("group")),
            "Регион": normalize_text(item.get("region") or item.get("regionName")),
        })

    df = pd.DataFrame(rows).drop_duplicates()
    if df.empty:
        return pd.DataFrame(columns=["Источник фраза", "Запрос", "Показы", "Тип", "Регион"])
    df["Запрос"] = df["Запрос"].map(normalize_text)
    df["Показы"] = df["Показы"].map(safe_float)
    df = df[df["Запрос"] != ""].copy()
    return df.sort_values(["Источник фраза", "Показы", "Запрос"], ascending=[True, False, True]).reset_index(drop=True)


def extract_yandex_dynamics_records(response_json: dict[str, Any], source_phrase: str) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for item in recursive_find_dicts(response_json):
        dt_value = item.get("date") or item.get("period") or item.get("month") or item.get("time")
        freq_value = (
            item.get("shows")
            or item.get("count")
            or item.get("value")
            or item.get("freq")
            or item.get("searches")
            or item.get("frequency")
        )
        if dt_value is None or freq_value is None:
            continue

        rows.append({
            "Источник фраза": source_phrase,
            "Период": normalize_text(dt_value),
            "Частотность": safe_float(freq_value),
        })

    df = pd.DataFrame(rows).drop_duplicates()
    if df.empty:
        return pd.DataFrame(columns=["Источник фраза", "Период", "Частотность"])
    return df.sort_values(["Источник фраза", "Период"]).reset_index(drop=True)


def load_yandex_wordstat(cfg: Config) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if not cfg.yandex_api_key or not cfg.yandex_folder_id:
        log("Yandex API секреты не заданы — блок Yandex пропускаем")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    raw_rows: list[dict[str, Any]] = []
    top_frames: list[pd.DataFrame] = []
    dyn_frames: list[pd.DataFrame] = []

    for phrase in cfg.yandex_phrases:
        phrase = phrase.strip()
        if not phrase:
            continue

        try:
            log(f"Yandex Wordstat: GetTop по фразе '{phrase}'")
            top_payload = build_yandex_payload(cfg, phrase, for_dynamics=False)
            top_json = yandex_post(cfg.yandex_top_url, top_payload, cfg)
            raw_rows.append({"Метод": "top", "Фраза": phrase, "JSON": json.dumps(top_json, ensure_ascii=False)})
            top_df = extract_yandex_top_records(top_json, phrase)
            if not top_df.empty:
                top_frames.append(top_df)
        except Exception as exc:
            raw_rows.append({"Метод": "top_error", "Фраза": phrase, "JSON": json.dumps({"error": str(exc)}, ensure_ascii=False)})
            log(f"Yandex Wordstat: ошибка GetTop по фразе '{phrase}': {exc}")

        try:
            log(f"Yandex Wordstat: GetDynamics по фразе '{phrase}'")
            dyn_json = try_yandex_dynamics(cfg, phrase)
            raw_rows.append({"Метод": "dynamics", "Фраза": phrase, "JSON": json.dumps(dyn_json, ensure_ascii=False)})
            dyn_df = extract_yandex_dynamics_records(dyn_json, phrase)
            if not dyn_df.empty:
                dyn_frames.append(dyn_df)
        except Exception as exc:
            raw_rows.append({"Метод": "dynamics_error", "Фраза": phrase, "JSON": json.dumps({"error": str(exc)}, ensure_ascii=False)})
            log(f"Yandex Wordstat: ошибка GetDynamics по фразе '{phrase}': {exc}")

    top_all = pd.concat(top_frames, ignore_index=True) if top_frames else pd.DataFrame()
    dyn_all = pd.concat(dyn_frames, ignore_index=True) if dyn_frames else pd.DataFrame()
    raw_df = pd.DataFrame(raw_rows)

    if not top_all.empty:
        top_all = (
            top_all.groupby(["Запрос"], as_index=False)
            .agg({
                "Показы": "max",
                "Источник фраза": lambda s: ", ".join(sorted(set(map(str, s)))),
                "Тип": lambda s: ", ".join(sorted({x for x in map(str, s) if x and x != 'nan'})),
                "Регион": lambda s: ", ".join(sorted({x for x in map(str, s) if x and x != 'nan'})),
            })
            .sort_values(["Показы", "Запрос"], ascending=[False, True])
            .reset_index(drop=True)
        )
    return top_all, dyn_all, raw_df


def build_yandex_weekly_table(yandex_dyn_df: pd.DataFrame) -> pd.DataFrame:
    if yandex_dyn_df is None or yandex_dyn_df.empty:
        return pd.DataFrame(columns=["Источник фраза", "Дата периода", "ISO год", "ISO неделя", "Неделя", "Частотность"])

    df = yandex_dyn_df.copy()
    df["Дата периода"] = df["Период"].map(parse_possible_date)
    df = df[df["Дата периода"].notna()].copy()
    if df.empty:
        return pd.DataFrame(columns=["Источник фраза", "Дата периода", "ISO год", "ISO неделя", "Неделя", "Частотность"])

    iso = df["Дата периода"].dt.isocalendar()
    df["ISO год"] = iso["year"].astype(int)
    df["ISO неделя"] = iso["week"].astype(int)
    df["Неделя"] = df["ISO год"].astype(str) + "-W" + df["ISO неделя"].astype(str).str.zfill(2)

    return (
        df.groupby(["Источник фраза", "Дата периода", "ISO год", "ISO неделя", "Неделя"], as_index=False)["Частотность"]
        .sum()
        .sort_values(["Источник фраза", "Дата периода"])
        .reset_index(drop=True)
    )

def build_yandex_last_full_week(yandex_weekly_df: pd.DataFrame) -> pd.DataFrame:
    if yandex_weekly_df is None or yandex_weekly_df.empty:
        return pd.DataFrame(columns=[
            "Источник фраза", "Последняя полная неделя", "Предыдущая неделя",
            "Частотность последняя неделя", "Частотность предыдущая неделя", "Δ", "Δ %"
        ])

    completed = yandex_weekly_df.sort_values(["Источник фраза", "Дата периода"]).copy()
    rows: list[dict[str, Any]] = []
    for phrase, grp in completed.groupby("Источник фраза"):
        grp = grp.sort_values("Дата периода").reset_index(drop=True)
        if grp.empty:
            continue
        last = grp.iloc[-1]
        prev = grp.iloc[-2] if len(grp) >= 2 else None
        last_val = safe_float(last.get("Частотность"))
        prev_val = safe_float(prev.get("Частотность")) if prev is not None else 0.0
        delta = last_val - prev_val
        delta_pct = (delta / prev_val * 100.0) if prev_val else None
        rows.append({
            "Источник фраза": phrase,
            "Последняя полная неделя": last.get("Неделя"),
            "Предыдущая неделя": prev.get("Неделя") if prev is not None else "",
            "Частотность последняя неделя": last_val,
            "Частотность предыдущая неделя": prev_val,
            "Δ": delta,
            "Δ %": round(delta_pct, 2) if delta_pct is not None else None,
        })
    return pd.DataFrame(rows).sort_values("Источник фраза").reset_index(drop=True)

def build_yandex_top_30_summary(yandex_top_df: pd.DataFrame) -> pd.DataFrame:
    if yandex_top_df is None or yandex_top_df.empty:
        return pd.DataFrame(columns=["Показатель", "Значение"])

    total_shows = safe_float(yandex_top_df["Показы"].sum()) if "Показы" in yandex_top_df.columns else 0.0
    return pd.DataFrame([
        {"Показатель": "Уникальных запросов за последние 30 дней", "Значение": int(len(yandex_top_df))},
        {"Показатель": "Сумма показов за последние 30 дней", "Значение": round(total_shows, 2)},
        {"Показатель": "Топ-1 запрос", "Значение": normalize_text(yandex_top_df.iloc[0]["Запрос"])},
        {"Показатель": "Топ-1 показы", "Значение": round(safe_float(yandex_top_df.iloc[0]["Показы"]), 2)},
    ])

# =========================
# Excel / Telegram / S3
# =========================

def autofit_worksheet(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name=FONT_NAME, size=FONT_SIZE)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER_THIN

    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)

    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def save_report(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            if df is None or df.empty:
                pd.DataFrame({"Пусто": ["Нет данных"]}).to_excel(writer, sheet_name=sheet_name[:31], index=False)
            else:
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

    wb = load_workbook(path)
    for ws in wb.worksheets:
        autofit_worksheet(ws)
    wb.save(path)


def upload_report(storage: S3Storage, cfg: Config, path: Path) -> str:
    key = f"{cfg.output_prefix}{path.name}"
    with open(path, "rb") as f:
        storage.write_bytes(key, f.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    log(f"Файл загружен в Object Storage: {key}")
    return key


def send_to_telegram(cfg: Config, path: Path, summary_df: pd.DataFrame, detected_variants: list[str], yandex_top_df: pd.DataFrame) -> None:
    if not cfg.telegram_bot_token or not cfg.telegram_chat_id:
        log("Telegram env не заданы — отправку пропускаем")
        return

    wb_last = summary_df.tail(1).to_dict(orient="records")
    wb_prev = summary_df.tail(2).head(1).to_dict(orient="records")
    wb_last = wb_last[0] if wb_last else {}
    wb_prev = wb_prev[0] if wb_prev else {}

    current_queries = int(wb_last.get("Уникальных брендовых запросов", 0) or 0)
    prev_queries = int(wb_prev.get("Уникальных брендовых запросов", 0) or 0)
    delta_queries = current_queries - prev_queries

    yandex_count = int(len(yandex_top_df)) if yandex_top_df is not None and not yandex_top_df.empty else 0
    variants_text = ", ".join(detected_variants[:12]) if detected_variants else "не найдены"

    caption = (
        f"🔎 Брендовый отчёт по запросам {cfg.store_name}\n"
        f"Дата: {cfg.run_date.strftime('%Y-%m-%d')}\n"
        f"WB брендовых запросов: {current_queries} ({delta_queries:+d} к прошлой неделе)\n"
        f"Yandex запросов: {yandex_count}\n"
        f"Варианты бренда: {variants_text}"
    )

    url = f"https://api.telegram.org/bot{cfg.telegram_bot_token}/sendDocument"
    with open(path, "rb") as f:
        resp = requests.post(
            url,
            data={"chat_id": cfg.telegram_chat_id, "caption": caption},
            files={"document": (path.name, f)},
            timeout=120,
        )
    resp.raise_for_status()
    log("Отчёт отправлен в Telegram")


# =========================
# Сводка
# =========================

def build_summary_sheet(
    cfg: Config,
    wb_summary: pd.DataFrame,
    detected_variants: list[str],
    wb_keys: list[str],
    yandex_top_df: pd.DataFrame,
    yandex_last_week_df: pd.DataFrame,
) -> pd.DataFrame:
    rows = [
        {"Параметр": "Магазин", "Значение": cfg.store_name},
        {"Параметр": "Дата запуска", "Значение": cfg.run_date.strftime("%Y-%m-%d")},
        {"Параметр": "Префикс WB weekly файлов", "Значение": cfg.wb_keywords_prefix},
        {"Параметр": "Обработано weekly файлов WB", "Значение": len(wb_keys)},
        {"Параметр": "Yandex фразы", "Значение": ", ".join(cfg.yandex_phrases)},
        {"Параметр": "Варианты бренда в WB", "Значение": ", ".join(detected_variants)},
    ]

    if not wb_summary.empty:
        last = wb_summary.iloc[-1].to_dict()
        rows.extend([
            {"Параметр": "Последняя неделя WB", "Значение": last.get("Неделя", "")},
            {"Параметр": "WB уникальных брендовых запросов", "Значение": int(last.get("Уникальных брендовых запросов", 0))},
            {"Параметр": "WB уникальных артикулов", "Значение": int(last.get("Уникальных артикулов", 0))},
            {"Параметр": "WB сумма частотности за неделю", "Значение": round(safe_float(last.get("Сумма частотности за неделю")), 2)},
            {"Параметр": "WB переходы в карточку", "Значение": round(safe_float(last.get("Переходы в карточку")), 2)},
            {"Параметр": "WB заказы", "Значение": round(safe_float(last.get("Заказы")), 2)},
        ])

    rows.append({"Параметр": "Yandex уникальных запросов за 30 дней", "Значение": int(len(yandex_top_df)) if not yandex_top_df.empty else 0})
    if yandex_last_week_df is not None and not yandex_last_week_df.empty:
        rows.append({"Параметр": "Yandex последняя полная неделя", "Значение": ", ".join(sorted(set(map(str, yandex_last_week_df["Последняя полная неделя"].dropna().tolist()))))})
        rows.append({"Параметр": "Yandex частотность последняя полная неделя", "Значение": round(float(yandex_last_week_df["Частотность последняя неделя"].sum()), 2)})
        rows.append({"Параметр": "Yandex частотность предыдущая неделя", "Значение": round(float(yandex_last_week_df["Частотность предыдущая неделя"].sum()), 2)})
    return pd.DataFrame(rows)


def build_top_wb_queries(wb_brand_df: pd.DataFrame, top_n: int = 200) -> pd.DataFrame:
    if wb_brand_df.empty:
        return pd.DataFrame()
    return (
        wb_brand_df.groupby("Поисковый запрос", as_index=False)
        .agg({
            "Частота запросов": "sum",
            "Частота за неделю": "sum",
            "Переходы в карточку": "sum",
            "Добавления в корзину": "sum",
            "Заказы": "sum",
            "Артикул WB": pd.Series.nunique,
        })
        .rename(columns={"Артикул WB": "Уникальных артикулов"})
        .sort_values(["Частота за неделю", "Заказы"], ascending=[False, False])
        .head(top_n)
        .reset_index(drop=True)
    )


def run() -> Path:
    cfg = get_config()
    storage = S3Storage(cfg)
    brand_regex = build_brand_regex(cfg.brand_variants)

    wb_brand_df, wb_keys = load_wb_brand_data(storage, cfg, brand_regex)
    detected_variants = extract_detected_variants(wb_brand_df if not wb_brand_df.empty else pd.DataFrame({"Поисковый запрос": []}), brand_regex)

    wb_summary = build_wb_weekly_summary(wb_brand_df)
    wb_compare, wb_growth, wb_decline = build_wb_compare_tables(wb_brand_df)
    wb_top = build_top_wb_queries(wb_brand_df)

    yandex_top_df, yandex_dyn_df, yandex_raw_df = load_yandex_wordstat(cfg)
    yandex_weekly_df = build_yandex_weekly_table(yandex_dyn_df)
    yandex_last_week_df = build_yandex_last_full_week(yandex_weekly_df)
    yandex_top_summary_df = build_yandex_top_30_summary(yandex_top_df)

    summary_df = build_summary_sheet(
        cfg=cfg,
        wb_summary=wb_summary,
        detected_variants=detected_variants,
        wb_keys=wb_keys,
        yandex_top_df=yandex_top_df,
        yandex_last_week_df=yandex_last_week_df,
    )

    report_name = f"Брендовые_запросы_{cfg.store_name}_{cfg.run_date.strftime('%Y%m%d')}.xlsx"
    report_path = Path("output") / report_name

    sheets = {
        "Сводка": summary_df,
        "WB_Недели": wb_summary,
        "WB_Топ_запросы": wb_top,
        "WB_Сравнение": wb_compare,
        "WB_Рост": wb_growth,
        "WB_Падение": wb_decline,
        "WB_Сырые_бренд": wb_brand_df,
        "Yandex_Last_Full_Week": yandex_last_week_df,
        "Yandex_Weekly_Dynamics": yandex_weekly_df,
        "Yandex_Top_30d_Summary": yandex_top_summary_df,
        "Yandex_Top_30d": yandex_top_df,
        "Yandex_Dynamics_Raw": yandex_dyn_df,
        "Yandex_Raw": yandex_raw_df,
    }
    save_report(report_path, sheets)
    log(f"Отчёт сохранён: {report_path}")

    upload_report(storage, cfg, report_path)

    if should_send_report(cfg):
        send_to_telegram(cfg, report_path, wb_summary, detected_variants, yandex_top_df)
    else:
        log("Сегодня не понедельник — отправка в Telegram пропущена")

    return report_path


if __name__ == "__main__":
    run()
