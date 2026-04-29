#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import calendar
import io
import os
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import boto3
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TARGET_SUBJECTS = [
    "Кисти косметические",
    "Помады",
    "Блески",
    "Косметические карандаши",
]
TARGET_SUBJECTS_NORM = {s.lower(): s for s in TARGET_SUBJECTS}
EXCLUDE_ARTICLES = {
    "CZ420", "CZ420БРОВИ", "CZ420ГЛАЗА", "DE49", "DE49ГЛАЗА", "PT901"
}
EXAMPLE_ARTICLES = ["901/5", "901/8", "901/14", "901/18"]

BLUE_HEADER = PatternFill("solid", fgColor="1F4E78")
BLUE_TITLE = PatternFill("solid", fgColor="2F75B5")
BLUES = {
    "Кисти косметические": PatternFill("solid", fgColor="D9EAF7"),
    "Помады": PatternFill("solid", fgColor="CFE2F3"),
    "Блески": PatternFill("solid", fgColor="BDD7EE"),
    "Косметические карандаши": PatternFill("solid", fgColor="9DC3E6"),
}
LIGHT_BLUE = PatternFill("solid", fgColor="EAF3FB")
TOTAL_FILL = PatternFill("solid", fgColor="D9E2F3")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN = Side(style="thin", color="D9E2F3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COMMON_ALIASES: Dict[str, List[str]] = {
    "date": ["Дата", "date", "dt", "Дата заказа", "Дата отчета"],
    "nm_id": ["Артикул WB", "Артикул ВБ", "nmId", "nmID", "nmid"],
    "supplier_article": ["Артикул продавца", "supplierArticle", "Артикул поставщика"],
    "subject": ["Предмет", "subject", "Название предмета", "Категория"],
    "title": ["Название", "Название товара", "Товар"],
    "finished_price": ["finishedPrice", "Цена покупателя", "Средняя цена покупателя", "Средняя цена продажи покупателю"],
    "price_with_disc": ["priceWithDisc", "Цена со скидкой продавца", "Средняя цена продажи", "Цена продажи"],
    "is_cancel": ["isCancel", "Отмена", "Отменен"],
    "orders_count": ["ordersCount", "Заказы", "Кол-во заказов"],
    "buyouts_count": ["buyoutsCount", "Выкупы", "Выкупленные заказы"],
    "spend": ["Расход", "spend", "Продвижение"],
    "week": ["Неделя", "week", "Период"],
    "commission_pct": ["Комиссия WB, %", "Комиссия WB %", "Комиссия, %"],
    "acquiring_pct": ["Эквайринг, %", "Эквайринг %"],
    "logistics_direct_unit": ["Логистика прямая, руб/ед", "Логистика прямая"],
    "logistics_return_unit": ["Логистика обратная, руб/ед", "Логистика обратная"],
    "storage_unit": ["Хранение, руб/ед", "Хранение"],
    "other_unit": ["Прочие расходы, руб/ед", "Прочие расходы"],
    "cost_unit": ["Себестоимость, руб", "Себестоимость"],
    "gross_profit": ["Валовая прибыль", "gross_profit"],
    "gross_revenue": ["Валовая выручка", "gross_revenue"],
}


def log(msg: str) -> None:
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}", flush=True)


def normalize_text(v: Any) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\xa0", " ")).strip()


def norm_key(v: Any) -> str:
    s = normalize_text(v).lower().replace("ё", "е")
    s = re.sub(r"[^0-9a-zа-я]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def canonical_subject(v: Any) -> str:
    s = normalize_text(v).lower()
    return TARGET_SUBJECTS_NORM.get(s, "")


def clean_article(v: Any) -> str:
    s = normalize_text(v)
    if s.lower() in {"nan", "none"}:
        return ""
    return s


def upper_article(v: Any) -> str:
    return clean_article(v).upper()


def is_excluded_article(v: Any) -> bool:
    return upper_article(v) in EXCLUDE_ARTICLES


def extract_code(article: Any) -> str:
    s = upper_article(article)
    if not s or is_excluded_article(s):
        return ""
    m = re.match(r"^PT(\d+)", s)
    if m:
        return m.group(1)
    m = re.match(r"^(\d+)", s)
    return m.group(1) if m else ""


def to_numeric(s: Any) -> pd.Series:
    return pd.to_numeric(s, errors="coerce")


def to_dt(s: Any) -> pd.Series:
    return pd.to_datetime(s, errors="coerce").dt.normalize()


def weighted_mean(values: pd.Series, weights: pd.Series) -> float:
    v = pd.to_numeric(values, errors="coerce")
    w = pd.to_numeric(weights, errors="coerce")
    mask = v.notna() & w.notna()
    if not mask.any():
        return np.nan
    v = v[mask]
    w = w[mask]
    if w.sum() == 0:
        return np.nan
    return float(np.average(v, weights=w))


def week_code_from_date(dt_value: Any) -> str:
    ts = pd.Timestamp(dt_value)
    iso = ts.isocalendar()
    return f"{int(iso.year)}-W{int(iso.week):02d}"


def week_bounds_from_code(week_code: str) -> Tuple[Optional[date], Optional[date]]:
    m = re.match(r"^(\d{4})-W(\d{2})$", str(week_code))
    if not m:
        return None, None
    y = int(m.group(1))
    w = int(m.group(2))
    return date.fromisocalendar(y, w, 1), date.fromisocalendar(y, w, 7)


def parse_abc_period_from_name(name: str) -> Tuple[Optional[date], Optional[date]]:
    m = re.search(r"__(\d{2})\.(\d{2})\.(\d{4})-(\d{2})\.(\d{2})\.(\d{4})__", name)
    if not m:
        return None, None
    return date(int(m.group(3)), int(m.group(2)), int(m.group(1))), date(int(m.group(6)), int(m.group(5)), int(m.group(4)))


def parse_week_code_from_text(v: Any) -> Optional[str]:
    s = normalize_text(v)
    if not s:
        return None
    m = re.search(r"(\d{4})-W(\d{2})", s)
    if m:
        return f"{m.group(1)}-W{m.group(2)}"
    dt = pd.to_datetime(s, errors="coerce", dayfirst=True)
    if pd.notna(dt):
        return week_code_from_date(dt)
    return None


def month_name_ru(month: int) -> str:
    names = {
        1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель", 5: "Май", 6: "Июнь",
        7: "Июль", 8: "Август", 9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
    }
    return names[month]


def ensure_aliases(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    norm_existing = {norm_key(c): c for c in out.columns}
    for target, variants in COMMON_ALIASES.items():
        if target in out.columns:
            continue
        for var in variants:
            k = norm_key(var)
            if k in norm_existing:
                out[target] = out[norm_existing[k]]
                break
    return out


def dedupe_columns(cols: Iterable[Any]) -> List[str]:
    out: List[str] = []
    cnt: Dict[str, int] = {}
    for c in cols:
        base = normalize_text(c) or "unnamed"
        cnt[base] = cnt.get(base, 0) + 1
        out.append(base if cnt[base] == 1 else f"{base}__{cnt[base]}")
    return out


def read_excel_flexible(data: bytes, preferred_sheets: Optional[List[str]] = None, header_candidates: Iterable[int] = (0, 1, 2)) -> pd.DataFrame:
    bio = io.BytesIO(data)
    xls = pd.ExcelFile(bio)
    sheet_name = None
    if preferred_sheets:
        norm_map = {norm_key(s): s for s in xls.sheet_names}
        for pref in preferred_sheets:
            if norm_key(pref) in norm_map:
                sheet_name = norm_map[norm_key(pref)]
                break
    if sheet_name is None:
        sheet_name = xls.sheet_names[0]
    best = None
    best_score = -10**9
    for hdr in header_candidates:
        try:
            df = xls.parse(sheet_name=sheet_name, header=hdr, dtype=object)
        except Exception:
            continue
        df.columns = dedupe_columns(df.columns)
        df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
        score = len(df.columns) - (1000 if df.empty else 0)
        if score > best_score:
            best_score = score
            best = df
    if best is None:
        raise ValueError(f"Не удалось прочитать Excel {sheet_name}")
    return ensure_aliases(best)


class BaseStorage:
    def list_files(self, prefix: str) -> List[str]:
        raise NotImplementedError
    def read_bytes(self, path: str) -> bytes:
        raise NotImplementedError
    def write_bytes(self, path: str, data: bytes) -> None:
        raise NotImplementedError
    def exists(self, path: str) -> bool:
        raise NotImplementedError


class LocalStorage(BaseStorage):
    def __init__(self, root: str):
        self.root = Path(root)
    def _abs(self, rel_path: str) -> Path:
        return self.root / rel_path
    def list_files(self, prefix: str) -> List[str]:
        prefix = prefix.replace("\\", "/").rstrip("/")
        base = self._abs(prefix)
        scan_root = base if base.exists() else base.parent
        if not scan_root.exists():
            return []
        out = []
        for p in scan_root.rglob("*"):
            if p.is_file():
                rel = str(p.relative_to(self.root)).replace("\\", "/")
                if rel.startswith(prefix):
                    out.append(rel)
        return sorted(out)
    def glob(self, pattern: str) -> List[str]:
        return sorted(str(p.relative_to(self.root)).replace("\\", "/") for p in self.root.glob(pattern) if p.is_file())
    def read_bytes(self, path: str) -> bytes:
        return self._abs(path).read_bytes()
    def write_bytes(self, path: str, data: bytes) -> None:
        out = self._abs(path)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_bytes(data)
    def exists(self, path: str) -> bool:
        return self._abs(path).exists()


class S3Storage(BaseStorage):
    def __init__(self, bucket: str, access_key: str, secret_key: str):
        self.bucket = bucket
        self.s3 = boto3.client(
            "s3",
            endpoint_url="https://storage.yandexcloud.net",
            aws_access_key_id=access_key,
            aws_secret_access_key=secret_key,
        )
    def list_files(self, prefix: str) -> List[str]:
        out: List[str] = []
        token = None
        while True:
            kwargs = {"Bucket": self.bucket, "Prefix": prefix}
            if token:
                kwargs["ContinuationToken"] = token
            resp = self.s3.list_objects_v2(**kwargs)
            out.extend([x["Key"] for x in resp.get("Contents", []) if not x["Key"].endswith("/")])
            if not resp.get("IsTruncated"):
                break
            token = resp.get("NextContinuationToken")
        return sorted(out)
    def read_bytes(self, path: str) -> bytes:
        return self.s3.get_object(Bucket=self.bucket, Key=path)["Body"].read()
    def write_bytes(self, path: str, data: bytes) -> None:
        self.s3.put_object(Bucket=self.bucket, Key=path, Body=data)
    def exists(self, path: str) -> bool:
        try:
            self.s3.head_object(Bucket=self.bucket, Key=path)
            return True
        except Exception:
            return False


def make_storage(root: str) -> BaseStorage:
    bucket = os.getenv("YC_BUCKET_NAME", "").strip()
    access = os.getenv("YC_ACCESS_KEY_ID", "").strip()
    secret = os.getenv("YC_SECRET_ACCESS_KEY", "").strip()
    if bucket and access and secret:
        log("Using Yandex Object Storage (S3)")
        return S3Storage(bucket, access, secret)
    log("Using local filesystem")
    return LocalStorage(root)


@dataclass
class LoadedData:
    orders: pd.DataFrame
    funnel: pd.DataFrame
    ads_daily: pd.DataFrame
    economics: pd.DataFrame
    abc_weekly: pd.DataFrame
    abc_monthly: pd.DataFrame
    plan: pd.DataFrame
    latest_day: pd.Timestamp
    source_paths: pd.DataFrame
    warnings: pd.DataFrame


class Loader:
    def __init__(self, storage: BaseStorage, reports_root: str = "Отчёты", store: str = "TOPFACE"):
        self.storage = storage
        self.reports_root = reports_root.rstrip("/")
        self.store = store
        self.paths: List[Dict[str, Any]] = []
        self.warnings: List[Dict[str, Any]] = []

    def _prefix(self, *parts: str) -> str:
        return "/".join([self.reports_root, *parts]).replace("//", "/")

    def _local_glob(self, patterns: List[str]) -> List[str]:
        if hasattr(self.storage, "glob"):
            out: List[str] = []
            for p in patterns:
                out.extend(self.storage.glob(p))
            return sorted(set(out))
        return []

    def _record_path(self, source: str, path: str, sheet: str = "") -> None:
        self.paths.append({"Источник": source, "Путь": path, "Лист": sheet})

    def _warn(self, source: str, path: str, message: str) -> None:
        self.warnings.append({"Источник": source, "Путь": path, "Сообщение": message})

    def _finalize(self, df: pd.DataFrame) -> pd.DataFrame:
        out = ensure_aliases(df)
        if "nm_id" in out.columns:
            out["nm_id"] = to_numeric(out["nm_id"])
        if "supplier_article" in out.columns:
            out["supplier_article"] = out["supplier_article"].map(clean_article)
        if "subject" in out.columns:
            out["subject"] = out["subject"].map(canonical_subject)
        if "title" in out.columns:
            out["title"] = out["title"].map(normalize_text)
        if "date" in out.columns:
            out["date"] = to_dt(out["date"])
        return out

    def load_orders(self) -> pd.DataFrame:
        files = self.storage.list_files(self._prefix("Заказы", self.store, "Недельные"))
        files += self.storage.list_files(self._prefix("Заказы", self.store))
        if not files:
            files = self._local_glob(["Заказы_*.xlsx"])
        files = sorted(set(files))
        frames = []
        for path in files:
            try:
                df = read_excel_flexible(self.storage.read_bytes(path), preferred_sheets=None)
                self._record_path("Заказы", path, "")
                df = self._finalize(df)
                if "date" not in df.columns or "nm_id" not in df.columns:
                    continue
                if "is_cancel" not in df.columns:
                    df["is_cancel"] = False
                else:
                    df["is_cancel"] = df["is_cancel"].fillna(False).astype(bool)
                if "orders_qty" in df.columns:
                    df["orders_qty"] = to_numeric(df["orders_qty"]).fillna(0)
                elif "Количество" in df.columns:
                    df["orders_qty"] = to_numeric(df["Количество"]).fillna(0)
                else:
                    df["orders_qty"] = 1.0
                for c in ["finished_price", "price_with_disc"]:
                    if c not in df.columns:
                        df[c] = np.nan
                    df[c] = to_numeric(df[c])
                if "supplier_article" not in df.columns:
                    df["supplier_article"] = ""
                if "subject" not in df.columns:
                    df["subject"] = ""
                df = df[df["subject"].isin(TARGET_SUBJECTS)].copy()
                df = df[~df["supplier_article"].map(is_excluded_article)].copy()
                frames.append(df[["date", "nm_id", "supplier_article", "subject", "title", "orders_qty", "finished_price", "price_with_disc", "is_cancel"]])
            except Exception as e:
                self._warn("Заказы", path, str(e))
        out = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=["date", "nm_id", "supplier_article", "subject", "title", "orders_qty", "finished_price", "price_with_disc", "is_cancel"])
        if not out.empty:
            out = out[out["date"].notna()].copy()
            log(f"Orders rows loaded: {len(out):,}; date range {out['date'].min().date()} .. {out['date'].max().date()}")
        else:
            log("Orders rows loaded: 0")
        return out

    def load_funnel(self) -> pd.DataFrame:
        candidates = [
            self._prefix("Воронка продаж", self.store, "Воронка продаж.xlsx"),
            self._prefix("Воронка продаж", "Воронка продаж.xlsx"),
            "Воронка продаж.xlsx",
            "Воронка продаж (1).xlsx",
        ]
        path = next((p for p in candidates if self.storage.exists(p)), None)
        if path is None:
            found = self._local_glob(["Воронка продаж*.xlsx"])
            path = found[0] if found else None
        if not path:
            return pd.DataFrame(columns=["date", "nm_id", "orders_count", "buyouts_count"])
        try:
            df = read_excel_flexible(self.storage.read_bytes(path), preferred_sheets=None)
            self._record_path("Воронка", path, "")
            df = self._finalize(df)
            if "date" not in df.columns:
                return pd.DataFrame(columns=["date", "nm_id", "orders_count", "buyouts_count"])
            if "orders_count" not in df.columns:
                for c in ["ordersCount", "Заказы"]:
                    if c in df.columns:
                        df["orders_count"] = to_numeric(df[c])
                        break
            if "buyouts_count" not in df.columns:
                for c in ["buyoutsCount", "Выкупы"]:
                    if c in df.columns:
                        df["buyouts_count"] = to_numeric(df[c])
                        break
            out = df[[c for c in ["date", "nm_id", "orders_count", "buyouts_count"] if c in df.columns]].copy()
            out["orders_count"] = to_numeric(out.get("orders_count", 0)).fillna(0)
            out["buyouts_count"] = to_numeric(out.get("buyouts_count", 0)).fillna(0)
            out = out[out["date"].notna()].copy()
            if not out.empty:
                log(f"Funnel rows loaded: {len(out):,}; date range {out['date'].min().date()} .. {out['date'].max().date()}")
            else:
                log("Funnel rows loaded: 0")
            return out
        except Exception as e:
            self._warn("Воронка", path, str(e))
            return pd.DataFrame(columns=["date", "nm_id", "orders_count", "buyouts_count"])

    def load_ads(self) -> pd.DataFrame:
        files = self.storage.list_files(self._prefix("Реклама", self.store, "Недельные"))
        files += self.storage.list_files(self._prefix("Реклама", self.store))
        if not files:
            files = self._local_glob(["Реклама_*.xlsx", "Анализ рекламы.xlsx"])
        files = sorted(set([f for f in files if Path(f).suffix.lower() in {".xlsx", ".xlsm"}]))
        frames = []
        for path in files:
            try:
                raw = self.storage.read_bytes(path)
                xls = pd.ExcelFile(io.BytesIO(raw))
                for sh in xls.sheet_names:
                    if "статист" not in norm_key(sh):
                        continue
                    try:
                        df = read_excel_flexible(raw, preferred_sheets=[sh], header_candidates=(0, 1, 2))
                        self._record_path("Реклама", path, sh)
                        df = self._finalize(df)
                        if "date" not in df.columns or "nm_id" not in df.columns:
                            continue
                        if "spend" not in df.columns:
                            continue
                        out = df[["date", "nm_id", "spend"]].copy()
                        out["spend"] = to_numeric(out["spend"]).fillna(0)
                        out = out[out["date"].notna() & out["nm_id"].notna()].copy()
                        frames.append(out)
                    except Exception as e:
                        self._warn("Реклама", f"{path}::{sh}", str(e))
            except Exception as e:
                self._warn("Реклама", path, str(e))
        out = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=["date", "nm_id", "spend"])
        if not out.empty:
            out = out.groupby(["date", "nm_id"], as_index=False)["spend"].sum()
            log(f"Ads rows loaded: {len(out):,}; date range {out['date'].min().date()} .. {out['date'].max().date()}; spend sum {out['spend'].sum():,.0f}")
        else:
            log("Ads rows loaded: 0")
        return out

    def load_economics(self) -> pd.DataFrame:
        candidates = [
            self._prefix("Финансовые показатели", self.store, "Экономика.xlsx"),
            self._prefix("Финансовые показатели", self.store, "Недельные", "Экономика.xlsx"),
            "Экономика.xlsx",
            "Экономика (4).xlsx",
        ]
        path = next((p for p in candidates if self.storage.exists(p)), None)
        if not path:
            found = self._local_glob(["Экономика*.xlsx"])
            path = found[0] if found else None
        if not path:
            return pd.DataFrame()
        try:
            df = read_excel_flexible(self.storage.read_bytes(path), preferred_sheets=["Юнит экономика"], header_candidates=(0, 1, 2))
            self._record_path("Экономика", path, "Юнит экономика")
            df = self._finalize(df)
            if "week" not in df.columns:
                df["week"] = df.get("Неделя", "")
            df["week"] = df["week"].map(parse_week_code_from_text)
            if "supplier_article" not in df.columns:
                df["supplier_article"] = ""
            if "subject" not in df.columns:
                df["subject"] = ""
            for c in ["commission_pct", "acquiring_pct", "logistics_direct_unit", "logistics_return_unit", "storage_unit", "other_unit", "cost_unit"]:
                if c not in df.columns:
                    df[c] = np.nan
                df[c] = to_numeric(df[c])
            df = df[df["subject"].isin(TARGET_SUBJECTS)].copy()
            df = df[~df["supplier_article"].map(is_excluded_article)].copy()
            df["code"] = df["supplier_article"].map(extract_code)
            log(f"Economics rows loaded: {len(df):,}; weeks {', '.join(sorted([x for x in df['week'].dropna().unique()])[:10])}")
            return df[["week", "nm_id", "supplier_article", "subject", "title", "code", "commission_pct", "acquiring_pct", "logistics_direct_unit", "logistics_return_unit", "storage_unit", "other_unit", "cost_unit"]]
        except Exception as e:
            self._warn("Экономика", path, str(e))
            return pd.DataFrame()

    def load_abc(self) -> Tuple[pd.DataFrame, pd.DataFrame]:
        files = self.storage.list_files(self._prefix("ABC"))
        if not files:
            files = self._local_glob(["wb_abc_report_goods__*.xlsx"])
        files = [f for f in files if "wb_abc_report_goods__" in Path(f).name]
        weekly_frames: List[pd.DataFrame] = []
        monthly_frames: List[pd.DataFrame] = []
        for path in sorted(files):
            try:
                name = Path(path).name
                start, end = parse_abc_period_from_name(name)
                if not start or not end:
                    continue
                df = read_excel_flexible(self.storage.read_bytes(path), preferred_sheets=None, header_candidates=(0, 1, 2))
                self._record_path("ABC", path, "")
                df = self._finalize(df)
                if "gross_profit" not in df.columns:
                    continue
                if "gross_revenue" not in df.columns:
                    df["gross_revenue"] = np.nan
                for c in ["gross_profit", "gross_revenue"]:
                    df[c] = to_numeric(df[c]).fillna(0)
                if "supplier_article" not in df.columns:
                    df["supplier_article"] = ""
                if "subject" not in df.columns:
                    df["subject"] = ""
                df = df[df["subject"].isin(TARGET_SUBJECTS)].copy()
                df = df[~df["supplier_article"].map(is_excluded_article)].copy()
                df["code"] = df["supplier_article"].map(extract_code)
                df["week_code"] = week_code_from_date(start)
                df["week_label"] = f"{pd.Timestamp(start).strftime('%d.%m')}-{pd.Timestamp(end).strftime('%d.%m')}"
                df["month_key"] = pd.Timestamp(start).strftime("%Y-%m")
                df["vat"] = df["gross_revenue"] * 7.0 / 107.0
                df["gp_minus_nds"] = df["gross_profit"] - df["vat"]
                month_end = (pd.Timestamp(start).to_period("M").end_time.normalize()).date()
                keep_cols = ["week_code", "week_label", "month_key", "nm_id", "supplier_article", "subject", "title", "code", "gross_profit", "gross_revenue", "vat", "gp_minus_nds"]
                if start.day == 1 and end == month_end:
                    monthly_frames.append(df[[c for c in keep_cols if c in df.columns]])
                else:
                    weekly_frames.append(df[[c for c in keep_cols if c in df.columns]])
            except Exception as e:
                self._warn("ABC", path, str(e))
        weekly = pd.concat(weekly_frames, ignore_index=True) if weekly_frames else pd.DataFrame()
        monthly = pd.concat(monthly_frames, ignore_index=True) if monthly_frames else pd.DataFrame()
        if not weekly.empty:
            log(f"ABC weekly rows loaded: {len(weekly):,}; weeks {', '.join(sorted(weekly['week_code'].dropna().unique()))}")
        else:
            log("ABC weekly rows loaded: 0")
        if not monthly.empty:
            log(f"ABC monthly rows loaded: {len(monthly):,}; months {', '.join(sorted(monthly['month_key'].dropna().unique()))}")
        else:
            log("ABC monthly rows loaded: 0")
        return weekly, monthly

    def load_plan(self, current_month_key: str) -> pd.DataFrame:
        candidates = [self._prefix("Объединенный отчет", self.store, "План.xlsx"), "План.xlsx"]
        path = next((p for p in candidates if self.storage.exists(p)), None)
        if not path:
            return pd.DataFrame(columns=["supplier_article", "subject", "plan_month"])
        try:
            raw = self.storage.read_bytes(path)
            xls = pd.ExcelFile(io.BytesIO(raw))
            sheet_name = next((s for s in xls.sheet_names if "итог" in norm_key(s)), xls.sheet_names[0])
            best = None
            best_score = -10**9
            for hdr in (0, 1, 2, 3):
                try:
                    df = pd.read_excel(io.BytesIO(raw), sheet_name=sheet_name, header=hdr, dtype=object)
                except Exception:
                    continue
                df.columns = dedupe_columns(df.columns)
                df = ensure_aliases(df).dropna(axis=0, how="all").dropna(axis=1, how="all")
                score = len(df.columns)
                if score > best_score:
                    best = df
                    best_score = score
            if best is None:
                return pd.DataFrame(columns=["supplier_article", "subject", "plan_month"])
            self._record_path("План", path, sheet_name)
            df = self._finalize(best)
            target = f"ВП-НДС {month_name_ru(int(current_month_key[-2:]))} {current_month_key[:4]}"
            plan_col = None
            for c in df.columns:
                if norm_key(c) == norm_key(target) or norm_key(target) in norm_key(c):
                    plan_col = c
                    break
            if plan_col is None:
                for c in df.columns:
                    if "вп ндс" in norm_key(c):
                        plan_col = c
                        break
            if plan_col is None:
                return pd.DataFrame(columns=["supplier_article", "subject", "plan_month"])
            if "supplier_article" not in df.columns:
                return pd.DataFrame(columns=["supplier_article", "subject", "plan_month"])
            out = pd.DataFrame({
                "supplier_article": df["supplier_article"].map(clean_article),
                "subject": df.get("subject", "").map(canonical_subject),
                "plan_month": to_numeric(df[plan_col]).fillna(np.nan)
            })
            out = out[~out["supplier_article"].map(is_excluded_article)].copy()
            out["code"] = out["supplier_article"].map(extract_code)
            log(f"Plan rows loaded: {len(out):,}; non-null plan {out['plan_month'].notna().sum():,}")
            return out
        except Exception as e:
            self._warn("План", path, str(e))
            return pd.DataFrame(columns=["supplier_article", "subject", "plan_month"])

    def load_all(self) -> LoadedData:
        log("Loading data")
        log("Loading orders")
        orders = self.load_orders()
        log("Loading funnel")
        funnel = self.load_funnel()
        log("Loading ads")
        ads = self.load_ads()
        log("Loading economics")
        econ = self.load_economics()
        log("Loading ABC")
        abc_weekly, abc_monthly = self.load_abc()
        latest_candidates = [pd.to_datetime(df[col], errors="coerce").max() for df, col in [(orders, "date"), (funnel, "date"), (ads, "date")] if not df.empty and col in df.columns]
        latest_day = max([x for x in latest_candidates if pd.notna(x)], default=pd.Timestamp.today().normalize())
        log("Loading plan")
        plan = self.load_plan(latest_day.to_period("M").strftime("%Y-%m"))
        return LoadedData(
            orders=orders,
            funnel=funnel,
            ads_daily=ads,
            economics=econ,
            abc_weekly=abc_weekly,
            abc_monthly=abc_monthly,
            plan=plan,
            latest_day=pd.Timestamp(latest_day).normalize(),
            source_paths=pd.DataFrame(self.paths) if self.paths else pd.DataFrame(columns=["Источник", "Путь", "Лист"]),
            warnings=pd.DataFrame(self.warnings) if self.warnings else pd.DataFrame(columns=["Источник", "Путь", "Сообщение"]),
        )


class Builder:
    def __init__(self, data: LoadedData):
        self.data = data
        self.latest_day = pd.Timestamp(data.latest_day).normalize()
        self.current_week_start = self.latest_day - pd.Timedelta(days=self.latest_day.weekday())
        self.week_days = [self.current_week_start + pd.Timedelta(days=i) for i in range(7)]
        self.current_month_key = self.latest_day.to_period("M").strftime("%Y-%m")
        self.days_in_month = calendar.monthrange(self.latest_day.year, self.latest_day.month)[1]
        self.master = self.build_dictionary()
        self.buyout90 = self.build_buyout90()
        self.econ_article_map, self.subject_week_commission, self.subject_latest_commission = self.build_econ_maps()

    def build_dictionary(self) -> pd.DataFrame:
        frames = []
        for df in [self.data.orders, self.data.economics, self.data.abc_weekly, self.data.abc_monthly]:
            if df.empty:
                continue
            x = df.copy()
            for c in ["nm_id", "supplier_article", "subject", "title"]:
                if c not in x.columns:
                    x[c] = np.nan
            x = x[["nm_id", "supplier_article", "subject", "title"]].copy()
            frames.append(x)
        if not frames:
            return pd.DataFrame(columns=["nm_id", "supplier_article", "subject", "code", "title"])
        m = pd.concat(frames, ignore_index=True)
        m["nm_id"] = to_numeric(m["nm_id"])
        m["supplier_article"] = m["supplier_article"].map(clean_article)
        m["subject"] = m["subject"].map(canonical_subject)
        m["title"] = m["title"].map(normalize_text)
        m["code"] = m["supplier_article"].map(extract_code)
        m = m[(m["subject"].isin(TARGET_SUBJECTS)) & (m["code"] != "") & (~m["supplier_article"].map(is_excluded_article))].copy()
        m["quality"] = m["subject"].ne("").astype(int) * 4 + m["supplier_article"].ne("").astype(int) * 2 + m["title"].ne("").astype(int)
        m = m.sort_values("quality", ascending=False).drop_duplicates(subset=["supplier_article", "nm_id"], keep="first")
        return m[["nm_id", "supplier_article", "subject", "code", "title"]]

    def attach_dictionary(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty or self.master.empty:
            return df.copy()
        out = df.copy()
        if "nm_id" in out.columns:
            nm_map = self.master.dropna(subset=["nm_id"]).drop_duplicates(subset=["nm_id"])
            out = out.merge(nm_map[["nm_id", "supplier_article", "subject", "code", "title"]], on="nm_id", how="left", suffixes=("", "_m"))
            for c in ["supplier_article", "subject", "code", "title"]:
                if c not in out.columns:
                    out[c] = out[f"{c}_m"]
                else:
                    mask = out[c].isna() | (out[c] == "")
                    out.loc[mask, c] = out.loc[mask, f"{c}_m"]
                out.drop(columns=[f"{c}_m"], inplace=True, errors="ignore")
        if "supplier_article" in out.columns:
            art_map = self.master.drop_duplicates(subset=["supplier_article"])
            out = out.merge(art_map[["supplier_article", "nm_id", "subject", "code", "title"]], on="supplier_article", how="left", suffixes=("", "_a"))
            for c in ["nm_id", "subject", "code", "title"]:
                if c not in out.columns:
                    out[c] = out[f"{c}_a"]
                else:
                    mask = out[c].isna() | (out[c] == "")
                    out.loc[mask, c] = out.loc[mask, f"{c}_a"]
                out.drop(columns=[f"{c}_a"], inplace=True, errors="ignore")
        return out

    def build_buyout90(self) -> pd.DataFrame:
        f = self.data.funnel.copy()
        if f.empty:
            return pd.DataFrame(columns=["nm_id", "buyout_pct_90"])
        f = f[(f["date"] >= self.latest_day - pd.Timedelta(days=89)) & (f["date"] <= self.latest_day)].copy()
        out = f.groupby("nm_id", as_index=False).agg(orders_90=("orders_count", "sum"), buyouts_90=("buyouts_count", "sum"))
        out["buyout_pct_90"] = np.where(out["orders_90"] > 0, out["buyouts_90"] / out["orders_90"], np.nan)
        log(f"Buyout90 rows: {len(out):,}; non-null ratios {out['buyout_pct_90'].notna().sum():,}")
        return out[["nm_id", "buyout_pct_90"]]

    def build_econ_maps(self) -> Tuple[pd.DataFrame, Dict[Tuple[str, str], float], Dict[str, float]]:
        econ = self.attach_dictionary(self.data.economics)
        if econ.empty:
            return pd.DataFrame(), {}, {}
        econ = econ[(econ["subject"].isin(TARGET_SUBJECTS)) & (econ["supplier_article"].map(extract_code) != "")].copy()
        econ["week"] = econ["week"].astype(str)
        econ["week_start"] = econ["week"].map(lambda x: pd.Timestamp(week_bounds_from_code(x)[0]) if week_bounds_from_code(x)[0] else pd.NaT)
        econ = econ.sort_values(["supplier_article", "week_start"], ascending=[True, False])
        by_subject_week: Dict[Tuple[str, str], float] = {}
        tmp = econ[(to_numeric(econ["commission_pct"]).fillna(0) > 0)].copy()
        if not tmp.empty:
            g = tmp.groupby(["subject", "week"], as_index=False)["commission_pct"].median()
            by_subject_week = {(r.subject, r.week): float(r.commission_pct) for r in g.itertuples(index=False)}
        by_subject_latest: Dict[str, float] = {}
        if not tmp.empty:
            tmp = tmp.sort_values(["subject", "week_start"], ascending=[True, False]).drop_duplicates(subset=["subject"], keep="first")
            by_subject_latest = {r.subject: float(r.commission_pct) for r in tmp.itertuples(index=False)}
        log(f"Economics usable rows: {len(econ):,}; articles {econ['supplier_article'].nunique():,}")
        return econ, by_subject_week, by_subject_latest

    def pick_econ_for_daily(self, keys: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
        if keys.empty or self.econ_article_map.empty:
            empty = pd.DataFrame(columns=["day", "nm_id", "supplier_article", "econ_week", "commission_pct", "acquiring_pct", "logistics_direct_unit", "logistics_return_unit", "storage_unit", "other_unit", "cost_unit"])
            return empty, pd.DataFrame(columns=["day", "supplier_article", "subject", "target_week", "econ_week", "commission_source"]) 
        rows = []
        diag = []
        exact = fallback = missing = 0
        for rec in keys.itertuples(index=False):
            day = rec.day
            nm_id = rec.nm_id
            art = rec.supplier_article
            subject = rec.subject
            target_week = week_code_from_date(day)
            g = self.econ_article_map[self.econ_article_map["supplier_article"] == art].copy()
            if g.empty and pd.notna(nm_id):
                g = self.econ_article_map[self.econ_article_map["nm_id"] == nm_id].copy()
            if g.empty:
                missing += 1
                continue
            exact_g = g[g["week"].astype(str) == str(target_week)]
            if not exact_g.empty:
                chosen = exact_g.iloc[0]
                exact += 1
            else:
                chosen = g.iloc[0]
                fallback += 1
            commission_pct = float(pd.to_numeric(pd.Series([chosen.get("commission_pct")]), errors="coerce").fillna(0).iloc[0])
            source = "article_week" if commission_pct > 0 else ""
            if commission_pct <= 0:
                commission_pct = float(self.subject_week_commission.get((subject, str(target_week)), np.nan)) if subject else np.nan
                source = "subject_week" if pd.notna(commission_pct) and commission_pct > 0 else source
            if pd.isna(commission_pct) or commission_pct <= 0:
                commission_pct = float(self.subject_latest_commission.get(subject, np.nan)) if subject else np.nan
                source = "subject_latest" if pd.notna(commission_pct) and commission_pct > 0 else source
            row = {
                "day": day,
                "nm_id": nm_id,
                "supplier_article": art,
                "econ_week": chosen.get("week", ""),
                "commission_pct": commission_pct,
                "acquiring_pct": float(pd.to_numeric(pd.Series([chosen.get("acquiring_pct")]), errors="coerce").fillna(0).iloc[0]),
                "logistics_direct_unit": float(pd.to_numeric(pd.Series([chosen.get("logistics_direct_unit")]), errors="coerce").fillna(0).iloc[0]),
                "logistics_return_unit": float(pd.to_numeric(pd.Series([chosen.get("logistics_return_unit")]), errors="coerce").fillna(0).iloc[0]),
                "storage_unit": float(pd.to_numeric(pd.Series([chosen.get("storage_unit")]), errors="coerce").fillna(0).iloc[0]),
                "other_unit": float(pd.to_numeric(pd.Series([chosen.get("other_unit")]), errors="coerce").fillna(0).iloc[0]),
                "cost_unit": float(pd.to_numeric(pd.Series([chosen.get("cost_unit")]), errors="coerce").fillna(0).iloc[0]),
            }
            rows.append(row)
            diag.append({
                "day": day, "supplier_article": art, "subject": subject,
                "target_week": target_week, "econ_week": chosen.get("week", ""),
                "commission_source": source or "missing"
            })
        log(f"Economics matching: exact week = {exact:,}, fallback latest = {fallback:,}, missing = {missing:,}")
        return pd.DataFrame(rows), pd.DataFrame(diag)

    def build_current_week_daily(self) -> Tuple[pd.DataFrame, Dict[str, pd.DataFrame]]:
        orders = self.attach_dictionary(self.data.orders)
        orders = orders[(orders["date"] >= self.current_week_start) & (orders["date"] <= self.latest_day) & (~orders["is_cancel"])].copy()
        orders = orders[(orders["subject"].isin(TARGET_SUBJECTS)) & (~orders["supplier_article"].map(is_excluded_article))].copy()
        if orders.empty:
            return pd.DataFrame(), {"orders_used": orders}
        log(f"Current week order rows: {len(orders):,}; day range {orders['date'].min().date()} .. {orders['date'].max().date()}")
        daily = orders.groupby(["date", "nm_id", "supplier_article", "subject", "code", "title"], dropna=False).agg(
            orders_day=("orders_qty", "sum"),
            finished_price_avg=("finished_price", lambda s: weighted_mean(s, orders.loc[s.index, "orders_qty"].fillna(1))),
            price_with_disc_avg=("price_with_disc", lambda s: weighted_mean(s, orders.loc[s.index, "orders_qty"].fillna(1))),
        ).reset_index().rename(columns={"date": "day"})
        daily = daily.merge(self.buyout90, on="nm_id", how="left")
        econ_pick, econ_diag = self.pick_econ_for_daily(daily[["day", "nm_id", "supplier_article", "subject"]].drop_duplicates())
        daily = daily.merge(econ_pick, on=["day", "nm_id", "supplier_article"], how="left")
        ads = self.data.ads_daily.copy()
        ads = ads[(ads["date"] >= self.current_week_start) & (ads["date"] <= self.latest_day)].copy()
        ads = ads.rename(columns={"date": "day", "spend": "ad_spend_day"})
        daily = daily.merge(ads[["day", "nm_id", "ad_spend_day"]], on=["day", "nm_id"], how="left")
        daily["ad_spend_day"] = to_numeric(daily["ad_spend_day"]).fillna(0)
        daily["buyout_factor"] = to_numeric(daily["buyout_pct_90"]).fillna(1.0)
        daily["buyout_qty"] = daily["orders_day"] * daily["buyout_factor"]
        daily["price_with_disc_avg"] = to_numeric(daily["price_with_disc_avg"]).fillna(0)
        daily["finished_price_avg"] = to_numeric(daily["finished_price_avg"]).fillna(0)
        daily["revenue_pwd"] = daily["buyout_qty"] * daily["price_with_disc_avg"]
        daily["commission_rub"] = daily["revenue_pwd"] * to_numeric(daily["commission_pct"]).fillna(0) / 100.0
        daily["acquiring_rub"] = daily["revenue_pwd"] * to_numeric(daily["acquiring_pct"]).fillna(0) / 100.0
        daily["logistics_direct_rub"] = daily["buyout_qty"] * to_numeric(daily["logistics_direct_unit"]).fillna(0)
        daily["logistics_return_rub"] = daily["buyout_qty"] * to_numeric(daily["logistics_return_unit"]).fillna(0)
        daily["storage_rub"] = daily["buyout_qty"] * to_numeric(daily["storage_unit"]).fillna(0)
        daily["other_rub"] = daily["buyout_qty"] * to_numeric(daily["other_unit"]).fillna(0)
        daily["cost_rub"] = daily["buyout_qty"] * to_numeric(daily["cost_unit"]).fillna(0)
        daily["vat_rub"] = daily["buyout_qty"] * daily["finished_price_avg"] * 7.0 / 107.0
        daily["gp_minus_nds_raw"] = (
            daily["revenue_pwd"] - daily["commission_rub"] - daily["acquiring_rub"] - daily["logistics_direct_rub"] -
            daily["logistics_return_rub"] - daily["storage_rub"] - daily["other_rub"] - daily["cost_rub"] - daily["ad_spend_day"] - daily["vat_rub"]
        )
        daily["negative_flag"] = daily["gp_minus_nds_raw"] < 0
        # Main sheet must not show negatives without diagnostics
        daily["gp_minus_nds"] = daily["gp_minus_nds_raw"].clip(lower=0)
        log(f"Ads matching to daily rows: matched rows = {(daily['ad_spend_day'] > 0).sum():,} из {len(daily):,}; spend matched = {daily['ad_spend_day'].sum():,.0f}")
        tech = {
            "orders_used": orders,
            "ads_used": ads,
            "economics_used": econ_pick,
            "daily_formula": daily,
            "diagnostics": daily[daily["negative_flag"]].copy(),
            "econ_diag": econ_diag,
        }
        return daily, tech

    def build_weekly_facts(self) -> pd.DataFrame:
        abc = self.attach_dictionary(self.data.abc_weekly)
        if abc.empty:
            return pd.DataFrame()
        current_week = week_code_from_date(self.latest_day)
        weeks = sorted([w for w in abc["week_code"].dropna().unique() if w != current_week])[-4:]
        out = abc[abc["week_code"].isin(weeks)].copy()
        return out

    def build_monthly_facts(self) -> pd.DataFrame:
        month_keys = [(self.latest_day.to_period("M") - 2).strftime("%Y-%m"), (self.latest_day.to_period("M") - 1).strftime("%Y-%m"), self.current_month_key]
        monthly = self.attach_dictionary(self.data.abc_monthly)
        weekly = self.attach_dictionary(self.data.abc_weekly)
        frames = []
        if not monthly.empty:
            frames.append(monthly[monthly["month_key"].isin(month_keys)].copy())
        have_current = not monthly.empty and self.current_month_key in set(monthly["month_key"].astype(str))
        if not have_current and not weekly.empty:
            curm = weekly[pd.to_datetime(weekly["week_code"].map(lambda x: week_bounds_from_code(x)[0])).dt.to_period("M").astype(str) == self.current_month_key].copy()
            if not curm.empty:
                curm = curm.groupby(["month_key", "nm_id", "supplier_article", "subject", "title", "code"], as_index=False).agg(
                    gross_profit=("gross_profit", "sum"), gross_revenue=("gross_revenue", "sum"), vat=("vat", "sum"), gp_minus_nds=("gp_minus_nds", "sum")
                )
                curm["month_key"] = self.current_month_key
                if frames:
                    frames = [f[f["month_key"] != self.current_month_key] for f in frames]
                frames.append(curm)
        return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

    def build_plan_maps(self) -> Tuple[Dict[str, float], Dict[Tuple[str, str], float], Dict[str, float]]:
        plan = self.attach_dictionary(self.data.plan)
        if plan.empty:
            return {}, {}, {}
        art_map: Dict[str, float] = {}
        for r in plan.dropna(subset=["supplier_article"]).itertuples(index=False):
            if pd.notna(r.plan_month):
                art_map[r.supplier_article] = float(r.plan_month)
        prod = plan.groupby(["subject", "code"], as_index=False)["plan_month"].sum(min_count=1)
        prod_map = {(r.subject, r.code): float(r.plan_month) for r in prod.itertuples(index=False) if pd.notna(r.plan_month)}
        cat = plan.groupby("subject", as_index=False)["plan_month"].sum(min_count=1)
        cat_map = {r.subject: float(r.plan_month) for r in cat.itertuples(index=False) if pd.notna(r.plan_month)}
        return art_map, prod_map, cat_map

    def build_blocks(self) -> Dict[str, pd.DataFrame]:
        daily, tech = self.build_current_week_daily()
        weekly = self.build_weekly_facts()
        monthly = self.build_monthly_facts()
        art_plan, prod_plan, cat_plan = self.build_plan_maps()
        month_fact_art = monthly[monthly["month_key"] == self.current_month_key].groupby("supplier_article", as_index=False)["gp_minus_nds"].sum().set_index("supplier_article")["gp_minus_nds"].to_dict() if not monthly.empty else {}
        month_fact_prod = monthly[monthly["month_key"] == self.current_month_key].groupby(["subject", "code"], as_index=False)["gp_minus_nds"].sum().set_index(["subject", "code"])["gp_minus_nds"].to_dict() if not monthly.empty else {}
        month_fact_cat = monthly[monthly["month_key"] == self.current_month_key].groupby("subject", as_index=False)["gp_minus_nds"].sum().set_index("subject")["gp_minus_nds"].to_dict() if not monthly.empty else {}

        day_labels = [d.strftime("%a %d.%m").replace("Mon", "Пн").replace("Tue", "Вт").replace("Wed", "Ср").replace("Thu", "Чт").replace("Fri", "Пт").replace("Sat", "Сб").replace("Sun", "Вс") for d in self.week_days]
        daily_map = {d.normalize(): lbl for d, lbl in zip(self.week_days, day_labels)}
        if not daily.empty:
            daily = daily.copy()
            daily["period_label"] = daily["day"].map(daily_map)
        week_labels = sorted(weekly["week_label"].dropna().unique().tolist()) if not weekly.empty else []
        month_labels = [(self.latest_day.to_period("M") - 2).strftime("%m.%Y"), (self.latest_day.to_period("M") - 1).strftime("%m.%Y"), self.latest_day.to_period("M").strftime("%m.%Y")]
        if not monthly.empty:
            monthly = monthly.copy()
            monthly["period_label"] = pd.to_datetime(monthly["month_key"] + "-01").dt.strftime("%m.%Y")

        def agg_rows(base: pd.DataFrame, value_col: str, label_col: str, labels: List[str], plan_kind: str) -> pd.DataFrame:
            if base.empty:
                return pd.DataFrame()
            rows = []
            for subject in TARGET_SUBJECTS:
                sg = base[base["subject"] == subject].copy()
                if sg.empty:
                    continue
                cat_row = {"Категория": subject, "_kind": "category", "_subject": subject}
                for lbl in labels:
                    cat_row[lbl] = float(sg.loc[sg[label_col] == lbl, value_col].sum())
                if plan_kind == "daily":
                    p = cat_plan.get(subject, np.nan)
                    cat_row["План"] = float(sum(cat_row[l] for l in labels) / max(1, len([l for l in labels if sum(sg[label_col] == l) >= 0]))) if pd.isna(p) else float(p) / self.days_in_month
                elif plan_kind == "weekly":
                    p = cat_plan.get(subject, np.nan)
                    cat_row["План"] = float(month_fact_cat.get(subject, 0.0)) if pd.isna(p) else float(p) * 7.0 / self.days_in_month
                else:
                    p = cat_plan.get(subject, np.nan)
                    cat_row["План"] = float(month_fact_cat.get(subject, 0.0)) if pd.isna(p) else float(p)
                rows.append(cat_row)
                prod_order = sg.groupby("code", as_index=False)[value_col].sum().sort_values(value_col, ascending=False)["code"].tolist()
                for code in prod_order:
                    pg = sg[sg["code"] == code].copy()
                    prod_row = {"Категория": code, "_kind": "product", "_subject": subject, "_code": code}
                    for lbl in labels:
                        prod_row[lbl] = float(pg.loc[pg[label_col] == lbl, value_col].sum())
                    if plan_kind == "daily":
                        p = prod_plan.get((subject, code), np.nan)
                        prod_row["План"] = float(month_fact_prod.get((subject, code), 0.0)) / max(1, self.days_in_month) if pd.isna(p) else float(p) / self.days_in_month
                    elif plan_kind == "weekly":
                        p = prod_plan.get((subject, code), np.nan)
                        prod_row["План"] = float(month_fact_prod.get((subject, code), 0.0)) if pd.isna(p) else float(p) * 7.0 / self.days_in_month
                    else:
                        p = prod_plan.get((subject, code), np.nan)
                        prod_row["План"] = float(month_fact_prod.get((subject, code), 0.0)) if pd.isna(p) else float(p)
                    rows.append(prod_row)
                    art_order = pg.groupby("supplier_article", as_index=False)[value_col].sum().sort_values(value_col, ascending=False)["supplier_article"].tolist()
                    for art in art_order:
                        ag = pg[pg["supplier_article"] == art].copy()
                        art_row = {"Категория": art, "_kind": "article", "_subject": subject, "_code": code, "_article": art}
                        for lbl in labels:
                            art_row[lbl] = float(ag.loc[ag[label_col] == lbl, value_col].sum())
                        if plan_kind == "daily":
                            p = art_plan.get(art, np.nan)
                            art_row["План"] = float(month_fact_art.get(art, 0.0)) / max(1, self.days_in_month) if pd.isna(p) else float(p) / self.days_in_month
                        elif plan_kind == "weekly":
                            p = art_plan.get(art, np.nan)
                            art_row["План"] = float(month_fact_art.get(art, 0.0)) if pd.isna(p) else float(p) * 7.0 / self.days_in_month
                        else:
                            p = art_plan.get(art, np.nan)
                            art_row["План"] = float(month_fact_art.get(art, 0.0)) if pd.isna(p) else float(p)
                        rows.append(art_row)
            total = {"Категория": "Итого по всем 4 категориям", "_kind": "grand_total"}
            for lbl in labels:
                total[lbl] = float(base.loc[base[label_col] == lbl, value_col].sum())
            if plan_kind == "daily":
                total["План"] = float(sum(cat_plan.values())) / self.days_in_month if cat_plan else float(sum(total[l] for l in labels)) / max(1, len(labels))
            elif plan_kind == "weekly":
                total["План"] = float(sum(cat_plan.values())) * 7.0 / self.days_in_month if cat_plan else float(sum(month_fact_cat.values()))
            else:
                total["План"] = float(sum(cat_plan.values())) if cat_plan else float(sum(month_fact_cat.values()))
            rows.append(total)
            return pd.DataFrame(rows)

        daily_block = agg_rows(daily, "gp_minus_nds", "period_label", day_labels, "daily") if not daily.empty else pd.DataFrame()
        weekly_block = agg_rows(weekly, "gp_minus_nds", "week_label", week_labels, "weekly") if not weekly.empty else pd.DataFrame()
        monthly_block = agg_rows(monthly, "gp_minus_nds", "period_label", month_labels, "monthly") if not monthly.empty else pd.DataFrame()

        # keep only requested tech tabs
        blocks = {
            "daily": daily_block,
            "weekly": weekly_block,
            "monthly": monthly_block,
            "dictionary": self.master,
            "orders_used": tech.get("orders_used", pd.DataFrame()),
            "funnel_used": self.buyout90,
            "ads_used": tech.get("ads_used", pd.DataFrame()),
            "economics_used": tech.get("economics_used", pd.DataFrame()),
            "abc_weekly_used": weekly,
            "abc_monthly_used": monthly,
            "plan_used": self.data.plan,
            "daily_formula": tech.get("daily_formula", pd.DataFrame()),
            "diagnostics": pd.concat([tech.get("diagnostics", pd.DataFrame()), tech.get("econ_diag", pd.DataFrame())], ignore_index=True),
            "paths": self.data.source_paths,
            "warnings": self.data.warnings,
        }
        blocks["example"] = self.build_examples(blocks)
        return blocks

    def build_examples(self, blocks: Dict[str, pd.DataFrame]) -> pd.DataFrame:
        daily = blocks.get("daily_formula", pd.DataFrame())
        weekly = blocks.get("abc_weekly_used", pd.DataFrame())
        if daily.empty and weekly.empty:
            return pd.DataFrame()
        rows = []
        for art in EXAMPLE_ARTICLES:
            d = daily[daily["supplier_article"] == art].copy() if not daily.empty else pd.DataFrame()
            if not d.empty:
                for r in d.itertuples(index=False):
                    rows.append({
                        "Артикул": art,
                        "Период": pd.Timestamp(r.day).strftime("%Y-%m-%d"),
                        "Тип": "День",
                        "Заказы": r.orders_day,
                        "Выкуп": r.buyout_factor,
                        "Выручка": r.revenue_pwd,
                        "Комиссия": r.commission_rub,
                        "Эквайринг": r.acquiring_rub,
                        "Логистика прямая": r.logistics_direct_rub,
                        "Логистика обратная": r.logistics_return_rub,
                        "Хранение": r.storage_rub,
                        "Прочие": r.other_rub,
                        "Себестоимость": r.cost_rub,
                        "Реклама": r.ad_spend_day,
                        "НДС": r.vat_rub,
                        "Валовая Прибыль-НДС": r.gp_minus_nds_raw,
                    })
            w = weekly[weekly["supplier_article"] == art].copy() if not weekly.empty else pd.DataFrame()
            if not w.empty:
                ww = w.groupby("week_label", as_index=False).agg(gp_minus_nds=("gp_minus_nds", "sum"))
                for r in ww.itertuples(index=False):
                    rows.append({"Артикул": art, "Период": r.week_label, "Тип": "ABC неделя", "Валовая Прибыль-НДС": r.gp_minus_nds})
        return pd.DataFrame(rows)


def money_fmt(cell) -> None:
    cell.number_format = '# ##0 "₽"'


def set_header(cell, fill=BLUE_HEADER):
    cell.fill = fill
    cell.font = WHITE_FONT
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_body(cell, bold: bool = False):
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if bold:
        cell.font = BOLD_FONT


def autofit(ws) -> None:
    widths: Dict[int, int] = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is None:
                continue
            widths[c.column] = max(widths.get(c.column, 0), len(str(c.value)) + 2)
    for col, w in widths.items():
        if col == 1:
            ws.column_dimensions[get_column_letter(col)].width = 28
        else:
            ws.column_dimensions[get_column_letter(col)].width = min(max(w, 12), 18)


def write_block(ws, start_row: int, title: str, df: pd.DataFrame) -> int:
    if df.empty:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=9)
        cell = ws.cell(start_row, 1, title)
        cell.fill = BLUE_TITLE
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.cell(start_row + 1, 1, "Нет данных")
        return start_row + 3

    display_cols = [c for c in df.columns if not c.startswith("_")]
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(display_cols))
    cell = ws.cell(start_row, 1, title)
    cell.fill = BLUE_TITLE
    cell.font = WHITE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    header_row = start_row + 1
    for j, col in enumerate(display_cols, start=1):
        set_header(ws.cell(header_row, j, col))

    row = header_row + 1
    category_group_start = None
    product_group_start = None
    for rec in df.to_dict("records"):
        kind = rec.get("_kind", "")
        if kind == "category":
            if product_group_start and row - 1 >= product_group_start[0]:
                ws.row_dimensions.group(product_group_start[0], product_group_start[1], outline_level=2, hidden=True)
                product_group_start = None
            if category_group_start and row - 1 >= category_group_start[0]:
                ws.row_dimensions.group(category_group_start[0], category_group_start[1], outline_level=1, hidden=True)
            category_group_start = [row + 1, row]
        elif kind == "product":
            if product_group_start and row - 1 >= product_group_start[0]:
                ws.row_dimensions.group(product_group_start[0], product_group_start[1], outline_level=2, hidden=True)
            product_group_start = [row + 1, row]
            if category_group_start:
                category_group_start[1] = row
        elif kind == "article":
            if product_group_start:
                product_group_start[1] = row
            if category_group_start:
                category_group_start[1] = row
        else:
            if product_group_start and row - 1 >= product_group_start[0]:
                ws.row_dimensions.group(product_group_start[0], product_group_start[1], outline_level=2, hidden=True)
                product_group_start = None
            if category_group_start and row - 1 >= category_group_start[0]:
                ws.row_dimensions.group(category_group_start[0], category_group_start[1], outline_level=1, hidden=True)
                category_group_start = None

        for j, col in enumerate(display_cols, start=1):
            c = ws.cell(row, j, rec.get(col, ""))
            set_body(c, bold=(col == "План" or kind == "grand_total"))
            if j > 1 and isinstance(rec.get(col), (int, float, np.integer, np.floating)) and not pd.isna(rec.get(col)):
                money_fmt(c)
        # styling
        if kind == "category":
            fill = BLUES.get(rec.get("_subject", ""), LIGHT_BLUE)
            for j in range(1, len(display_cols) + 1):
                ws.cell(row, j).fill = fill
                ws.cell(row, j).font = BOLD_FONT if j in {1, len(display_cols)} else Font(bold=False)
        elif kind == "product":
            fill = LIGHT_BLUE
            for j in range(1, len(display_cols) + 1):
                ws.cell(row, j).fill = fill
                if j == len(display_cols):
                    ws.cell(row, j).font = BOLD_FONT
        elif kind == "grand_total":
            for j in range(1, len(display_cols) + 1):
                ws.cell(row, j).fill = TOTAL_FILL
                ws.cell(row, j).font = BOLD_FONT
        else:
            if len(display_cols) >= 1:
                ws.cell(row, len(display_cols)).font = BOLD_FONT
        row += 1

    if product_group_start and product_group_start[1] >= product_group_start[0]:
        ws.row_dimensions.group(product_group_start[0], product_group_start[1], outline_level=2, hidden=True)
    if category_group_start and category_group_start[1] >= category_group_start[0]:
        ws.row_dimensions.group(category_group_start[0], category_group_start[1], outline_level=1, hidden=True)
    ws.sheet_properties.outlinePr.summaryBelow = False
    return row + 2


def write_df_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title=name[:31])
    if df is None or df.empty:
        ws.cell(1, 1, "Нет данных")
        return
    for j, col in enumerate(df.columns, start=1):
        set_header(ws.cell(1, j, col))
    for i, rec in enumerate(df.to_dict("records"), start=2):
        for j, col in enumerate(df.columns, start=1):
            c = ws.cell(i, j, rec.get(col, ""))
            set_body(c)
            if isinstance(rec.get(col), (int, float, np.integer, np.floating)) and not pd.isna(rec.get(col)):
                money_fmt(c)
    autofit(ws)


def export_main(main_path: str, blocks: Dict[str, pd.DataFrame]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    c = ws.cell(1, 1, "Валовая Прибыль-НДС")
    c.fill = BLUE_TITLE
    c.font = WHITE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    row = 3
    row = write_block(ws, row, "Текущая неделя", blocks.get("daily", pd.DataFrame()))
    row = write_block(ws, row, "Прошлые недели", blocks.get("weekly", pd.DataFrame()))
    row = write_block(ws, row, "Последние 3 месяца", blocks.get("monthly", pd.DataFrame()))
    ws.freeze_panes = "B4"
    autofit(ws)
    wb.save(main_path)


def export_tech(tech_path: str, blocks: Dict[str, pd.DataFrame]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    order = ["dictionary", "orders_used", "funnel_used", "ads_used", "economics_used", "abc_weekly_used", "abc_monthly_used", "plan_used", "daily_formula", "diagnostics", "paths", "warnings"]
    labels = {
        "dictionary": "Словарь",
        "orders_used": "Заказы",
        "funnel_used": "Выкуп90",
        "ads_used": "Реклама",
        "economics_used": "Экономика",
        "abc_weekly_used": "ABC_недели",
        "abc_monthly_used": "ABC_месяцы",
        "plan_used": "План",
        "daily_formula": "Формула_день",
        "diagnostics": "Диагностика",
        "paths": "Пути",
        "warnings": "Предупреждения",
    }
    for key in order:
        write_df_sheet(wb, labels[key], blocks.get(key, pd.DataFrame()))
    wb.save(tech_path)


def export_example(example_path: str, blocks: Dict[str, pd.DataFrame]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Примеры"
    df = blocks.get("example", pd.DataFrame())
    if df.empty:
        ws.cell(1, 1, "Нет данных")
    else:
        for j, col in enumerate(df.columns, start=1):
            set_header(ws.cell(1, j, col))
        for i, rec in enumerate(df.to_dict("records"), start=2):
            for j, col in enumerate(df.columns, start=1):
                c = ws.cell(i, j, rec.get(col, ""))
                set_body(c)
                if isinstance(rec.get(col), (int, float, np.integer, np.floating)) and not pd.isna(rec.get(col)):
                    money_fmt(c)
        autofit(ws)
    wb.save(example_path)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--root", default=".")
    p.add_argument("--reports-root", default="Отчёты")
    p.add_argument("--store", default="TOPFACE")
    p.add_argument("--out-subdir", default="Отчёты/Объединенный отчет/TOPFACE")
    return p.parse_args()


def main() -> int:
    args = parse_args()
    storage = make_storage(args.root)
    data = Loader(storage, args.reports_root, args.store).load_all()
    builder = Builder(data)
    blocks = builder.build_blocks()
    stamp = datetime.now().strftime("%Y-%m-%d")
    local_main = Path("/tmp") / f"wb_main_{stamp}.xlsx"
    local_tech = Path("/tmp") / f"wb_tech_{stamp}.xlsx"
    local_example = Path("/tmp") / f"wb_example_{stamp}.xlsx"
    export_main(str(local_main), blocks)
    export_tech(str(local_tech), blocks)
    export_example(str(local_example), blocks)
    out_main = f"{args.out_subdir}/Объединенный_отчет_{args.store}_{stamp}.xlsx"
    out_tech = f"{args.out_subdir}/Технические_расчеты_{args.store}_{stamp}.xlsx"
    out_example = f"{args.out_subdir}/Пример_расчета_901_{args.store}_{stamp}.xlsx"
    storage.write_bytes(out_main, local_main.read_bytes())
    storage.write_bytes(out_tech, local_tech.read_bytes())
    storage.write_bytes(out_example, local_example.read_bytes())
    log(f"Saved report: {out_main}")
    log(f"Saved technical workbook: {out_tech}")
    log(f"Saved example workbook: {out_example}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
