#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import io
import math
import os
import re
import tempfile
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Set, Tuple

import boto3
import numpy as np
import pandas as pd
import requests
from botocore.client import Config as BotoConfig
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def env_bool(name: str, default: bool = False) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


@dataclass
class AppConfig:
    bucket_name: str = os.getenv("WB_S3_BUCKET", "")
    access_key: str = os.getenv("WB_S3_ACCESS_KEY", "")
    secret_key: str = os.getenv("WB_S3_SECRET_KEY", "")
    endpoint_url: str = os.getenv("WB_S3_ENDPOINT", "https://storage.yandexcloud.net")
    region_name: str = os.getenv("WB_S3_REGION", "ru-central1")

    store_name: str = os.getenv("WB_STORE", "TOPFACE").strip()
    run_date: datetime = datetime.strptime(
        os.getenv("WB_RUN_DATE", datetime.now().strftime("%Y-%m-%d")),
        "%Y-%m-%d",
    )
    season_coeff: float = float(os.getenv("WB_SEASON_COEFF", "1.0"))

    lookback_days: int = 90
    recent_days: int = 14
    target_days: int = 74

    price_drop_threshold: float = 0.10
    sales_drop_threshold: float = 0.30
    growth_threshold: float = float(os.getenv("WB_GROWTH_THRESHOLD", "1.5"))

    coeff_local: float = 1.0
    coeff_okrug: float = 1.1
    coeff_far: float = 1.3

    low_turnover_sales_threshold: int = 100
    low_turnover_network_stock: int = 50

    output_dir: str = os.getenv("WB_OUTPUT_DIR", "output")
    upload_result_to_s3: bool = env_bool("WB_UPLOAD_RESULT_TO_S3", False)

    telegram_bot_token: str = os.getenv("TELEGRAM_BOT_TOKEN", "")
    telegram_chat_id: str = os.getenv("TELEGRAM_CHAT_ID", "")
    send_telegram: bool = env_bool("WB_SEND_TELEGRAM", True)

    orders_prefix_tpl: str = "Отчёты/Заказы/{store}/Недельные/"
    stocks_prefix_tpl: str = "Отчёты/Остатки/{store}/Недельные/"
    stocks_1c_key: str = "Отчёты/Остатки/1С/Остатки 1С.xlsx"
    article_map_1c_key: str = "Отчёты/Остатки/1С/Артикулы 1с.xlsx"
    template_key: str = os.getenv("WB_TEMPLATE_KEY", "Служебные файлы/Согласование поставки WB.xlsm")

    excluded_subjects: Tuple[str, ...] = ("Лаки для ногтей",)
    economy_subjects: Tuple[str, ...] = (
        "Помады",
        "Блески",
        "Кисти косметические",
        "Косметические карандаши",
    )
    strategy_mode: str = os.getenv("WB_STRATEGY_MODE", "default").strip().lower()

    district_primary_warehouse: Dict[str, str] = field(default_factory=dict)

    def __post_init__(self) -> None:
        if not self.district_primary_warehouse:
            self.district_primary_warehouse = {
                "Центральный федеральный округ": "Коледино",
                "Северо-Западный федеральный округ": "Санкт-Петербург Уткина Заводь",
                "Южный федеральный округ": "Краснодар",
                "Северо-Кавказский федеральный округ": "Невинномысск",
                "Приволжский федеральный округ": "Казань",
                "Уральский федеральный округ": "Екатеринбург - Перспективная 14",
                "Сибирский федеральный округ": "Новосибирск",
                "Дальневосточный федеральный округ": "Екатеринбург - Перспективная 14",
            }


CONFIG = AppConfig()

WAREHOUSE_ALIASES: Dict[str, str] = {
    "Коледино/Электросталь (Москва)": "MOSCOW_CLUSTER",
    "Новосемейкино": "Самара (Новосемейкино)",
    "Самара Новосемейкино": "Самара (Новосемейкино)",
    "Екатеринбург - Перспективный 12": "Екатеринбург - Перспективная 14",
    "Екатеринбург Перспективная 14": "Екатеринбург - Перспективная 14",
    "Санкт Петербург Уткина Заводь": "Санкт-Петербург Уткина Заводь",
    "Москва": "Коледино",
}

TEMPLATE_WAREHOUSE_ALIASES: Dict[str, str] = {
    "Коледино": "Коледино",
    "Тула": "Тула",
    "Электросталь": "Электросталь",
    "Казань": "Казань",
    "Новосемейкино": "Самара (Новосемейкино)",
    "Самара (Новосемейкино)": "Самара (Новосемейкино)",
    "Краснодар": "Краснодар",
    "Невинномысск": "Невинномысск",
    "Волгоград": "Волгоград",
    "Рязань": "Рязань (Тюшевское)",
    "Рязань (Тюшевское)": "Рязань (Тюшевское)",
    "Сарапул": "Сарапул",
    "Екатеринбург": "Екатеринбург - Перспективная 14",
    "Екатеринбург - Перспективная 14": "Екатеринбург - Перспективная 14",
    "Екатеринбург - Перспективный 12": "Екатеринбург - Перспективная 14",
    "Владимир": "Владимир",
    "Котовск": "Котовск",
    "Воронеж": "Воронеж",
    "Москва": "Коледино",
    "СПб Уткина Заводь": "Санкт-Петербург Уткина Заводь",
    "СПБ Уткина Заводь": "Санкт-Петербург Уткина Заводь",
    "Санкт-Петербург Уткина Заводь": "Санкт-Петербург Уткина Заводь",
    "Новосибирск": "Новосибирск",
}

ECONOMY_REPLACEMENT_MAP: Dict[str, str] = {
    "Коледино": "Рязань (Тюшевское)",
    "Электросталь": "Владимир",
    "Краснодар": "Невинномысск",
    "Казань": "Самара (Новосемейкино)",
    "Тула": "Котовск",
}

WAREHOUSE_TO_DISTRICT: Dict[str, str] = {
    "Коледино": "Центральный федеральный округ",
    "Электросталь": "Центральный федеральный округ",
    "Белые Столбы": "Центральный федеральный округ",
    "Подольск": "Центральный федеральный округ",
    "Тула": "Центральный федеральный округ",
    "Владимир": "Центральный федеральный округ",
    "Котовск": "Центральный федеральный округ",
    "Воронеж": "Центральный федеральный округ",
    "Рязань (Тюшевское)": "Центральный федеральный округ",
    "Санкт-Петербург Уткина Заводь": "Северо-Западный федеральный округ",
    "Краснодар": "Южный федеральный округ",
    "Волгоград": "Южный федеральный округ",
    "Невинномысск": "Северо-Кавказский федеральный округ",
    "Казань": "Приволжский федеральный округ",
    "Самара (Новосемейкино)": "Приволжский федеральный округ",
    "Сарапул": "Приволжский федеральный округ",
    "Екатеринбург - Перспективная 14": "Уральский федеральный округ",
    "Новосибирск": "Сибирский федеральный округ",
}

ALL_TARGET_WAREHOUSES: List[str] = list(WAREHOUSE_TO_DISTRICT.keys())
MOSCOW_CLUSTER: List[str] = ["Коледино", "Электросталь", "Белые Столбы", "Подольск"]
MOSCOW_SPLIT_TARGET: Dict[str, float] = {"Коледино": 0.5, "Электросталь": 0.5}

ONE_C_STOCK_COLUMNS: List[str] = [
    'Оптовый склад Луганск- ООО "Хайлер"',
    'Основной склад - ООО "Хайлер"',
    'Адресный склад',
    'Основной склад - ИП Куканянц И.Ю.',
]


def build_region_to_group() -> Dict[str, str]:
    mapping: Dict[str, str] = {}

    def add(group: str, regions: Sequence[str]) -> None:
        for region in regions:
            mapping[region] = group

    add("MOSCOW_CLUSTER", ["Москва", "Московская область"])
    add("Краснодар", [
        "Краснодарский край", "Ростовская область", "Республика Крым", "Севастополь",
        "Республика Адыгея", "Ереван", "Котайкская область", "Лорийская область",
        "Гехаркуникская область", "Ширакская область", "федеральная территория Сириус",
        "Тавушская область", "Республика Каракалпакстан", "Вайоцдзорская область",
    ])
    add("Санкт-Петербург Уткина Заводь", ["Санкт-Петербург", "Ленинградская область", "Новгородская область", "Республика Карелия"])
    add("Невинномысск", [
        "Ставропольский край", "Республика Дагестан", "Чеченская Республика",
        "Республика Северная Осетия — Алания", "Кабардино-Балкарская Республика",
        "Карачаево-Черкесская Республика", "Республика Ингушетия", "Республика Калмыкия",
        "Армавирская область", "Араратская область", "Сюникская область", "Арагацотнская область",
    ])
    add("Казань", ["Республика Татарстан", "Ульяновская область", "Кировская область", "Чувашская Республика", "Республика Коми", "Республика Марий Эл"])
    add("Владимир", ["Нижегородская область", "Владимирская область", "Ярославская область", "Ивановская область", "Костромская область", "Бухарская область"])
    add("Екатеринбург - Перспективная 14", [
        "Свердловская область", "Иркутская область", "Красноярский край", "Челябинская область",
        "Новосибирская область", "Кемеровская область", "Ханты-Мансийский автономный округ",
        "Тюменская область", "Алтайский край", "Омская область", "Томская область",
        "Республика Саха (Якутия)", "Республика Бурятия", "Забайкальский край", "Амурская область",
        "Ямало-Ненецкий автономный округ", "Курганская область", "Республика Алтай", "Алматы",
        "Карагандинская область", "Костанайская область", "Восточно-Казахстанская область",
        "город республиканского значения Астана", "Астана", "Павлодарская область",
        "город республиканского подчинения Бишкек", "Акмолинская область", "Северо-Казахстанская область",
        "город Бишкек", "Алматинская область", "область Абай", "область Жетысу",
        "Джалал-Абадская область", "область Улытау", "Абайская область",
        "город республиканского подчинения Ош", "Ошская область", "Иссык-Кульская область",
        "Улутауская область", "Нарынская область", "город республиканского значения Нур-Султан",
    ])
    add("Самара (Новосемейкино)", [
        "Самарская область", "Оренбургская область", "Западно-Казахстанская область",
        "Актюбинская область", "Чуйская область", "Жамбылская область", "Шымкент",
        "Туркестанская область", "город республиканского значения Байконур",
        "Баткенская область", "Кызылординская область", "город Ош",
    ])
    add("Сарапул", ["Республика Башкортостан", "Пермский край", "Удмуртская Республика", "Республика Хакасия"])
    add("Воронеж", ["Воронежская область", "Хатлонская область"])
    add("Тула", ["Тульская область", "Белгородская область", "Курская область", "Брянская область", "Орловская область", "Гомельская область", "Могилёвская область", "Витебская область"])
    add("Волгоград", ["Саратовская область", "Волгоградская область", "Астраханская область", "Атырауская область", "Мангистауская область"])
    add("Котовск", ["Липецкая область", "Пензенская область", "Тамбовская область", "Республика Мордовия", "Сурхандарьинская область"])
    add("Рязань (Тюшевское)", ["Рязанская область", "Навоийская область"])
    add("Новосибирск", ["Республика Тыва", "Ташкентская область"])
    add("MOSCOW_CLUSTER", [
        "Приморский край", "Калужская область", "Вологодская область", "Архангельская область",
        "Минск", "Тверская область", "Мурманская область", "Смоленская область",
        "Калининградская область", "Хабаровский край", "Сахалинская область", "Псковская область",
        "Минская область", "Гродненская область", "Брестская область", "Камчатский край",
        "Магаданская область", "Ташкент", "Еврейская автономная область", "Тбилиси",
        "Ненецкий автономный округ", "Душанбе", "Квемо Картли", "Чукотский автономный округ",
        "Аджарская Автономная Республика", "Самаркандская область", "Хорезмская область",
        "муниципалитет Тбилиси", "Районы республиканского подчинения", "Самцхе-Джавахети",
        "Ферганская область", "Согдийская область", "Горно-Бадахшанская автономная область", "Имеретия",
    ])
    return mapping


REGION_TO_GROUP: Dict[str, str] = build_region_to_group()


def log(msg: str) -> None:
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}", flush=True)


def normalize_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_nmid(value: object) -> str:
    s = normalize_text(value)
    return s[:-2] if s.endswith(".0") else s


def normalize_warehouse(name: object) -> str:
    s = normalize_text(name)
    return WAREHOUSE_ALIASES.get(s, s)


def normalize_template_header(name: object) -> str:
    return normalize_text(name)


def floor_int(value: float) -> int:
    if pd.isna(value):
        return 0
    return int(math.floor(float(value)))


def ceil_int(value: float) -> int:
    if pd.isna(value):
        return 0
    return int(math.ceil(float(value)))


def parse_week_key_date(key: str) -> Optional[datetime]:
    match = re.search(r"_(\d{4}-W\d{2})\.xlsx$", key)
    if not match:
        return None
    year, week = match.group(1).split("-W")
    return datetime.fromisocalendar(int(year), int(week), 1)


def largest_remainder_allocation(total: int, weights: Dict[str, float], minimum_one_for_nonzero: bool = False) -> Dict[str, int]:
    keys = list(weights.keys())
    if total <= 0 or not keys:
        return {key: 0 for key in keys}

    clean = {k: max(float(v), 0.0) for k, v in weights.items()}
    if sum(clean.values()) == 0:
        equal = total // len(keys)
        rem = total - equal * len(keys)
        out = {k: equal for k in keys}
        for k in keys[:rem]:
            out[k] += 1
        return out

    base = {}
    reserved = 0
    if minimum_one_for_nonzero:
        for k, v in clean.items():
            if v > 0 and reserved < total:
                base[k] = 1
                reserved += 1
            else:
                base[k] = 0
    else:
        base = {k: 0 for k in keys}

    remaining = max(total - reserved, 0)
    total_weight = sum(clean.values())
    raw = {k: remaining * clean[k] / total_weight for k in keys}
    floored = {k: int(math.floor(raw[k])) for k in keys}
    out = {k: base[k] + floored[k] for k in keys}
    rest = total - sum(out.values())

    remainders = sorted(((raw[k] - floored[k], k) for k in keys), reverse=True)
    for _, k in remainders[:rest]:
        out[k] += 1

    return out


class S3Storage:
    def __init__(self, cfg: AppConfig):
        if not cfg.bucket_name or not cfg.access_key or not cfg.secret_key:
            raise ValueError("Не заданы параметры Object Storage. Нужны env: WB_S3_BUCKET, WB_S3_ACCESS_KEY, WB_S3_SECRET_KEY.")

        self.bucket = cfg.bucket_name
        self.s3 = boto3.client(
            "s3",
            endpoint_url=cfg.endpoint_url,
            aws_access_key_id=cfg.access_key,
            aws_secret_access_key=cfg.secret_key,
            region_name=cfg.region_name,
            config=BotoConfig(signature_version="s3v4", read_timeout=300, connect_timeout=60),
        )

    def list_keys(self, prefix: str) -> List[str]:
        keys: List[str] = []
        continuation_token = None

        while True:
            kwargs = {"Bucket": self.bucket, "Prefix": prefix, "MaxKeys": 1000}
            if continuation_token:
                kwargs["ContinuationToken"] = continuation_token

            resp = self.s3.list_objects_v2(**kwargs)
            keys.extend([obj["Key"] for obj in resp.get("Contents", [])])

            if not resp.get("IsTruncated"):
                break
            continuation_token = resp.get("NextContinuationToken")

        return keys

    def read_bytes(self, key: str) -> bytes:
        obj = self.s3.get_object(Bucket=self.bucket, Key=key)
        return obj["Body"].read()

    def read_excel(self, key: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
        return pd.read_excel(io.BytesIO(self.read_bytes(key)), sheet_name=sheet_name)

    def download_file(self, key: str, local_path: str) -> str:
        Path(local_path).parent.mkdir(parents=True, exist_ok=True)
        Path(local_path).write_bytes(self.read_bytes(key))
        return local_path

    def upload_file(self, local_path: str, key: str) -> None:
        self.s3.upload_file(local_path, self.bucket, key)


def load_weekly_window(storage: S3Storage, prefix: str, run_date: datetime, lookback_days: int, expected_sheet: Optional[str] = None) -> pd.DataFrame:
    keys = storage.list_keys(prefix)
    if not keys:
        raise FileNotFoundError(f"В Object Storage не найдено файлов по префиксу: {prefix}")

    cutoff = run_date - timedelta(days=lookback_days + 21)
    selected = [(parse_week_key_date(k) or datetime.min, k) for k in keys if (parse_week_key_date(k) or datetime.min) >= cutoff]
    if not selected:
        selected = [(parse_week_key_date(k) or datetime.min, k) for k in keys]

    parts: List[pd.DataFrame] = []
    for _, key in sorted(selected, key=lambda x: x[0]):
        try:
            df = storage.read_excel(key, sheet_name=expected_sheet)
            if isinstance(df, dict):
                df = next(iter(df.values()))
            if not df.empty:
                parts.append(df)
                log(f"Загружен файл: {key}")
        except Exception as exc:
            log(f"⚠️ Не удалось прочитать {key}: {exc}")

    if not parts:
        raise ValueError(f"Не удалось прочитать ни одного файла по префиксу {prefix}")

    return pd.concat(parts, ignore_index=True)


def load_orders(storage: S3Storage, cfg: AppConfig) -> pd.DataFrame:
    df = load_weekly_window(storage, cfg.orders_prefix_tpl.format(store=cfg.store_name), cfg.run_date, cfg.lookback_days, expected_sheet="Заказы")
    required = ["date", "warehouseName", "oblastOkrugName", "regionName", "supplierArticle", "nmId", "subject", "finishedPrice"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise KeyError(f"В заказах отсутствуют обязательные колонки: {missing}")

    df = df.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.normalize()
    df = df[(df["date"] >= (cfg.run_date - timedelta(days=cfg.lookback_days - 1))) & (df["date"] <= cfg.run_date)].copy()

    if "isCancel" in df.columns:
        df = df[~df["isCancel"].fillna(False)].copy()

    df["warehouseName"] = df["warehouseName"].map(normalize_warehouse)
    df["subject"] = df["subject"].map(normalize_text)
    df["regionName"] = df["regionName"].map(normalize_text)
    df["oblastOkrugName"] = df["oblastOkrugName"].map(normalize_text)
    df["supplierArticle"] = df["supplierArticle"].map(normalize_text)
    df["nmId"] = df["nmId"].map(normalize_nmid)
    df["finishedPrice"] = pd.to_numeric(df["finishedPrice"], errors="coerce")
    df["qty"] = 1

    before = len(df)
    df = df[~df["subject"].isin(cfg.excluded_subjects)].copy()
    log(f"Исключено строк заказов по запрещённым категориям: {before - len(df)}")
    return df


def load_stocks(storage: S3Storage, cfg: AppConfig) -> pd.DataFrame:
    df = load_weekly_window(storage, cfg.stocks_prefix_tpl.format(store=cfg.store_name), cfg.run_date, cfg.lookback_days, expected_sheet="Остатки")
    required = ["Склад", "Артикул WB", "Полное количество"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise KeyError(f"В остатках отсутствуют обязательные колонки: {missing}")

    df = df.copy()
    date_col = "Дата запроса" if "Дата запроса" in df.columns else "Дата сбора"
    df["stock_date"] = pd.to_datetime(df[date_col], errors="coerce").dt.normalize()
    df = df[(df["stock_date"] >= (cfg.run_date - timedelta(days=cfg.lookback_days - 1))) & (df["stock_date"] <= cfg.run_date)].copy()

    df["warehouse"] = df["Склад"].map(normalize_warehouse)
    df["nmId"] = df["Артикул WB"].map(normalize_nmid)
    df["qty_full"] = pd.to_numeric(df["Полное количество"], errors="coerce").fillna(0)
    df["qty_available"] = pd.to_numeric(df["Доступно для продажи"], errors="coerce").fillna(0) if "Доступно для продажи" in df.columns else df["qty_full"]
    df["subject"] = df["Предмет"].map(normalize_text) if "Предмет" in df.columns else ""

    before = len(df)
    df = df[~df["subject"].isin(cfg.excluded_subjects)].copy()
    log(f"Исключено строк остатков по запрещённым категориям: {before - len(df)}")
    return df


def load_article_map_1c(storage: S3Storage, cfg: AppConfig) -> Dict[str, str]:
    df = storage.read_excel(cfg.article_map_1c_key)
    if isinstance(df, dict):
        df = next(iter(df.values()))

    wb_col = df.columns[0]
    col_1c = df.columns[2] if len(df.columns) >= 3 else df.columns[-1]

    out: Dict[str, str] = {}
    for _, row in df.iterrows():
        wb_article = normalize_nmid(row.get(wb_col))
        article_1c = normalize_text(row.get(col_1c))
        if wb_article:
            out[wb_article] = article_1c

    log(f"Загружено соответствий WB -> 1С: {len(out)}")
    return out


def load_1c_stocks(storage: S3Storage, cfg: AppConfig) -> pd.DataFrame:
    df = storage.read_excel(cfg.stocks_1c_key)
    if isinstance(df, dict):
        df = next(iter(df.values()))
    df = df.copy()
    df.columns = [normalize_text(c) for c in df.columns]
    return df


def prepare_daily_orders(orders: pd.DataFrame) -> pd.DataFrame:
    df = orders.copy()
    df["region_group"] = df["regionName"].map(REGION_TO_GROUP)

    unmapped = int(df["region_group"].isna().sum())
    if unmapped:
        log(f"⚠️ Регионов без привязки к группе обслуживания: {unmapped}. Они будут отброшены.")
        df = df[df["region_group"].notna()].copy()

    return (
        df.groupby(["nmId", "supplierArticle", "subject", "date", "region_group", "oblastOkrugName"], as_index=False)
        .agg(qty=("qty", "sum"), avg_price=("finishedPrice", "mean"))
    )


def build_daily_grid(daily_orders: pd.DataFrame, cfg: AppConfig) -> pd.DataFrame:
    all_dates = pd.date_range(cfg.run_date - timedelta(days=cfg.lookback_days - 1), cfg.run_date, freq="D")
    keys = daily_orders[["nmId", "supplierArticle", "subject", "region_group", "oblastOkrugName"]].drop_duplicates()
    keys["__tmp"] = 1
    dates_df = pd.DataFrame({"date": all_dates, "__tmp": 1})

    grid = keys.merge(dates_df, on="__tmp", how="outer").drop(columns="__tmp")
    grid = grid.merge(
        daily_orders,
        on=["nmId", "supplierArticle", "subject", "region_group", "oblastOkrugName", "date"],
        how="left",
    )
    grid["qty"] = grid["qty"].fillna(0)
    return grid


def prepare_stock_presence(stocks: pd.DataFrame):
    per_wh = (
        stocks.groupby(["stock_date", "nmId", "warehouse"], as_index=False)
        .agg(qty_full=("qty_full", "sum"), qty_available=("qty_available", "sum"))
    )

    history_dates = set(pd.to_datetime(per_wh["stock_date"]).dt.normalize().unique()) if not per_wh.empty else set()

    per_wh["district"] = per_wh["warehouse"].map(WAREHOUSE_TO_DISTRICT)
    per_district = (
        per_wh.groupby(["stock_date", "nmId", "district"], as_index=False)
        .agg(qty_full=("qty_full", "sum"), qty_available=("qty_available", "sum"))
    )

    latest_stock_date = per_wh["stock_date"].max() if not per_wh.empty else pd.NaT
    current_wh = per_wh[per_wh["stock_date"] == latest_stock_date].copy() if pd.notna(latest_stock_date) else per_wh.iloc[0:0].copy()
    current_wh = current_wh.rename(columns={"stock_date": "latest_stock_date"})

    return current_wh, history_dates, per_wh, per_district, latest_stock_date


def attach_presence_flags(grid: pd.DataFrame, per_wh: pd.DataFrame, per_district: pd.DataFrame, history_dates: Set[pd.Timestamp]) -> pd.DataFrame:
    grid = grid.copy()
    wh_presence = per_wh.copy()
    wh_presence["is_positive"] = wh_presence["qty_full"] > 0

    local_records: List[pd.DataFrame] = []
    for group in grid["region_group"].dropna().unique():
        needed = MOSCOW_CLUSTER if group == "MOSCOW_CLUSTER" else [group]
        tmp = (
            wh_presence[wh_presence["warehouse"].isin(needed)]
            .groupby(["stock_date", "nmId"], as_index=False)["is_positive"]
            .max()
        )
        tmp["region_group"] = group
        local_records.append(tmp)

    local_presence = pd.concat(local_records, ignore_index=True) if local_records else pd.DataFrame(columns=["stock_date", "nmId", "is_positive", "region_group"])
    local_presence = local_presence.rename(columns={"stock_date": "date", "is_positive": "local_positive"})
    grid = grid.merge(local_presence, on=["date", "nmId", "region_group"], how="left")

    district_presence = per_district.copy()
    district_presence["district_positive"] = district_presence["qty_full"] > 0
    district_presence = district_presence.rename(columns={"stock_date": "date"})

    def region_target_district(region_group: str) -> str:
        if region_group == "MOSCOW_CLUSTER":
            return "Центральный федеральный округ"
        return WAREHOUSE_TO_DISTRICT.get(region_group, "")

    grid["target_district"] = grid["region_group"].map(region_target_district)

    grid = grid.merge(
        district_presence[["date", "nmId", "district", "district_positive"]],
        left_on=["date", "nmId", "target_district"],
        right_on=["date", "nmId", "district"],
        how="left",
    ).drop(columns=["district"])

    known_dates = set(pd.to_datetime(pd.Series(list(history_dates))).dt.normalize()) if history_dates else set()
    grid["history_date_exists"] = grid["date"].isin(known_dates)

    grid["local_positive"] = grid["local_positive"].astype("boolean")
    grid["district_positive"] = grid["district_positive"].astype("boolean")

    grid["local_positive"] = np.where(grid["history_date_exists"], grid["local_positive"].fillna(False), True)
    grid["district_positive"] = np.where(grid["history_date_exists"], grid["district_positive"].fillna(False), True)

    grid["status"] = np.select(
        [
            grid["local_positive"] == True,
            (grid["local_positive"] == False) & (grid["district_positive"] == True),
            (grid["local_positive"] == False) & (grid["district_positive"] == False),
        ],
        ["local", "okrug", "far"],
        default="local",
    )

    return grid


def mark_valid_days(group_df: pd.DataFrame, cfg: AppConfig) -> pd.DataFrame:
    df = group_df.copy()
    base_price = df.loc[df["qty"] > 0, "avg_price"].mean()

    df["promo_invalid"] = False
    if pd.notna(base_price):
        df.loc[df["qty"] > 0, "promo_invalid"] = (
            df.loc[df["qty"] > 0, "avg_price"] < (1 - cfg.price_drop_threshold) * base_price
        )

    prelim_mean = df.loc[~df["promo_invalid"], "qty"].mean()
    prelim_mean = 0.0 if pd.isna(prelim_mean) else float(prelim_mean)
    sales_floor = prelim_mean * (1 - cfg.sales_drop_threshold)

    df["drop_invalid"] = (df["qty"] > 0) & (~df["promo_invalid"]) & (df["qty"] < sales_floor)
    df["valid_day"] = ~(df["promo_invalid"] | df["drop_invalid"])
    return df


def aggregate_region_metrics(grid: pd.DataFrame, cfg: AppConfig) -> pd.DataFrame:
    rows: List[Dict[str, object]] = []
    group_cols = ["nmId", "supplierArticle", "subject", "region_group", "oblastOkrugName"]

    for keys, part in grid.groupby(group_cols, dropna=False):
        part = mark_valid_days(part.sort_values("date"), cfg)

        base90 = part.loc[part["valid_day"], "qty"].mean()
        if pd.isna(base90):
            base90 = part["qty"].mean()
        base90 = 0.0 if pd.isna(base90) else float(base90)

        last14 = part[part["date"] >= (cfg.run_date - timedelta(days=cfg.recent_days - 1))]
        base14 = last14.loc[last14["valid_day"], "qty"].mean()
        if pd.isna(base14):
            base14 = last14["qty"].mean()
        base14 = 0.0 if pd.isna(base14) else float(base14)

        d_local90 = int((part["status"] == "local").sum())
        d_okrug90 = int((part["status"] == "okrug").sum())
        d_far90 = int((part["status"] == "far").sum())

        d_local14 = int((last14["status"] == "local").sum())
        d_okrug14 = int((last14["status"] == "okrug").sum())
        d_far14 = int((last14["status"] == "far").sum())

        avg_adj90 = base90 * (cfg.coeff_local * d_local90 + cfg.coeff_okrug * d_okrug90 + cfg.coeff_far * d_far90) / max(cfg.lookback_days, 1)
        avg_adj14 = base14 * (cfg.coeff_local * d_local14 + cfg.coeff_okrug * d_okrug14 + cfg.coeff_far * d_far14) / max(cfg.recent_days, 1)

        price90 = part.loc[part["qty"] > 0, "avg_price"].mean()
        price14 = last14.loc[last14["qty"] > 0, "avg_price"].mean()

        rows.append(
            {
                "nmId": keys[0],
                "supplierArticle": keys[1],
                "subject": keys[2],
                "region_group": keys[3],
                "oblastOkrugName": keys[4],
                "sales_90": int(part["qty"].sum()),
                "sales_14": int(last14["qty"].sum()),
                "avg_adj90_region": float(avg_adj90),
                "avg_adj14_region": float(avg_adj14),
                "avg_price90_region": float(price90) if pd.notna(price90) else np.nan,
                "avg_price14_region": float(price14) if pd.notna(price14) else np.nan,
            }
        )

    return pd.DataFrame(rows)


def choose_final_daily_demand(region_metrics: pd.DataFrame, cfg: AppConfig) -> pd.DataFrame:
    rows: List[Dict[str, object]] = []

    for (nmid, supplier_article, subject), part in region_metrics.groupby(["nmId", "supplierArticle", "subject"], dropna=False):
        avg90 = float(part["avg_adj90_region"].sum())
        avg14 = float(part["avg_adj14_region"].sum())
        sales90 = int(part["sales_90"].sum())
        sales14 = int(part["sales_14"].sum())

        w90 = np.maximum(part["sales_90"].to_numpy(dtype=float), 1)
        w14 = np.maximum(part["sales_14"].to_numpy(dtype=float), 1)
        p90 = part["avg_price90_region"].fillna(0).to_numpy(dtype=float)
        p14 = part["avg_price14_region"].fillna(0).to_numpy(dtype=float)

        price90 = float(np.average(p90, weights=w90)) if len(part) else np.nan
        price14 = float(np.average(p14, weights=w14)) if len(part) else np.nan

        can_use_14 = avg14 > avg90 * cfg.growth_threshold and (
            pd.isna(price90) or pd.isna(price14) or price14 >= (1 - cfg.price_drop_threshold) * price90
        )

        base_daily = avg14 if can_use_14 else avg90

        rows.append(
            {
                "nmId": nmid,
                "supplierArticle": supplier_article,
                "subject": subject,
                "sales_90_total": sales90,
                "sales_14_total": sales14,
                "daily_demand_final": float(base_daily * cfg.season_coeff),
            }
        )

    return pd.DataFrame(rows)


def build_warehouse_shares(region_metrics: pd.DataFrame) -> pd.DataFrame:
    rows: List[Dict[str, object]] = []

    for (nmid, supplier_article, subject), part in region_metrics.groupby(["nmId", "supplierArticle", "subject"], dropna=False):
        total_sales = float(part["sales_90"].sum())
        warehouse_sales: Dict[str, float] = defaultdict(float)

        for _, row in part.iterrows():
            group = row["region_group"]
            sales = float(row["sales_90"])

            if group == "MOSCOW_CLUSTER":
                for warehouse, share in MOSCOW_SPLIT_TARGET.items():
                    warehouse_sales[warehouse] += sales * share
            else:
                warehouse_sales[group] += sales

        if total_sales <= 0 and warehouse_sales:
            for warehouse in warehouse_sales:
                warehouse_sales[warehouse] = 1.0
            total_sales = sum(warehouse_sales.values())

        for warehouse, sales in warehouse_sales.items():
            rows.append(
                {
                    "nmId": nmid,
                    "supplierArticle": supplier_article,
                    "subject": subject,
                    "warehouse": warehouse,
                    "warehouse_share": (sales / total_sales) if total_sales > 0 else 0.0,
                }
            )

    shares = pd.DataFrame(rows)
    if not shares.empty:
        sums = shares.groupby("nmId")["warehouse_share"].transform("sum").replace(0, 1)
        shares["warehouse_share"] = shares["warehouse_share"] / sums

    return shares


def apply_strategy(shares_df: pd.DataFrame, cfg: AppConfig) -> pd.DataFrame:
    if cfg.strategy_mode != "economy":
        return shares_df.copy()

    df = shares_df.copy()
    mask = df["subject"].isin(cfg.economy_subjects) & df["warehouse"].isin(ECONOMY_REPLACEMENT_MAP)
    df.loc[mask, "warehouse"] = df.loc[mask, "warehouse"].map(ECONOMY_REPLACEMENT_MAP)

    df = (
        df.groupby(["nmId", "supplierArticle", "subject", "warehouse"], as_index=False)
        .agg(warehouse_share=("warehouse_share", "sum"))
    )

    sums = df.groupby("nmId")["warehouse_share"].transform("sum").replace(0, 1)
    df["warehouse_share"] = df["warehouse_share"] / sums
    return df


def current_stock_by_warehouse(current_wh: pd.DataFrame) -> pd.DataFrame:
    return current_wh.groupby(["nmId", "warehouse"], as_index=False).agg(current_stock_full=("qty_full", "sum"))


def allocate_low_turnover(shares_sku: pd.DataFrame, cfg: AppConfig) -> Dict[str, int]:
    weights = {row["warehouse"]: float(row["warehouse_share"]) for _, row in shares_sku.iterrows() if row["warehouse"]}

    selected = dict(weights)
    for _, warehouse in cfg.district_primary_warehouse.items():
        selected.setdefault(warehouse, 0.0001)

    selected = {warehouse: weight for warehouse, weight in selected.items() if warehouse in ALL_TARGET_WAREHOUSES}
    if not selected:
        selected = {warehouse: 1.0 for warehouse in cfg.district_primary_warehouse.values()}

    return largest_remainder_allocation(cfg.low_turnover_network_stock, selected, minimum_one_for_nonzero=True)


def calculate_supply_plan(
    sku_df: pd.DataFrame,
    shares_df: pd.DataFrame,
    current_stock_df: pd.DataFrame,
    article_1c_map: Dict[str, str],
    cfg: AppConfig,
) -> pd.DataFrame:
    current_lookup = (
        current_stock_df.set_index(["nmId", "warehouse"]).to_dict("index")
        if not current_stock_df.empty
        else {}
    )

    rows: List[Dict[str, object]] = []

    for _, sku in sku_df.iterrows():
        nmid = sku["nmId"]
        sku_shares = shares_df[shares_df["nmId"] == nmid].copy()
        if sku_shares.empty:
            continue

        article_1c = article_1c_map.get(nmid, "")

        if int(sku["sales_90_total"]) < cfg.low_turnover_sales_threshold:
            allocations = allocate_low_turnover(sku_shares, cfg)
            for warehouse, target_qty in allocations.items():
                current_stock = float(current_lookup.get((nmid, warehouse), {}).get("current_stock_full", 0))
                to_supply = max(0, int(target_qty - current_stock))

                rows.append(
                    {
                        "nmId": nmid,
                        "supplierArticle": sku["supplierArticle"],
                        "subject": sku["subject"],
                        "Артикул 1С": article_1c,
                        "warehouse": warehouse,
                        "warehouse_share": float(
                            sku_shares.loc[sku_shares["warehouse"] == warehouse, "warehouse_share"].sum()
                        ) if warehouse in sku_shares["warehouse"].values else 0.0,
                        "daily_demand_final": float(sku["daily_demand_final"]),
                        "target_stock": float(target_qty),
                        "current_stock_full": float(current_stock),
                        "to_supply": float(to_supply),
                        "calc_mode": "low_turnover",
                    }
                )
            continue

        for _, share_row in sku_shares.iterrows():
            warehouse = share_row["warehouse"]
            share = float(share_row["warehouse_share"])
            target_stock = float(sku["daily_demand_final"]) * share * cfg.target_days
            current_stock = float(current_lookup.get((nmid, warehouse), {}).get("current_stock_full", 0))
            to_supply = max(0, ceil_int(target_stock - current_stock))

            rows.append(
                {
                    "nmId": nmid,
                    "supplierArticle": sku["supplierArticle"],
                    "subject": sku["subject"],
                    "Артикул 1С": article_1c,
                    "warehouse": warehouse,
                    "warehouse_share": share,
                    "daily_demand_final": float(sku["daily_demand_final"]),
                    "target_stock": float(target_stock),
                    "current_stock_full": float(current_stock),
                    "to_supply": float(to_supply),
                    "calc_mode": "regular",
                }
            )

    plan = pd.DataFrame(rows)
    if plan.empty:
        return plan

    plan = (
        plan.groupby(
            ["nmId", "supplierArticle", "subject", "Артикул 1С", "warehouse", "calc_mode"],
            as_index=False,
        )
        .agg(
            warehouse_share=("warehouse_share", "sum"),
            daily_demand_final=("daily_demand_final", "first"),
            target_stock=("target_stock", "sum"),
            current_stock_full=("current_stock_full", "sum"),
            to_supply=("to_supply", "sum"),
        )
    )

    plan = plan[plan["to_supply"] > 0].copy()
    return plan


def prepare_1c_stocks_map(df_1c_stocks: pd.DataFrame) -> pd.DataFrame:
    df = df_1c_stocks.copy()
    if "Артикул" not in df.columns:
        raise KeyError('В файле "Остатки 1С.xlsx" нет колонки "Артикул".')

    for col in ONE_C_STOCK_COLUMNS:
        if col not in df.columns:
            df[col] = 0

    df["Артикул"] = df["Артикул"].map(normalize_text)
    return df[["Артикул"] + ONE_C_STOCK_COLUMNS].drop_duplicates(subset=["Артикул"])


def resolve_template(storage: S3Storage, cfg: AppConfig, workdir: str) -> str:
    target = Path(workdir) / "Согласование поставки WB.xlsm"
    storage.download_file(cfg.template_key, str(target))
    return str(target)


def pick_sheet_name(store_name: str) -> str:
    name = store_name.strip().upper()
    if "TOPFACE" in name or name == "TF":
        return "ТФ ВБ"
    if "MISS" in name or "TAIS" in name or name == "MT":
        return "МТ ВБ"
    return "ТФ ВБ"


def build_template_dataset(plan_df: pd.DataFrame, stocks_1c_map: pd.DataFrame) -> pd.DataFrame:
    wide = plan_df.pivot_table(
        index=["supplierArticle", "Артикул 1С", "nmId", "subject"],
        columns="warehouse",
        values="to_supply",
        aggfunc="sum",
        fill_value=0,
    ).reset_index()

    warehouse_cols = [
        c for c in wide.columns
        if c not in ["supplierArticle", "Артикул 1С", "nmId", "subject"]
    ]

    if warehouse_cols:
        wide["Общий итог"] = wide[warehouse_cols].sum(axis=1)
    else:
        wide["Общий итог"] = 0

    stocks_1c_map = stocks_1c_map.rename(columns={"Артикул": "Артикул 1С"})
    out = wide.merge(stocks_1c_map, on="Артикул 1С", how="left")

    out["missing_1c_article"] = (
        out["Артикул 1С"].astype(str).str.strip().eq("")
        | out[ONE_C_STOCK_COLUMNS].isna().all(axis=1)
    )

    for col in ONE_C_STOCK_COLUMNS:
        out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0)

    out = out[out["Артикул 1С"].astype(str).str.strip() != ""].copy()

    if warehouse_cols:
        out["Общий итог"] = out[warehouse_cols].sum(axis=1)
        out = out[out["Общий итог"] > 0].copy()

    return out


def fill_template_file(template_path: str, output_path: str, data_df: pd.DataFrame, cfg: AppConfig) -> str:
    wb = load_workbook(template_path, keep_vba=True)
    ws = wb[pick_sheet_name(cfg.store_name)]

    header_row = 1

    def get_header_map() -> Dict[str, int]:
        return {
            normalize_template_header(ws.cell(header_row, col).value): col
            for col in range(1, ws.max_column + 1)
            if normalize_template_header(ws.cell(header_row, col).value)
        }

    header_map = get_header_map()

    article_1c_col = header_map.get("Артикул 1С") or header_map.get("Артикул 1с")
    if article_1c_col is None:
        raise KeyError(f"Не найдена колонка 'Артикул 1С'. Заголовки шаблона: {list(header_map.keys())}")

    total_col = header_map.get("Общий итог")
    if total_col is None:
        raise KeyError(f"Не найдена колонка 'Общий итог'. Заголовки шаблона: {list(header_map.keys())}")

    one_c_col_map = {
        col_name: header_map[col_name]
        for col_name in ONE_C_STOCK_COLUMNS
        if col_name in header_map
    }
    sobrat_col = header_map.get("Собрать всего")

    technical_cols = {"supplierArticle", "Артикул 1С", "nmId", "subject", "Общий итог", "missing_1c_article"}
    source_warehouse_cols = [
        c for c in data_df.columns
        if c not in technical_cols and c not in ONE_C_STOCK_COLUMNS
    ]

    preferred_order = [
        "Коледино",
        "Тула",
        "Электросталь",
        "Казань",
        "Самара (Новосемейкино)",
        "Краснодар",
        "Невинномысск",
        "Волгоград",
        "Рязань (Тюшевское)",
        "Сарапул",
        "Екатеринбург - Перспективная 14",
        "Владимир",
        "Котовск",
        "Воронеж",
        "Санкт-Петербург Уткина Заводь",
        "Новосибирск",
    ]
    ordered_warehouses = [w for w in preferred_order if w in source_warehouse_cols]
    ordered_warehouses += [w for w in source_warehouse_cols if w not in ordered_warehouses]

    # Полностью перестраиваем складскую зону:
    # удаляем все колонки между A и "Общий итог", затем вставляем нужное число складских колонок с B
    if total_col > article_1c_col + 1:
        ws.delete_cols(article_1c_col + 1, total_col - article_1c_col - 1)

    header_map = get_header_map()
    total_col = header_map.get("Общий итог")
    if total_col is None:
        raise KeyError("После перестройки не найдена колонка 'Общий итог'")

    if ordered_warehouses:
        ws.insert_cols(article_1c_col + 1, amount=len(ordered_warehouses))

    reverse_template_names = {
        "Самара (Новосемейкино)": "Новосемейкино",
        "Рязань (Тюшевское)": "Рязань",
        "Екатеринбург - Перспективная 14": "Екатеринбург",
        "Санкт-Петербург Уткина Заводь": "СПб Уткина Заводь",
    }

    for i, wh in enumerate(ordered_warehouses, start=article_1c_col + 1):
        ws.cell(header_row, i, reverse_template_names.get(wh, wh))

    # после вставки пересчитываем карту заголовков
    header_map = get_header_map()
    total_col = header_map.get("Общий итог")
    article_1c_col = header_map.get("Артикул 1С") or header_map.get("Артикул 1с")
    one_c_col_map = {
        col_name: header_map[col_name]
        for col_name in ONE_C_STOCK_COLUMNS
        if col_name in header_map
    }
    sobrat_col = header_map.get("Собрать всего")

    warehouse_write_map: Dict[str, int] = {}
    for col in range(article_1c_col + 1, total_col):
        header = normalize_template_header(ws.cell(header_row, col).value)
        if not header:
            continue
        internal_name = TEMPLATE_WAREHOUSE_ALIASES.get(header, header)
        warehouse_write_map[internal_name] = col

    template_warehouse_cols = [c for c in warehouse_write_map.keys() if c in data_df.columns]
    if template_warehouse_cols:
        data_df = data_df.copy()
        data_df["template_total"] = data_df[template_warehouse_cols].sum(axis=1)
        data_df = data_df[data_df["template_total"] > 0].copy()

    # очищаем только нужные области
    clear_cols = {article_1c_col, total_col}
    clear_cols.update(warehouse_write_map.values())
    clear_cols.update(one_c_col_map.values())
    # Собрать всего НЕ трогаем

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in clear_cols:
            ws.cell(row_idx, col_idx).value = None
            ws.cell(row_idx, col_idx).fill = PatternFill(fill_type=None)

    log(f"Строк после фильтрации под шаблон: {len(data_df)}")

    if len(data_df) > ws.max_row - 1:
        raise ValueError(
            f"В шаблоне недостаточно строк. Доступно: {ws.max_row - 1}, нужно: {len(data_df)}"
        )

    red_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")

    for i, (_, row) in enumerate(data_df.iterrows(), start=2):
        ws.cell(i, article_1c_col, row["Артикул 1С"])

        row_total = 0
        for warehouse, col_idx in warehouse_write_map.items():
            value = floor_int(row.get(warehouse, 0))
            ws.cell(i, col_idx, value)
            row_total += value

        ws.cell(i, total_col, row_total)

        for col_name, col_idx in one_c_col_map.items():
            ws.cell(i, col_idx, floor_int(row.get(col_name, 0)))

        # Собрать всего не заполняем

        if bool(row.get("missing_1c_article", False)):
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(i, col_idx).fill = red_fill

    wb.save(output_path)
    wb.close()
    return output_path


def send_telegram_document(bot_token: str, chat_id: str, file_path: str, caption: str = "") -> None:
    if not bot_token or not chat_id:
        raise ValueError("Не заданы TELEGRAM_BOT_TOKEN или TELEGRAM_CHAT_ID")

    url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
    with open(file_path, "rb") as f:
        files = {"document": (Path(file_path).name, f)}
        data = {"chat_id": chat_id, "caption": caption[:1024]}
        response = requests.post(url, data=data, files=files, timeout=300)

    if response.status_code != 200:
        raise RuntimeError(f"Ошибка отправки в Telegram: {response.status_code} {response.text}")


def send_results_to_telegram(cfg: AppConfig, files: List[str]) -> None:
    if not cfg.send_telegram:
        log("Отправка в Telegram отключена.")
        return

    for file_path in files:
        caption = f"{cfg.store_name} | {cfg.run_date:%Y-%m-%d} | {Path(file_path).name}"
        send_telegram_document(cfg.telegram_bot_token, cfg.telegram_chat_id, file_path, caption)
        log(f"Файл отправлен в Telegram: {file_path}")


def save_debug_files(
    output_dir: str,
    sku_df: pd.DataFrame,
    region_metrics: pd.DataFrame,
    shares_df: pd.DataFrame,
    plan_df: pd.DataFrame,
    filled_template_path: str,
) -> None:
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    debug_path = Path(output_dir) / "wb_supply_debug.xlsx"
    with pd.ExcelWriter(debug_path, engine="openpyxl") as writer:
        sku_df.to_excel(writer, index=False, sheet_name="SKU")
        region_metrics.to_excel(writer, index=False, sheet_name="RegionMetrics")
        shares_df.to_excel(writer, index=False, sheet_name="WarehouseShares")
        plan_df.to_excel(writer, index=False, sheet_name="Plan")

    log(f"Сохранён debug-файл: {debug_path}")
    log(f"Сохранён шаблон: {filled_template_path}")


def main(cfg: AppConfig = CONFIG) -> str:
    Path(cfg.output_dir).mkdir(parents=True, exist_ok=True)

    log(
        f"Старт расчёта. store={cfg.store_name}, season_coeff={cfg.season_coeff}, "
        f"run_date={cfg.run_date:%Y-%m-%d}, strategy={cfg.strategy_mode}"
    )

    storage = S3Storage(cfg)

    orders = load_orders(storage, cfg)
    stocks = load_stocks(storage, cfg)
    article_1c_map = load_article_map_1c(storage, cfg)
    stocks_1c_raw = load_1c_stocks(storage, cfg)
    stocks_1c_map = prepare_1c_stocks_map(stocks_1c_raw)

    stocks_1c_local = str(Path(cfg.output_dir) / "Остатки 1С.xlsx")
    storage.download_file(cfg.stocks_1c_key, stocks_1c_local)

    daily_orders = prepare_daily_orders(orders)
    grid = build_daily_grid(daily_orders, cfg)

    current_stock_df, history_dates, per_wh, per_district, latest_stock_date = prepare_stock_presence(stocks)
    if pd.notna(latest_stock_date):
        log(f"Актуальная дата остатков: {latest_stock_date:%Y-%m-%d}")

    current_stock = current_stock_by_warehouse(current_stock_df)
    log(f"Текущих остатков по складам в расчёте: {len(current_stock):,} строк")
    log(f"Сумма текущих остатков по складам: {current_stock['current_stock_full'].sum():,.0f}")

    grid = attach_presence_flags(grid, per_wh, per_district, history_dates)
    log(f"GRID SIZE: {len(grid):,}")

    log("Начинаем aggregate_region_metrics...")
    region_metrics = aggregate_region_metrics(grid, cfg)
    log(f"Region metrics рассчитаны: {len(region_metrics):,}")

    sku_df = choose_final_daily_demand(region_metrics, cfg)
    shares_df = build_warehouse_shares(region_metrics)
    shares_df = apply_strategy(shares_df, cfg)

    plan_df = calculate_supply_plan(sku_df, shares_df, current_stock, article_1c_map, cfg)
    if plan_df.empty:
        raise ValueError("После расчёта план поставки пуст.")

    template_data = build_template_dataset(plan_df, stocks_1c_map)

    with tempfile.TemporaryDirectory() as tmpdir:
        template_path = resolve_template(storage, cfg, tmpdir)
        output_path = str(Path(cfg.output_dir) / f"Согласование поставки WB_{cfg.store_name}_{cfg.run_date:%Y%m%d}.xlsm")
        fill_template_file(template_path, output_path, template_data, cfg)

    save_debug_files(cfg.output_dir, sku_df, region_metrics, shares_df, plan_df, output_path)

    if cfg.upload_result_to_s3:
        result_key = f"Отчёты/Поставки/{cfg.store_name}/{Path(output_path).name}"
        storage.upload_file(output_path, result_key)
        log(f"Результат загружен в Object Storage: {result_key}")

    files_to_send = [
        output_path,
        str(Path(cfg.output_dir) / "wb_supply_debug.xlsx"),
        stocks_1c_local,
    ]
    send_results_to_telegram(cfg, files_to_send)

    log("Готово")
    return output_path


if __name__ == "__main__":
    main()
